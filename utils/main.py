import asyncio
import io
import json
import sys
import tempfile
from urllib.parse import urlparse
from typing import Optional
from fpdf import FPDF
import matplotlib.pyplot as plt
import xlsxwriter
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton,
    Update, ForceReply,
)
from telegram.constants import ParseMode, ChatAction
from telegram.error import BadRequest
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    filters,
    ConversationHandler,
    ContextTypes,
)
from pathlib import Path

# Add the project root to Python path
project_root = Path(__file__).resolve().parent.parent
sys.path.append(str(project_root))

# Now import your Database class
from db.database import Database
from utils.validation import validate_job_post_data
from utils.validation import validate_job_post
from utils.validation import validate_job_post_data_for_job_preview


import os
# Directory where language files are stored
translations_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "translations")

# Dictionary to hold all translations
translations = {}

# Load all language files
try:
    for filename in os.listdir(translations_dir):
        if filename.endswith('.json'):
            lang_code = filename.split('.')[0]  # e.g., "english" from "english.json"
            file_path = os.path.join(translations_dir, filename)
            with open(file_path, "r", encoding="utf-8") as file:
                translations[lang_code] = json.load(file)
except FileNotFoundError:
    print(f"Error: translations directory not found at {translations_dir}")
    print("Current working directory:", os.getcwd())
    print("Directory contents:", os.listdir('.'))
    raise
except json.JSONDecodeError as e:
    print(f"Error decoding JSON from translation file: {e}")
    raise

# Initialize Database
db = Database()

# Enable logging
import logging
import os

# Ensure the user_documents directory exists
if not os.path.exists("user_documents"):
    os.makedirs("user_documents")

logging.basicConfig(level=logging.WARNING)
# Define states
(LANGUAGE, MOBILE, REGISTRATION_TYPE, JOB_SEEKER, JOB_SEEKER_DOB,JOB_SEEKER_GENDER, JOB_SEEKER_CONTACT_NUMBERS, JOB_SEEKER_LANGUAGES, JOB_SEEKER_QUALIFICATION, JOB_SEEKER_FIELD_OF_STUDY,
 JOB_SEEKER_CGPA, JOB_SEEKER_SKILLS, JOB_SEEKER_PROFILE_SUMMARY, JOB_SEEKER_SUPPORTING_DOCUMENTS, JOB_SEEKER_PORTFOLIO_LINK, MAIN_MENU, EDIT_PROFILE, EDIT_FIELD_VALUE, PROFILE_COMPLETION, EDIT_PROFILE_FIELD_VALUE,
 EMPLOYER_NAME, EMPLOYER_LOCATION,EMPLOYER_TYPE, ABOUT_COMPANY, EMPLOYER_DOCUMENTS,EMPLOYER_MAIN_MENU, EDIT_EMPLOYER_PROFILE, EDIT_EMPLOYER_FIELD_VALUE, VERIFICATION_DOCUMENTS,
 ADMIN_LOGIN, ADMIN_MAIN_MENU,BROADCAST_TYPE, BROADCAST_MESSAGE, CONFIRM_BROADCAST,
 POST_JOB_TITLE, POST_EMPLOYMENT_TYPE, POST_GENDER, POST_QUANTITY, POST_LEVEL, POST_DESCRIPTION, POST_QUALIFICATION, POST_SKILLS, POST_SALARY, POST_BENEFITS,
 POST_DEADLINE, JOB_PREVIEW, CONFIRM_POST,REJECT_JOB_REASON , CONFIRM_SUBMISSION, VIEW_APPLICATIONS, SELECT_JOB_POSTS_TO_SHARE,SELECT_JOB_TO_MANAGE,HANDLE_JOB_ACTIONS,
 CONFIRM_CLOSE, RESUBMIT_CONFIRMATION, DISPLAY_VACANCIES,SELECT_VACANCY,CONFIRM_SELECTION, WRITE_COVER_LETTER,
 MANAGE_USERS,REMOVE_JOB_SEEKERS, REMOVE_EMPLOYERS, REMOVE_APPLICATIONS, CLEAR_ALL_DATA, CONFIRM_REMOVAL, CLEAR_CONFIRMATION,
 DATABASE_MANAGEMENT, MANAGE_JOBS, AD_MANAGE_VACANCIES, LIST_JOBS, SEARCH_JOB_SEEKERS, REMOVE_JOB_SEEKERS_PAGINATED,
 SEARCH_EMPLOYERS, REMOVE_EMPLOYERS_PAGINATED, SEARCH_APPLICATIONS, REMOVE_APPLICATIONS_PAGINATED, EXPORT_DATA, SEARCH_JOBS, REMOVE_JOBS_PAGINATED,
 LIST_VACANCIES_PAGINATED, REMOVE_VACANCIES_PAGINATED, SEARCH_VACANCIES,  MANAGE_APPLICATIONS, LIST_APPLICATIONS_PAGINATED, SAVE_EDITED_FIELD,
 PROFILE_MENU, CONFIRM_DELETE_ACCOUNT, CONFIRM_CHANGE_LANGUAGE, SELECT_LANGUAGE, VIEW_PROFILE, EMPLOYER_PROFILE_MENU, CONFIRM_DELETE_MY_ACCOUNT, SELECT_EMPLOYER_LANGUAGE, CONFIRM_CHANGE_EMPLOYER_LANGUAGE,
 ACCEPT_REJECT_CONFIRMATION, REJECTION_REASON_INPUT, EMPLOYER_MESSAGE_INPUT,BAN_JOB_SEEKERS, BAN_EMPLOYERS, UNBAN_USERS, VIEW_BANNED_USERS,
 REASON_FOR_BAN, APPEAL_SUBMIT, REVIEW_APPEALS, HANDLE_APPEAL, BAN_JOB_SEEKERS_PAGINATED, SUBMIT_APPEAL, BAN_EMPLOYERS_PAGINATED,
 SEARCH_JOB_SEEKERS_FOR_BAN, REASON_FOR_BAN_JOB_SEEKER,  SEARCH_EMPLOYERS_FOR_BAN, REASON_FOR_BAN_EMPLOYER, UNBAN_USERS_MENU, UNBAN_SELECTION, UNBAN_ALL_CONFIRMATION, APPEAL_START, APPEAL_INPUT, APPEAL_REVIEW, BANNED_STATE,
 SEARCH_OPTIONS, ADVANCED_FILTERS, FILTER_JOB_TYPE, SEARCH_RESULTS, KEYWORD_SEARCH , FILTER_SALARY, FILTER_EXPERIENCE , VIEW_JOB_DETAILS, VIEWING_APPLICATIONS, APPLICATION_DETAILS, EXPORTING_APPLICATIONS, RENEW_VACANCY,  CONFIRM_RENEWAL,
 ANALYTICS_VIEW , ANALYTICS_TRENDS , ANALYTICS_DEMOGRAPHICS , ANALYTICS_RESPONSE , ANALYTICS_BENCHMARK , ANALYTICS_EXPORT,
 HELP_MENU, FAQ_SECTION, FAQ_CATEGORY, FAQ_QUESTION, JS_FAQ_SECTION,  JS_FAQ_CATEGORY, JS_FAQ_QUESTION, ADMIN_FAQ_SECTION, ADMIN_FAQ_CATEGORY, ADMIN_FAQ_QUESTION,
 RATE_OPTIONS, SEARCH_USER_FOR_RATING , RATE_DIMENSION, CONFIRM_REVIEW, SELECT_USER_FOR_RATING, ADD_COMMENT_OPTIONAL, REVIEW_SETTINGS,  MY_REVIEWS, REVIEW_DETAILS, SEARCH_REVIEWS, SUBMIT_COMMENT, PROMPT_FOR_COMMENT, POST_REVIEW,
 CONTACT_CATEGORY, CONTACT_PRIORITY, CONTACT_MESSAGE , ADMIN_REPLY_STATE, CONTACT_MANAGEMENT,  CONTACT_INBOX,  CONTACT_VIEW_MESSAGE, CONTACT_OUTBOX, CONTACT_PENDING, CONTACT_ANSWERED, CONTACT_STATS, CONTACT_CONFIRM_DELETE, VIEW_ERRORS,  ERROR_DETAIL,
 DB_STATS_VIEW, SYSTEM_CONFIGURATIONS_MENU, DATABASE_STORAGE_OVERVIEW, TABLE_SIZE_ANALYSIS, OPTIMIZE_DATABASE, VACUUM_DATABASE, QUERY_PERFORMANCE_INSIGHTS, ERROR_LOGS, SHARE_JOBS_NAVIGATION,CONFIRM_REMOVE_VACANCY, JOB_DETAIL_VIEW, CONFIRM_JOB_REMOVAL,
 LIST_JOBS_PAGINATED, APPLICATION_DETAIL_VIEW, CONFIRM_APPLICATION_REMOVAL, TABLE_CLEANUP, CONFIRM_TABLE_DELETION, PROFILE_COMPLETION_VIEW,
 REPORT_MAIN_MENU, SELECT_REPORT_ENTITY, SEARCH_REPORT_ENTITY, SELECT_REPORT_RESULT, CONFIRM_REPORT_DETAILS, SUBMIT_REPORT, USER_INTERACTIONS_MENU,
 VIOLATION_REPORTS_DASHBOARD,  VIEW_ALL_REPORTS, VIEW_REPORTED_USERS, VIEW_REPORTED_VACANCIES, VIEW_REPORTED_APPLICATIONS, VIEW_REPORTED_EMPLOYERS, VACANCY_DISPLAY_OPTION, FILTER_SELECTION,  EMPLOYMENT_FILTER, LEVEL_FILTER, QUALIFICATION_FILTER, GENDER_FILTER, SELECT_SEARCH_RESULT,EMPLOYER_MANAGE_VACANCIES,
 ADMIN_RATINGS_MENU, ADMIN_REVIEW_LIST, ADMIN_REVIEW_SEARCH, ADMIN_REVIEW_STATISTICS, ADMIN_FLAGGED_REVIEWS, ADMIN_DELETE_REVIEW, ADMIN_REVIEW_DETAILS) = range(221)



# Helper function to fetch translations
def get_translation(user_id, key, **kwargs):
    # Retrieve the user's language preference from the database
    user_language = db.get_user_language(user_id)  # Assume this returns 'english', 'amharic', etc.

    # Fallback to English if the user's language isn't available
    if user_language not in translations:
        user_language = "english"

    # Get the translations for the selected language
    language_translations = translations.get(user_language, {})

    # Get the translation, or return a detailed missing translation message
    translation = language_translations.get(key)
    if translation is None:
        # Try English as final fallback
        english_translation = translations.get("english", {}).get(key)
        if english_translation is None:
            return f"Translation not found for '{key}'"
        return english_translation.format(**kwargs) if kwargs else english_translation

    # If kwargs are provided, format the translation string using them
    return translation.format(**kwargs) if kwargs else translation


def is_profile_complete(user_profile, db):
    """Check if the user's profile is complete based on their registration type."""
    required_fields = ["registration_type", "full_name"]

    if user_profile.get("registration_type") == "employer":
        employer_profile = db.get_employer_profile(user_profile["user_id"])
        if not employer_profile:  # If no employer profile exists, it's incomplete
            return False
        # Ensure the employer profile has all required fields, including employer_id
        required_employer_fields = ["company_name", "city", "employer_type", "employer_id"]
        return all(employer_profile.get(field) for field in required_employer_fields)

    return all(user_profile.get(field) for field in required_fields)



async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle the /start command with modular structure."""
    user_id = get_user_id(update)

    # ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ CHECK BANS ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
    ban_status = await _check_user_bans(user_id, update, context)
    if ban_status == BANNED_STATE:
        return BANNED_STATE

    # ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ INITIAL SETUP ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
    context.user_data.clear()
    await _handle_start_parameter(update, context)

    # ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ USER PROFILE CHECK ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
    user_profile = db.get_user_profile(user_id)
    if not user_profile:
        await show_language_selection(update, context)
        return LANGUAGE

    # ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ PROFILE COMPLETENESS CHECK ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
    if not is_profile_complete(user_profile, db):
        await _handle_incomplete_profile(user_id, update, context)
        return LANGUAGE

    # ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ USER TYPE HANDLING ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
    try:
        registration_type = user_profile.get("registration_type")

        if registration_type == "employer":
            return await _handle_employer_start(user_id, update, context, user_profile)
        else:
            return await _handle_job_seeker_start(user_id, update, context, user_profile)

    except Exception as e:
        logging.error(f"Error in start function: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An unexpected error occurred. Please try again.",
            parse_mode="HTML"
        )
        return ConversationHandler.END


# ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ HELPER FUNCTIONS ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬

async def _check_user_bans(user_id: int, update: Update, context: ContextTypes.DEFAULT_TYPE) -> Optional[int]:
    """Check if user is banned and handle ban message."""
    is_banned = False
    ban_reason = ""

    # Check job seeker ban
    if db.is_user_banned(user_id=user_id, employer_id=None):
        is_banned = True
        ban_reason = db.get_ban_reason(user_id=user_id, employer_id=None)

    # Check employer ban
    employer_profile = db.get_employer_profile(user_id)
    if employer_profile and db.is_user_banned(user_id=None, employer_id=employer_profile.get("employer_id")):
        is_banned = True
        ban_reason = db.get_ban_reason(user_id=None, employer_id=employer_profile.get("employer_id"))

    if is_banned:
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"🚫 {get_translation(user_id, 'ban_title')}\n\n"
                f"📝 {get_translation(user_id, 'ban_reason')}: {ban_reason}\n\n"
                f"ℹ️ {get_translation(user_id, 'ban_appeal_info')}"
            ),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    f"✍️ {get_translation(user_id, 'appeal_ban_button')}",
                    callback_data="appeal_start"
                )]
            ]),
            parse_mode="HTML"
        )
        return BANNED_STATE
    return None


async def _handle_start_parameter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle the start parameter if user came via a deep link."""
    start_param = update.effective_message.text.split(' ', 1)[-1] if update.effective_message.text else None
    if start_param and start_param.startswith("apply_"):
        job_id = start_param.split("_")[-1]
        context.user_data["job_id"] = job_id


async def _handle_incomplete_profile(user_id: int, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle incomplete user profile."""
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "profile_incomplete"),
        parse_mode="HTML"
    )
    await show_language_selection(update, context)


async def _handle_employer_start(user_id: int, update: Update, context: ContextTypes.DEFAULT_TYPE,
                               user_profile: dict) -> int:
    """Handle start flow for employers."""
    employer_profile = db.get_employer_profile(user_id)
    if not employer_profile:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "employer_profile_not_found"),
            parse_mode="HTML"
        )
        return LANGUAGE

    company_name = employer_profile.get("company_name") or user_profile.get("full_name") or get_translation(user_id, "default_employer_name")
    employer_id = employer_profile.get("employer_id")
    if not employer_id:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "employer_id_missing"),
            parse_mode="HTML"
        )
        return ConversationHandler.END

    context.user_data["employer_id"] = employer_id

    if "job_id" in context.user_data:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "employer_cannot_apply"),
            parse_mode="HTML"
        )

    await employer_main_menu(update, context)
    return EMPLOYER_MAIN_MENU

async def _handle_job_seeker_start(user_id: int, update: Update, context: ContextTypes.DEFAULT_TYPE,
                                   user_profile: dict) -> int:
    """Handle start flow for job seekers."""
    if "job_id" in context.user_data:
        return await _handle_job_application(user_id, update, context)

    await main_menu(update, context)
    return MAIN_MENU


async def _handle_job_application(user_id: int, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle job application flow when user starts with apply link."""
    job_id = context.user_data["job_id"]
    selected_job = db.get_job_by_id(job_id)

    if not selected_job:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "job_not_exist"),
            parse_mode="HTML"
        )
        return MAIN_MENU

    job_details = {
        "job_id": selected_job.get("job_id", ""),
        "employer_id": selected_job.get("employer_id", ""),
        "job_title": selected_job.get("job_title", get_translation(user_id, "title_not_available")),
        "company_name": selected_job.get("company_name", get_translation(user_id, "company_not_provided")),
        "employment_type": get_translation(user_id, selected_job.get("employment_type", "not_specified")),
        "deadline": selected_job.get("application_deadline", get_translation(user_id, "no_deadline")),
        "gender": get_translation(user_id, selected_job.get("gender", "any")),
        "quantity": selected_job.get("quantity", "N/A"),
        "level": get_translation(user_id, selected_job.get("level", "not_specified")),
        "description": selected_job.get("description", get_translation(user_id, "no_description")),
        "qualification": get_translation(user_id, selected_job.get("qualification", "not_specified")),
        "skills": selected_job.get("skills", get_translation(user_id, "no_skills")),
        "salary": selected_job.get("salary", get_translation(user_id, "negotiable")),
        "benefits": selected_job.get("benefits", get_translation(user_id, "no_benefits"))
    }

    context.user_data["selected_job"] = job_details
    await _send_job_details(user_id, context, job_details)

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "write_cover_letter_prompt"),
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="HTML"
    )
    return WRITE_COVER_LETTER


async def _send_job_details(user_id: int, context: ContextTypes.DEFAULT_TYPE, job_details: dict) -> None:
    """Send formatted job details to user."""
    job_card = [
        f"<b>{get_translation(user_id, 'applying_for_V')}</b>\n\n"
        f"🌟 <b>{get_translation(user_id, 'job_details_title')}</b> 🌟",
        "",
        f"📌 <b>{job_details['job_title']}</b>",
        "",
        f"⏳ <b>{get_translation(user_id, 'deadline')}:</b> {job_details['deadline']}",
        f"💼 <b>{get_translation(user_id, 'type')}:</b> {job_details['employment_type']}",
        f"👥 <b>{get_translation(user_id, 'positions')}:</b> {job_details['quantity']}",
        f"📊 <b>{get_translation(user_id, 'level')}:</b> {job_details['level']}",
        f"🚻 <b>{get_translation(user_id, 'gender')}:</b> {job_details['gender']}",
        "",
        f"💰 <b>{get_translation(user_id, 'salary')}:</b> {job_details['salary']}",
        "",
        f"📝 <b>{get_translation(user_id, 'description')}:</b>",
        f"{job_details['description']}",
        "",
        f"🎓 <b>{get_translation(user_id, 'qualification')}:</b>",
        f"{job_details['qualification']}",
        "",
        f"🛠️ <b>{get_translation(user_id, 'skills')}:</b>",
        f"{job_details['skills']}",
        "",
        f"🎁 <b>{get_translation(user_id, 'benefits')}:</b>",
        f"{job_details['benefits']}",
        "",
        "▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬"
    ]

    await context.bot.send_message(
        chat_id=user_id,
        text="\n".join(job_card),
        parse_mode="HTML"
    )

def escape_html(text):
    """Escape special characters for Telegram HTML formatting."""
    escape_chars = {'&': '&amp;', '<': '<', '>': '>'}
    return ''.join(escape_chars.get(char, char) for char in text)

# Total steps in registration process (for progress tracking)
TOTAL_REGISTRATION_STEPS = 12
def get_progress_bar(user_id: int,
                    context: ContextTypes.DEFAULT_TYPE,
                    width: int = 20,
                    show_percentage: bool = True) -> str:

    # user_id = get_user_id(update)
    current_progress = context.user_data.get('progress', 1)
    percentage = min(100, int((current_progress / TOTAL_REGISTRATION_STEPS) * 100))  # Ensure we don't exceed 100%
    language = context.user_data.get('language', 'english')

    # Create a clean line-based progress bar
    bar_length = 25  # Reduced length for better visibility
    filled_length = int(bar_length * percentage / 100)

    # Using full block characters with better spacing
    progress_bar = (
        f"[{''.join('█' for _ in range(filled_length))}"
        f"{''.join('░' for _ in range(bar_length - filled_length))}]"
    )
    # Format the progress message
    progress_message = (
        f"\n📊 *{get_translation(user_id, 'registration_progress')}:* "
        f"{percentage}%\n"
        f"`[{progress_bar}]`\n"
        f"*{get_translation(user_id, 'step')} {current_progress} "
        f"{get_translation(user_id, 'of')} {TOTAL_REGISTRATION_STEPS}*"
    )

    return progress_message

async def show_language_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show language selection with enhanced layout and preview."""
    user_id = get_user_id(update)


    # Language selection keyboard with 2 buttons per row
    keyboard = [
        [
            InlineKeyboardButton("🇬🇧 English", callback_data="english"),
            InlineKeyboardButton("🇪🇹 አማርኛ", callback_data="amharic")
        ],
        [
            InlineKeyboardButton(" Afaan Oromoo", callback_data="oromia"),
            InlineKeyboardButton(" ትግርኛ", callback_data="tigrigna")
        ],
        [
            InlineKeyboardButton(" Qafar af", callback_data="afar"),
            InlineKeyboardButton(" Soomaali", callback_data="somalia")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send welcome message with progress indicator
    welcome_message = (
        f"🌍 *{get_translation('en', 'welcome_message')}*\n\n"
        f"Please select your preferred language:\n\n"
    )

    await update.message.reply_text(
        welcome_message,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    return LANGUAGE

async def handle_language_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    selected_language = query.data

    # Update the user's language in the database
    db.update_user_language(user_id, selected_language)

    return await main_menu(update, context)

async def set_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    language = query.data
    user_id = get_user_id(update)  # Use get_user_id for consistency
    # Default to English if invalid choice
    language = language if language in ["amharic", "oromia", "english", "tigrigna", "somalia", "afar"] else "english"
    # Update progress

    context.user_data['language'] = language
    # Save language to database
    db.insert_user(user_id, language)
    # Fetch translation for the next prompt
    # Prepare contact sharing message with progress
    text = (
        f"✅ {get_translation(user_id, 'language_set_success')}\n\n"
        f"{get_translation(user_id, 'share_mobile')}\n\n"
    )
    # Contact sharing keyboard
    keyboard = [[
        KeyboardButton(
            f"📱 {get_translation(user_id, 'share_contact_button')}",
            request_contact=True
        )
    ]]

    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        one_time_keyboard=True,
        resize_keyboard=True
    )

    # Send new message with contact prompt
    await context.bot.send_message(
        chat_id=user_id,
        text=text,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )

    return MOBILE


async def save_mobile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.contact:
        contact = update.message.contact
        user_id = get_user_id(update)
        language = context.user_data.get('language', 'english')

        # Validate and format phone number
        phone_number = contact.phone_number
        if not phone_number.startswith('+'):
            phone_number = f"+{phone_number}"

        # Save to database
        db.update_user_profile(user_id, contact_number=phone_number)

        # Prepare registration type prompt with progress
        text = (
            f"✅ {get_translation(user_id, 'contact_saved_success')}\n\n"
            f"{get_translation(user_id, 'register_prompt')}\n\n"
        )

        # Registration type keyboard
        job_seeker = get_translation(user_id, "job_seeker")
        employer = get_translation(user_id, "employer")

        keyboard = [
            [InlineKeyboardButton(f"👨‍💼 {job_seeker}", callback_data="job_seeker")],
            [InlineKeyboardButton(f"🏢 {employer}", callback_data="employer")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        # Send registration prompt
        await update.message.reply_text(
            text,
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

        return REGISTRATION_TYPE
    else:
        # Handle case where contact isn't shared
        user_id = get_user_id(update)
        language = context.user_data.get('language', 'english')

        text = (
            f"⚠️ {get_translation(user_id, 'contact_required')}\n\n"
            f"{get_translation(user_id, 'share_mobile')}\n\n"
        )

        keyboard = [[
            KeyboardButton(
                f"📱 {get_translation(user_id, 'share_contact_button')}",
                request_contact=True
            )
        ]]

        reply_markup = ReplyKeyboardMarkup(
            keyboard,
            one_time_keyboard=True,
            resize_keyboard=True
        )

        await update.message.reply_text(
            text,
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

        return MOBILE


async def registration_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    user_id = get_user_id(update)
    language = context.user_data.get('language', 'english')
    choice = query.data

    if choice == "job_seeker":
        # Update progress
        context.user_data['progress'] = 4

        # Save registration type
        db.update_user_profile(user_id, registration_type="job_seeker")
        db.record_user_creation(user_id, 'job_seeker')
        context.user_data["registration"] = "job_seeker"

        # Send welcome message with progress
        welcome_message = (
            f"👨‍💼 *{get_translation(user_id, 'job_seeker_welcome')}*\n\n"
            f"{get_translation(user_id, 'job_seeker_start')}\n\n"
        )

        await query.edit_message_text(
            welcome_message,
            parse_mode="Markdown"
        )

        # Start job seeker flow
        await update_job_seeker_flow(user_id, context)
        return JOB_SEEKER

    elif choice == "employer":



        # Save registration type
        db.update_user_profile(user_id, registration_type="employer")
        db.record_user_creation(user_id, 'employer')
        context.user_data["registration"] = "employer"

        # Send welcome message with progress
        welcome_message = (
            f"🏢 *{get_translation(user_id, 'employer_welcome')}*\n\n"
            f"{get_translation(user_id, 'employer_registration_start')}\n\n"

        )

        await context.bot.send_message(
            chat_id=user_id,
            text=welcome_message,
            parse_mode="Markdown"
        )

        return EMPLOYER_NAME
    else:
        # Handle invalid choice
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ {get_translation(user_id, 'invalid_choice')}",
            parse_mode="Markdown"
        )
        return REGISTRATION_TYPE

# Helper function to fetch the user's registration type from the database
def get_user_registration_type(user_id):
    try:
        user_profile = db.get_user_profile(user_id)
        if user_profile:
            return user_profile[12]  # Assuming registration_type is the 13th column in the user profile
        return None
    except Exception as e:
        print(f"Error fetching registration type: {e}")
        return None

async def update_job_seeker_flow(user_id: int, context: ContextTypes.DEFAULT_TYPE):
    """Initiate Job Seeker registration workflow."""
    await context.bot.send_message(
        chat_id=user_id, text=get_translation(user_id, "job_seeker_full_name_prompt")
    )

async def job_seeker_full_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    full_name = update.message.text.strip()
    user_id = update.message.from_user.id
    context.user_data['progress'] = 1

    # Validation checks with detailed error messages
    errors = []

    # 1. Empty or too short check
    if not full_name or len(full_name) < 3:
        errors.append(get_translation(user_id, 'invalid_full_name'))

    # 2. Minimum two parts check
    name_parts = [part for part in full_name.split() if part]  # Handle multiple spaces
    if len(name_parts) < 2:
        errors.append(get_translation(user_id, 'full_name_two_parts'))

    # 3. Check for repeating characters (like "mmmm mmmm")
    if any(all(c == part[0] for c in part) for part in name_parts):
        errors.append(get_translation(user_id, 'name_no_repeating_chars'))

    # 4. Check for numeric characters
    if any(any(c.isdigit() for c in part) for part in name_parts):
        errors.append(get_translation(user_id, 'name_no_numbers'))

    # 5. Check for special characters (allow hyphens and apostrophes)
    if any(not all(c.isalpha() or c in ("'", "-") for c in part) for part in name_parts):
        errors.append(get_translation(user_id, 'name_invalid_chars'))

    # 7. Check for reasonable length of each part (2-25 characters)
    if any(len(part) < 2 or len(part) > 25 for part in name_parts):
        errors.append(get_translation(user_id, 'name_part_length'))

    # 8. Check for mixed scripts (like Cyrillic + Latin)
    scripts = set()
    for part in name_parts:
        for char in part:
            if char.isalpha():  # Only check alphabetic characters
                script = 'Cyrillic' if '\u0400' <= char <= '\u04FF' else 'Latin'
                scripts.add(script)
    if len(scripts) > 1:
        errors.append(get_translation(user_id, 'name_mixed_scripts'))

    # 9. Enhanced pattern detection (within parts and across parts)

    def has_repeating_pattern(s, min_repeats=3, max_segment_length=4):
        # Check for patterns within a single string
        for l in range(2, max_segment_length + 1):
            if len(s) < l * min_repeats:
                continue
            for i in range(len(s) - l * min_repeats + 1):
                segment = s[i:i + l]
                if s.count(segment) >= min_repeats:
                    if segment * min_repeats in s:
                        return True
        return False

        # Check each part individually

    if any(has_repeating_pattern(part) for part in name_parts):
        errors.append(get_translation(user_id, 'name_repeated_pattern'))

        # Check combined name for cross-part patterns (like afsafs gfsgfs)
    combined = ''.join(name_parts).lower()
    if has_repeating_pattern(combined):
        errors.append(get_translation(user_id, 'name_repeated_pattern'))


    # If any errors found, return them
    if errors:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'name_validation_errors')}:\n\n"
            f"• {', '.join(errors)}\n\n"
            f"ℹ️ {get_translation(user_id, 'full_name_instructions')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(
            error_msg,
            parse_mode="Markdown"
        )
        return JOB_SEEKER

    # Save to database
    db.update_user_profile(user_id, full_name=full_name)

    # Prepare DOB prompt with progress
    prompt = (
        f"✅ {get_translation(user_id, 'full_name_saved')}\n\n"
        f"📅 *{get_translation(user_id, 'job_seeker_dob_prompt')}*\n"
        f"_ {get_translation(user_id, 'dob_format_instructions')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await update.message.reply_text(
        prompt,
        parse_mode="Markdown"
    )
    return JOB_SEEKER_DOB


async def job_seeker_dob(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    dob = update.message.text.strip()
    user_id = update.message.from_user.id
    # Update progress
    context.user_data['progress'] = 2

    if not re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$", dob):
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_dob_format')}\n\n"
            f"ℹ️ {get_translation(user_id, 'dob_format_instructions')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(
            error_msg,
            parse_mode="Markdown"
        )
        return JOB_SEEKER_DOB

    try:
        dob_normalized = dob.replace("/", "-").replace(".", "-")
        birth_date = datetime.strptime(dob_normalized, "%Y-%m-%d")

        # Validate age is reasonable (between 16 and 100)
        age = (datetime.now() - birth_date).days / 365
        if age < 16 or age > 100:
            error_msg = (
                f"⚠️ {get_translation(user_id, 'invalid_dob_age')}\n\n"
                f"ℹ️ {get_translation(user_id, 'dob_format_instructions')}\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(
                error_msg,
                parse_mode="Markdown"
            )
            return JOB_SEEKER_DOB

    except ValueError:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_dob_format')}\n\n"
            f"ℹ️ {get_translation(user_id, 'dob_format_instructions')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(
            error_msg,
            parse_mode="Markdown"
        )
        return JOB_SEEKER_DOB

        # Save valid DOB
    db.update_user_profile(user_id, dob=dob_normalized)

    # Prepare gender prompt with progress
    prompt = (
        f"✅ {get_translation(user_id, 'dob_saved_success')}\n\n"
        f"🚻 *{get_translation(user_id, 'job_seeker_gender_prompt')}*\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    keyboard = [
        [
            InlineKeyboardButton(f"♂️ {get_translation(user_id, 'male')}", callback_data="male"),
            InlineKeyboardButton(f"♀️ {get_translation(user_id, 'female')}", callback_data="female")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        prompt,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    return JOB_SEEKER_GENDER


from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes


async def job_seeker_gender(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    gender = query.data
    context.user_data['progress'] = 3

    # Save gender
    db.update_user_profile(user_id, gender=gender)

    # Prepare contact prompt with progress
    prompt = (
        f"✅ {get_translation(user_id, 'gender_saved')}\n\n"
        f"📱 *{get_translation(user_id, 'job_seeker_contact_numbers_prompt')}*\n"
        f"_{get_translation(user_id, 'contact_numbers_instructions')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await query.edit_message_text(
        text=prompt,
        parse_mode="Markdown"
    )
    return JOB_SEEKER_CONTACT_NUMBERS


async def job_seeker_contact_numbers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    contact_numbers = update.message.text.strip()
    context.user_data['progress'] = 4

    # Validation
    try:
        numbers_list = [num.strip() for num in contact_numbers.split(",") if num.strip()]

        if not numbers_list:
            raise ValueError("empty_input")

        if len(numbers_list) > 3:
            error_msg = (
                f"⚠️ {get_translation(user_id, 'contact_numbers_limit_exceeded')}\n"
                f"_{get_translation(user_id, 'contact_numbers_instructions')}_\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(error_msg, parse_mode="Markdown")
            return JOB_SEEKER_CONTACT_NUMBERS

        phone_pattern = re.compile(r"^\+?[\d\s\-]{10,15}$")
        invalid_numbers = [num for num in numbers_list if not phone_pattern.match(num)]

        if invalid_numbers:
            error_msg = (
                f"⚠️ {get_translation(user_id, 'contact_numbers_invalid_format')}\n"
                f"• {', '.join(invalid_numbers)}\n\n"
                f"_{get_translation(user_id, 'contact_numbers_instructions')}_\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(error_msg, parse_mode="Markdown")
            return JOB_SEEKER_CONTACT_NUMBERS

        # Save valid numbers
        db.update_user_profile(user_id, contact_number=", ".join(numbers_list))

        # Prepare languages prompt
        prompt = (
            f"✅ {get_translation(user_id, 'contact_numbers_saved')}\n\n"
            f"🗣 *{get_translation(user_id, 'job_seeker_languages_prompt')}*\n"
            f"_{get_translation(user_id, 'languages_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )

        await update.message.reply_text(prompt, parse_mode="Markdown")
        return JOB_SEEKER_LANGUAGES

    except ValueError:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'contact_numbers_empty')}\n"
            f"_{get_translation(user_id, 'contact_numbers_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(error_msg, parse_mode="Markdown")
        return JOB_SEEKER_CONTACT_NUMBERS


async def job_seeker_languages(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Validate and save languages, then proceed to qualifications."""
    user_id = get_user_id(update)
    languages = update.message.text.strip()
    context.user_data['progress'] = 5

    # Validation
    if not languages or len(languages) < 2:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_languages')}\n"
            f"_{get_translation(user_id, 'languages_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(error_msg, parse_mode="Markdown")
        return JOB_SEEKER_LANGUAGES

    # Save languages
    db.update_user_profile(user_id, languages=languages)

    # Prepare qualification prompt with keyboard
    keyboard = [
        [
            InlineKeyboardButton("📜 " + get_translation(user_id, "certificate"), callback_data="certificate"),
            InlineKeyboardButton("📃 " + get_translation(user_id, "diploma"), callback_data="diploma")
        ],
        [
            InlineKeyboardButton("🎓 " + get_translation(user_id, "degree"), callback_data="degree"),
            InlineKeyboardButton("🎓 " + get_translation(user_id, "ma"), callback_data="ma")
        ],
        [
            InlineKeyboardButton("👨‍🔬 " + get_translation(user_id, "phd"), callback_data="phd"),
            InlineKeyboardButton("❓ " + get_translation(user_id, "other"), callback_data="other")
        ],
        [
            InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")
        ]
    ]

    prompt = (
        f"✅ {get_translation(user_id, 'languages_saved')}\n\n"
        f"📚 *{get_translation(user_id, 'job_seeker_qualification_prompt')}*\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await update.message.reply_text(
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_QUALIFICATION


async def job_seeker_qualification(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle qualification selection and proceed to field of study."""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    qualification = query.data
    context.user_data['progress'] = 6  # Update progress

    if qualification == "skip":
        message = f"⏭ {get_translation(user_id, 'job_seeker_qualification_skipped')}"
    else:
        db.update_user_profile(user_id, qualification=qualification)
        message = f"✅ {get_translation(user_id, 'job_seeker_qualification_selected')}"

    # Prepare field of study prompt
    prompt = (
        f"{message}\n\n"
        f"📖 *{get_translation(user_id, 'job_seeker_field_of_study_prompt')}*\n"
        f"_{get_translation(user_id, 'field_of_study_instructions')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    keyboard = [
        [InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")]
    ]

    await query.edit_message_text(
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_FIELD_OF_STUDY


async def job_seeker_field_of_study(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle field of study input and proceed to CGPA prompt."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 7

    if update.callback_query and update.callback_query.data == "skip":
        query = update.callback_query
        await query.answer()

        prompt = (
            f"⏭ {get_translation(user_id, 'job_seeker_field_of_study_skipped')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await query.edit_message_text(text=prompt, parse_mode="Markdown")
        return await job_seeker_cgpa_prompt(update, context)

    # Handle text input
    field_of_study = update.message.text.strip()

    if not field_of_study or len(field_of_study) < 3:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_field_of_study')}\n"
            f"_{get_translation(user_id, 'field_of_study_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(error_msg, parse_mode="Markdown")
        return JOB_SEEKER_FIELD_OF_STUDY

    # Save valid input
    db.update_user_profile(user_id, field_of_study=field_of_study)

    success_msg = (
        f"✅ {get_translation(user_id, 'field_of_study_saved')}\n\n"
        f"{get_progress_bar(user_id, context)}"
    )
    await update.message.reply_text(success_msg, parse_mode="Markdown")
    return await job_seeker_cgpa_prompt(update, context)


async def job_seeker_cgpa_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user for CGPA input with skip option."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 8

    keyboard = [
        [InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")]
    ]

    prompt = (
        f"📊 *{get_translation(user_id, 'job_seeker_cgpa_prompt')}*\n"
        f"_{get_translation(user_id, 'cgpa_instructions')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_CGPA


async def job_seeker_cgpa(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle CGPA input and proceed to skills prompt."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 8

    if update.callback_query and update.callback_query.data == "skip":
        query = update.callback_query
        await query.answer()

        prompt = (
            f"⏭ {get_translation(user_id, 'job_seeker_cgpa_skipped')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await query.edit_message_text(text=prompt, parse_mode="Markdown")
        return await job_seeker_skills_prompt(update, context)

    # Handle text input
    cgpa = update.message.text.strip()

    try:
        cgpa_value = float(cgpa)
        if not (0 <= cgpa_value <= 4):
            raise ValueError

        # Save valid CGPA
        db.update_user_profile(user_id, cgpa=cgpa_value)
        context.user_data['progress'] = 9  # Update progress after validation

        success_msg = (
            f"✅ {get_translation(user_id, 'cgpa_saved')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(success_msg, parse_mode="Markdown")
        return await job_seeker_skills_prompt(update, context)

    except ValueError:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_cgpa')}\n"
            f"_{get_translation(user_id, 'cgpa_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(error_msg, parse_mode="Markdown")
        return JOB_SEEKER_CGPA

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

# Helper function to get user_id safely
def get_user_id(update: Update) -> int:
    if update.callback_query:
        return update.callback_query.from_user.id
    return update.message.from_user.id

# Job Seeker Skills Prompt
async def job_seeker_skills_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user for skills input with skip option."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 9

    keyboard = [
        [InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")]
    ]

    prompt = (
        f"🛠 *{get_translation(user_id, 'job_seeker_skills_prompt')}*\n"
        f"_{get_translation(user_id, 'skills_instructions')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_SKILLS


async def job_seeker_skills(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle skills input and proceed to profile summary."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 9  # Maintain progress until valid input

    if update.callback_query and update.callback_query.data == "skip":
        query = update.callback_query
        await query.answer()

        prompt = (
            f"⏭ {get_translation(user_id, 'job_seeker_skills_skipped')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await query.edit_message_text(text=prompt, parse_mode="Markdown")
        return await job_seeker_profile_summary_prompt(update, context)

    # Handle text input
    skills = update.message.text.strip()

    if not skills or len(skills) < 3:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_skills')}\n"
            f"_{get_translation(user_id, 'skills_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(error_msg, parse_mode="Markdown")
        return JOB_SEEKER_SKILLS

    # Save valid skills
    db.update_user_profile(user_id, skills_experience=skills)
    context.user_data['progress'] = 10  # Update progress

    success_msg = (
        f"✅ {get_translation(user_id, 'skills_saved')}\n\n"
        f"{get_progress_bar(user_id, context)}"
    )
    await update.message.reply_text(success_msg, parse_mode="Markdown")
    return await job_seeker_profile_summary_prompt(update, context)

# Job Seeker Profile Summary Prompt
async def job_seeker_profile_summary_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user for profile summary with skip option."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 10

    keyboard = [
        [InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")]
    ]

    prompt = (
        f"📝 *{get_translation(user_id, 'job_seeker_profile_summary_prompt')}*\n"
        f"_{get_translation(user_id, 'profile_summary_instructions')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_PROFILE_SUMMARY


async def job_seeker_profile_summary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle profile summary input and proceed to supporting documents."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 10  # Maintain progress until valid input

    if update.callback_query and update.callback_query.data == "skip":
        query = update.callback_query
        await query.answer()

        prompt = (
            f"⏭ {get_translation(user_id, 'job_seeker_profile_summary_skipped')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await query.edit_message_text(text=prompt, parse_mode="Markdown")
        return await job_seeker_supporting_documents_prompt(update, context)

    # Handle text input
    profile_summary = update.message.text.strip()

    if not profile_summary or len(profile_summary) < 50:
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_profile_summary')}\n"
            f"_{get_translation(user_id, 'profile_summary_instructions')}_\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(error_msg, parse_mode="Markdown")
        return JOB_SEEKER_PROFILE_SUMMARY

    # Save valid profile summary
    db.update_user_profile(user_id, profile_summary=profile_summary)
    context.user_data['progress'] = 11  # Update progress

    success_msg = (
        f"✅ {get_translation(user_id, 'profile_summary_saved')}\n\n"
        f"{get_progress_bar(user_id, context)}"
    )
    await update.message.reply_text(success_msg, parse_mode="Markdown")
    return await job_seeker_supporting_documents_prompt(update, context)


async def job_seeker_supporting_documents_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user for supporting documents with skip option."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 11  # Update progress

    keyboard = [
        [InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")]
    ]

    prompt = (
        f"📎 *{get_translation(user_id, 'job_seeker_supporting_documents_prompt')}*\n"
        f"_{get_translation(user_id, 'supporting_documents_instructions')}_\n\n"
        f"• {get_translation(user_id, 'supported_formats')}: PDF, DOCX, PPTX\n"
        f"• {get_translation(user_id, 'max_size')}: 20MB\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_SUPPORTING_DOCUMENTS


async def job_seeker_supporting_documents(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle supporting documents upload and proceed to portfolio link."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 11  # Maintain progress until valid input

    if update.callback_query and update.callback_query.data == "skip":
        query = update.callback_query
        await query.answer()

        prompt = (
            f"⏭ {get_translation(user_id, 'job_seeker_supporting_documents_skipped')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await query.edit_message_text(text=prompt, parse_mode="Markdown")
        return await job_seeker_portfolio_link_prompt(update, context)

    if update.message and update.message.document:
        document = update.message.document
        file_size = document.file_size

        # Check file size (20MB limit)
        if file_size > 20 * 1024 * 1024:
            error_msg = (
                f"⚠️ {get_translation(user_id, 'file_too_large')}\n"
                f"_{get_translation(user_id, 'max_size')}: 20MB_\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(error_msg, parse_mode="Markdown")
            return JOB_SEEKER_SUPPORTING_DOCUMENTS

        # Check file extension
        file_ext = os.path.splitext(document.file_name)[1].lower()
        if file_ext not in ['.pdf', '.jpg', '.png', '.jpeg', '.doc', '.docx', '.ppt', '.pptx']:
            error_msg = (
                f"⚠️ {get_translation(user_id, 'unsupported_format')}\n"
                f"_{get_translation(user_id, 'supported_formats')}: PDF, JPG, PNG, DOC/DOCX, PPT/PPTX_\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(error_msg, parse_mode="Markdown")
            return JOB_SEEKER_SUPPORTING_DOCUMENTS

        try:
            # Save document reference in database
            db.save_user_document(user_id, document.file_id)
            context.user_data['progress'] = 12  # Update progress

            success_msg = (
                f"✅ {get_translation(user_id, 'document_uploaded_successfully')}\n"
                f"📄 {document.file_name}\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(success_msg, parse_mode="Markdown")
            return await job_seeker_portfolio_link_prompt(update, context)

        except Exception as e:
            error_msg = (
                f"⚠️ {get_translation(user_id, 'document_upload_failed')}\n"
                f"_{str(e)}_\n\n"
                f"{get_progress_bar(user_id, context)}"
            )
            await update.message.reply_text(error_msg, parse_mode="Markdown")
            return JOB_SEEKER_SUPPORTING_DOCUMENTS

    # If no valid document was provided
    error_msg = (
        f"⚠️ {get_translation(user_id, 'no_document_provided')}\n"
        f"_{get_translation(user_id, 'please_upload_a_document_or_skip')}_\n\n"
        f"{get_progress_bar(user_id, context)}"
    )
    await update.message.reply_text(error_msg, parse_mode="Markdown")
    return JOB_SEEKER_SUPPORTING_DOCUMENTS

async def get_file_from_telegram(job_seeker_file_id, context: ContextTypes.DEFAULT_TYPE):
    try:
        # Use the job_seeker_file_id to get the file object
        file = await context.bot.get_file(job_seeker_file_id)
        # Download the file to a specific path or process it as needed
        file_path = await file.download_to_drive("downloaded_file.pdf")
        return file_path
    except Exception as e:
        print(f"Error retrieving file: {e}")
        return None


from urllib.parse import urlparse


async def job_seeker_portfolio_link_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user for portfolio link with skip option."""
    user_id = get_user_id(update)
    context.user_data['progress'] = 12  # Update progress

    keyboard = [
        [InlineKeyboardButton("⏩ " + get_translation(user_id, "skip"), callback_data="skip")]
    ]

    prompt = (
        f"🔗 *{get_translation(user_id, 'job_seeker_portfolio_link_prompt')}*\n"
        f"_{get_translation(user_id, 'portfolio_link_instructions')}_\n\n"
        f"🌐 {get_translation(user_id, 'acceptable_formats')}:\n"
        f"- https://linkedin.com/in/yourprofile\n"
        f"- www.github.com/username\n"
        f"- example.com\n"
        f"- http://personalwebsite.com\n\n"
        f"{get_translation(user_id, 'lenient_url_note')}\n\n"
        f"{get_progress_bar(user_id, context)}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=prompt,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return JOB_SEEKER_PORTFOLIO_LINK


async def job_seeker_portfolio_link(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle portfolio link with lenient validation."""
    user_id = get_user_id(update)
    language = context.user_data.get('language', 'english')

    if update.callback_query and update.callback_query.data == "skip":
        # Handle skip
        await update.callback_query.answer()

        # Update progress to complete
        context.user_data['progress'] = TOTAL_REGISTRATION_STEPS

        await update.callback_query.edit_message_text(
            f"⏭️ {get_translation(user_id, 'job_seeker_portfolio_link_skipped')}",
            parse_mode="Markdown"
        )
        await finalize_registration(user_id, context)
        return MAIN_MENU

    # Handle text input
    portfolio_link = update.message.text.strip()

    # Update progress to complete
    context.user_data['progress'] = TOTAL_REGISTRATION_STEPS

    # Input Cleaning
    portfolio_link = portfolio_link.lower()  # Normalize case
    if not portfolio_link.startswith(("http://", "https://")):
        if portfolio_link.startswith("www."):
            portfolio_link = "https://" + portfolio_link
        else:
            portfolio_link = "https://" + portfolio_link

    # Basic Checks
    if not re.match(r"^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}(/.*)?$", urlparse(portfolio_link).netloc):
        error_msg = (
            f"⚠️ {get_translation(user_id, 'invalid_portfolio_link')}\n\n"
            f"ℹ️ {get_translation(user_id, 'portfolio_validation_help')}\n"
            f"- https://linkedin.com/in/yourprofile\n"
            f"- www.github.com/username\n"
            f"- example.com\n\n"
            f"{get_translation(user_id, 'lenient_url_note')}\n\n"
            f"{get_progress_bar(user_id, context)}"
        )
        await update.message.reply_text(
            error_msg,
            parse_mode="Markdown"
        )
        return JOB_SEEKER_PORTFOLIO_LINK

    # Save valid URL
    db.update_user_profile(user_id, portfolio_link=portfolio_link)

    # Confirm save
    success_message = (
        f"✅ {get_translation(user_id, 'portfolio_link_saved')}\n"
        f"🔗 {get_translation(user_id, 'final_saved_url')}: `{portfolio_link}`"
    )
    await update.message.reply_text(
        success_message,
        parse_mode="Markdown"
    )

    await finalize_registration(user_id, context)
    return MAIN_MENU


async def finalize_registration(update_or_user_id: Update | int, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Finalize registration with comprehensive summary and next steps."""
    # Determine user_id based on input type
    if isinstance(update_or_user_id, int):
        user_id = update_or_user_id
    else:
        user_id = get_user_id(update_or_user_id)

    # Mark registration as complete
    # db.update_user_registration_status(user_id, complete=True)
    context.user_data['progress'] = 100  # Complete progress

    # Prepare welcome message
    welcome_msg = (
        f"🎉 *{get_translation(user_id, 'job_seeker_registration_complete')}*\n\n"
        f"✅ {get_translation(user_id, 'registration_success_message')}\n"
        f"📌 {get_translation(user_id, 'next_steps_instructions')}\n\n"
    )

    # Send completion message
    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_msg,
        parse_mode="Markdown"
    )

    # Display profile summary
    await display_user_profile(user_id, context)


from telegram import KeyboardButton, ReplyKeyboardMarkup

async def display_user_profile(user_id, context):
    try:
        # Retrieve the user's profile as a dictionary
        user_profile = db.get_user_profile(user_id)
        if not user_profile:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "profile_not_found")
            )
            return

        # Extract fields dynamically, using "Not Provided" as a fallback
        full_name = user_profile.get("full_name") or get_translation(user_id, "not_provided")
        contact_number = user_profile.get("contact_number") or get_translation(user_id, "not_provided")
        dob = user_profile.get("dob") or get_translation(user_id, "not_provided")
        gender = user_profile.get("gender") or get_translation(user_id, "not_provided")
        languages = user_profile.get("languages") or get_translation(user_id, "not_provided")
        qualification = user_profile.get("qualification") or get_translation(user_id, "not_provided")
        field_of_study = user_profile.get("field_of_study") or get_translation(user_id, "not_provided")
        cgpa = user_profile.get("cgpa") or get_translation(user_id, "not_provided")
        skills_experience = user_profile.get("skills_experience") or get_translation(user_id, "not_provided")
        profile_summary = user_profile.get("profile_summary") or get_translation(user_id, "not_provided")
        cv_path = user_profile.get("cv_path")  # This is the file_id of the CV
        portfolio_link = user_profile.get("portfolio_link")

        # Define a separator for better readability
        separator = "────────────────────"

        # Format the profile message with enhanced visuals
        profile_message = (
            f"📌 *{get_translation(user_id, 'profile_header')}*\n"
            f"{separator}\n"
            f"👤 *{get_translation(user_id, 'full_name')}:* `{full_name}`\n"
            f"📞 *{get_translation(user_id, 'contact_number')}:* `{contact_number}`\n"
            f"🎂 *{get_translation(user_id, 'dob')}:* `{dob}`\n"
            f"🚻 *{get_translation(user_id, 'gender')}:* `{gender}`\n"
            f"🗣️ *{get_translation(user_id, 'languages')}:* `{languages}`\n"
            f"{separator}\n"
            f"🎓 *{get_translation(user_id, 'qualification')}:* `{qualification}`\n"
            f"📚 *{get_translation(user_id, 'field_of_study')}:* `{field_of_study}`\n"
            f"📊 *{get_translation(user_id, 'cgpa')}:* `{cgpa}`\n"
            f"{separator}\n"
            f"💼 *{get_translation(user_id, 'skills_experience')}:*\n `{skills_experience}`\n"
            f"{separator}\n"
            f"📝 *{get_translation(user_id, 'profile_summary')}:*\n"
            f"```{profile_summary}```\n"
            f"{separator}\n"
        )

        # Add portfolio link if available
        if portfolio_link:
            profile_message += (
                f"🌐 *{get_translation(user_id, 'portfolio_link')}:* "
                f"[{get_translation(user_id, 'view_portfolio')}]({portfolio_link})\n"
            )

        # Fallback message if all fields are "Not Provided"
        if all(value == get_translation(user_id, "not_provided") for value in [
            full_name, contact_number, dob, gender, languages,
            qualification, field_of_study, cgpa, skills_experience, profile_summary
        ]):
            profile_message = (
                f"📌 *{get_translation(user_id, 'profile_header')}*\n"
                f"{separator}\n"
                f"⚠️ {get_translation(user_id, 'no_profile_info_available')}\n"
                f"{separator}\n"
            )

        # Send the profile message
        await context.bot.send_message(
            chat_id=user_id,
            text=profile_message,
            parse_mode="Markdown",
            disable_web_page_preview=True
        )

        # Send the CV file if available
        if cv_path:
            try:
                # Send the CV file using the file_id
                await context.bot.send_document(
                    chat_id=user_id,
                    document=cv_path,
                    caption=get_translation(user_id, "your_cv_file")
                )
            except Exception as e:
                print(f"Error sending CV file: {e}")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "error_sending_cv")
                )
        else:
            # Inform the user if no CV is available
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_cv_available")
            )

        # Show the "Go to Main Menu" button
        keyboard = [[KeyboardButton(get_translation(user_id, "proceed_to_main_menu"))]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "what_next"),
            reply_markup=reply_markup
        )

    except Exception as e:
        print(f"Error displaying user profile: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_retrieving_profile")
        )


# Job Seeker Main Menu
async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Get user data for personalization
    user_data = db.get_user_profile(user_id)
    first_name = user_data.get('full_name', '').split()[0] if user_data else ''

    # Check for pending applications
    pending_apps = db.get_pending_applications_count(user_id)
    approved_apps = db.get_approved_applications_count(user_id)

    # Check profile completion status
    profile_completion = calculate_profile_completion(user_data)
    keyboard = [
        [KeyboardButton(f"👤 {get_translation(user_id, 'profile_button')} ({profile_completion}%)")],
        [KeyboardButton(f"✉️ {get_translation(user_id, 'applications_button')} ({pending_apps}) ({approved_apps})")],
        [KeyboardButton(f"✍️ {get_translation(user_id, 'apply_vacancy_button')}"),
         KeyboardButton(f"🔍 {get_translation(user_id, 'search_vacancies_button')}")],
        [KeyboardButton(f"ℹ️ {get_translation(user_id, 'help_button')}"),
         KeyboardButton(f"⭐ {get_translation(user_id, 'rate_button')}"),
         KeyboardButton(f"📝 {get_translation(user_id, 'report_button')}")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "main_menu_prompt"),
        reply_markup=reply_markup
    )

    # Create personalized welcome message
    welcome_msg = (
        f"🌟 <b>{get_translation(user_id, 'welcome_back')}, {first_name or get_translation(user_id, 'valued_user')}!</b>\n\n"
        f"📊 <i>{get_translation(user_id, 'profile_completion')}:</i> {profile_completion}%\n"
        f"📨 <i>{get_translation(user_id, 'pending_applications')}:</i> {pending_apps}\n"
        f"📨 <i>{get_translation(user_id, 'approved_applications')}:</i> {approved_apps}\n\n"

    )

    # Add tip of the day (optional)
    tip_of_day = get_tip_of_the_day(user_id)
    if tip_of_day:
        welcome_msg += f"\n💡 <b>{get_translation(user_id, 'tip_of_day')}:</b> {tip_of_day}"

    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_msg,
        parse_mode="HTML",
        reply_markup=reply_markup
    )


    return MAIN_MENU


def calculate_profile_completion(user_data: dict) -> int:
    """Calculate profile completion percentage counting skips as incomplete."""
    all_fields = [
        'full_name', 'contact_number', 'dob', 'gender', 'languages',
        'qualification', 'field_of_study', 'cgpa',
        'skills_experience', 'profile_summary', 'cv_path', 'portfolio_link'
    ]

    if not user_data:
        return 0

    filled = 0
    for field in all_fields:
        value = user_data.get(field)

        # Field counts as filled ONLY if:
        # 1. It has a value AND
        # 2. The value is not "skip" or empty
        if value and str(value).lower() != "skip":
            filled += 1

    return min(100, (filled * 100) // len(all_fields))

def get_tip_of_the_day(user_id: int) -> str:
    """Get a random career tip for the user with translations."""
    tips = [
        get_translation(user_id, "customize_resume"),
        get_translation(user_id, "follow_up_applications"),
        get_translation(user_id, "highlight_achievements"),
        get_translation(user_id, "network_with_professionals"),
        get_translation(user_id, "keep_skills_updated"),
        get_translation(user_id, "prepare_for_interviews"),
        get_translation(user_id, "set_clear_goals"),
        get_translation(user_id, "stay_positive"),
        get_translation(user_id, "seek_mentorship"),
        get_translation(user_id, "use_linkedin_effectively")
    ]
    return random.choice(tips)


async def view_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    profile_data = db.get_user_profile(user_id)
    completion = calculate_profile_completion(profile_data)

    # Enhanced keyboard with visual indicators
    keyboard = [
        [KeyboardButton(f"👤 {get_translation(user_id, 'view_profile_button')} ({completion}%)")],  # First row
        [KeyboardButton(f"✏️ {get_translation(user_id, 'edit_profile_button')}"),  # Second row
         KeyboardButton(f"📊 {get_translation(user_id, 'profile_stats')}")],
        [KeyboardButton(f"🗑️ {get_translation(user_id, 'delete_account_button')}"),  # Third row
         KeyboardButton(f"🌐 {get_translation(user_id, 'change_language_button')}")],
        [KeyboardButton(f"🔙 {get_translation(user_id, 'back_to_main_menu')}")]  # Fourth row
    ]

    # Create the ReplyKeyboardMarkup
    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        one_time_keyboard=True,
        resize_keyboard=True
    )

    # Send the message with the new keyboard
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "profile_menu_prompt"),
        reply_markup=reply_markup
    )

    return PROFILE_MENU


def get_profile_strength_tip_job_seeker(completion: int, user_id: int) -> str:
    """Generate personalized tips to improve profile strength for job seekers."""
    profile_data = db.get_user_profile(user_id)
    missing_components = []

    # Check which critical components are missing
    if not profile_data.get('skills_experience'):
        missing_components.append("skills and experience")
    if not profile_data.get('profile_summary'):
        missing_components.append("profile summary")
    if not profile_data.get('cv_path'):
        missing_components.append("CV/resume")
    if not profile_data.get('portfolio_link'):
        missing_components.append("portfolio link")

    # Base message on completion percentage
    if completion >= 90:
        base_msg = get_translation(user_id, 'profile_strength_excellent')
    elif completion >= 70:
        base_msg = get_translation(user_id, 'profile_strength_good')
    elif completion >= 50:
        base_msg = get_translation(user_id, 'profile_strength_needs_improvement')
    else:
        base_msg = get_translation(user_id, 'profile_strength_incomplete')

    # Specific recommendations based on missing components
    recommendations = []
    if missing_components:
        if 'CV/resume' in missing_components:
            recommendations.append(get_translation(user_id, 'profile_recommend_cv'))
        if 'skills and experience' in missing_components:
            recommendations.append(get_translation(user_id, 'profile_recommend_skills'))
        if 'profile_summary' in missing_components:
            recommendations.append(get_translation(user_id, 'profile_recommend_summary'))
        if 'portfolio_link' in missing_components:
            recommendations.append(get_translation(user_id, 'profile_recommend_portfolio'))

    # Compose and return the tip message
    if recommendations:
        return f"💡 <b>{get_translation(user_id, 'profile_tip_title')}</b> {base_msg}\n\n🔧 <b>{get_translation(user_id, 'recommendations_title')}</b>\n" + "\n".join(recommendations)
    return f"💡 <b>{get_translation(user_id, 'profile_tip_title')}</b> {base_msg}"

async def handle_profile_actions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text  # Get the text sent by the user (button press)

    # Dynamically construct button texts with emojis
    view_profile_button_text = f"👤 {get_translation(user_id, 'view_profile_button')} ({calculate_profile_completion(db.get_user_profile(user_id))}%)"
    edit_profile_button_text = f"✏️ {get_translation(user_id, 'edit_profile_button')}"
    profile_stats_button_text = f"📊 {get_translation(user_id, 'profile_stats')}"
    delete_account_button_text = f"🗑️ {get_translation(user_id, 'delete_account_button')}"
    change_language_button_text = f"🌐 {get_translation(user_id, 'change_language_button')}"
    back_to_main_menu_button_text = f"🔙 {get_translation(user_id, 'back_to_main_menu')}"

    if choice == view_profile_button_text:
        # Fetch and display the user's profile
        await display_user_profile(user_id, context)
        return await view_profile(update, context)

    elif choice == edit_profile_button_text:
        # Redirect to the edit profile menu
        return await edit_profile(update, context)

    elif choice == profile_stats_button_text:
        # Get profile data and metrics
        profile_data = db.get_user_profile(user_id)
        profile_completion = calculate_profile_completion(profile_data)
        strength = analyze_profile_strength(profile_data)

        # Get application metrics
        total_applications = db.get_total_applications_count(user_id)
        pending_apps = db.get_pending_applications_count(user_id)
        approved_apps = db.get_approved_applications_count(user_id)




        # Build the comprehensive stats message
        stats_message = (
            f"📊 <b>{get_translation(user_id, 'dashboard_title')}</b>\n\n"
            f"👤 <b>{get_translation(user_id, 'name_label')}:</b> {profile_data.get('full_name', get_translation(user_id, 'not_provided'))}\n"
            f"📅 <b>{get_translation(user_id, 'member_since_label')}:</b> {db.get_member_since_date(user_id)}\n\n"

            f"🔍 <b>{get_translation(user_id, 'profile_status_title')}:</b>\n"
            f"   • {get_translation(user_id, 'completion_label')}: {profile_completion}%\n"
            

            f"{generate_profile_strength_bar(profile_completion)}\n\n"

            f"📈 <b>{get_translation(user_id, 'application_metrics_title')}:</b>\n"
            f"   • {get_translation(user_id, 'total_applications_label')}: {total_applications}\n"
            f"   • {get_translation(user_id, 'pending_label')}: {pending_apps}\n"
            f"   • {get_translation(user_id, 'approved_label')}: {approved_apps}\n"

            f"📝 <b>{get_translation(user_id, 'profile_components_title')}:</b>\n"

            f"   • {get_translation(user_id, 'skills_experience_label')}: {'✅ ' + get_translation(user_id, 'complete') if profile_data.get('skills_experience') else '❌ ' + get_translation(user_id, 'missing')}\n"
            f"   • {get_translation(user_id, 'profile_summary_label')}: {'✅ ' + get_translation(user_id, 'complete') if profile_data.get('profile_summary') else '❌ ' + get_translation(user_id, 'missing')}\n"
            f"   • {get_translation(user_id, 'portfolio_link_label')}: {'✅ ' + get_translation(user_id, 'provided') if profile_data.get('portfolio_link') else '❌ ' + get_translation(user_id, 'missing')}\n"
            f"   • {get_translation(user_id, 'cv_resume_label')}: {'✅ ' + get_translation(user_id, 'uploaded') if profile_data.get('cv_path') else '❌ ' + get_translation(user_id, 'missing')}\n\n"

            f"💪 <i>{get_translation(user_id, 'profile_strength_title')}:</i> {strength}\n\n"

            f"{get_profile_strength_tip_job_seeker(profile_completion, user_id)}"
        )
        # Send the enhanced stats message
        await context.bot.send_message(
            chat_id=user_id,
            text=stats_message,
            parse_mode="HTML"
        )
        return await view_profile(update, context)
    elif choice == delete_account_button_text:
        return await confirm_delete_account(update, context)

    elif choice == change_language_button_text:
        # Show language selection menu
        return await show_job_seeker_language_selection(update, context)

    elif choice == back_to_main_menu_button_text:
        # Return to the main menu
        return await main_menu(update, context)
    else:
        # Check if the current state is CONFIRM_DELETE_ACCOUNT
        current_state = context.user_data.get('current_state')
        if current_state == CONFIRM_DELETE_ACCOUNT:
            # Delegate invalid input handling to handle_delete_confirmation
            return await handle_delete_confirmation(update, context)

        # Handle invalid input for other states
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_choice")
        )
        return PROFILE_MENU

def analyze_profile_strength(profile_data: dict) -> str:
    strength = 0
    if profile_data.get('skills_experience'): strength += 25
    if profile_data.get('profile_summary'): strength += 20
    if profile_data.get('portfolio_link'): strength += 20
    if profile_data.get('cv_path'): strength += 35
    return f"{min(100, strength)}% Strong"


async def confirm_delete_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Store the expected responses in context for validation
    context.user_data['expected_responses'] = {
        'delete': get_translation(user_id, 'confirm_delete_yes'),
        'keep': get_translation(user_id, 'confirm_delete_no')
    }

    # Set the current state to CONFIRM_DELETE_ACCOUNT
    context.user_data['current_state'] = CONFIRM_DELETE_ACCOUNT

    keyboard = [
        [KeyboardButton(f"❌ {context.user_data['expected_responses']['delete']}")],
        [KeyboardButton(f"✅ {context.user_data['expected_responses']['keep']}")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "delete_confirmation"),
        reply_markup=reply_markup
    )
    return CONFIRM_DELETE_ACCOUNT


async def handle_delete_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip()

    # Get expected responses from context
    expected = context.user_data.get('expected_responses', {})
    delete_text = f"❌ {expected.get('delete', 'Delete my account')}"
    keep_text = f"✅ {expected.get('keep', 'Keep my account')}"

    # Check both button text and raw text (case insensitive)
    if choice.lower() in [delete_text.lower(), expected.get('delete', '').lower(), 'delete', 'del']:
        db.delete_user_account(user_id)
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "account_deleted_message")
        )
        return ConversationHandler.END

    elif choice.lower() in [keep_text.lower(), expected.get('keep', '').lower(), 'keep', 'no']:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "delete_account_cancelled")
        )
        return await view_profile(update, context)

    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_delete_choice")  # Correct invalid choice message
        )
        return CONFIRM_DELETE_ACCOUNT  # Keep the user in the same state


async def delete_account_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Delete the user's account from the database
    db.delete_user_account(user_id)

    # Notify the user and end the conversation
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "account_deleted_message")
    )
    return ConversationHandler.END  # End the conversation

async def cancel_delete_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Notify the user that the deletion was cancelled
    await query.edit_message_text(
        text=get_translation(user_id, "delete_account_cancelled")
    )
    return await view_profile(update, context)  # Return to the profile menu

async def show_job_seeker_language_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Define the inline keyboard for language selection
    keyboard = [
        [
            InlineKeyboardButton("🇬🇧 English", callback_data="english"),
            InlineKeyboardButton("🇪🇹 አማርኛ", callback_data="amharic")
        ],
        [
            InlineKeyboardButton(" Afaan Oromoo", callback_data="oromia"),
            InlineKeyboardButton(" ትግርኛ", callback_data="tigrigna")
        ],
        [
            InlineKeyboardButton(" Qafar af", callback_data="afar"),
            InlineKeyboardButton(" Soomaali", callback_data="somalia")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send the language selection prompt
    await update.message.reply_text(
        get_translation(user_id, "employer_select_language_prompt"),
        reply_markup=reply_markup
    )
    return SELECT_LANGUAGE

async def handle_job_seeker_language_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    selected_language = query.data

    # Update the employer's language in the database
    db.update_user_language(user_id, selected_language)

    # Notify the user about the language change
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "employer_language_updated_message")
    )

    # Return to the employer main menu
    return await main_menu(update, context)


async def confirm_change_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Show confirmation buttons
    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, "yes_button"), callback_data="change_language_confirmed")],
        [InlineKeyboardButton(get_translation(user_id, "no_button"), callback_data="cancel_change_language")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        text=get_translation(user_id, "confirm_change_language_prompt"),
        reply_markup=reply_markup
    )
    return CONFIRM_CHANGE_LANGUAGE

async def change_language_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Update the user's language preference
    selected_language = context.user_data.get("selected_language", "english")
    db.update_user_language(user_id, selected_language)

    # Notify the user about the language change
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "language_updated_message")
    )
    return await main_menu(update, context)

async def cancel_change_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Notify the user that the language change was cancelled
    await query.edit_message_text(
        text=get_translation(user_id, "change_language_cancelled")
    )
    return await view_profile(update, context)  # Return to the profile menu


async def edit_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    user_data = db.get_user_profile(user_id)
    profile_completion = calculate_profile_completion(user_data)  # You'll need to implement this

    # Create a grid layout for profile fields
    keyboard = []
    row = []
    valid_fields = [
        "full_name", "contact_number", "dob", "gender", "languages",
        "qualification", "field_of_study", "cgpa", "skills_experience",
        "profile_summary", "cv_path", "portfolio_link"
    ]

    for i, field in enumerate(valid_fields):
        # Check if field is completed
        is_completed = user_data and user_data.get(field) and str(user_data.get(field)).lower() != "skip"
        completion_icon = "✅" if is_completed else "🟡"

        row.append(InlineKeyboardButton(
            f"{get_translation(user_id, f'edit_{field}')} {completion_icon}",
            callback_data=f"edit_{field}"
        ))

        # Create new row every 2 buttons
        if (i + 1) % 2 == 0:
            keyboard.append(row)
            row = []

    # Add remaining buttons and navigation
    if row:
        keyboard.append(row)

    keyboard.append([
        InlineKeyboardButton(
            f"📊 {get_translation(user_id, 'view_profile_completion')} ({profile_completion}%)",
            callback_data="view_completion"
        )
    ])


    keyboard.append([
        InlineKeyboardButton(
            f"🔙 {get_translation(user_id, 'back_to_main_menu')}",
            callback_data="back_to_main_menu"
        )
    ])

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Enhanced message with profile summary
    message = (
        f"✨ <b>{get_translation(user_id, 'edit_profile')}</b> ✨\n\n"
        f"📌 <i>{get_translation(user_id, 'profile_completion')}:</i> <b>{profile_completion}%</b>\n"
        f"{generate_profile_strength_bar(profile_completion)}\n\n" 
        f"ℹ️ {get_translation(user_id, 'edit_profile_instructions')}\n\n"
        f"{get_profile_edit_tip(user_id, profile_completion)}"
    )

    # Edit existing message if possible, otherwise send new
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )

    return EDIT_PROFILE


async def handle_edit_profile_field(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    callback_data = query.data

    # Handle navigation cases
    if callback_data == "back_to_main_menu":
        await query.edit_message_text(text=get_translation(user_id, "returning_to_main_menu"))
        return await main_menu(update, context)
    elif callback_data == "view_completion":
        return await show_profile_completion_details_js(update, context)

    # Extract field name
    if not callback_data.startswith("edit_"):
        await query.edit_message_text(text=get_translation(user_id, "invalid_field_selected"))
        return EDIT_PROFILE

    field_name = callback_data.replace("edit_", "")
    valid_fields = [
        "full_name", "contact_number", "dob", "gender", "languages",
        "qualification", "field_of_study", "cgpa", "skills_experience",
        "profile_summary", "cv_path", "portfolio_link"
    ]

    if field_name not in valid_fields:
        await query.edit_message_text(text=get_translation(user_id, "invalid_field_selected"))
        return EDIT_PROFILE

    # Store field in context
    context.user_data["editing_field"] = field_name

    # Special handling for CV upload
    if field_name == "cv_path":
        # Get current CV info if exists
        user_data = db.get_user_profile(user_id)
        current_cv = ""
        if user_data and user_data.get("cv_path") and str(user_data.get("cv_path")).lower() != "skip":
            current_cv = f"\n\n📌 <b>{get_translation(user_id, 'current_value')}:</b> {get_translation(user_id, 'cv_uploaded')}"

        prompt = (
            f"📄 <b>{get_translation(user_id, 'upload_cv_title')}</b>\n\n"
            f"ℹ️ {get_translation(user_id, 'upload_cv_instructions')}\n"
            f"▸ {get_translation(user_id, 'accepted_formats')}: PDF, JPG, PNG, DOC/DOCX, PPT/PPTX \n"
            f"▸ {get_translation(user_id, 'max_size')}: 20MB"
            f"{current_cv}\n\n"
            f"⬇️ {get_translation(user_id, 'upload_now_prompt')}"
        )

        keyboard = [
            [
                InlineKeyboardButton(
                    f"❌ {get_translation(user_id, 'cancel_editing')}",
                    callback_data="cancel_editing"
                ),

            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            text=prompt,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
        return EDIT_PROFILE_FIELD_VALUE

    # Default handling for other fields
    user_data = db.get_user_profile(user_id)
    current_value = ""
    if user_data and field_name in user_data:
        current_value = user_data[field_name]
        if current_value and str(current_value).lower() != "skip":
            current_value = f"\n\n📌 <b>{get_translation(user_id, 'current_value')}:</b>\n<code>{current_value[:100]}{'...' if len(current_value) > 100 else ''}</code>"

    prompt = (
        f"✏️ <b>{get_translation(user_id, f'edit_{field_name}_title')}</b>\n\n"
        f"{get_translation(user_id, f'edit_{field_name}_instructions')}"
        f"{current_value}\n\n"
        f"⬇️ {get_translation(user_id, 'enter_new_value_prompt')}"
    )

    # Send the prompt with a cancel button
    keyboard = [
        [InlineKeyboardButton(
            f"❌ {get_translation(user_id, 'cancel_editing')}",
            callback_data="cancel_editing"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=prompt,
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    return EDIT_PROFILE_FIELD_VALUE

async def save_edited_field(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    field_name = context.user_data.get("editing_field")

    if not field_name:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'session_expired_start_again'),
            parse_mode="HTML"
        )
        return await edit_profile(update, context)

    try:
        # Handle CV document upload
        if field_name == "cv_path":
            if update.callback_query and update.callback_query.data == "skip_cv_upload":
                query = update.callback_query
                await query.answer()
                new_value = "skip"
                success_message = f"⏭ {get_translation(user_id, 'cv_upload_skipped')}"
            elif update.message and update.message.document:
                document = update.message.document
                file_ext = os.path.splitext(document.file_name)[1].lower() if document.file_name else ''

                # Validate file
                if file_ext not in ['.pdf', '.jpg', '.png', '.jpeg', '.doc', '.docx', '.ppt', '.pptx']:
                    await update.message.reply_text(
                        f"❌ {get_translation(user_id, 'invalid_file_format')}\n\n"
                        f"📌 {get_translation(user_id, 'accepted_formats')}: PDF, JPG, PNG, DOC/DOCX, PPT/PPTX\n"
                        f"💾 {get_translation(user_id, 'please_try_again')}",
                        parse_mode="HTML"
                    )
                    return EDIT_PROFILE_FIELD_VALUE

                if document.file_size > 20 * 1024 * 1024:  # 20MB
                    await update.message.reply_text(
                        f"❌ {get_translation(user_id, 'document_too_large')}\n"
                        f"📏 {get_translation(user_id, 'max_size')}: 20MB",
                        parse_mode="HTML"
                    )
                    return EDIT_PROFILE_FIELD_VALUE

                new_value = document.file_id
                success_message = f"✅ {get_translation(user_id, 'cv_updated_successfully')}"

                # Save document reference in database
                db.save_user_document(user_id, document.file_id)
            else:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=f"❌ {get_translation(user_id, 'no_document_uploaded')}",
                    parse_mode="HTML"
                )
                return EDIT_PROFILE_FIELD_VALUE
        else:
            # Handle text fields (original implementation)
            new_value = update.message.text.strip()
            success_message = f"✅ {get_translation(user_id, 'field_updated_success')}"

        # Update profile in database
        db.update_user_profile(user_id, **{field_name: new_value})

        # Send success message with updated field preview
        preview = new_value[:50] + "..." if len(str(new_value)) > 50 else new_value
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"✅ {get_translation(user_id, 'field_updated_successfully')}\n\n"
                f"📋 <b>{get_translation(user_id, 'updated_value')}:</b>\n"
                f"<code>{preview}</code>\n\n"
                f"{get_random_positive_emoji()} {get_translation(user_id, 'what_next_prompt')}"
            ),
            parse_mode="HTML"
        )

        # Update profile completion percentage
        user_data = db.get_user_profile(user_id)
        new_completion = calculate_profile_completion(user_data)  # You'll need to implement this

        # Send completion update if significant change
        prev_completion = context.user_data.get("profile_completion", 0)
        if abs(new_completion - prev_completion) >= 10:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"📈 <b>{get_translation(user_id, 'profile_completion_updated')}</b>\n"
                    f"{generate_profile_strength_bar(new_completion)}\n"
                    f"<b>{new_completion}%</b> {get_completion_message(new_completion, user_id)}"
                ),
                parse_mode="HTML"
            )
            context.user_data["profile_completion"] = new_completion

    except Exception as e:
        logging.error(f"Error updating field: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"⚠️ <b>{get_translation(user_id, 'update_error_title')}</b>\n\n"
                f"{get_translation(user_id, 'update_error_message')}\n"
                f"🔧 {get_translation(user_id, 'technical_details')}: <code>{str(e)[:100]}</code>"
            ),
            parse_mode="HTML"
        )

    return await edit_profile(update, context)


async def show_profile_completion(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show detailed profile completion breakdown"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    user_data = db.get_user_profile(user_id)

    if not user_data:
        await query.edit_message_text(text=get_translation(user_id, "profile_not_found"))
        return EDIT_PROFILE

    completion = calculate_profile_completion(user_data)
    valid_fields = [
        "full_name", "contact_number", "dob", "gender", "languages",
        "qualification", "field_of_study", "cgpa", "skills_experience",
        "profile_summary", "cv_path", "portfolio_link"
    ]

    missing_fields = []
    completed_fields = []

    for field in valid_fields:
        value = user_data.get(field)
        if value and str(value).lower() != "skip":
            completed_fields.append(field)
        else:
            missing_fields.append(field)

    # Create detailed message
    message_parts = [
        f"📊 <b>{get_translation(user_id, 'profile_completion_details')}</b>\n\n",
        f"{generate_profile_strength_bar(completion)}\n",
        f"<b>{completion}%</b> {get_completion_message(completion, user_id)}\n\n"
    ]

    if completed_fields:
        completed_list = [f'▸ {get_translation(user_id, f"edit_{f}")}\n' for f in completed_fields]
        message_parts.append(f"✅ <b>{get_translation(user_id, 'completed_sections')}:</b>\n{''.join(completed_list)}\n")

    if missing_fields:
        missing_list = []
        for f in missing_fields:
            field_name = get_translation(user_id, f"edit_{f}")
            missing_list.append(f'▸ {field_name}\n')

        message_parts.extend([
            f"⚠️ <b>{get_translation(user_id, 'missing_sections')}:</b>\n",
            ''.join(missing_list),
            f"\n💡 <i>{get_translation(user_id, 'completion_tip')}</i>"
        ])

    # Combine all message parts
    message = ''.join(message_parts)

    # Add back button
    keyboard = [
        [InlineKeyboardButton(
            f"🔙 {get_translation(user_id, 'back_to_editing')}",
            callback_data="back_to_editing"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=message,
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    return PROFILE_COMPLETION

def validate_cgpa(cgpa_str: str) -> bool:
    """Validate CGPA format (0.00-4.00)"""
    try:
        cgpa = float(cgpa_str)
        return 0 <= cgpa <= 4
    except:
        return False


def validate_date_format(date_str: str) -> bool:
    """
    Validate date of birth format with the same rules as registration flow.
    Accepts formats: YYYY/MM/DD, YYYY-MM-DD, YYYY.MM.DD
    Validates:
    - Correct format (4-digit year, 1-2 digit month/day)
    - Valid date values (month 1-12, day 1-31)
    - Reasonable age (16-100 years old)
    """
    if not date_str or not isinstance(date_str, str):
        return False

    # Check basic format pattern first
    if not re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$", date_str.strip()):
        return False

    try:
        # Normalize separators to hyphens
        normalized_date = date_str.replace("/", "-").replace(".", "-")

        # Parse the date
        birth_date = datetime.strptime(normalized_date, "%Y-%m-%d")

        # Validate date components
        year, month, day = birth_date.year, birth_date.month, birth_date.day
        if not (1 <= month <= 12 and 1 <= day <= 31):
            return False

        # Validate age range (16-100 years)
        today = datetime.now()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        if not (16 <= age <= 100):
            return False

        return True

    except ValueError:
        return False


def validate_url(url: str) -> bool:
    """Validate URL with lenient rules matching registration flow"""
    if not url:
        return False

    # Normalize the URL
    url = url.lower().strip()

    # Add https:// if missing
    if not url.startswith(("http://", "https://")):
        if url.startswith("www."):
            url = "https://" + url
        else:
            url = "https://" + url

    # Basic domain validation
    try:
        netloc = urlparse(url).netloc
        if not netloc:
            return False

        # Check domain pattern (matches your registration validation)
        return bool(re.match(r"^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}(/.*)?$", netloc))
    except:
        return False


JOB_SEEKER_EDITABLE_FIELDS = {
    "full_name": {
        "emoji": "👤",
        "max_length": 100,
        "validation": lambda x: len(x) >= 3
    },
    "contact_number": {
        "emoji": "📞",
        "max_length": 20,
        "validation": lambda x: x.isdigit() and len(x) >= 7
    },
    "dob": {
        "emoji": "🎂",
        "max_length": 10,
        "validation": lambda x: validate_date_format(x)  # You'll need to implement this
    },
    "gender": {
        "emoji": "⚥",
        "max_length": 20,
        "validation": lambda x: len(x) >= 1
    },
    "languages": {
        "emoji": "🗣️",
        "max_length": 200,
        "validation": lambda x: len(x) >= 2
    },
    "qualification": {
        "emoji": "🎓",
        "max_length": 100,
        "validation": lambda x: len(x) >= 2
    },
    "field_of_study": {
        "emoji": "📚",
        "max_length": 100,
        "validation": lambda x: len(x) >= 2
    },
    "cgpa": {
        "emoji": "📊",
        "max_length": 5,
        "validation": lambda x: validate_cgpa(x)  # You'll need to implement this
    },
    "skills_experience": {
        "emoji": "🛠️",
        "max_length": 1000,
        "validation": lambda x: len(x) >= 10
    },
    "profile_summary": {
        "emoji": "📝",
        "max_length": 500,
        "validation": lambda x: len(x) >= 20
    },
    "cv_path": {
        "emoji": "📄",
        "file_types": ['.pdf', '.jpg', '.png', '.jpeg', '.doc', '.docx', '.ppt', '.pptx'],
        "max_size": 20 * 1024 * 1024  # 20MB
    },
    "portfolio_link": {
        "emoji": "🔗",
        "max_length": 200,
        "validation": lambda x: validate_url(x)  # You'll need to implement this
    }
}

async def show_profile_completion_details_js(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show detailed job seeker profile completion breakdown"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    user_data = db.get_user_profile(user_id)

    if not user_data:
        await query.edit_message_text(text=get_translation(user_id, "profile_not_found"))
        return EDIT_PROFILE

    completion = calculate_profile_completion(user_data)
    missing_fields = []
    completed_fields = []

    for field in JOB_SEEKER_EDITABLE_FIELDS:
        value = user_data.get(field)
        if value and str(value).lower() != "skip":
            completed_fields.append(field)
        else:
            missing_fields.append(field)

    # Helper function to format field entries
    def format_field_entries(fields):
        entries = []
        for field in fields:
            emoji = JOB_SEEKER_EDITABLE_FIELDS[field]["emoji"]
            translation = get_translation(user_id, "edit_{}".format(field))
            entries.append(f"▸ {emoji} {translation}\n")
        return "".join(entries)

    # Build message parts
    message_parts = [
        f"📊 <b>{get_translation(user_id, 'profile_completion_details')}</b>\n\n",
        f"{generate_profile_strength_bar(completion)}\n",
        f"<b>{completion}%</b> {get_completion_message(completion, user_id)}\n\n"
    ]

    if completed_fields:
        message_parts.extend([
            f"✅ <b>{get_translation(user_id, 'completed_sections')}:</b>\n",
            format_field_entries(completed_fields),
            "\n"
        ])

    if missing_fields:
        message_parts.extend([
            f"⚠️ <b>{get_translation(user_id, 'missing_sections')}:</b>\n",
            format_field_entries(missing_fields),
            f"\n💡 <i>{get_translation(user_id, 'completion_tip')}</i>"
        ])

    # Combine all parts
    message = "".join(message_parts)

    # Add back button
    keyboard = [
        [InlineKeyboardButton(
            f"🔙 {get_translation(user_id, 'back_to_editing')}",
            callback_data="back_to_editing"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=message,
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    return PROFILE_COMPLETION

async def handle_document_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    field_name = context.user_data.get("editing_field")

    if field_name != "cv_path":
        await update.message.reply_text(get_translation(user_id, "unexpected_document"))
        return EDIT_PROFILE_FIELD_VALUE

    document = update.message.document
    file_ext = os.path.splitext(document.file_name)[1].lower() if document.file_name else ''

    # Validate file type
    if file_ext not in ['.pdf', '.jpg', '.png', '.jpeg', '.doc', '.docx', '.ppt', '.pptx' ]:
        await update.message.reply_text(
            f"❌ {get_translation(user_id, 'invalid_file_format')}\n\n"
            f"📌 {get_translation(user_id, 'accepted_formats')}: PDF, JPG, PNG, DOC/DOCX, PPT/PPTX\n"
            f"💾 {get_translation(user_id, 'please_try_again')}",
            parse_mode="HTML"
        )
        return EDIT_PROFILE_FIELD_VALUE

    # Validate file size (20MB)
    if document.file_size > 20 * 1024 * 1024:
        await update.message.reply_text(
            f"❌ {get_translation(user_id, 'document_too_large')}\n"
            f"📏 {get_translation(user_id, 'max_size')}: 20MB",
            parse_mode="HTML"
        )
        return EDIT_PROFILE_FIELD_VALUE

    try:
        # Save file reference
        db.update_user_profile(user_id, cv_path=document.file_id)

        # Also save document details properly
        db.save_user_document(user_id, document.file_id)

        await update.message.reply_text(
            f"✅ {get_translation(user_id, 'cv_updated_successfully')}\n\n"
            f"📄 {document.file_name}",
            parse_mode="HTML"
        )

        # Return to edit profile menu
        return await edit_profile(update, context)

    except Exception as e:
        logging.error(f"Error saving CV: {e}")
        await update.message.reply_text(
            get_translation(user_id, "error_saving_cv"),
            parse_mode="HTML"
        )
        return EDIT_PROFILE_FIELD_VALUE

async def cancel_editing_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle cancellation of editing process"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    await query.edit_message_text(
        text=f"❌ {get_translation(user_id, 'editing_cancelled')}",
        parse_mode="HTML"
    )

    return await edit_profile(update, context)


async def handle_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    # Get user data for personalization
    user_data = db.get_user_profile(user_id)

    # Check for pending applications
    pending_apps = db.get_pending_applications_count(user_id)
    approved_apps = db.get_approved_applications_count(user_id)

    # Check profile completion status
    profile_completion = calculate_profile_completion(user_data)

    # Define button texts dynamically
    profile_button_text = f"👤 {get_translation(user_id, 'profile_button')} ({profile_completion}%)"
    applications_button_text = f"✉️ {get_translation(user_id, 'applications_button')} ({pending_apps}) ({approved_apps})"
    apply_vacancy_button_text = f"✍️ {get_translation(user_id, 'apply_vacancy_button')}"
    search_vacancies_button_text = f"🔍 {get_translation(user_id, 'search_vacancies_button')}"
    help_button_text = f"ℹ️ {get_translation(user_id, 'help_button')}"
    rate_button_text = f"⭐ {get_translation(user_id, 'rate_button')}"
    report_button_text = f"📝 {get_translation(user_id, 'report_button')}"

    # Ensure text matching is correct
    proceed_button_text = get_translation(user_id, "proceed_to_main_menu")

    if choice == proceed_button_text:  # Handling "Go to Main Menu" button
        return await main_menu(update, context)

    if choice == profile_button_text:
        return await view_profile(update, context)
    elif choice == applications_button_text:
        return await view_my_applications(update, context)
    elif choice == apply_vacancy_button_text:
        return await display_vacancies(update, context)
    elif choice == search_vacancies_button_text:
        return await start_job_search(update, context)
    elif choice == help_button_text:
        return await show_help(update, context)
    elif choice == rate_button_text:
        return await show_rate_options(update, context)
    elif choice == report_button_text:
        return await handle_report(update, context)

    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_choice")
        )
        return MAIN_MENU


async def handle_proceed_to_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    # Call the main_menu function directly to show the main menu options
    return await main_menu(update, context)
import os


async def view_my_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    applications = db.get_user_applications(user_id)

    if not applications:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "no_applications_found"),
            parse_mode="HTML"
        )
        return await view_profile(update, context)

    # Group by status
    status_groups = {}
    for app in applications:
        status_groups.setdefault(app['status'], []).append(app)

    # Create message with tabs
    message = f"📬 <b>{get_translation(user_id, 'your_applications')}</b>\n\n"
    keyboard = []
    for status, apps in status_groups.items():
        localized_status = get_translation(user_id, f"application_status_{status.lower()}")
        message += f"📌 <b>{localized_status}</b> ({len(apps)})\n"
        for app in apps[:3]:  # Show max 3 per status
            message += (
                f"├─ 🏢 {app['company_name']}\n"
                f"├─ 💼 {app['job_title']}\n"
                f"├─ 📅 {get_translation(user_id, 'applied_on')}: {app['application_date'].split()[0]}\n"
                f"└─ 🔍 {get_translation(user_id, 'view_details')}\n\n"
            )
            # Add InlineKeyboardButton for each application
            keyboard.append([InlineKeyboardButton(
                f"{get_translation(user_id, 'view_details_button')} {app['job_title']}",
                callback_data=f"app_{app['application_id']}"
            )])

    # Add navigation buttons
    keyboard.append([InlineKeyboardButton(get_translation(user_id, "refresh_button"), callback_data="refresh_applications")])
    keyboard.append([InlineKeyboardButton(get_translation(user_id, "export_all_button"), callback_data="export_applications")])
    keyboard.append([InlineKeyboardButton(get_translation(user_id, "back_to_profile_button"), callback_data="back_to_profile")])

    await context.bot.send_message(
        chat_id=user_id,
        text=message,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML",
        disable_web_page_preview=True
    )
    return VIEWING_APPLICATIONS

async def show_application_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    query = update.callback_query
    await query.answer()

    app_id = int(query.data.split('_')[1])
    application = db.get_application_details(app_id)

    # Status emoji mapping
    status_emoji = {
        'pending': '🕒',
        'reviewed': '🔍',
        'approved': '✅',
        'rejected': '❌'
    }

    message = (
        f"{status_emoji.get(application['status'], '📄')} <b>{get_translation(user_id, 'application_details_title')}</b>\n\n"
        f"🏢 <b>{get_translation(user_id, 'company')}:</b> {application['company_name']}\n"
        f"💼 <b>{get_translation(user_id, 'position')}:</b> {application['job_title']}\n"
        f"📅 <b>{get_translation(user_id, 'applied_on')}:</b> {application['application_date']}\n"
        f"📌 <b>{get_translation(user_id, 'status')}:</b> {application['status'].capitalize()}\n\n"
        f"📝 <b>{get_translation(user_id, 'cover_letter')}:</b>\n{application['cover_letter'][:300]}...\n\n"
        f"📋 <b>{get_translation(user_id, 'job_description')}:</b>\n{application['description'][:300]}..."
    )

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'withdraw_application'), callback_data=f"withdraw_{app_id}")],
        [InlineKeyboardButton(get_translation(user_id, 'back_to_list'), callback_data="back_to_applications")]
    ]

    await query.edit_message_text(
        text=message,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )
    return APPLICATION_DETAILS

async def handle_application_actions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    action = query.data

    if action == "refresh_applications":
        return await view_my_applications(update, context)

    elif action == "export_applications":
        try:
            # Generate the Excel file and get the filename
            filename = await export_applications_csv(user_id)

            # Send the file with a custom caption
            await context.bot.send_document(
                chat_id=user_id,
                document=open(filename, 'rb'),
                caption=get_translation(user_id, "applications_exported"),
                filename=f"Job_Applications_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )

            # Clean up the file after sending
            os.remove(filename)

        except Exception as e:
            logging.error(f"Export failed: {e}")
            await query.edit_message_text(
                text=get_translation(user_id, "export_failed_error"),
                parse_mode="HTML"
            )
        return VIEWING_APPLICATIONS

    elif action.startswith("app_"):
        return await show_application_details(update, context)

    elif action.startswith("withdraw_"):
        app_id = int(action.split('_')[1])
        db.update_application_status(app_id, "withdrawn")
        await query.edit_message_text(
            text=get_translation(user_id, "application_withdrawn"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to List", callback_data="back_to_applications")]
            ])
        )
        return VIEWING_APPLICATIONS

    elif action == "back_to_applications":
        return await view_my_applications(update, context)

    elif action == "back_to_profile":
        return await view_profile(update, context)


import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


async def export_applications_csv(user_id: int):
    """Generate professional Excel report with formatting"""
    applications = db.get_user_applications(user_id)
    filename = f"job_applications_{user_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    status_colors = {
        'pending': 'FFF2CC',
        'approved': 'C6EFCE',
        'rejected': 'FFC7CE',
        'withdrawn': 'D9D9D9'
    }
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Headers
    headers = [
        "Application ID", "Company", "Job Title",
        "Position Type", "Applied Date", "Status",
        "Cover Letter Excerpt", "Job Description Excerpt"
    ]
    ws.append(headers)

    # Apply header styles
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Add data rows
    for app in applications:
        ws.append([
            app['application_id'],
            app['company_name'],
            app['job_title'],
            app['employment_type'],
            app['application_date'].split()[0],  # Just the date part
            app['status'].capitalize(),
            app['cover_letter'][:100] + "..." if app['cover_letter'] else "",
            app['description'][:100] + "..." if app.get('description') else ""
        ])

        # Apply row styling
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            # Status-based coloring
            if col == 6:  # Status column
                cell.fill = PatternFill(
                    start_color=status_colors.get(app['status'], 'FFFFFF'),
                    end_color=status_colors.get(app['status'], 'FFFFFF'),
                    fill_type="solid"
                )

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(applications) + 1}"

    # Adjust column widths
    column_widths = [12, 20, 25, 15, 12, 12, 40, 40]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add freeze panes and conditional formatting
    ws.freeze_panes = "A2"

    # Add summary sheet
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.append(["Application Status", "Count"])

    status_counts = {}
    for app in applications:
        status_counts[app['status']] = status_counts.get(app['status'], 0) + 1

    for status, count in status_counts.items():
        summary_sheet.append([status.capitalize(), count])

    # Add chart to summary sheet
    from openpyxl.chart import PieChart, Reference
    chart = PieChart()
    labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=len(status_counts) + 1)
    data = Reference(summary_sheet, min_col=2, min_row=1, max_row=len(status_counts) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Application Status Distribution"
    summary_sheet.add_chart(chart, "D2")

    # Save the file
    wb.save(filename)
    return filename

async def handle_export_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = get_user_id(update)
    query = update.callback_query
    await query.answer()

    try:
        # Generate the Excel file and get the filename
        filename = await export_applications_csv(user_id)

        # Send the file with a custom caption
        await context.bot.send_document(
            chat_id=user_id,
            document=open(filename, 'rb'),
            caption=(
                f"📊 <b>{get_translation(user_id, 'report_title')}</b>\n\n"
                f"📅 {get_translation(user_id, 'generated_on')}: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                f"📝 {get_translation(user_id, 'contains_active_applications')}: {db.get_pending_applications_count(user_id)}\n"
                f"💼 {get_translation(user_id, 'export_includes')}:\n"
                f"- {get_translation(user_id, 'application_details')}\n"
                f"- {get_translation(user_id, 'status_overview')}\n"
                f"- {get_translation(user_id, 'interactive_filters')}\n\n"
                f"🔍 {get_translation(user_id, 'open_excel_full_features')}"
            ),
            parse_mode="HTML",
            filename=f"Job_Applications_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

        # Clean up the file after sending
        os.remove(filename)

    except Exception as e:
        logging.error(f"Export failed: {e}")
        await query.edit_message_text(
            text=get_translation(user_id, "export_failed_error"),
            parse_mode="HTML"
        )



# Ensure the employer_documents directory exists
if not os.path.exists("employer_documents"):
    os.makedirs("employer_documents")
#employer registration
TOTAL_REGISTRATION_STEP = 6  # Adjust based on your actual steps


def generate_registration_progress(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> str:
    """
    Generates a visually appealing progress bar with step information
    """
    current_step = context.user_data.get('current_step', 1)
    percentage = min(100, int((current_step / TOTAL_REGISTRATION_STEP) * 100))
    bar_width = 15  # Width of the progress bar

    # Create the progress bar
    filled = '█' * int(bar_width * (percentage / 100))
    empty = '░' * (bar_width - len(filled))

    return (f"\n🚀 *{get_translation(user_id, 'registration_progress')}*\n"
            f"`[{filled}{empty}]` {percentage}%\n"
            f"▸ {get_translation(user_id, 'step')} {current_step} {get_translation(user_id, 'of')} {TOTAL_REGISTRATION_STEP}\n")


async def employer_name_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    await update.callback_query.answer()

    # Initialize registration progress
    context.user_data.clear()
    context.user_data['current_step'] = 1

    # Build the welcome message
    welcome_msg = (
        f"✨ *{get_translation(user_id, 'employer_welcome_message')}* ✨\n\n"
        f"{get_translation(user_id, 'employer_registration_intro')}\n"
        f"{generate_registration_progress(user_id, context)}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_msg,
        parse_mode=ParseMode.MARKDOWN
    )

    # Name prompt with formatting
    name_prompt = (
        f"📛 *{get_translation(user_id, 'employer_name_prompt')}*\n\n"
        f"▸ {get_translation(user_id, 'example')}: `TechSolutions Inc.`\n"
        f"▸ {get_translation(user_id, 'requirements')}: "
        f"{get_translation(user_id, 'company_name_requirements')}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=name_prompt,
        parse_mode=ParseMode.MARKDOWN
    )
    return EMPLOYER_NAME


async def save_employer_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    company_name = update.message.text.strip()

    if not company_name or len(company_name) < 3:
        await update.message.reply_text(
            f"❌ {get_translation(user_id, 'invalid_company_name')}\n\n"
            f"▸ {get_translation(user_id, 'minimum_length')}: 3 {get_translation(user_id, 'characters')}\n"
            f"▸ {get_translation(user_id, 'try_again_prompt')}",
            parse_mode=ParseMode.MARKDOWN
        )
        return EMPLOYER_NAME

    # Save data and update progress
    context.user_data["company_name"] = company_name
    context.user_data['current_step'] = 2

    db.cursor.execute("""
        INSERT INTO employers (employer_id, company_name) 
        VALUES (?, ?)
    """, (user_id, company_name))
    db.connection.commit()

    # Build confirmation message
    confirmation_msg = (
        f"✅ *{get_translation(user_id, 'company_name_saved')}*\n\n"
        f"{generate_registration_progress(user_id, context)}\n"
        f"📍 *{get_translation(user_id, 'employer_location_prompt')}*\n\n"
        f"▸ {get_translation(user_id, 'requirements')}: "
        f"{get_translation(user_id, 'location_requirements')}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=confirmation_msg,
        parse_mode=ParseMode.MARKDOWN
    )
    return EMPLOYER_LOCATION


async def save_employer_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    location = update.message.text.strip()

    if not location or len(location) < 3:
        await update.message.reply_text(
            f"❌ {get_translation(user_id, 'invalid_location')}\n\n"
            f"▸ {get_translation(user_id, 'minimum_length')}: 3 {get_translation(user_id, 'characters')}\n"
            f"▸ {get_translation(user_id, 'try_again_prompt')}",
            parse_mode=ParseMode.MARKDOWN
        )
        return EMPLOYER_LOCATION

    # Save data and update progress
    context.user_data["location"] = location
    context.user_data['current_step'] = 3

    db.cursor.execute("""
        UPDATE employers
        SET city = ?
        WHERE employer_id = ?
    """, (location, user_id))
    db.connection.commit()

    # Build type selection message
    keyboard = [
        [InlineKeyboardButton(f"🏢 {get_translation(user_id, 'company')}", callback_data="company")],
        [InlineKeyboardButton(f"👔 {get_translation(user_id, 'private_client')}", callback_data="private_client")],
        [InlineKeyboardButton(f"👤 {get_translation(user_id, 'individual')}", callback_data="individual")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    type_prompt = (
        f"✅ *{get_translation(user_id, 'location_saved')}*\n\n"
        f"{generate_registration_progress(user_id, context)}\n"
        f"🏷 *{get_translation(user_id, 'employer_type_prompt')}*\n\n"
        f"{get_translation(user_id, 'employer_type_explanation')}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=type_prompt,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    return EMPLOYER_TYPE


async def save_employer_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    employer_type = query.data

    # Validate the employer type
    valid_types = ["company", "private_client", "individual"]
    if employer_type not in valid_types:
        await query.edit_message_text(
            text=f"❌ {get_translation(user_id, 'invalid_employer_type')}",
            parse_mode=ParseMode.MARKDOWN
        )
        return EMPLOYER_TYPE

    # Save data and update progress
    context.user_data["employer_type"] = employer_type
    context.user_data['current_step'] = 4

    # Update database
    db.cursor.execute("""
        UPDATE employers
        SET employer_type = ?
        WHERE employer_id = ?
    """, (employer_type, user_id))
    db.connection.commit()

    # Prepare about company prompt with skip option
    keyboard = [[InlineKeyboardButton(f"⏩ {get_translation(user_id, 'skip')}", callback_data="skip")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    about_prompt = (
        f"✅ *{get_translation(user_id, 'employer_type_saved')}: "
        f"{get_translation(user_id, employer_type)}*\n\n"
        f"{generate_registration_progress(user_id, context)}\n"
        f"📝 *{get_translation(user_id, 'about_company_prompt')}*\n\n"
        f"▸ {get_translation(user_id, 'example')}: `We are a tech startup specializing in AI solutions...`\n"
        f"▸ {get_translation(user_id, 'recommendation')}: "
        f"{get_translation(user_id, 'about_company_recommendation')}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=about_prompt,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    return ABOUT_COMPANY


async def save_about_company(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    if update.callback_query and update.callback_query.data == "skip":
        await update.callback_query.answer()
        context.user_data["about_company"] = None
        await update.callback_query.edit_message_text(
            text=f"⏭ {get_translation(user_id, 'about_company_skipped')}",
            parse_mode=ParseMode.MARKDOWN
        )
    else:
        about_company = update.message.text.strip()
        if not about_company or len(about_company) < 10:
            await update.message.reply_text(
                f"❌ {get_translation(user_id, 'invalid_about_company')}\n\n"
                f"▸ {get_translation(user_id, 'minimum_length')}: 10 {get_translation(user_id, 'characters')}\n"
                f"▸ {get_translation(user_id, 'try_again_prompt')}",
                parse_mode=ParseMode.MARKDOWN
            )
            return ABOUT_COMPANY
        context.user_data["about_company"] = about_company

    # Update progress
    context.user_data['current_step'] = 5

    # Prepare documents prompt
    keyboard = [[InlineKeyboardButton(f"⏩ {get_translation(user_id, 'skip')}", callback_data="skip")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    docs_prompt = (
        f"✅ *{get_translation(user_id, 'about_company_saved')}*\n\n"
        f"{generate_registration_progress(user_id, context)}\n"
        f"📎 *{get_translation(user_id, 'verification_documents_prompt')}*\n\n"
        f"▸ {get_translation(user_id, 'accepted_formats')}: PDF, DOC/DOCX, PPT/PPTX\n"
        f"▸ {get_translation(user_id, 'verification_documents_benefits')}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=docs_prompt,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    return VERIFICATION_DOCUMENTS


async def upload_verification_documents(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    if update.callback_query and update.callback_query.data == "skip":
        await update.callback_query.answer()
        context.user_data["verification_docs"] = None
        await update.callback_query.edit_message_text(
            text=f"⏭ {get_translation(user_id, 'verification_documents_skipped')}",
            parse_mode=ParseMode.MARKDOWN
        )
        return await finalize_employer_registration(update, context)

    elif update.message and update.message.document:
        document = update.message.document
        file_ext = os.path.splitext(document.file_name)[1].lower() if document.file_name else ''

        if file_ext not in ['.pdf', '.jpg', '.png', '.jpeg', '.doc', '.docx', '.ppt', '.pptx']:
            await update.message.reply_text(
                f"❌ {get_translation(user_id, 'invalid_file_format')}\n\n"
                f"▸ {get_translation(user_id, 'accepted_formats')}: PDF, DOC/DOCX, PPT/PPTX",
                parse_mode=ParseMode.MARKDOWN
            )
            return VERIFICATION_DOCUMENTS

        try:
            context.user_data["verification_docs"] = document.file_id
            await update.message.reply_text(
                f"✅ {get_translation(user_id, 'document_uploaded_successfully')}",
                parse_mode=ParseMode.MARKDOWN
            )
            return await finalize_employer_registration(update, context)
        except Exception as e:
            print(f"Error uploading document: {e}")
            await update.message.reply_text(
                f"❌ {get_translation(user_id, 'document_upload_failed')}",
                parse_mode=ParseMode.MARKDOWN
            )
            return VERIFICATION_DOCUMENTS

    else:
        keyboard = [[InlineKeyboardButton(f"⏩ {get_translation(user_id, 'skip')}", callback_data="skip")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await context.bot.send_message(
            chat_id=user_id,
            text=f"📎 {get_translation(user_id, 'verification_documents_prompt')}",
            reply_markup=reply_markup,
            parse_mode=ParseMode.MARKDOWN
        )
        return VERIFICATION_DOCUMENTS

async def get_employer_file_from_telegram(employer_file_id, context: ContextTypes.DEFAULT_TYPE):
    try:
        # Use the employer_file_id to get the file object
        file = await context.bot.get_file(employer_file_id)
        # Download the file to a specific path or process it as needed
        file_path = await file.download_to_drive("downloaded_employer_file.pdf")
        return file_path
    except Exception as e:
        print(f"Error retrieving employer file: {e}")
        return None


async def finalize_employer_registration(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    context.user_data['current_step'] = 6

    # Extract all collected data
    profile_data = {
        "user_id": user_id,
        "company_name": context.user_data.get("company_name"),
        "location": context.user_data.get("location"),
        "employer_type": context.user_data.get("employer_type"),
        "about_company": context.user_data.get("about_company"),
        "verification_docs": context.user_data.get("verification_docs")
    }

    try:
        # Save to database
        employer_id = db.save_employer_profile(**profile_data)
        context.user_data["employer_id"] = employer_id

        # Prepare completion message
        completion_msg = (
            f"🎉 *{get_translation(user_id, 'employer_registration_complete')}* 🎉\n\n"
            f"{generate_registration_progress(user_id, context)}\n"
            f"*{get_translation(user_id, 'registration_completion_message')}*\n\n"
            f"▸ {get_translation(user_id, 'company_name')}: `{profile_data['company_name']}`\n"
            f"▸ {get_translation(user_id, 'location')}: `{profile_data['location']}`\n"
            f"▸ {get_translation(user_id, 'employer_type')}: "
            f"{get_translation(user_id, profile_data['employer_type'])}"
        )

        # Send completion message
        await context.bot.send_message(
            chat_id=user_id,
            text=completion_msg,
            parse_mode=ParseMode.MARKDOWN
        )

        # Add main menu button
        keyboard = [[KeyboardButton(get_translation(user_id, "proceed_to_employer_main_menu"))]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "proceed_to_employer_main_menu_prompt"),
            reply_markup=reply_markup
        )

        return EMPLOYER_MAIN_MENU

    except Exception as e:
        print(f"Error finalizing registration: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=f"❌ {get_translation(user_id, 'registration_failed')}",
            parse_mode=ParseMode.MARKDOWN
        )
        return ConversationHandler.END


import logging
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler()]
)


def escape_markdown(text):
    if text is None:
        return ""
    text = str(text)
    # Fixed regex pattern (handles =, -, and other reserved chars)
    markdown_reserved_chars = r"([_*[\]()~`>#+=|!.\\-])"
    return re.sub(markdown_reserved_chars, r"\\\1", text)


async def display_employer_profile(user_id, context):
    try:
        # First try to get employer profile directly by user_id
        employer_profile = db.get_employer_profile_by_user_id(user_id)
        if not employer_profile:
            logging.error(f"No employer profile found for user_id={user_id}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "employer_not_found_error")
            )
            return

        employer_id = employer_profile.get("employer_id")
        context.user_data["employer_id"] = employer_id

        logging.debug(f"Retrieved employer_id={employer_id} for user_id={user_id}")
        # Get employer profile from database
        employer_profile = db.get_employer_profile(employer_id)
        if not employer_profile:
            logging.error(f"No employer profile found in DB for employer_id={employer_id}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "employer_profile_not_found")
            )
            return

        logging.debug(f"Retrieved employer_profile={employer_profile} for employer_id={employer_id}")

        # Extract fields with fallbacks and escape Markdown special characters
        company_name = escape_markdown(employer_profile.get("company_name")) or get_translation(user_id, "not_provided")
        city = escape_markdown(employer_profile.get("city")) or get_translation(user_id, "not_provided")
        contact_number = escape_markdown(employer_profile.get("contact_number")) or get_translation(user_id, "not_provided")
        employer_type = escape_markdown(employer_profile.get("employer_type")) or get_translation(user_id, "not_provided")
        about_company = escape_markdown(employer_profile.get("about_company")) or get_translation(user_id, "not_provided")
        verification_docs = employer_profile.get("verification_docs")  # File ID

        separator = "────────────────────"

        # Format profile message with professional style
        profile_message = (
            f" *{escape_markdown(get_translation(user_id, 'employer_profile_header'))}*\n"
            f"{separator}\n"
            f"📛 *{escape_markdown(get_translation(user_id, 'company_name'))}:* `{company_name}`\n"
            f"📍 *{escape_markdown(get_translation(user_id, 'location'))}:* `{city}`\n"
            f"📞 *{escape_markdown(get_translation(user_id, 'contact_number'))}:* `{contact_number}`\n"
            f"💼 *{escape_markdown(get_translation(user_id, 'employer_type'))}:* `{employer_type}`\n"
            f"{separator}\n"
            f"📝 *{escape_markdown(get_translation(user_id, 'about_company'))}:*\n"
            f"```{about_company}```\n"
            f"{separator}\n"
        )

        # Add verification document status
        if verification_docs:
            profile_message += (
                f"📄 *{escape_markdown(get_translation(user_id, 'verification_documents'))}:* "
                f"{escape_markdown(get_translation(user_id, 'document_available'))}\n"
            )
        else:
            profile_message += (
                f"⚠️ *{escape_markdown(get_translation(user_id, 'verification_documents'))}:* "
                f"{escape_markdown(get_translation(user_id, 'not_provided'))}\n"
            )
        profile_message += separator

        # Fallback for empty profiles
        if all(value == get_translation(user_id, "not_provided") for value in
               [company_name, city, contact_number, employer_type, about_company]):
            profile_message = (
                f"🏢 *{escape_markdown(get_translation(user_id, 'employer_profile_header'))}*\n"
                f"{separator}\n"
                f"⚠️ {escape_markdown(get_translation(user_id, 'no_employer_info_available'))}\n"
                f"{separator}\n"
            )

        # Send profile message
        await context.bot.send_message(
            chat_id=user_id,
            text=profile_message,
            parse_mode="MarkdownV2",  # Use MarkdownV2 for stricter parsing
            disable_web_page_preview=True
        )

        # Send verification document if available
        if verification_docs:
            try:
                await context.bot.send_document(
                    chat_id=user_id,
                    document=verification_docs,
                    caption=get_translation(user_id, "verification_document_caption")
                )
            except Exception as e:
                logging.error(f"Error sending verification document for employer_id={employer_id}: {e}")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "document_send_error")
                )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_verification_document")
            )

        # Show main menu button
        keyboard = [[KeyboardButton(get_translation(user_id, "employer_main_menu"))]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "what_next"),
            reply_markup=reply_markup
        )

    except KeyError as ke:
        logging.error(f"KeyError in display_employer_profile: {ke}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "internal_error")
        )
    except Exception as e:
        logging.error(f"Unexpected error in display_employer_profile: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "profile_retrieval_error")
        )
# In the employer flow
async def view_employer_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    # Fetch and display the employer's profile
    await display_employer_profile(user_id, context)
    return EMPLOYER_MAIN_MENU

# Employer Main Menu
async def employer_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Get employer data for personalization
    employer_data = db.get_employer_profile(user_id)
    company_name = employer_data.get('company_name', '') if employer_data else ''

    # Check for active vacancies and applications
    counts = db.get_active_vacancies_with_applications(user_id)
    active_vacancies = counts['active_vacancies']
    new_applications = counts['new_applications']

    # Check profile completion status
    profile_completion = calculate_employer_profile_completion(employer_data)

    # Modified keyboard layout as requested
    keyboard = [
        [KeyboardButton(f"🏢 {get_translation(user_id, 'employer_profile_button')} ({profile_completion}%)")],
        # Profile in its own row
        [KeyboardButton(f"📢 {get_translation(user_id, 'post_vacancy_button')}"),  # Post and manage vacancies together
         KeyboardButton(f"📊 {get_translation(user_id, 'manage_vacancies_button')} ({active_vacancies})")],
        [KeyboardButton(f"📈 {get_translation(user_id, 'view_analytics_button')}")],  # Analytics in its own row
        [KeyboardButton(f"ℹ️ {get_translation(user_id, 'help_button')}"),
         KeyboardButton(f"⭐ {get_translation(user_id, 'rate_button')}"),
         KeyboardButton(f"📝 {get_translation(user_id, 'report_button')}")]
    ]

    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    # Create personalized welcome message (keeping all enhancements)
    welcome_msg = (
        f"👔 <b>{get_translation(user_id, 'welcome_employer')}, {company_name or get_translation(user_id, 'valued_employer')}!</b>\n\n"
        f"📊 <i>{get_translation(user_id, 'profile_completion')}:</i> {profile_completion}%\n"
        f"📢 <i>{get_translation(user_id, 'active_vacancies')}:</i> {active_vacancies} {'✨' if active_vacancies > 0 else '➡️ Post new vacancies to attract candidates'}\n"
        f"📨 <i>{get_translation(user_id, 'new_applications')}:</i> {new_applications} {'🔥' if new_applications > 5 else '💤' if new_applications == 0 else '👀'}\n\n"
    )

    # Add quick action suggestions based on current stats
    if new_applications > 0:
        welcome_msg += f"💼 <b>Quick Action:</b> {get_translation(user_id, 'review_new_applications_prompt')}\n"
    elif active_vacancies == 0:
        welcome_msg += f"💼 <b>Quick Action:</b> {get_translation(user_id, 'post_first_vacancy_prompt')}\n"

    # Add employer-specific tip of the day
    employer_tip = get_employer_tip_of_the_day(user_id)
    if employer_tip:
        welcome_msg += f"\n💼 <b>{get_translation(user_id, 'employer_tip_of_day')}:</b> {employer_tip}"

    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_msg,
        parse_mode="HTML",
        reply_markup=reply_markup
    )

    return EMPLOYER_MAIN_MENU

def get_employer_tip_of_the_day(user_id: int) -> str:
    """Get a random employer tip for the user with translations."""
    tips = [
        get_translation(user_id, "employer_tip_write_effective_job_descriptions"),
        get_translation(user_id, "employer_tip_improve_employer_brand"),
        get_translation(user_id, "employer_tip_streamline_hire_process"),
        get_translation(user_id, "employer_tip_diversity_hiring"),
        get_translation(user_id, "employer_tip_candidate_experience"),
        get_translation(user_id, "employer_tip_effective_interview_techniques"),
        get_translation(user_id, "employer_tip_retention_strategies"),
        get_translation(user_id, "employer_tip_market_trends"),
        get_translation(user_id, "employer_tip_ai_in_recruiting"),
        get_translation(user_id, "employer_tip_employee_referral_programs")
    ]
    return random.choice(tips)


def calculate_employer_profile_completion(employer_data: dict) -> int:
    """Calculate employer profile completion percentage."""
    required_fields = [
        'company_name', 'city', 'contact_number',
        'employer_type', 'about_company', 'verification_docs'
    ]

    if not employer_data:
        return 0

    filled = 0
    for field in required_fields:
        value = employer_data.get(field)
        if value and str(value).lower() != "skip":
            filled += 1

    return min(100, (filled * 100) // len(required_fields))


async def display_employer_profile_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    employer_data = db.get_employer_profile(user_id)
    profile_completion = calculate_employer_profile_completion(employer_data)
    verification_status = "✅" if employer_data.get('verification_docs') else "⚠️"


    keyboard = [
        [KeyboardButton(f"🏢 {get_translation(user_id, 'employer_view_profile_button')} ({profile_completion}%)")],
        [KeyboardButton(f"✏️ {get_translation(user_id, 'employer_edit_profile_button')}"),
         KeyboardButton(f"📈 {get_translation(user_id, 'employer_profile_stats_button')}")],
        [KeyboardButton(f"🔒 {get_translation(user_id, 'verification_status_button')}"),
         KeyboardButton(f"🌐 {get_translation(user_id, 'employer_change_language_button')}")],
        [KeyboardButton(f"🗑️ {get_translation(user_id, 'delete_my_account_button')}")],
        [KeyboardButton(f"🔙 {get_translation(user_id, 'back_to_employer_main_menu')}")]
    ]

    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "employer_profile_menu_prompt"),
        reply_markup=reply_markup
    )
    return EMPLOYER_PROFILE_MENU


async def handle_employer_profile_actions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip()
    employer_data = db.get_employer_profile(user_id)

    # Button text templates
    view_profile_text = f"🏢 {get_translation(user_id, 'employer_view_profile_button')}"
    edit_profile_text = f"✏️ {get_translation(user_id, 'employer_edit_profile_button')}"
    profile_stats_text = f"📈 {get_translation(user_id, 'employer_profile_stats_button')}"
    verification_text = f"🔒 {get_translation(user_id, 'verification_status_button')}"
    delete_account_text = f"🗑️ {get_translation(user_id, 'delete_my_account_button')}"
    change_lang_text = f"🌐 {get_translation(user_id, 'employer_change_language_button')}"
    back_text = f"🔙 {get_translation(user_id, 'back_to_employer_main_menu')}"

    if view_profile_text in choice:
        await display_employer_profile(user_id, context)
        return await display_employer_profile_menu(update, context)

    elif edit_profile_text in choice:
        return await edit_employer_profile(update, context)

    elif profile_stats_text in choice:
        # Enhanced employer profile stats
        active_vacancies = db.get_active_vacancies_count(user_id)
        total_applications = db.get_total_applications_count(user_id)
        hire_rate = db.get_employer_hire_rate(user_id)
        avg_response_time = db.get_avg_response_time(user_id)

        verification_status = "✅ Verified" if employer_data.get('verification_docs') else "⚠️ Unverified"
        profile_completion = calculate_employer_profile_completion(employer_data)

        stats_msg = (
            f"📊 <b>{get_translation(user_id, 'dashboard_title_employer')}</b>\n\n"
            f"🏢 <b>{get_translation(user_id, 'company')}:</b> {employer_data.get('company_name', get_translation(user_id, 'not_provided'))}\n"
            f"📅 <b>{get_translation(user_id, 'member_since')}:</b> {db.get_member_since_date(user_id)}\n\n"

            f"🔍 <b>{get_translation(user_id, 'profile_status')}:</b>\n"
            f"   • {get_translation(user_id, 'completion')}: {profile_completion}%\n"
            f"   • {get_translation(user_id, 'verification')}: {verification_status}\n\n"

            f"📈 <b>{get_translation(user_id, 'recruitment_metrics')}:</b>\n"
            f"   • {get_translation(user_id, 'active_vacancies')}: {active_vacancies}\n"
            f"   • {get_translation(user_id, 'total_applications')}: {total_applications}\n"
            f"   • {get_translation(user_id, 'hire_rate')}: {hire_rate}%\n"
            f"   • {get_translation(user_id, 'avg_response_time')}: {avg_response_time} {get_translation(user_id, 'days')}\n\n"

            f"🏆 <b>{get_translation(user_id, 'profile_strength')}:</b>\n"
            f"{generate_profile_strength_bar(profile_completion)}\n"
            f"{get_profile_strength_tip(profile_completion, user_id)}"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text=stats_msg,
            parse_mode="HTML"
        )
        return EMPLOYER_PROFILE_MENU

    elif verification_text in choice:
        if employer_data.get('verification_docs'):
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'account_already_verified')
            )
            return await display_employer_profile_menu(update, context)
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'upload_verification_documents')
            )
            return await display_employer_profile_menu(update, context)


    elif delete_account_text in choice:
        return await confirm_delete_my_account(update, context)

    elif change_lang_text in choice:
        return await show_employer_language_selection(update, context)

    elif back_text in choice:
        return await employer_main_menu(update, context)

    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_choice")
        )
        return EMPLOYER_PROFILE_MENU


def generate_profile_strength_bar(completion: int) -> str:
    """Generate visual progress bar for profile strength"""
    filled = '█' * (completion // 10)
    empty = '░' * (10 - (completion // 10))
    return f"{filled}{empty} {completion}%"


def get_profile_strength_tip(completion: int, user_id: int) -> str:
    """Get personalized tip based on profile completion"""
    if completion == 100:
        return get_translation(user_id, "profile_complete_tip")
    elif completion >= 75:
        return get_translation(user_id, "profile_almost_complete_tip")
    elif completion >= 50:
        return get_translation(user_id, "profile_half_complete_tip")
    else:
        return get_translation(user_id, "profile_beginner_tip")

async def confirm_delete_my_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Show confirmation buttons
    keyboard = [
        [KeyboardButton(get_translation(user_id, "yes_delete_my_account_button"))],
        [KeyboardButton(get_translation(user_id, "no_keep_my_account_button"))]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "confirm_delete_my_account_prompt"),
        reply_markup=reply_markup
    )
    return CONFIRM_DELETE_MY_ACCOUNT

async def delete_my_account_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Delete the account from the database
    db.delete_employer_account(user_id)

    # Notify the user that their account has been deleted
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "account_deleted_message")
    )
    return ConversationHandler.END  # End the conversation

async def cancel_delete_my_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Notify the user that the deletion was cancelled
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "delete_account_cancelled")
    )

    # Return to the profile menu
    return await display_employer_profile_menu(update, context)


async def handle_my_delete_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip()

    if choice == get_translation(user_id, "yes_delete_my_account_button"):
        # Delete the employer's account
        db.delete_employer_account(user_id)  # Ensure this function exists in your database module
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "my_account_deleted_message")
        )
        return ConversationHandler.END  # End the conversation

    elif choice == get_translation(user_id, "no_keep_my_account_button"):
        # Notify the user that the deletion was cancelled
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "delete_my_account_cancelled")
        )
        return await display_employer_profile_menu(update, context)  # Return to the profile menu

    else:
        # Invalid choice
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_choice")
        )
        return CONFIRM_DELETE_MY_ACCOUNT

async def handle_employer_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Ensure update.message exists and has text
    if not update.message or not update.message.text:
        return await employer_main_menu(update, context)

    choice = update.message.text.strip()

    # Get employer data for personalization
    employer_data = db.get_employer_profile(user_id)
    company_name = employer_data.get('company_name', '') if employer_data else ''

    # Check for active vacancies and applications
    active_vacancies = db.get_active_vacancies_count(user_id)

    # Check profile completion status
    profile_completion = calculate_employer_profile_completion(employer_data)

    # Dynamically generate button texts
    profile_button_text = f"🏢 {get_translation(user_id, 'employer_profile_button')} ({profile_completion}%)"
    post_vacancy_button_text = f"📢 {get_translation(user_id, 'post_vacancy_button')}"
    manage_vacancies_button_text = f"📊 {get_translation(user_id, 'manage_vacancies_button')} ({active_vacancies})"
    view_analytics_button_text = f"📈 {get_translation(user_id, 'view_analytics_button')}"
    help_button_text = f"ℹ️ {get_translation(user_id, 'help_button')}"
    rate_button_text = f"⭐ {get_translation(user_id, 'rate_button')}"
    report_button_text = f"📝 {get_translation(user_id, 'report_button')}"

    # Handle user's choice
    if choice == profile_button_text:
        return await display_employer_profile_menu(update, context)
    elif choice == post_vacancy_button_text:
        return await post_vacancy_start(update, context)
    elif choice == manage_vacancies_button_text:
        return await manage_vacancies(update, context)
    elif choice == view_analytics_button_text:
        return await view_analytics(update, context)
    elif choice == rate_button_text:
        return await show_rate_options(update, context)
    elif choice == report_button_text:
        return await handle_report(update, context)
    elif choice == help_button_text:
        return await show_help(update, context)

    else:
        # Redirect to the main menu for invalid choices
        return await employer_main_menu(update, context)

async def show_employer_language_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Define the inline keyboard for language selection
    keyboard = [
        [
            InlineKeyboardButton("🇬🇧 English", callback_data="english"),
            InlineKeyboardButton("🇪🇹 አማርኛ", callback_data="amharic")
        ],
        [
            InlineKeyboardButton(" Afaan Oromoo", callback_data="oromia"),
            InlineKeyboardButton(" ትግርኛ", callback_data="tigrigna")
        ],
        [
            InlineKeyboardButton(" Qafar af", callback_data="afar"),
            InlineKeyboardButton(" Soomaali", callback_data="somalia")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send the language selection prompt
    await update.message.reply_text(
        get_translation(user_id, "employer_select_language_prompt"),
        reply_markup=reply_markup
    )
    return SELECT_EMPLOYER_LANGUAGE

async def handle_employer_language_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    selected_language = query.data

    # Update the employer's language in the database
    db.update_user_language(user_id, selected_language)

    # Notify the user about the language change
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "employer_language_updated_message")
    )

    # Return to the employer main menu
    return await employer_main_menu(update, context)

async def confirm_change_employer_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Store the selected language temporarily in context.user_data
    context.user_data["selected_language"] = query.data

    # Show confirmation buttons
    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, "yes_button"), callback_data="confirm_change_employer_language")],
        [InlineKeyboardButton(get_translation(user_id, "no_button"), callback_data="cancel_change_employer_language")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=get_translation(user_id, "confirm_change_employer_language_prompt"),
        reply_markup=reply_markup
    )
    return CONFIRM_CHANGE_EMPLOYER_LANGUAGE

async def change_employer_language_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Retrieve the selected language from context.user_data
    selected_language = context.user_data.get("selected_language", "english")

    # Update the employer's language in the database
    db.update_user_language(user_id, selected_language)

    # Notify the user about the language change
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "employer_language_updated_message")
    )

    # Return to the employer main menu
    return await employer_main_menu(update, context)

async def cancel_change_employer_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Notify the user that the language change was cancelled
    await query.edit_message_text(
        text=get_translation(user_id, "employer_change_language_cancelled")
    )

    # Return to the employer profile menu
    return await display_employer_profile_menu(update, context)


# Constants for better maintainability
EMPLOYER_EDITABLE_FIELDS = {
    "company_name": {
        "emoji": "🏢",
        "max_length": 100,
        "validation": lambda x: len(x) >= 3
    },
    "city": {
        "emoji": "📍",
        "max_length": 100,
        "validation": lambda x: len(x) >= 3
    },
    "contact_number": {
        "emoji": "📞",
        "max_length": 20,
        "validation": lambda x: x.isdigit() and len(x) >= 7
    },
    "employer_type": {
        "emoji": "🏷️",
        "max_length": 50,
        "validation": lambda x: len(x) >= 2
    },
    "about_company": {
        "emoji": "📝",
        "max_length": 2000,
        "validation": lambda x: len(x) >= 50
    },
    "verification_docs": {
        "emoji": "📄",
        "file_types": ['.pdf', '.jpg', '.png', '.jpeg', '.jpeg', '.doc', '.docx', '.ppt', '.pptx'],
        "max_size": 10 * 1024 * 1024  # 10MB
    }
}


# Enhanced Edit Profile Menu
async def edit_employer_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    employer_data = db.get_employer_profile(user_id)
    profile_completion = calculate_employer_profile_completion(employer_data)

    # Create a grid layout with emojis and completion indicators
    keyboard = []
    row = []
    for i, (field, config) in enumerate(EMPLOYER_EDITABLE_FIELDS.items()):
        # Check if field is completed
        is_completed = employer_data and employer_data.get(field) and str(employer_data.get(field)).lower() != "skip"
        completion_icon = "✅" if is_completed else "🟡"

        row.append(InlineKeyboardButton(
            f"{config['emoji']} {get_translation(user_id, f'edit_{field}')} {completion_icon}",
            callback_data=f"edit_{field}"
        ))

        # Create new row every 2 buttons
        if (i + 1) % 2 == 0:
            keyboard.append(row)
            row = []

    # Add remaining buttons and navigation
    if row:
        keyboard.append(row)

    keyboard.extend([
        [InlineKeyboardButton(
            f"📊 {get_translation(user_id, 'view_profile_completion')} ({profile_completion}%)",
            callback_data="view_completion"
        )],
        [InlineKeyboardButton(
            f"🔙 {get_translation(user_id, 'back_to_main_menu')}",
            callback_data="back_to_employer_main_menu"
        )]
    ])

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Enhanced message with profile summary
    message = (
        f"✨ <b>{get_translation(user_id, 'edit_employer_profile_title')}</b> ✨\n\n"
        f"📌 <i>{get_translation(user_id, 'profile_completion')}:</i> <b>{profile_completion}%</b>\n"
        f"{generate_profile_strength_bar(profile_completion)}\n\n"
        f"ℹ️ {get_translation(user_id, 'edit_profile_instructions')}\n\n"
        f"{get_profile_edit_tip(user_id, profile_completion)}"
    )

    # Edit existing message if possible, otherwise send new
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )

    return EDIT_EMPLOYER_PROFILE


def get_profile_edit_tip(user_id: int, completion: int) -> str:
    """Get contextual tip for profile editing"""
    if completion < 30:
        return f"💡 <i>{get_translation(user_id, 'profile_tip_low_completion')}</i>"
    elif completion < 70:
        return f"💡 <i>{get_translation(user_id, 'profile_tip_medium_completion')}</i>"
    else:
        return f"💡 <i>{get_translation(user_id, 'profile_tip_high_completion')}</i>"


# Field Selection Handler with Enhanced UX
async def handle_edit_employer_field(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    callback_data = query.data

    # Handle navigation cases
    if callback_data == "back_to_employer_main_menu":
        await query.edit_message_text(text=get_translation(user_id, "returning_to_main_menu"))
        return EMPLOYER_MAIN_MENU
    elif callback_data == "view_completion":
        return await show_profile_completion_details(update, context)

    # Extract field name
    if not callback_data.startswith("edit_"):
        await query.edit_message_text(text=get_translation(user_id, "invalid_field_selected"))
        return EDIT_EMPLOYER_PROFILE

    field_name = callback_data.replace("edit_", "")
    if field_name not in EMPLOYER_EDITABLE_FIELDS:
        await query.edit_message_text(text=get_translation(user_id, "invalid_field_selected"))
        return EDIT_EMPLOYER_PROFILE

    # Store field in context
    context.user_data["editing_field"] = field_name
    field_config = EMPLOYER_EDITABLE_FIELDS[field_name]

    # Prepare specialized prompts
    if field_name == "verification_docs":
        prompt = (
            f"📄 <b>{get_translation(user_id, 'upload_verification_docs_title')}</b>\n\n"
            f"ℹ️ {get_translation(user_id, 'upload_verification_docs_instructions')}\n"
            f"▸ {get_translation(user_id, 'accepted_formats')}: PDF, JPG, PNG\n"
            f"▸ {get_translation(user_id, 'max_size')}: 10MB\n\n"
            f"⬇️ {get_translation(user_id, 'upload_now_prompt')}"
        )
    else:
        current_value = ""
        employer_data = db.get_employer_profile(user_id)
        if employer_data and field_name in employer_data:
            current_value = employer_data[field_name]
            if current_value and str(current_value).lower() != "skip":
                current_value = f"\n\n📌 <b>{get_translation(user_id, 'current_value')}:</b>\n<code>{current_value[:100]}{'...' if len(current_value) > 100 else ''}</code>"

        prompt = (
            f"{field_config['emoji']} <b>{get_translation(user_id, f'edit_{field_name}_title')}</b>\n\n"
            f"✏️ {get_translation(user_id, f'edit_{field_name}_instructions')}\n"
            f"▸ {get_translation(user_id, 'max_length')}: {field_config['max_length']} {get_translation(user_id, 'characters')}"
            f"{current_value}\n\n"
            f"⬇️ {get_translation(user_id, 'enter_new_value_prompt')}"
        )

    # Send the prompt with a cancel button
    keyboard = [
        [InlineKeyboardButton(
            f"❌ {get_translation(user_id, 'cancel_editing')}",
            callback_data="cancel_editing"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=prompt,
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    return EDIT_EMPLOYER_FIELD_VALUE


# Enhanced Field Saving with Validation
async def save_updated_employer_field(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    field_name = context.user_data.get("editing_field")

    if not field_name:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'session_expired_start_again'),
            parse_mode="HTML"
        )
        return await edit_employer_profile(update, context)

    field_config = EMPLOYER_EDITABLE_FIELDS.get(field_name, {})

    try:
        # Handle document upload
        if field_name == "verification_docs":
            if not update.message or not update.message.document:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=f"❌ {get_translation(user_id, 'no_document_uploaded')}",
                    parse_mode="HTML"
                )
                return EDIT_EMPLOYER_FIELD_VALUE

            document = update.message.document
            file_ext = os.path.splitext(document.file_name)[1].lower() if document.file_name else ''

            # Validate file
            if file_ext not in field_config.get("file_types", []):
                await update.message.reply_text(
                    f"❌ {get_translation(user_id, 'invalid_file_format')}\n\n"
                    f"📌 {get_translation(user_id, 'accepted_formats')}: {', '.join(field_config.get('file_types', []))}\n"
                    f"💾 {get_translation(user_id, 'please_try_again')}",
                    parse_mode="HTML"
                )
                return EDIT_EMPLOYER_FIELD_VALUE

            if document.file_size > field_config.get("max_size", 0):
                await update.message.reply_text(
                    f"❌ {get_translation(user_id, 'document_too_large')}\n"
                    f"📏 {get_translation(user_id, 'max_size')}: {field_config.get('max_size', 0) // (1024 * 1024)}MB",
                    parse_mode="HTML"
                )
                return EDIT_EMPLOYER_FIELD_VALUE

            new_value = document.file_id
            success_message = f"✅ {get_translation(user_id, 'verification_docs_updated')}"

        # Handle text fields
        else:
            new_value = update.message.text.strip()

            # Validate text input
            if not field_config.get("validation", lambda x: True)(new_value):
                await context.bot.send_message(
                    chat_id=user_id,
                    text=(
                        f"❌ {get_translation(user_id, 'invalid_input_for_field')}\n\n"
                        f"ℹ️ {get_translation(user_id, f'edit_{field_name}_requirements')}\n"
                        f"▸ {get_translation(user_id, 'current_length')}: {len(new_value)}\n"
                        f"▸ {get_translation(user_id, 'required_length')}: 3-{field_config.get('max_length', 100)}"
                    ),
                    parse_mode="HTML"
                )
                return EDIT_EMPLOYER_FIELD_VALUE

            if len(new_value) > field_config.get("max_length", 100):
                await context.bot.send_message(
                    chat_id=user_id,
                    text=(
                        f"❌ {get_translation(user_id, 'input_too_long')}\n"
                        f"📏 {get_translation(user_id, 'max_length')}: {field_config.get('max_length', 100)} {get_translation(user_id, 'characters')}\n"
                        f"✂️ {get_translation(user_id, 'please_shorten')}"
                    ),
                    parse_mode="HTML"
                )
                return EDIT_EMPLOYER_FIELD_VALUE

            success_message = f"✅ {get_translation(user_id, 'field_updated_success')}"

        # Update profile in database
        db.update_employer_profile_field(user_id, field_name, new_value)

        # Send success message with updated field preview
        preview = new_value[:50] + "..." if len(str(new_value)) > 50 else new_value
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"{success_message}\n\n"
                f"📋 <b>{get_translation(user_id, 'updated_value')}:</b>\n"
                f"<code>{preview}</code>\n\n"
                f"{get_random_positive_emoji()} {get_translation(user_id, 'what_next_prompt')}"
            ),
            parse_mode="HTML"
        )

        # Update profile completion percentage
        employer_data = db.get_employer_profile(user_id)
        new_completion = calculate_employer_profile_completion(employer_data)

        # Send completion update if significant change
        prev_completion = context.user_data.get("profile_completion", 0)
        if abs(new_completion - prev_completion) >= 10:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"📈 <b>{get_translation(user_id, 'profile_completion_updated')}</b>\n"
                    f"{generate_profile_strength_bar(new_completion)}\n"
                    f"<b>{new_completion}%</b> {get_completion_message(new_completion, user_id)}"
                ),
                parse_mode="HTML"
            )
            context.user_data["profile_completion"] = new_completion

    except Exception as e:
        logging.error(f"Error updating employer field: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"⚠️ <b>{get_translation(user_id, 'update_error_title')}</b>\n\n"
                f"{get_translation(user_id, 'update_error_message')}\n"
                f"🔧 {get_translation(user_id, 'technical_details')}: <code>{str(e)[:100]}</code>"
            ),
            parse_mode="HTML"
        )

    return await edit_employer_profile(update, context)


def get_random_positive_emoji() -> str:
    """Return a random positive emoji for success messages"""
    emojis = ["🎉", "✨", "👍", "👏", "💫", "🌟", "🏆", "🥇", "🚀"]
    return random.choice(emojis)


def get_completion_message(percentage: int, user_id: int) -> str:
    """Get a motivational message based on completion percentage"""
    if percentage < 30:
        return get_translation(user_id, "completion_message_low")
    elif percentage < 70:
        return get_translation(user_id, "completion_message_medium")
    elif percentage < 100:
        return get_translation(user_id, "completion_message_high")
    else:
        return get_translation(user_id, "completion_message_complete")


async def show_profile_completion_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show detailed profile completion breakdown"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    employer_data = db.get_employer_profile(user_id)

    if not employer_data:
        await query.edit_message_text(text=get_translation(user_id, "profile_not_found"))
        return EDIT_EMPLOYER_PROFILE

    completion = calculate_employer_profile_completion(employer_data)
    missing_fields = []
    completed_fields = []

    for field in EMPLOYER_EDITABLE_FIELDS:
        value = employer_data.get(field)
        if value and str(value).lower() != "skip":
            completed_fields.append(field)
        else:
            missing_fields.append(field)

    # Create message parts
    message_parts = [
        f"📊 <b>{get_translation(user_id, 'profile_completion_details')}</b>\n\n",
        f"{generate_profile_strength_bar(completion)}\n",
        f"<b>{completion}%</b> {get_completion_message(completion, user_id)}\n\n"
    ]

    # Helper function to format field entries
    def format_field_entry(field):
        emoji = EMPLOYER_EDITABLE_FIELDS[field].get("emoji", "")
        field_name = get_translation(user_id, "edit_" + field)
        return f"▸ {emoji} {field_name}\n"

    # Add completed fields if any
    if completed_fields:
        message_parts.append(f"✅ <b>{get_translation(user_id, 'completed_sections')}:</b>\n")
        message_parts.extend(format_field_entry(field) for field in completed_fields)
        message_parts.append("\n")

    # Add missing fields if any
    if missing_fields:
        message_parts.append(f"⚠️ <b>{get_translation(user_id, 'missing_sections')}:</b>\n")
        message_parts.extend(format_field_entry(field) for field in missing_fields)
        message_parts.append(f"\n💡 <i>{get_translation(user_id, 'completion_tip')}</i>")

    # Combine all parts
    message = "".join(message_parts)

    # Add back button
    keyboard = [
        [InlineKeyboardButton(
            get_translation(user_id, "back_to_editing"),
            callback_data="back_to_editing"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=message,
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    return PROFILE_COMPLETION_VIEW

async def cancel_editing(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle cancellation of editing process"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    await query.edit_message_text(
        text=f"❌ {get_translation(user_id, 'editing_cancelled')}",
        parse_mode="HTML"
    )

    return await edit_employer_profile(update, context)
# Admin credentials
ADMIN_USERNAME = "Arefat"
ADMIN_PASSWORD = "1234"



# (global variable)
active_admins = set()


# Admin login process with improved UI and security
async def admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    # Clear any existing session data
    context.user_data.clear()

    # Send styled login prompt
    await context.bot.send_message(
        chat_id=user_id,
        text="""🔐 <b>Admin Authentication Required</b>
━━━━━━━━━━━━━━━━
Welcome to the <i>Secure Admin Portal</i>.
Please enter your username:""",
        parse_mode=ParseMode.HTML,
        reply_markup=ReplyKeyboardRemove()
    )
    return ADMIN_LOGIN


async def check_admin_credentials(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    message_text = update.message.text.strip()

    # Initialize login attempts if not exists
    if "login_attempts" not in context.user_data:
        context.user_data["login_attempts"] = 0

    # First step - username entry
    if "username" not in context.user_data:
        # Basic validation
        if not message_text or len(message_text) < 3:
            await update.message.reply_text(
                "⚠️ Username must be at least 3 characters. Please try again:"
            )
            return ADMIN_LOGIN

        # Validate username before proceeding to password
        if message_text != ADMIN_USERNAME:
            await update.message.reply_text(
                "⚠️ Invalid username. Please enter your username again:"
            )
            return ADMIN_LOGIN

        context.user_data["username"] = message_text
        context.user_data["login_attempts"] = 0  # Reset attempts for new username

        # Password prompt with security notice
        await context.bot.send_message(
            chat_id=user_id,
            text="""🔒 <b>Password Required</b>
━━━━━━━━━━━━━━━━
Please enter your admin password:
<i>Note: Your password will not be displayed</i>""",
            parse_mode=ParseMode.HTML
        )
        return ADMIN_LOGIN

    # Second step - password verification
    else:
        password = message_text
        login_attempts = context.user_data.get("login_attempts", 0) + 1
        context.user_data["login_attempts"] = login_attempts

        # Verify password only (username already validated)
        if password == ADMIN_PASSWORD and login_attempts <= 3:
            # Successful login
            active_admins.add(user_id)

            # Send success message with login details
            login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            await context.bot.send_message(
                chat_id=user_id,
                text=f"""✅ <b>Login Successful</b>
━━━━━━━━━━━━━━━━
Welcome back, <code>{update.effective_user.full_name}</code>!
⏱️ Login time: {login_time}
🆔 User ID: <code>{user_id}</code>

⚡ <b>{context.bot.first_name} Admin Panel</b> ⚡""",
                parse_mode=ParseMode.HTML
            )
            return await show_admin_menu(update, context)
        else:
            # Failed attempt handling
            remaining_attempts = 3 - login_attempts

            if remaining_attempts <= 0:
                await context.bot.send_message(
                    chat_id=user_id,
                    text="🚫 Maximum attempts reached. Session locked.",
                    parse_mode=ParseMode.HTML
                )
                context.user_data.clear()
                return ConversationHandler.END

            # Prompt to try again
            await context.bot.send_message(
                chat_id=user_id,
                text=f"""⚠️ Access denied. {remaining_attempts} attempt(s) remaining.
━━━━━━━━━━━━━━━━
Please enter your password again:""",
                parse_mode=ParseMode.HTML
            )
            return ADMIN_LOGIN
def get_all_admins() -> list:
    """Get all currently active admin user IDs."""
    return list(active_admins)

# Show the admin menu
ADMIN_MENU_OPTIONS = [
    "📝 Manage Job Posts",
    "📤 Share Job Posts",
    "🔧 User Interactions",
    "📢 Broadcast Message",
    "📂 Database Tools",
    "⚙️ System Configurations",

    "🔙 Exit Admin Panel"
]


async def show_admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    welcome_text = (
            f"Select an option..."
    )

    # Create keyboard using our constants
    keyboard = [
        [ADMIN_MENU_OPTIONS[0], ADMIN_MENU_OPTIONS[1]],  # First two options
        [ADMIN_MENU_OPTIONS[2], ADMIN_MENU_OPTIONS[3]],  # Next two
        [ADMIN_MENU_OPTIONS[4], ADMIN_MENU_OPTIONS[5]],
        [ADMIN_MENU_OPTIONS[6]]  # Exit button
    ]

    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        one_time_keyboard=True,
        resize_keyboard=True,
        input_field_placeholder="Select an option..."
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_text,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    return ADMIN_MAIN_MENU


# And update your handler to use the same constants
async def handle_admin_menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    choice = update.message.text

    if choice == ADMIN_MENU_OPTIONS[0]:  # Manage Job Posts
        await manage_job_posts(update, context)
        return ADMIN_MAIN_MENU
    elif choice == ADMIN_MENU_OPTIONS[1]:  # Share Job Posts
        await handle_share_job_posts(update, context)
        return SHARE_JOBS_NAVIGATION
    elif choice == ADMIN_MENU_OPTIONS[2]:  # User Interaction
        return await show_user_interactions_menu(update, context)
    # elif choice == ADMIN_MENU_OPTIONS[3]:  # Analytics
    #     await show_analytics_dashboard(update, context)
    #     return ADMIN_MAIN_MENU
    elif choice == ADMIN_MENU_OPTIONS[3]:  # Broadcast
        await handle_broadcast_choice(update, context)
        return BROADCAST_TYPE
    elif choice == ADMIN_MENU_OPTIONS[4]:  # Database Tools
        return await show_database_menu(update, context)
    elif choice == ADMIN_MENU_OPTIONS[5]:  # Bot Settings
        return await show_system_configurations_menu(update, context)
    # elif choice == ADMIN_MENU_OPTIONS[7]:  # Maintenance
    #     return await show_maintenance_menu(update, context)
    elif choice == ADMIN_MENU_OPTIONS[6]:  # Exit
        await context.bot.send_message(
            chat_id=user_id,
            text="✅ Admin session completed. Returning to user mode.",
            reply_markup=ReplyKeyboardRemove()
        )
        context.user_data.clear() # Clear admin session data
        return await start(update, context)
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Invalid option selected. Please choose from the menu.",
            reply_markup=ReplyKeyboardMarkup(
                [
                    [ADMIN_MENU_OPTIONS[0], ADMIN_MENU_OPTIONS[1]],
                    [ADMIN_MENU_OPTIONS[2], ADMIN_MENU_OPTIONS[3]],
                    [ADMIN_MENU_OPTIONS[4], ADMIN_MENU_OPTIONS[5]],
                    [ADMIN_MENU_OPTIONS[6], ADMIN_MENU_OPTIONS[7]],
                    [ADMIN_MENU_OPTIONS[8]]
                ],
                one_time_keyboard=True,
                resize_keyboard=True
            )
        )
        return ADMIN_MAIN_MENU

# New sub-menu options for User Interactions
USER_INTERACTION_OPTIONS = [
    "📩 Contact Management",
    "⚠️ Violation Reports",
    "⭐ Ratings & Reviews",
    "🔙 Back to Admin"
]


# NEW FUNCTION: User Interactions Sub-Menu
async def show_user_interactions_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    reply_markup = ReplyKeyboardMarkup(
        [
            [USER_INTERACTION_OPTIONS[0], USER_INTERACTION_OPTIONS[1]],
            [USER_INTERACTION_OPTIONS[2]],
            [USER_INTERACTION_OPTIONS[3]]
        ],
        one_time_keyboard=True,
        resize_keyboard=True
    )

    await context.bot.send_message(
        chat_id=user_id,
        text="User Interaction Management:",
        reply_markup=reply_markup
    )
    return USER_INTERACTIONS_MENU


# NEW HANDLER: User Interactions Sub-Menu Choices
async def handle_user_interactions_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    choice = update.message.text

    if choice == USER_INTERACTION_OPTIONS[0]:  # Contact Management
        return await show_contact_management_dashboard(update, context)
    elif choice == USER_INTERACTION_OPTIONS[1]:  # Violation Reports
        return await violation_reports_dashboard(update, context)
    elif choice == USER_INTERACTION_OPTIONS[2]:  # Ratings & Reviews
        return await ratings_dashboard(update, context)
    elif choice == USER_INTERACTION_OPTIONS[3]:  # Back to Admin
        return await show_admin_menu(update, context)
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Invalid option selected.",
            reply_markup=ReplyKeyboardMarkup(
                [
                    [USER_INTERACTION_OPTIONS[0], USER_INTERACTION_OPTIONS[1]],
                    [USER_INTERACTION_OPTIONS[2]],
                    [USER_INTERACTION_OPTIONS[3]]
                ],
                one_time_keyboard=True,
                resize_keyboard=True
            )
        )
        return USER_INTERACTIONS_MENU

async def show_database_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show the new Database Management menu with expanded options"""
    user_id = update.effective_user.id

    keyboard = [
        [InlineKeyboardButton("👥 User Management", callback_data="manage_users"),
         InlineKeyboardButton("💼 Job Management", callback_data="manage_jobs")],

        [InlineKeyboardButton("📋 Vacancy Control", callback_data="ad_manage_vacancies"),
         InlineKeyboardButton("📨 Application Hub", callback_data="manage_applications")],

        [InlineKeyboardButton("📤 Export Tools", callback_data="export_data_menu"),
         InlineKeyboardButton("🧹 Data Cleanup", callback_data="clear_data")],
        [InlineKeyboardButton("🗑️ Table Cleanup", callback_data="table_cleanup")],
        [InlineKeyboardButton("📊 Database Stats", callback_data="db_stats"),
         InlineKeyboardButton("⚠️ Error Logs", callback_data="view_system_errors")],
        [InlineKeyboardButton("🔙 Back to Admin", callback_data="back_to_admin_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text="Database Management Menu:",
        reply_markup=reply_markup
    )
    return DATABASE_MANAGEMENT

async def manage_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show submenu for managing jobs"""
    user_id = update.effective_user.id
    keyboard = [
        [InlineKeyboardButton("List Jobs", callback_data="list_jobs")],
        [InlineKeyboardButton("Remove Jobs", callback_data="remove_jobs")],
        [InlineKeyboardButton("Export Jobs", callback_data="export_jobs")],
        [InlineKeyboardButton("Back to Database Menu", callback_data="back_to_database_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await context.bot.send_message(
        chat_id=user_id,
        text="Job Management Options:",
        reply_markup=reply_markup
    )
    return MANAGE_JOBS
async def ad_manage_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show submenu for managing vacancies (renamed to avoid conflicts)"""
    user_id = update.effective_user.id
    keyboard = [
        [InlineKeyboardButton("List Vacancies", callback_data="list_vacancies")],
        [InlineKeyboardButton("Remove Vacancies", callback_data="remove_vacancies")],
        [InlineKeyboardButton("Export Vacancies", callback_data="export_vacancies")],
        [InlineKeyboardButton("Back to Database Menu", callback_data="back_to_database_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await context.bot.send_message(
        chat_id=user_id,
        text="Vacancy Management Options:",
        reply_markup=reply_markup
    )
    return AD_MANAGE_VACANCIES
async def list_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show paginated list of all jobs"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    page = 1  # Default to first page

    try:
        jobs = db.get_all_jobs(page=page)
        total_pages = db.get_total_pages_jobs()

        if not jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text="No jobs found in the database."
            )
            return await back_to_manage_jobs(update, context)

        keyboard = create_paginated_keyboard(jobs, page, total_pages, "job")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_jobs")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"All Jobs (Page {page}/{total_pages}):",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Error listing jobs: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching jobs."
        )
        return await back_to_manage_jobs(update, context)

    return LIST_JOBS_PAGINATED
async def manage_job_posts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Send typing action for better UX
    await context.bot.send_chat_action(chat_id=user_id, action=ChatAction.TYPING)

    try:
        # Normalize job post statuses before fetching data
        db.normalize_job_post_statuses()

        # Fetch job post statistics (counts by status)
        stats = db.get_job_post_status_counts()
        stats_vac = db.get_job_post_status_counts_vaccancy()
        pending_count = stats.get('pending', 0)
        approved_count = stats_vac.get('approved', 0)
        rejected_count = stats.get('rejected', 0)

        # Prepare dashboard message
        dashboard_message = (
            f"📊 <b>Job Post Management Dashboard</b>\n"
            f"━━━━━━━━━━━━━━━━\n\n"
            f"📋 <i>Pending Review:</i> {pending_count}\n"
            f"✅ <i>Approved:</i> {approved_count}\n"
            f"❌ <i>Rejected:</i> {rejected_count}\n"
            f"━━━━━━━━━━━━━━━━\n\n"
        )

        # Check if there are any pending job posts
        if pending_count == 0:
            await context.bot.send_message(
                chat_id=user_id,
                text=dashboard_message + get_translation(user_id, "no_pending_jobs"),
                parse_mode="HTML"
            )
            await show_admin_menu(update, context)
            return ADMIN_MAIN_MENU

        # Fetch and display pending jobs
        dashboard_message += get_translation(user_id, "review_pending_jobs")
        await context.bot.send_message(
            chat_id=user_id,
            text=dashboard_message,
            parse_mode="HTML"
        )

        # Display detailed pending job posts
        await fetch_and_display_pending_jobs(user_id, context)

    except ValueError as ve:
        logging.error(f"ValueError in manage_job_posts: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_fetching_jobs", error=str(ve)),
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Unexpected error in manage_job_posts: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e)),
            parse_mode="HTML"
        )

    return ADMIN_MAIN_MENU


from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes



async def fetch_and_display_pending_jobs(user_id: int, context: ContextTypes.DEFAULT_TYPE, page: int = 1,
                                         per_page: int = 5) -> None:
    try:
        # Store pagination state in context
        context.user_data.setdefault('pagination', {'current_page': page})

        # Fetch pending job posts
        pending_jobs = db.get_pending_job_posts()

        if not pending_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text="✨ <b>No pending jobs found!</b>\nAll caught up!",
                parse_mode=ParseMode.HTML
            )
            return

        # Create pagination
        total_pages = (len(pending_jobs) + per_page - 1) // per_page
        jobs_page = pending_jobs[(page - 1) * per_page: page * per_page]

        # Update context with total pages
        context.user_data['pagination']['total_pages'] = total_pages
        context.user_data['pagination']['current_page'] = page

        # Clear previous messages (optional - you might want to keep some)
        if 'pagination_messages' in context.user_data:
            for msg_id in context.user_data['pagination_messages']:
                try:
                    await context.bot.delete_message(chat_id=user_id, message_id=msg_id)
                except:
                    pass
        context.user_data['pagination_messages'] = []

        # Send jobs with enhanced formatting
        sent_messages = []
        for job in jobs_page:
            job_text = (
                f"<b>📋 Job Title:</b> {escape_html(job['job_title']) if 'job_title' in job.keys() else 'N/A'}\n"
                f"<b>📅 Deadline:</b> {escape_html(job['deadline']) if 'deadline' in job.keys() else 'N/A'}\n"
                f"<b>📍 Status:</b> {escape_html(job['status'].capitalize() if 'status' in job.keys() else 'Pending')}\n"
                f"<b>🆔 Job ID:</b> <code>{escape_html(str(job['id'])) if 'id' in job.keys() else 'N/A'}</code>\n"
            )
            # Check if employer has verification docs
            employer_id = job['employer_id'] if 'employer_id' in job.keys() else None
            has_docs = db.has_verification_docs(employer_id) if employer_id else False
            # Create inline keyboard for job actions
            keyboard = [
                [InlineKeyboardButton("✅ Approve", callback_data=f"approve_{job['id']}")],
                [InlineKeyboardButton("❌ Reject", callback_data=f"reject_{job['id']}"),
                 InlineKeyboardButton("🔍 Preview", callback_data=f"preview_{job['id']}")]
            ]
            # Always include contact and docs buttons
            if employer_id:
                contact_button = InlineKeyboardButton("📞 Contact Employer", callback_data=f"contact_{employer_id}")
                docs_button_text = "📄 No Docs Available" if not has_docs else "📄 View Docs"
                docs_button = InlineKeyboardButton(docs_button_text, callback_data=f"docs_{employer_id}")

                keyboard.append([contact_button, docs_button])

            msg = await context.bot.send_message(
                chat_id=user_id,
                text=job_text,
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(keyboard),
                disable_web_page_preview=True
            )
            sent_messages.append(msg.message_id)

        # Add reply keyboard for pagination
        reply_markup = ReplyKeyboardMarkup(
            [
                ["⬅️ Previous", f"Page {page}/{total_pages}", "Next ➡️"],
                ["🔄 Refresh", "🔙 Back"]
            ],
            resize_keyboard=True,
            one_time_keyboard=False
        )

        msg = await context.bot.send_message(
            chat_id=user_id,
            text=f"📄 Showing page {page} of {total_pages}",
            reply_markup=reply_markup
        )
        sent_messages.append(msg.message_id)

        # Store sent message IDs for cleanup
        context.user_data['pagination_messages'] = sent_messages

    except Exception as e:
        logging.error(f"Error displaying jobs: {str(e)}", exc_info=True)


async def handle_pagination_job(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle pagination using reply keyboard."""
    message = update.message
    user_id = message.from_user.id
    text = message.text

    # Get pagination state from context
    pagination = context.user_data.get('pagination', {})
    current_page = pagination.get('current_page', 1)
    total_pages = pagination.get('total_pages', 1)

    # Calculate new page
    if text == "⬅️ Previous":
        new_page = max(1, current_page - 1)
    elif text == "Next ➡️":
        new_page = min(total_pages, current_page + 1)
    elif text == "🔄 Refresh":
        new_page = current_page
    else:
        return  # Unknown command

    # Redisplay jobs with new page
    await fetch_and_display_pending_jobs(user_id, context, page=new_page)

async def handle_job_preview(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    try:
        # Extract job_id from callback data
        job_id = int(query.data.split("_")[1])
        job = db.get_job_by_id(job_id)

        if not job:
            await query.edit_message_text(text="⚠️ Job post not found!")
            return

        # Format job details
        deadline = escape_html(job.get('application_deadline', 'N/A'))
        # Fetch employer contact link
        employer_link = ""
        employer_id = job.get("employer_id")
        if employer_id:
            username = await get_username_from_id(employer_id, context)
            if username:
                employer_link = f'<a href="https://t.me/{username}">Contact Employer</a>'
            else:
                employer_link = f'<a href="tg://user?id={employer_id}">Contact Employer</a>'

        job_details = (
            f"<b>📋 Job Title:</b> {escape_html(job.get('job_title', 'N/A'))}\n"
            f"<b>🏢 Employer:</b> {escape_html(job.get('employer_name', 'Not provided'))}\n"
            f"<b>📅 Deadline:</b> {deadline}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>📝 Description:</b>\n{escape_html(job.get('description', ''))}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>💼 Employment Type:</b> {escape_html(job.get('employment_type', 'N/A'))}\n"
            f"<b>🚻 Gender:</b> {escape_html(job.get('gender', 'N/A'))}\n"
            f"<b>👥 Quantity:</b> {escape_html(str(job.get('quantity', 'N/A')))}\n"
            f"<b>🎓 Qualification:</b> {escape_html(job.get('qualification', 'N/A'))}\n"
            f"<b>📊 Level:</b> {escape_html(job.get('level', 'N/A'))}\n"
            f"<b>🔑 Skills:</b> {escape_html(job.get('skills', 'N/A'))}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>💰 Salary:</b> {escape_html(job.get('salary', 'Negotiable'))}\n"
            f"<b>🎁 Benefits:</b> {escape_html(job.get('benefits', 'Negotiable'))}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>🆔 Job ID:</b> <code>{escape_html(str(job.get('job_id', 'N/A')))}</code>\n"
            + (f"<b>👤 Employer Contact:</b> {employer_link}" if employer_link else "")
        )

        # Check if employer has verification docs
        has_docs = db.has_verification_docs(employer_id) if employer_id else False

        # Create inline keyboard with Approve, Reject, and Go Back buttons
        keyboard = [
            [
                InlineKeyboardButton("✅ Approve", callback_data=f"approve_{job['job_id']}"),
                InlineKeyboardButton("❌ Reject", callback_data=f"reject_{job['job_id']}")
            ],
            [
                InlineKeyboardButton("📄 View Docs", callback_data=f"docs_{employer_id}"),
                InlineKeyboardButton("⬅️ Back", callback_data="back_to_pending_jobs")
            ]
        ]

        # Update message with job details and buttons
        await query.edit_message_text(
            text=job_details,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(keyboard),
            disable_web_page_preview=True
        )

    except Exception as e:
        logging.error(f"Error handling job preview: {str(e)}", exc_info=True)
        # await send_error_message(context, query.from_user.id, "job_preview_error")

async def go_back_to_pending_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id

    # Fetch and display pending jobs again
    await fetch_and_display_pending_jobs(user_id, context)


async def handle_docs_view(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    try:
        employer_id = int(query.data.split("_")[1])
        docs_info = db.get_employer_docs(employer_id)

        if not docs_info or not docs_info.get('file_id'):
            # Answer with a temporary message that doesn't affect the buttons
            await query.answer("⚠️ No verification documents available", show_alert=True)
            return

        try:
            # Send the document as a new message
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=docs_info['file_id'],
                caption=f"📄 Verification documents for employer {employer_id}"
            )
            await query.answer("Documents sent to your chat", show_alert=False)
        except BadRequest as e:
            if "There is no document in the request" in str(e):
                await query.answer("⚠️ Documents not found (invalid file reference)", show_alert=True)
            else:
                raise

    except Exception as e:
        logging.error(f"Error handling docs view: {str(e)}", exc_info=True)
        await query.answer("⚠️ Failed to retrieve documents", show_alert=True)
        
async def get_username_from_id(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> str | None:
    try:
        user = await context.bot.get_chat(user_id)
        return user.username if user.username else None
    except Exception as e:
        logging.warning(f"Could not fetch username for ID {user_id}: {e}")
        return None

async def handle_contact_employer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    try:
        parts = query.data.split("_")
        if len(parts) < 2:
            raise ValueError("Invalid callback data format")

        employer_id = int(parts[1])
        logging.info(f"Attempting to contact employer with ID: {employer_id}")

        # Optional: Fetch username if available
        username = await get_username_from_id(employer_id, context)  # Implement this
        if username:
            link = f"https://t.me/{username}"
        else:
            link = f"tg://user?id={employer_id}"

        contact_msg = (
            "📞 <b>Contact Employer</b>\n\n"
            f"You can contact the employer directly by clicking here: "
            f"<a href='{link}'>Contact Employer</a>\n\n"
            "<i>Note: This will open a chat with the employer in Telegram</i>"
        )

        await query.edit_message_text(
            text=contact_msg,
            parse_mode=ParseMode.HTML,
            disable_web_page_preview=True
        )
        logging.info(f"Successfully sent contact link for employer ID: {employer_id}")

    except Exception as e:
        logging.error(f"Error in handle_contact_employer: {str(e)}", exc_info=True)
        await query.edit_message_text(
            text="⚠️ <b>Error:</b> Unable to process your request. Please try again later.",
            parse_mode=ParseMode.HTML
        )

        # await send_error_message(context, query.from_user.id, "contact_employer_error")

# Handle the "Approve & Share" action
async def handle_admin_job_approval(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        # Log raw callback_data
        logging.debug(f"Received callback_data: {query.data}")

        # Validate callback data
        job_id = validate_callback_data(query.data, ("approve_", "reject_"))
        if not job_id:
            logging.error(f"Invalid or missing job_id in callback_data: {query.data}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "invalid_data_detected", error="Missing or invalid job_id.")
            )
            return ADMIN_MAIN_MENU

        logging.debug(f"Extracted job_id: {job_id}")

        # Fetch and validate the job post
        validated_job_post = await fetch_and_validate_job_post(job_id, user_id, context)
        if not validated_job_post:
            logging.error(f"Failed to fetch and validate job post with ID: {job_id}")
            return ADMIN_MAIN_MENU

        try:
            validate_job_post_data(validated_job_post)
        except ValueError as ve:
            logging.error(f"Validation error in job post {job_id}: {ve}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "error_validating_job", job_id=job_id, error=str(ve))
            )
            return ADMIN_MAIN_MENU

        current_status = validated_job_post.get("status", "Unknown")
        logging.info(f"Admin {user_id} is processing job {job_id} with current status: {current_status}")

        action = query.data.split("_")[0]
        if action == "approve":
            try:
                # Delete the message with the inline keyboard first
                try:
                    await query.delete_message()
                except Exception as delete_error:
                    logging.warning(f"Could not delete message: {delete_error}")

                # Get employer ID before approval
                employer_id = validated_job_post.get("employer_id")
                if not employer_id:
                    logging.error(f"No employer_id found for job {job_id}")
                    raise ValueError("No employer_id associated with this job post")

                db.approve_job_post(job_id)
                new_status = db.get_job_post_status(job_id)
                logging.info(f"Job {job_id} approved. New status: {new_status}")

                # Notify admin with a new message
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "job_approved", job_id=job_id)
                )

                # Notify employer - since employer_id is the same as user_id in your schema
                try:
                    await context.bot.send_message(
                        chat_id=employer_id,  # Directly use employer_id as it's the same as user_id
                        text=get_translation(employer_id, "your_job_approved",
                                             job_title=validated_job_post.get("job_title", "N/A"),
                                             job_id=job_id)
                    )
                except Exception as e:
                    logging.error(f"Failed to notify employer {employer_id} about approval: {e}")

            except Exception as e:
                logging.error(f"Error approving job post {job_id}: {e}")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "error_approving_job", job_id=job_id, error=str(e))
                )
                return ADMIN_MAIN_MENU

        elif action == "reject":
            try:
                try:
                    await query.delete_message()
                except Exception as delete_error:
                    logging.warning(f"Could not delete message: {delete_error}")
                # Get employer ID from the job post
                employer_id = validated_job_post.get("employer_id")
                if not employer_id:
                    logging.error(f"No employer_id found for job {job_id}")
                    raise ValueError("No employer_id associated with this job post")

                # Store necessary information in context for the rejection process
                context.user_data.update({
                    "pending_rejection": {
                        "job_id": job_id,
                        "employer_id": employer_id,
                        "job_title": validated_job_post.get("job_title", "N/A")
                    }
                })

                # Prompt the admin for a rejection reason
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "ask_rejection_reason"),
                    reply_markup=ReplyKeyboardRemove()
                )
                return REJECT_JOB_REASON  # Transition to the next state for rejection

            except Exception as e:
                logging.error(f"Error handling job post rejection {job_id}: {e}")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
                )
                return ADMIN_MAIN_MENU

    except Exception as e:
        logging.error(f"Unexpected error in handle_admin_job_approval: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return ADMIN_MAIN_MENU
#share job
def format_job_post(user_id, job, bot_username, for_sharing=False):
    """Formats the job post using HTML with premium styling."""
    # Random selection of motivational messages
    share_messages = [
        get_translation(user_id, "share_variation_1"),
        get_translation(user_id, "share_variation_2"),
        get_translation(user_id, "share_variation_3"),
        get_translation(user_id, "share_variation_4"),
        get_translation(user_id, "share_variation_5"),
        get_translation(user_id, "share_variation_6"),
        get_translation(user_id, "share_variation_7"),
        get_translation(user_id, "share_variation_8"),
        get_translation(user_id, "share_variation_9"),
        get_translation(user_id, "share_variation_10")  # Fixed duplicate
    ]

    opportunity_messages = [
        get_translation(user_id, "opportunity_variation_1"),
        get_translation(user_id, "opportunity_variation_2"),
        get_translation(user_id, "opportunity_variation_3"),
        get_translation(user_id, "opportunity_variation_4"),
        get_translation(user_id, "opportunity_variation_5"),
        get_translation(user_id, "opportunity_variation_6"),
        get_translation(user_id, "opportunity_variation_7"),
        get_translation(user_id, "opportunity_variation_8"),
        get_translation(user_id, "opportunity_variation_9"),
        get_translation(user_id, "opportunity_variation_10")
    ]

    share_text = random.choice(share_messages)
    opportunity_text = random.choice(opportunity_messages)

    # Generate a random emoji banner for visual appeal
    banner_emojis = ["✨", "🌟", "⚡", "🔥", "💎", "🚀", "💼", "👔", "🎯", "🏆"]
    banner = "".join(random.choices(banner_emojis, k=11))

    job_title = escape_html(job['job_title'])
    company_name = escape_html(job.get('company_name', 'Not provided'))
    employment_type = escape_html(job['employment_type'])
    gender = escape_html(job['gender'])
    quantity = job['quantity']
    level = escape_html(job['level'])
    full_description = escape_html(job['description'])  # Store full description
    qualification = escape_html(job['qualification'])
    skills = escape_html(job['skills'])
    salary = escape_html(job['salary'])
    benefits = escape_html(job['benefits'])
    deadline = escape_html(job['deadline'])
    status = escape_html(job['status'].capitalize())
    job_id = job['job_id']

    # --- Header and Job Title ---
    header = f"✨ <b>{job_title}</b> ✨\n"
    header += f"<i>{get_translation(user_id, 'employer')}: {company_name}</i>\n"
    header += "—" * 17 + "\n\n"

    # --- Key Details (arranged for conciseness) ---
    details = ""
    details += f"📌 <b>{get_translation(user_id, 'employment_type')}:</b> {employment_type}\n"
    details += f"👤 <b>{get_translation(user_id, 'gender')}:</b> {gender} | 👥 <b>{get_translation(user_id, 'quantity')}:</b> {quantity}\n"
    details += f"📈 <b>{get_translation(user_id, 'level')}:</b> {level}\n\n"

    # --- Description (with truncation) ---
    description_limit = 250  # Character limit for the preview description
    description_section = f"📜 <b>{get_translation(user_id, 'description')}:</b>\n"
    if len(full_description) > description_limit:
        truncated_description = full_description[:description_limit].rsplit(' ', 1)[
                                    0] + "..."  # Truncate at a word boundary
        description_section += f"{truncated_description}\n"
        description_section += f"<i>({get_translation(user_id, 'view_details_prompt')})</i>\n\n"  # "view_details_prompt" needs to be added to your translations
    else:
        description_section += f"{full_description}\n\n"

    # --- Requirements (Qualification & Skills) ---
    requirements_section = f"🎯 <b>{get_translation(user_id, 'qualification')}:</b>\n"
    requirements_section += f"• {qualification.replace('\n', '\n• ')}\n\n"
    requirements_section += f"🛠️ <b>{get_translation(user_id, 'skills')}:</b>\n"
    requirements_section += f"• {skills.replace('\n', '\n• ')}\n\n"

    # --- Salary and Benefits ---
    salary_benefits_section = f"💰 <b>{get_translation(user_id, 'salary')}:</b> {salary}\n"
    salary_benefits_section += f"🎁 <b>{get_translation(user_id, 'benefits')}:</b>\n"
    salary_benefits_section += f"• {benefits.replace('\n', '\n• ')}\n\n"

    # --- Urgency and CTA ---
    urgency_cta_section = f"⏳ <b>{get_translation(user_id, 'deadline')}:</b> {deadline}\n"
    urgency_cta_section += f"🟢 <b>{get_translation(user_id, 'status')}:</b> {status}\n\n"
    urgency_cta_section += f"{opportunity_text}\n"
    urgency_cta_section += f"👉 <a href='https://t.me/{bot_username}?start=apply_{job_id}'><b>🚀 Apply Now</b></a> 👈\n\n"

    # --- Footer ---
    footer = f"<i>{share_text}</i>\n"
    footer += f"<i>{get_translation(user_id, 'posted_via_brand').format(bot_username=bot_username)}</i>"

    # --- Combine all sections ---
    job_details = (
            header + details + description_section +
            requirements_section + salary_benefits_section + urgency_cta_section + footer
    )

    return job_details


def escape_html(text):
    """Escape special characters for Telegram HTML formatting."""
    escape_chars = {'&': '&amp;', '<': '<', '>': '>'}
    return ''.join(escape_chars.get(char, char) for char in text)


async def handle_share_job_posts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handles sharing approved job posts with reply keyboard navigation."""
    user_id = update.message.from_user.id

    try:
        db = Database()
        validated_jobs = db.fetch_approved_vacancies()

        if not validated_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_job_posts_found")
            )
            return ADMIN_MAIN_MENU

        # Store jobs and initialize navigation state
        context.user_data["share_jobs"] = {
            "all_jobs": validated_jobs,
            "current_index": 0,
            "jobs_per_page": 5  # Show 5 jobs at a time
        }

        # Send initial batch information
        batch_info = (
            f"<b>📦 Job Post Batch Ready</b>\n\n"
            f"• Total Jobs: {len(validated_jobs)}\n"
            f"• Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
            f"• Please manually forward the job posts to share them.\n"
            f"<b>━━━━━━━━━━━━━━━━━━━━━━</b>"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text=batch_info,
            parse_mode='HTML',
            reply_markup=ReplyKeyboardMarkup(
                _get_navigation_buttons(context),
                resize_keyboard=True,
                one_time_keyboard=False
            )
        )

        # Send the first batch of jobs
        await _send_current_jobs_batch(user_id, context)

    except Exception as e:
        logging.error(f"Error in handle_share_job_posts: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )

    return SHARE_JOBS_NAVIGATION # New state for handling navigation


def _get_navigation_buttons(context):
    """Returns the appropriate navigation buttons based on current position"""
    share_data = context.user_data.get("share_jobs", {})
    jobs = share_data.get("all_jobs", [])
    current_index = share_data.get("current_index", 0)
    per_page = share_data.get("jobs_per_page", 5)
    total_pages = (len(jobs) + per_page - 1) // per_page  # Calculate total pages
    current_page = (current_index // per_page) + 1  # Calculate current page

    buttons = []

    # Only show Back if not on first page
    if current_index > 0:
        buttons.append("⬅️ Back")

        # Add page indicator (non-clickable)
    buttons.append(f"{current_page}/{total_pages}")

    # Only show Forward if more jobs available
    if current_index + per_page < len(jobs):
        buttons.append("➡️ Forward")

    # Return as list of lists for ReplyKeyboardMarkup
    return [buttons, ["🏠 Main Menu"]] if buttons else [["🏠 Main Menu"]]


async def _send_current_jobs_batch(user_id, context):
    """Helper function to send current batch of jobs"""
    share_data = context.user_data["share_jobs"]
    jobs = share_data["all_jobs"]
    current_index = share_data["current_index"]
    per_page = share_data["jobs_per_page"]

    # Calculate current batch
    batch = jobs[current_index:current_index + per_page]

    # Send each job in the current batch
    for job in batch:
        job_text = format_job_post(user_id, job, context.bot.username)
        await context.bot.send_message(
            chat_id=user_id,
            text=job_text,
            parse_mode='HTML',
            reply_markup=ReplyKeyboardMarkup(
                _get_navigation_buttons(context),
                resize_keyboard=True,
                one_time_keyboard=False
            )
        )
        await asyncio.sleep(0.1)  # Small delay between sends


async def handle_share_jobs_navigation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handles navigation through job posts using reply keyboard"""
    user_id = update.message.from_user.id
    text = update.message.text

    # Ignore the page number button press
    if "/" in text:  # This is our page number indicator
        return SHARE_JOBS_NAVIGATION

    share_data = context.user_data.get("share_jobs")
    if not share_data:
        await update.message.reply_text("Session expired. Please start again.")
        return await show_admin_menu(update, context)

    per_page = share_data["jobs_per_page"]

    if text == "⬅️ Back":
        share_data["current_index"] = max(0, share_data["current_index"] - per_page)
        await _send_current_jobs_batch(user_id, context)
    elif text == "➡️ Forward":
        share_data["current_index"] = min(
            len(share_data["all_jobs"]) - per_page,
            share_data["current_index"] + per_page
        )
        await _send_current_jobs_batch(user_id, context)
    elif text == "🏠 Main Menu":
        return await show_admin_menu(update, context)

    return SHARE_JOBS_NAVIGATION

async def fetch_and_validate_job_post(job_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE) -> dict:
    try:
        # Fetch job from job_posts first
        job_post = db.get_job_post_by_id(job_id)
        if not job_post:
            # If not found, check vacancies table
            job_post = db.get_vacancy_by_id(job_id)
        if not job_post:
            raise ValueError(f"Job post with ID {job_id} not found in job_posts or vacancies.")

        # Ensure job_id and source are present
        if "job_id" not in job_post:
            raise ValueError(f"Missing job_id in fetched job post: {job_post}")
        if "source" not in job_post:
            raise ValueError(f"Missing source in fetched job post: {job_post}")

        # Validate status
        valid_statuses = {"pending", "approved", "rejected", "closed", "open"}
        job_status = job_post.get("status", "").strip().lower()
        if job_status not in valid_statuses:
            raise ValueError(f"Invalid status: '{job_status}'. Expected one of {valid_statuses}.")

        return job_post
    except ValueError as ve:
        logging.error(f"Invalid job post data: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_fetching_job", error=str(ve))
        )
        return None
    except Exception as e:
        logging.error(f"Error fetching job post {job_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return None

def validate_job_post_for_rejection(job_id: int):
    """
    Validate that the job post exists and its status is 'pending'.
    """
    job_post = db.get_job_post_by_id(job_id)
    if not job_post:
        raise ValueError(f"Job post with ID {job_id} not found.")
    if job_post["status"] != "pending":
        raise ValueError(f"Cannot reject job post with ID {job_id}. Status must be 'pending'.")
    return job_post

async def handle_rejection_reason(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    reason = update.message.text.strip()

    try:
        # Validate the rejection reason
        if not reason or len(reason) > 500:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "invalid_rejection_reason")
            )
            return REJECT_JOB_REASON

        # Retrieve the job ID from user_data
        pending_data = context.user_data.get("pending_rejection", {})
        job_id = pending_data.get("job_id")
        employer_id = pending_data.get("employer_id")
        job_title = pending_data.get("job_title")

        if not job_id:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "error_processing_rejection_no_job_id")
            )
            return ADMIN_MAIN_MENU

        # Validate the job post before rejection
        try:
            validate_job_post_for_rejection(job_id)
        except ValueError as ve:
            logging.error(f"Invalid job post for rejection: {ve}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "error_rejecting_job", job_id=job_id, error=str(ve))
            )
            return ADMIN_MAIN_MENU

        # Reject the job post with the provided reason
        db.reject_job_post(job_id, reason)
        updated_status = db.get_job_post_status(job_id)

        # Confirm the rejection to the admin
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "job_rejected", job_id=job_id, reason=reason)
        )

        # Notify employer
        if employer_id:
            try:
                await context.bot.send_message(
                    chat_id=employer_id,
                    text=get_translation(employer_id, "your_job_rejected", job_title=job_title, job_id=job_id,
                                         reason=reason)
                )
            except Exception as e:
                logging.error(f"Failed to notify employer {employer_id} about rejection: {e}")


    except Exception as e:
        logging.error(f"Unexpected error rejecting job post {job_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_rejecting_job", job_id=job_id, error=str(e))
        )
        return ADMIN_MAIN_MENU

    # Refresh the list of pending job posts
    try:
        await manage_job_posts(update, context)
    except Exception as e:
        logging.error(f"Error refreshing job posts: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_refreshing_job_posts")
        )

    return ADMIN_MAIN_MENU

# Cancel the admin session
async def cancel_admin_session(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    await context.bot.send_message(
        chat_id=user_id,
        text="Admin session canceled."
    )
    context.user_data.clear()
    return ConversationHandler.END

def update_user_data_and_proceed(context: ContextTypes.DEFAULT_TYPE, key: str, value, next_step: int):
    context.user_data[key] = value
    return next_step

def validate_quantity(quantity: str) -> bool:
    """Validate that the quantity is a positive integer."""
    try:
        value = int(quantity)
        return value > 0
    except ValueError:
        return False

from datetime import datetime

def validate_date(date_str: str) -> bool:
    """Validate that the date is in the format YYYY-MM-DD."""
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def validate_salary(salary: str) -> bool:
    """Validate that the salary is numeric or a valid range."""
    if "-" in salary:
        parts = salary.split("-")
        return len(parts) == 2 and all(part.strip().isdigit() for part in parts)
    return salary.strip().isdigit()


async def post_vacancy_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Send a welcome message with animated typing action
    await context.bot.send_chat_action(chat_id=user_id, action=ChatAction.TYPING)

    # Check if employer is verified
    employer_id = context.user_data.get("employer_id")
    if not employer_id:
        employer_id = db.get_employer_id(user_id)

    if not employer_id:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "employer_id_not_found")


        )
        return await employer_main_menu(update, context)


    # Initialize job post data with progress tracking
    context.user_data["job_post"] = {
        "employer_id": employer_id,
        "progress": {
            "current_step": 1,
            "total_steps": 12,
            "completed_steps": []
        }
    }

    # Create a visually appealing progress bar
    progress_text = generate_progress_bar(
        current=1,
        total=12,
        user_id=user_id
    )

    # Send job title prompt with formatting and progress
    await context.bot.send_message(
        chat_id=user_id,
        text=f"🎯 {get_translation(user_id, 'vacancy_post_welcome')}\n\n"
             f"{progress_text}\n\n"
             f"📌 {get_translation(user_id, 'enter_job_title_instructions')}\n"
             f"💡 _{get_translation(user_id, 'job_title_tips')}_",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    return POST_JOB_TITLE


def generate_progress_bar(current: int, total: int, user_id: int) -> str:
    """
    Generate a visual progress bar with emoji indicators
    """
    filled = '🔵'
    empty = '⚪'
    progress_length = 12  # Number of segments in the progress bar

    # Calculate filled segments
    filled_segments = round((current / total) * progress_length)
    progress_bar = (filled * filled_segments) + (empty * (progress_length - filled_segments))

    return (f"📊 {get_translation(user_id, 'progress')}:\n           {progress_bar}\n"
            f"⏳ {get_translation(user_id, 'step')} {current}/{total}")


async def post_job_title(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Validate input
    if not update.message or not update.message.text or not update.message.text.strip():
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_job_title"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_job_title")]
            ])
        )
        return POST_JOB_TITLE

    job_title = update.message.text.strip()

    # Validate length
    if len(job_title) > 100:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "job_title_too_long")
        )
        return POST_JOB_TITLE

    # Store job title and update progress
    context.user_data["job_post"]["job_title"] = job_title
    context.user_data["job_post"]["progress"]["current_step"] = 2
    context.user_data["job_post"]["progress"]["completed_steps"].append("job_title")

    # Generate confirmation message with formatting
    progress_text = generate_progress_bar(2, 12, user_id)

    await context.bot.send_message(
        chat_id=user_id,
        text=f"✅ {get_translation(user_id, 'job_title_confirmation')}\n"
             f" *{job_title}*\n\n"
             f"{progress_text}\n\n"
             f"{get_translation(user_id, 'next_step_employment_type')}",
        parse_mode="Markdown"
    )

    # Display employment type options with improved layout
    keyboard = [
        [
            InlineKeyboardButton("🏢 " + get_translation(user_id, 'full_time'), callback_data="full_time"),
            InlineKeyboardButton("🏠 " + get_translation(user_id, 'remote'), callback_data="remote")
        ],
        [
            InlineKeyboardButton("⏱ " + get_translation(user_id, 'part_time'), callback_data="part_time"),
            InlineKeyboardButton("🔀 " + get_translation(user_id, 'hybrid'), callback_data="hybrid")
        ],
        [
            InlineKeyboardButton("🖥 " + get_translation(user_id, 'freelance'), callback_data="freelance")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=f"📋 *{get_translation(user_id, 'select_employment_type_question')}*\n"
             f"ℹ️ {get_translation(user_id, 'employment_type_help')}",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    return POST_EMPLOYMENT_TYPE


async def post_employment_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    # Validate callback data
    employment_types = {
        "full_time": "🏢 Full-time",
        "part_time": "⏱ Part-time",
        "remote": "🏠 Remote",
        "hybrid": "🔀 Hybrid",
        "freelance": "🖥 Freelance"
    }

    if query.data not in employment_types:
        await query.edit_message_text(
            text=get_translation(user_id, "invalid_employment_type"),
            reply_markup=query.message.reply_markup  # Keep the same keyboard
        )
        return POST_EMPLOYMENT_TYPE

    # Store employment type and update progress
    context.user_data["job_post"]["employment_type"] = query.data
    context.user_data["job_post"]["progress"]["current_step"] = 3
    context.user_data["job_post"]["progress"]["completed_steps"].append("employment_type")

    # Generate progress text
    progress_text = generate_progress_bar(3, 12, user_id)

    # Edit the original message to show selection
    await query.edit_message_text(
        text=f"✅ {get_translation(user_id, 'employment_type_selected')}\n"
             f" *{employment_types[query.data]}*\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_gender')}*",
        parse_mode="Markdown"
    )

    # Display gender options with improved layout
    keyboard = [
        [
            InlineKeyboardButton("👨 " + get_translation(user_id, 'male'), callback_data="male"),
            InlineKeyboardButton("👩 " + get_translation(user_id, 'female'), callback_data="female")
        ],
        [
            InlineKeyboardButton("👥 " + get_translation(user_id, 'both'), callback_data="any")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=f"👤 {get_translation(user_id, 'select_gender_question')}\n"
             f"ℹ️ {get_translation(user_id, 'gender_selection_help')}",
        reply_markup=reply_markup
    )

    return POST_GENDER


async def post_gender(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    # Validate callback data
    allowed_genders = {"male", "female", "any"}
    if query.data not in allowed_genders:
        await query.message.reply_text(get_translation(user_id, "invalid_gender"))
        return POST_GENDER

    # Map to database-compatible values
    gender_mapping = {
        "male": "Male",
        "female": "Female",
        "any": "Any"  # Database expects "Any" but we'll show "Both" to users
    }

    # Store the properly capitalized gender value for database
    db_gender = gender_mapping.get(query.data, "Any")
    context.user_data["job_post"]["gender"] = db_gender

    # Debugging log
    logging.debug(f"Received gender: {query.data}, mapped to: {db_gender}")

    # For display purposes, show "Both" instead of "Any"
    display_gender = {
        "male": "👨 Male",
        "female": "👩 Female",
        "any": "👥 Both"  # Visual representation shows "Both"
    }.get(query.data, "👥 Both")

    # Confirm receipt and ask for quantity
    await query.edit_message_text(
        text=get_translation(user_id, "gender_received").format(gender=display_gender)
             + "\n\n" + get_translation(user_id, "enter_quantity")
    )
    return POST_QUANTITY


async def post_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Validate input
    if not update.message or not update.message.text.strip():
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_quantity"),


        )
        return POST_QUANTITY

    try:
        quantity = int(update.message.text.strip())
        if quantity <= 0 or quantity > 100:
            raise ValueError()
    except ValueError:
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_quantity_range"),

        )
        return POST_QUANTITY

    # Store quantity and update progress
    context.user_data["job_post"]["quantity"] = quantity
    context.user_data["job_post"]["progress"]["current_step"] = 5
    context.user_data["job_post"]["progress"]["completed_steps"].append("quantity")

    # Generate progress text
    progress_text = generate_progress_bar(5, 12, user_id)

    await update.message.reply_text(
        text=f"✅ {get_translation(user_id, 'quantity_confirmation')}\n"
             f" *{quantity}* {get_translation(user_id, 'positions')}\n\n"
             f"{progress_text}\n\n"
             f"{get_translation(user_id, 'next_step_level')}",
        parse_mode="Markdown"
    )

    # Display job level options with improved layout
    keyboard = [
        [
            InlineKeyboardButton("👶 " + get_translation(user_id, 'entry_level'), callback_data="entry_level"),
            InlineKeyboardButton("🧑 " + get_translation(user_id, 'mid_level'), callback_data="mid_level")
        ],
        [
            InlineKeyboardButton("👴 " + get_translation(user_id, 'senior_level'), callback_data="senior_level")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=f"📈 *{get_translation(user_id, 'select_level_question')}*\n"
             f"ℹ️ {get_translation(user_id, 'level_selection_help')}",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    return POST_LEVEL


async def post_level(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    # Validate callback data
    level_options = {
        "entry_level": {"emoji": "👶", "name": "Entry-level"},
        "mid_level": {"emoji": "🧑", "name": "Mid-level"},
        "senior_level": {"emoji": "👴", "name": "Senior-level"}
    }

    if query.data not in level_options:
        await query.edit_message_text(
            text=get_translation(user_id, "invalid_level"),
            reply_markup=query.message.reply_markup
        )
        return POST_LEVEL

    # Store level and update progress
    level_info = level_options[query.data]
    context.user_data["job_post"]["level"] = query.data
    context.user_data["job_post"]["progress"]["current_step"] = 6
    context.user_data["job_post"]["progress"]["completed_steps"].append("level")

    # Generate progress text
    progress_text = generate_progress_bar(6, 12, user_id)

    # Edit the original message to show selection
    await query.edit_message_text(
        text=f"✅ {get_translation(user_id, 'level_selected')}\n"
             f"{level_info['emoji']} *{level_info['name']}*\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_description')}*",
        parse_mode="Markdown"
    )

    # Ask for description with formatting tips
    await context.bot.send_message(
        chat_id=user_id,
        text=f"📝 *{get_translation(user_id, 'enter_description_question')}*\n\n"
             f"✨ {get_translation(user_id, 'description_formatting_tips')}:\n"
             f"• {get_translation(user_id, 'description_tip_responsibilities')}\n"
             f"• {get_translation(user_id, 'description_tip_requirements')}\n"
             f"• {get_translation(user_id, 'description_tip_environment')}\n\n"
             f"ℹ️ {get_translation(user_id, 'description_character_limit')}",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    return POST_DESCRIPTION


async def post_description(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Validate input
    if not update.message or not update.message.text.strip():
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_description"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_description")]
            ])
        )
        return POST_DESCRIPTION

    description = update.message.text.strip()

    # Validate length
    if len(description) > 2000:
        await update.message.reply_text(
            text=get_translation(user_id, "description_too_long"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_description")]
            ])
        )
        return POST_DESCRIPTION

    # Store description and update progress
    context.user_data["job_post"]["description"] = description
    context.user_data["job_post"]["progress"]["current_step"] = 7
    context.user_data["job_post"]["progress"]["completed_steps"].append("description")

    # Generate progress text
    progress_text = generate_progress_bar(7, 12, user_id)

    # Create a preview of the first few lines
    preview_lines = description.split('\n')[:3]
    preview = '\n'.join(preview_lines) + ('...' if len(description) > 100 else '')

    await update.message.reply_text(
        text=f"✅ {get_translation(user_id, 'description_received')}\n"
             f"📝 *Preview*:\n{preview}\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_qualification')}*",
        parse_mode="Markdown"
    )

    # Display qualification options with improved layout
    keyboard = [
        [InlineKeyboardButton("🎓 " + get_translation(user_id, 'training_certificate'), callback_data="training")],
        [InlineKeyboardButton("🎓 " + get_translation(user_id, 'degree'), callback_data="degree")],
        [InlineKeyboardButton("🎓 " + get_translation(user_id, 'ma'), callback_data="ma")],
        [InlineKeyboardButton("🎓 " + get_translation(user_id, 'phd'), callback_data="phd")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=f"📚 *{get_translation(user_id, 'select_qualification_question')}*\n"
             f"ℹ️ {get_translation(user_id, 'qualification_selection_help')}",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    return POST_QUALIFICATION


async def post_qualification(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    # Validate callback data
    qualification_options = {
        "training": {"emoji": "🎓", "name": "Training/Certificate"},
        "degree": {"emoji": "🎓", "name": "Bachelor's Degree"},
        "ma": {"emoji": "🎓", "name": "Master's Degree"},
        "phd": {"emoji": "🎓", "name": "PhD"}
    }

    if query.data not in qualification_options:
        await query.edit_message_text(
            text=get_translation(user_id, "invalid_qualification"),
            reply_markup=query.message.reply_markup
        )
        return POST_QUALIFICATION

    # Store qualification and update progress
    qual_info = qualification_options[query.data]
    context.user_data["job_post"]["qualification"] = query.data
    context.user_data["job_post"]["progress"]["current_step"] = 8
    context.user_data["job_post"]["progress"]["completed_steps"].append("qualification")

    # Generate progress text
    progress_text = generate_progress_bar(8, 12, user_id)

    # Edit the original message to show selection
    await query.edit_message_text(
        text=f"✅ {get_translation(user_id, 'qualification_selected')}\n"
             f"{qual_info['emoji']} *{qual_info['name']}*\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_skills')}*",
        parse_mode="Markdown"
    )

    # Ask for skills with formatting tips
    await context.bot.send_message(
        chat_id=user_id,
        text=f"🛠️*{get_translation(user_id, 'enter_skills_question')}*\n\n"
             f"✨ {get_translation(user_id, 'skills_formatting_tips')}:\n"
             f"• {get_translation(user_id, 'skills_tip_separate')}\n"
             f"• {get_translation(user_id, 'skills_tip_specific')}\n"
             f"• {get_translation(user_id, 'skills_tip_level')}\n\n"
             f"ℹ️ {get_translation(user_id, 'skills_character_limit')}",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    return POST_SKILLS


async def post_skills(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Validate input
    if not update.message or not update.message.text.strip():
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_skills"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_skills")]
            ])
        )
        return POST_SKILLS

    skills = update.message.text.strip()

    # Validate length
    if len(skills) > 500:
        await update.message.reply_text(
            text=get_translation(user_id, "skills_too_long"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_skills")]
            ])
        )
        return POST_SKILLS

    # Store skills and update progress
    context.user_data["job_post"]["skills"] = skills
    context.user_data["job_post"]["progress"]["current_step"] = 9
    context.user_data["job_post"]["progress"]["completed_steps"].append("skills")

    # Generate progress text
    progress_text = generate_progress_bar(9, 12, user_id)

    # Create a preview of the skills
    preview = skills[:100] + ('...' if len(skills) > 100 else '')

    await update.message.reply_text(
        text=f"✅ {get_translation(user_id, 'skills_received')}\n"
             f"🛠️ *Preview*:\n{preview}\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_salary')}*",
        parse_mode="Markdown"
    )


    await context.bot.send_message(
        chat_id=user_id,
        text=f"💰 *{get_translation(user_id, 'enter_salary_question')}*\n\n"
             f"• {get_translation(user_id, 'salary_tip_negotiable')}",

        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    return POST_SALARY


async def post_salary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Validate input
    if not update.message or not update.message.text.strip():
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_salary"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_salary")]
            ])
        )
        return POST_SALARY

    salary = update.message.text.strip()

    # Validate length
    if len(salary) > 100:
        await update.message.reply_text(
            text=get_translation(user_id, "salary_too_long"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_salary")]
            ])
        )
        return POST_SALARY

    # Store salary and update progress
    context.user_data["job_post"]["salary"] = salary
    context.user_data["job_post"]["progress"]["current_step"] = 10
    context.user_data["job_post"]["progress"]["completed_steps"].append("salary")

    # Generate progress text
    progress_text = generate_progress_bar(10, 12, user_id)

    await update.message.reply_text(
        text=f"✅ {get_translation(user_id, 'salary_received')}\n"
             f"💰 *{salary}*\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_benefits')}*",
        parse_mode="Markdown"
    )

    # Ask for benefits with examples
    await context.bot.send_message(
        chat_id=user_id,
        text=f"🎁 *{get_translation(user_id, 'enter_benefits_question')}*\n\n"
             f"✨ {get_translation(user_id, 'benefits_formatting_tips')}:\n"
             f"• {get_translation(user_id, 'benefits_tip_common')}\n"
             f"• {get_translation(user_id, 'benefits_tip_unique')}\n"
             f"• {get_translation(user_id, 'benefits_tip_quantify')}\n\n"
             f"ℹ️ {get_translation(user_id, 'benefits_character_limit')}",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    return POST_BENEFITS


async def post_benefits(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Validate input
    if not update.message or not update.message.text.strip():
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_benefits"),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, "try_again"),
                                      callback_data="retry_benefits")]
            ])
        )
        return POST_BENEFITS

    benefits = update.message.text.strip()

    # Validate length
    if len(benefits) > 500:
        await update.message.reply_text(
            text=get_translation(user_id, "benefits_too_long"),

        )
        return POST_BENEFITS

    # Store benefits and update progress
    context.user_data["job_post"]["benefits"] = benefits
    context.user_data["job_post"]["progress"]["current_step"] = 11
    context.user_data["job_post"]["progress"]["completed_steps"].append("benefits")

    # Generate progress text
    progress_text = generate_progress_bar(11, 12, user_id)

    # Create a preview of the benefits
    preview = benefits[:100] + ('...' if len(benefits) > 100 else '')

    await update.message.reply_text(
        text=f"✅ {get_translation(user_id, 'benefits_received')}\n"
             f"🎁 *Preview*:\n{preview}\n\n"
             f"{progress_text}\n\n"
             f"*{get_translation(user_id, 'next_step_deadline')}*",
        parse_mode="Markdown"
    )

    # Ask for deadline with examples
    await context.bot.send_message(
        chat_id=user_id,
        text=f"⏰*{get_translation(user_id, 'enter_deadline_question')}*\n\n"
             f"📅 {get_translation(user_id, 'deadline_format_examples')}:\n"
             f"- {get_translation(user_id, 'deadline_example1')}\n\n"
             f"⚠️ {get_translation(user_id, 'deadline_requirements')}",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )

    return POST_DEADLINE


from datetime import datetime, timedelta
import re


async def post_deadline(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    if not update.message or not update.message.text.strip():
        await update.message.reply_text(
            text=get_translation(user_id, "invalid_deadline"),

        )
        return POST_DEADLINE

    deadline_input = update.message.text.strip()

    # Normalize date separators
    normalized_deadline = re.sub(r"[/.]", "-", deadline_input)

    # Try to parse the date in different formats
    try:
        deadline_date = datetime.strptime(normalized_deadline, "%Y-%m-%d").date()
    except ValueError:
        try:
            deadline_date = datetime.strptime(normalized_deadline, "%d-%m-%Y").date()
        except ValueError:
            await update.message.reply_text(
                text=get_translation(user_id, "invalid_date_format"),

            )
            return POST_DEADLINE

    # Validate date is in the future
    if deadline_date < datetime.now().date():
        await update.message.reply_text(
            text=get_translation(user_id, "deadline_must_be_future"),

        )
        return POST_DEADLINE

    # Store deadline in standardized format (YYYY-MM-DD)
    formatted_deadline = deadline_date.strftime("%Y-%m-%d")
    context.user_data["job_post"]["deadline"] = formatted_deadline
    context.user_data["job_post"]["progress"]["current_step"] = 12
    context.user_data["job_post"]["progress"]["completed_steps"].append("deadline")

    # Generate progress text - completion!
    progress_text = generate_progress_bar(12, 12, user_id)

    await update.message.reply_text(
        text=f"🎉 {get_translation(user_id, 'all_info_received')}\n"
             f"📅 *Deadline*: {formatted_deadline}\n\n"
             f"{progress_text}\n\n"
             f"{get_translation(user_id, 'preparing_preview')}",
        parse_mode="Markdown"
    )

    # Generate and show the final preview
    return await job_preview(update, context)





async def job_preview(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    job_post = context.user_data.get("job_post", {})

    if not job_post:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "missing_job_post_data")
        )
        return await employer_main_menu(update, context)

    # Generate the formatted preview
    preview_text = generate_job_preview(job_post, user_id)

    # Create interactive buttons
    keyboard = [
        [InlineKeyboardButton("✅ " + get_translation(user_id, "confirm_post"),
                              callback_data="confirm")],
        [InlineKeyboardButton("✏️ " + get_translation(user_id, "edit_post"),
                              callback_data="edit")],
        [InlineKeyboardButton("❌ " + get_translation(user_id, "cancel_post"),
                              callback_data="cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send the preview with a nice header
    await context.bot.send_message(
        chat_id=user_id,
        text=f"📋 *{get_translation(user_id, 'job_post_preview')}*\n\n"
             f"{preview_text}\n\n"
             f"ℹ️ {get_translation(user_id, 'preview_instructions')}",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    return CONFIRM_POST


def generate_job_preview(job_details: dict, user_id: int) -> str:
    """Generate a beautifully formatted job preview with emojis and sections"""

    # Employment type mapping with emojis
    employment_types = {
        "full_time": "🏢 Full-time",
        "part_time": "⏱ Part-time",
        "remote": "🏠 Remote",
        "hybrid": "🔀 Hybrid",
        "freelance": "🖥 Freelance"
    }

    # Qualification mapping with emojis
    qualifications = {
        "training": "🎓 Training/Certificate",
        "degree": "🎓 Bachelor's Degree",
        "ma": "🎓 Master's Degree",
        "phd": "🎓 PhD"
    }

    # Level mapping with emojis
    levels = {
        "entry_level": "👶 Entry-level",
        "mid_level": "🧑 Mid-level",
        "senior_level": "👴 Senior-level"
    }

    # Gender mapping with emojis
    genders = {
        "male": "👨 Male",
        "Male": "👨 Male",
        "female": "👩 Female",
        "Female": "👩 Female",
        "any": "👥 Both",
        "Any": "👥 Both",
        "both": "👥 Both"
    }

    # Build the preview sections
    sections = [
        f"📌 *{get_translation(user_id, 'job_title')}:* {job_details.get('job_title', 'N/A')}",
        f"🕒 *{get_translation(user_id, 'employment_type')}:* {employment_types.get(job_details.get('employment_type'), 'N/A')}",
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 *{get_translation(user_id, 'gender')}:* {genders.get(job_details.get('gender'), 'N/A')}",
        f"🔢 *{get_translation(user_id, 'quantity')}:* {job_details.get('quantity', 'N/A')}",
        f"📈 *{get_translation(user_id, 'level')}:* {levels.get(job_details.get('level'), 'N/A')}",
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"\n📝 *{get_translation(user_id, 'description')}:*\n{job_details.get('description', 'N/A')}",
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"\n🎓 *{get_translation(user_id, 'qualification')}:* {qualifications.get(job_details.get('qualification'), 'N/A')}",
        f"\n🛠 *{get_translation(user_id, 'skills')}:*\n{job_details.get('skills', 'N/A')}",
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"\n💰 *{get_translation(user_id, 'salary')}:* {job_details.get('salary', 'N/A')}",
        f"\n🎁 *{get_translation(user_id, 'benefits')}:*\n{job_details.get('benefits', 'N/A')}",
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"\n⏰ *{get_translation(user_id, 'deadline')}:* {job_details.get('deadline', 'N/A')}"
    ]

    return "\n".join(sections)


async def confirm_post(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "confirm":
        try:
            job_post = context.user_data["job_post"]

            # Add timestamp and status
            job_post["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            job_post["status"] = "pending"

            # Save to database
            db.save_pending_job_post(job_post)

            # Send confirmation to employer
            await context.bot.send_message(
                chat_id=user_id,
                text=f"🎉 {get_translation(user_id, 'post_submitted_success')}\n\n"
                     f"⏳ {get_translation(user_id, 'under_review_message')}\n"
                     f"📩 {get_translation(user_id, 'notification_will_be_sent')}",
                parse_mode="Markdown"
            )

            # Notify admins
            await notify_admins_about_new_post(context, job_post, user_id)

        except Exception as e:
            logging.error(f"Error saving job post: {e}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "post_submission_error")
            )

        # Clear user data
        context.user_data.clear()
        return await employer_main_menu(update, context)


    elif query.data == "edit":
        # Restart the post vacancy process
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "restarting_post_process")
        )
        return await post_vacancy_start(update, context)
    elif query.data == "cancel":
        await query.edit_message_text(
            text=get_translation(user_id, "post_cancelled")
        )
        context.user_data.clear()
        return await employer_main_menu(update, context)


async def notify_admins_about_new_post(context: ContextTypes.DEFAULT_TYPE, job_post: dict, user_id: int):
    """Notify all admins about a new pending job post with employer details"""
    try:
        # Get employer details from database
        employer_id = job_post.get('employer_id')
        employer_info = db.get_employer_info(employer_id)  # New method to fetch employer details

        # Format company name or use fallback
        company_name = html.escape(employer_info.get('company_name', 'N/A')) if employer_info else 'N/A'

        admin_message = (
            f"📢 <b>New Job Post for Review</b>\n\n"
            f"🏢 <b>Employer:</b> {company_name}\n"
            f"👤 <b>Employer ID:</b> {html.escape(str(employer_id))} (User ID: {user_id})\n"
            f"📌 <b>Position:</b> {html.escape(job_post.get('job_title', 'N/A'))}\n"
            f"🕒 <b>Type:</b> {html.escape(job_post.get('employment_type', 'N/A').capitalize())}\n"
            f"⏰ <b>Submitted:</b> {html.escape(str(job_post.get('created_at', 'N/A')))}\n\n"
            f"Please review this post in the admin panel."
        )

        for admin_id in get_all_admins():
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=admin_message,
                    parse_mode="HTML"
                )
            except Exception as e:
                logging.error(f"Failed to notify admin {admin_id}: {e}")

    except Exception as e:
        logging.error(f"Error preparing admin notification: {e}")
        # Fallback to original message if there's an error
        fallback_message = (
            f"📢 <b>New Job Post for Review</b>\n\n"
            f"👤 <b>Employer:</b> {html.escape(str(job_post.get('employer_id', 'N/A')))} (User ID: {user_id})\n"
            f"📌 <b>Position:</b> {html.escape(job_post.get('job_title', 'N/A'))}\n"
            f"🕒 <b>Type:</b> {html.escape(job_post.get('employment_type', 'N/A').capitalize())}\n"
            f"⏰ <b>Submitted:</b> {html.escape(str(job_post.get('created_at', 'N/A')))}\n\n"
            f"Please review this post in the admin panel."
        )
        for admin_id in get_all_admins():
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=fallback_message,
                    parse_mode="HTML"
                )
            except Exception as e:
                logging.error(f"Failed to notify admin {admin_id} with fallback message: {e}")

# Manage Vacancies for Employer
def get_main_menu_keyboard(user_id: int):
    return ReplyKeyboardMarkup(
        [[get_translation(user_id, 'back_to_main_menu')]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

async def manage_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    # Create reply keyboard with back button
    reply_keyboard = [[get_translation(user_id, 'back_to_main_menu')]]

    try:

        # Get enhanced job stats including application counts
        jobs_with_stats = db.get_jobs_with_stats(user_id)

        if not jobs_with_stats:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'no_job_listings'),
                parse_mode="HTML",
                reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True, one_time_keyboard=True)
            )
            return EMPLOYER_MAIN_MENU

        # Categorize jobs with visual indicators
        categorized = {
            'active': [],
            'pending': [],
            'expired': [],
            'closed': []
        }

        current_date = datetime.now().strftime('%Y-%m-%d')
        for job in jobs_with_stats:
            # Use the same status determination logic as display_job_post()
            status = job['status'].lower()

            # Check expiration for approved jobs
            if status == 'approved' and 'application_deadline' in job and job['application_deadline']:
                delta = (datetime.strptime(job['application_deadline'], '%Y-%m-%d') - datetime.now()).days
                if delta < 0:
                    status = 'expired'

            # Categorize based on final status
            if status == 'pending':
                categorized['pending'].append(job)
            elif status == 'closed':
                categorized['closed'].append(job)
            elif status == 'expired':
                categorized['expired'].append(job)
            else:  # approved and not expired
                categorized['active'].append(job)

        # Send overview stats
        stats_msg = (
            f"{get_translation(user_id, 'job_listings_overview_title')}\n\n"
            f"🟢 <b>{get_translation(user_id, 'active')}:</b> {len(categorized['active'])} {get_translation(user_id, 'jobs')}\n"
            f"🟡 <b>{get_translation(user_id, 'pending')}:</b> {len(categorized['pending'])} {get_translation(user_id, 'jobs')}\n"
            f"🔴 <b>{get_translation(user_id, 'expired')}:</b> {len(categorized['expired'])} {get_translation(user_id, 'jobs')}\n"
            f"⚫ <b>{get_translation(user_id, 'closed')}:</b> {len(categorized['closed'])} {get_translation(user_id, 'jobs')}\n\n"
            f"📨 <b>{get_translation(user_id, 'total_applications')}:</b> {sum(j['application_count'] for j in jobs_with_stats)}\n"
        )
        await context.bot.send_message(
            chat_id=user_id,
            text=stats_msg,
            parse_mode="HTML",
            reply_markup=get_main_menu_keyboard(user_id)
        )

        # Display jobs by category with rich formatting and pagination
        for category, jobs in categorized.items():
            if jobs:
                await display_job_category(
                    category=category,
                    jobs=jobs,
                    user_id=user_id,
                    context=context,
                    include_actions=(category == 'active')
                )

    except Exception as e:
        logging.error(f"Error in manage_vacancies: {str(e)}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'failed_to_load_vacancies'),
            reply_markup=get_main_menu_keyboard(user_id)
        )

    return EMPLOYER_MAIN_MENU

async def select_job_to_manage(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip()

    # Check if user wants to go back to main menu
    if choice == get_translation(user_id, 'back_to_main_menu'):
        return await employer_main_menu(update, context)

    try:
        # Retrieve stored approved jobs
        approved_jobs = context.user_data.get("approved_jobs")
        if not approved_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_approved_jobs_found"),
                reply_markup=get_main_menu_keyboard(user_id)
            )
            return EMPLOYER_MAIN_MENU

        # Validate the selected index
        selected_index = int(choice) - 1
        if not (0 <= selected_index < len(approved_jobs)):
            raise ValueError(get_translation(user_id, "invalid_selection"))

        # Retrieve selected job details
        selected_job = approved_jobs[selected_index]
        job_id = selected_job["id"]

        # Store the selected job in user_data
        context.user_data["selected_job_id"] = job_id

        # Display actions for the selected job
        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, "view_applications_button"), callback_data=f"view_apps {job_id}")],
            [InlineKeyboardButton(get_translation(user_id, "close_vacancy_button"), callback_data=f"close {job_id}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        # Send the job details and actions
        job_text = (
            f"{get_translation(user_id, 'job_title')}: {selected_job['job_title']}\n"
            f"{get_translation(user_id, 'deadline')}: {selected_job['deadline']}\n"
            f"{get_translation(user_id, 'status')}: {selected_job['status'].capitalize()}\n"
        )
        await context.bot.send_message(
            chat_id=user_id,
            text=job_text + get_translation(user_id, "select_action_for_job"),
            reply_markup=reply_markup
        )

        return HANDLE_JOB_ACTIONS  # Transition to job action handling state

    except ValueError as ve:
        logging.error(f"ValueError in select_job_to_manage: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_selection"),
            reply_markup=get_main_menu_keyboard(user_id)
        )
        return SELECT_JOB_TO_MANAGE

    except Exception as e:
        logging.error(f"Unexpected error in select_job_to_manage: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred"),
            reply_markup=get_main_menu_keyboard(user_id)
        )
        return EMPLOYER_MAIN_MENU

async def handle_job_actions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Check if this is a message (back button) or callback query
    if update.message:
        user_id = get_user_id(update)
        choice = update.message.text.strip()

        if choice == get_translation(user_id, 'back_to_main_menu'):
            return await employer_main_menu(update, context)
        else:
            # Add keyboard to unexpected message
            reply_keyboard = [[get_translation(user_id, 'back_to_main_menu')]]
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "please_use_buttons_provided"),
                reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True, one_time_keyboard=True)
            )
            return HANDLE_JOB_ACTIONS
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        # Parse the callback data
        action, job_id_str = query.data.split("_", 1)
        job_id = int(job_id_str)

        # Validate job ownership
        job_type, job_data = validate_job_ownership(db, job_id, user_id)
        if not job_type:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "unauthorized_access")
            )
            return EMPLOYER_MAIN_MENU

        # Normalize status
        try:
            validated_job = validate_job_post(job_data)
        except ValueError as ve:
            logging.error(f"Invalid job post data: {ve}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "invalid_job_post", error=str(ve))
            )
            return EMPLOYER_MAIN_MENU

        status = validated_job["status"]

        if action == "view_apps":
            if job_type != "vacancy":
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "cannot_view_apps_for_pending_job")
                )
                return EMPLOYER_MAIN_MENU

            # Fetch and display applicants
            await fetch_and_display_applicants(job_id, user_id, context)
            return VIEW_APPLICATIONS

        elif action == "close":
            if job_type == "vacancy" and status != "approved":
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "cannot_close_non_approved_vacancy")
                )
                return EMPLOYER_MAIN_MENU

            # Confirm closing the vacancy
            context.user_data["close_job_id"] = job_id
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "confirm_close_vacancy_prompt").format(job_id=job_id),
                reply_markup=ReplyKeyboardMarkup([["Yes", "No"]], one_time_keyboard=True, resize_keyboard=True)
            )
            return CONFIRM_CLOSE

        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "unknown_action_detected")
            )
            return EMPLOYER_MAIN_MENU

    except ValueError as ve:
        logging.error(f"ValueError in handle_job_actions: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_action_detected", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in handle_job_actions: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )

    return EMPLOYER_MAIN_MENU

async def display_job_category(category: str, jobs: list, user_id: int,
                               context: ContextTypes.DEFAULT_TYPE, include_actions: bool):
    """Display jobs in a category with enhanced formatting and pagination"""
    category_titles = {
        'active': f"🟢 {get_translation(user_id, 'active_vacancies').upper()}",
        'pending': f"🟡 {get_translation(user_id, 'pending_approval').upper()}",
        'expired': f"🔴 {get_translation(user_id, 'expired_vacancies').upper()}",
        'closed': f"⚫ {get_translation(user_id, 'closed_vacancies').upper()}"
    }

    # Store jobs in context for pagination
    context.user_data[f'manage_vacancies_{category}_jobs'] = jobs
    context.user_data[f'manage_vacancies_{category}_page'] = 0  # Start at page 0
    context.user_data['current_category'] = category

    await send_job_page(category, 0, jobs, user_id, context, include_actions)


async def send_job_page(category: str, page: int, jobs: list, user_id: int,
                        context: ContextTypes.DEFAULT_TYPE, include_actions: bool,
                        edit_message_id: int = None):
    """Send job page - returns the sent message"""
    category_titles = {
        'active': f"🟢 {get_translation(user_id, 'active_vacancy').upper()}",
        'pending': f"🟡 {get_translation(user_id, 'pending_approval').upper()}",
        'expired': f"🔴 {get_translation(user_id, 'expired_vacancies').upper()}",
        'closed': f"⚫ {get_translation(user_id, 'closed_vacancies').upper()}"
    }

    jobs_per_page = 4
    total_jobs = len(jobs)
    start_idx = page * jobs_per_page
    end_idx = min(start_idx + jobs_per_page, total_jobs)

    # Create message text
    text = (
        f"<b>{category_titles[category]}</b>\n"
        f"📋 <i>Showing {start_idx + 1}-{end_idx} of {total_jobs} vacancies</i>\n\n"
    )

    # Create pagination buttons
    reply_markup = None
    if len(jobs) > jobs_per_page:
        buttons = []
        if page > 0:
            buttons.extend([
                InlineKeyboardButton("⏪ First", callback_data=f"manage_vacancies_first_{category}"),
                InlineKeyboardButton("◀️ Prev", callback_data=f"manage_vacancies_prev_{category}")
            ])

        buttons.append(InlineKeyboardButton(
            f"Page {page + 1}/{max(1, (len(jobs) + jobs_per_page - 1) // jobs_per_page)}",
            callback_data="noop"
        ))

        if page < (len(jobs) // jobs_per_page):
            buttons.extend([
                InlineKeyboardButton("Next ▶️", callback_data=f"manage_vacancies_next_{category}"),
                InlineKeyboardButton("Last ⏭️", callback_data=f"manage_vacancies_last_{category}")
            ])

        reply_markup = InlineKeyboardMarkup([buttons])

    # Send or edit message
    if edit_message_id:
        msg = await context.bot.edit_message_text(
            chat_id=user_id,
            message_id=edit_message_id,
            text=text,
            parse_mode="HTML",
            reply_markup=reply_markup
        )
    else:
        msg = await context.bot.send_message(
            chat_id=user_id,
            text=text,
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    # Send job listings as separate messages
    for job in sorted(jobs[start_idx:end_idx], key=lambda x: x['application_deadline']):
        await display_job_post(job, user_id, context, include_actions)

    return msg


async def handle_pagination_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle pagination callback queries and delete old messages"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    data = query.data

    try:
        # Parse callback data
        if not data.startswith('manage_vacancies_'):
            return

        parts = data.split('_')
        if len(parts) < 4:
            raise ValueError("Invalid callback data")

        action = parts[2]  # "first", "prev", etc.
        category = '_'.join(parts[3:])  # The category

        # Get current state
        jobs = context.user_data.get(f'manage_vacancies_{category}_jobs', [])
        if not jobs:
            await context.bot.send_message(user_id, text=get_translation(user_id, 'no_vacancies_found'))
            return EMPLOYER_MANAGE_VACANCIES

        current_page = context.user_data.get(f'manage_vacancies_{category}_page', 0)
        old_message_id = context.user_data.get(f'manage_vacancies_{category}_message_id')
        total_pages = max(1, (len(jobs) + 3) // 4)

        # Calculate new page
        if action == 'first':
            new_page = 0
        elif action == 'prev':
            new_page = max(0, current_page - 1)
        elif action == 'next':
            new_page = min(total_pages - 1, current_page + 1)
        elif action == 'last':
            new_page = total_pages - 1
        else:
            return

        # Delete old message if exists
        if old_message_id:
            try:
                await context.bot.delete_message(chat_id=user_id, message_id=old_message_id)
            except Exception as e:
                logging.warning(f"Could not delete old message: {e}")

        # Update page in context
        context.user_data[f'manage_vacancies_{category}_page'] = new_page

        # Send new page (as fresh message)
        message = await send_job_page(
            category=category,
            page=new_page,
            jobs=jobs,
            user_id=user_id,
            context=context,
            include_actions=(category == 'active'),
            edit_message_id=None  # Force new message
        )

        # Store new message ID
        context.user_data[f'manage_vacancies_{category}_message_id'] = message.message_id

        return context.user_data.get('current_state', EMPLOYER_MANAGE_VACANCIES)

    except Exception as e:
        logging.error(f"Pagination error: {str(e)}")
        await context.bot.send_message(user_id, text="⚠️ Failed to update vacancies.")
        return EMPLOYER_MANAGE_VACANCIES


async def display_job_post(job: dict, user_id: int,
                           context: ContextTypes.DEFAULT_TYPE, include_actions: bool):
    """Display single job post with rich formatting"""
    try:
        context.user_data['last_job'] = job
        # Calculate days remaining (for active jobs)
        days_left = "N/A"
        current_status = job['status'].lower()
        if 'application_deadline' in job and job['application_deadline']:
            delta = (datetime.strptime(job['application_deadline'], '%Y-%m-%d') - datetime.now()).days
            if current_status == 'approved':
                days_left = f"{max(0, delta)} days" if delta >= 0 else "Expired"
                if delta < 0:
                    current_status = 'expired'
            elif current_status == 'expired':
                days_left = "Expired"

        # Status indicator
        status_icons = {
            'approved': '🟢',
            'pending': '🟡',
            'rejected': '🔴',
            'closed': '⚫',
            'expired': '🔴'
        }
        status_icon = status_icons.get(current_status, '⚪')

        # Build message with professional formatting
        msg = (
            f"{status_icon} <b>{escape_html(job['job_title'])}</b>\n"
            f"📅 <i>{get_translation(user_id, 'deadlined')}:</i> {job['application_deadline']} ({days_left})\n"
            f"👥 <i>{get_translation(user_id, 'applications')}:</i> {job.get('application_count', 0)}\n"
            f"💰 <i>{get_translation(user_id, 'salary')}:</i> {job.get('salary', get_translation(user_id, 'not_specified'))}\n\n"
            f"📝 <i>{get_translation(user_id, 'description')}:</i>\n{escape_html(job['description'][:150])}..."
        )

        # Add action buttons
        reply_markup = None
        if current_status == 'approved' and include_actions:
            reply_markup = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(get_translation(user_id, 'view_applicants'),
                                         callback_data=f"view_apps_{job['id']}"),
                    InlineKeyboardButton(get_translation(user_id, 'close_job'), callback_data=f"close_{job['id']}")
                ],
                [
                    InlineKeyboardButton(get_translation(user_id, 'stats'), callback_data=f"stats_{job['id']}"),
                    InlineKeyboardButton(get_translation(user_id, 'renew'), callback_data=f"renew_{job['id']}")
                ],
                [
                    InlineKeyboardButton(get_translation(user_id, 'preview_details'),
                                         callback_data=f"preview_{job['id']}")
                ]
            ])
        elif current_status == 'expired':
            reply_markup = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(get_translation(user_id, 'renew_vacancy'), callback_data=f"renew_{job['id']}"),
                    InlineKeyboardButton(get_translation(user_id, 'stats'), callback_data=f"stats_{job['id']}")
                ],
                [
                    InlineKeyboardButton(get_translation(user_id, 'view_applicants'),
                                         callback_data=f"view_apps_{job['id']}")
                ],
                [
                    InlineKeyboardButton(get_translation(user_id, 'preview_details'),
                                         callback_data=f"preview_{job['id']}")
                ]
            ])

        await context.bot.send_message(
            chat_id=user_id,
            text=msg,
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    except Exception as e:
        logging.error(f"Error displaying job post: {str(e)}")
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ Couldn't display job {job.get('id', '')}"
        )


async def preview_job_details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    job_id = int(query.data.split('_')[1])
    user_id = query.from_user.id

    # Fetch the job details from your database
    job = db.get_job_by_id(job_id)

    if not job:
        await query.edit_message_text(text="Job not found.")
        return

    # Format the deadline
    deadline = job.get('application_deadline', 'Not specified')
    if deadline:
        delta = (datetime.strptime(deadline, '%Y-%m-%d') - datetime.now()).days
        if delta >= 0:
            deadline = f"{deadline} ({delta} days remaining)"
        else:
            deadline = f"{deadline} (Expired)"

    # Format employer contact if available
    employer_link = ""
    if job.get('employer_contact'):
        employer_link = f"<a href='{job['employer_contact']}'>Contact Employer</a>"

    # Build the detailed message
    job_details = (
            f"<b>📋 Job Title:</b> {escape_html(job.get('job_title', 'N/A'))}\n"
            f"<b>🏢 Employer:</b> {escape_html(job.get('employer_name', 'Not provided'))}\n"
            f"<b>📅 Deadline:</b> {deadline}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>📝 Description:</b>\n{escape_html(job.get('description', ''))}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>💼 Employment Type:</b> {escape_html(job.get('employment_type', 'N/A'))}\n"
            f"<b>🚻 Gender:</b> {escape_html(job.get('gender', 'N/A'))}\n"
            f"<b>👥 Quantity:</b> {escape_html(str(job.get('quantity', 'N/A')))}\n"
            f"<b>🎓 Qualification:</b> {escape_html(job.get('qualification', 'N/A'))}\n"
            f"<b>📊 Level:</b> {escape_html(job.get('level', 'N/A'))}\n"
            f"<b>🔑 Skills:</b> {escape_html(job.get('skills', 'N/A'))}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>💰 Salary:</b> {escape_html(job.get('salary', 'Negotiable'))}\n"
            f"<b>🎁 Benefits:</b> {escape_html(job.get('benefits', 'Negotiable'))}\n"

    )

    # Add a back button to return to the previous view
    reply_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data=f"back_to_job_{job_id}")]
    ])

    await query.edit_message_text(
        text=job_details,
        parse_mode="HTML",
        reply_markup=reply_markup
    )


async def handle_back_to_job(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Get the stored job data
    job = context.user_data.get('last_job')

    if not job:
        await query.edit_message_text(text="Job data no longer available.")
        return

    # Redisplay the job post with actions
    return await display_job_post(
        job=job,
        user_id=query.from_user.id,
        context=context,
        include_actions=True
    )
def validate_and_format_job_post(job: dict, user_id: int, include_actions: bool) -> tuple:
    """
    Validate and format a job post for display.
    """
    required_fields = {
        "job_id", "status", "job_title", "employment_type", "gender", "quantity", "level",
        "description", "qualification", "skills", "salary", "benefits", "deadline"
    }

    # Fill missing fields with "Not provided" instead of failing
    validated_job = {field: job.get(field, "Not provided") for field in required_fields}

    # Check for any missing required fields
    missing_fields = [field for field, value in validated_job.items() if value == "Not provided"]
    if missing_fields:
        raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

    # Extract job details
    job_id = validated_job["job_id"]
    job_title = validated_job["job_title"]
    employment_type = validated_job["employment_type"]
    gender = validated_job["gender"]
    quantity = validated_job["quantity"]
    level = validated_job["level"]
    description = validated_job["description"][:50] + "..." if validated_job["description"] else "Not provided"
    qualification = validated_job["qualification"]
    skills = validated_job["skills"]
    salary = validated_job["salary"]
    benefits = validated_job["benefits"]
    deadline = validated_job["deadline"]
    status = validated_job["status"].lower()

    # Format the job post details
    job_text = (
        f"{get_translation(user_id, 'job_title')}: {job_title}\n"
        f"{get_translation(user_id, 'employment_type')}: {employment_type}\n"
        f"{get_translation(user_id, 'gender')}: {gender}\n"
        f"{get_translation(user_id, 'quantity')}: {quantity}\n"
        f"{get_translation(user_id, 'level')}: {level}\n"
        f"{get_translation(user_id, 'description')}: {description}\n"
        f"{get_translation(user_id, 'qualification')}: {qualification}\n"
        f"{get_translation(user_id, 'skills')}: {skills}\n"
        f"{get_translation(user_id, 'salary')}: {salary}\n"
        f"{get_translation(user_id, 'benefits')}: {benefits}\n"
        f"{get_translation(user_id, 'deadline')}: {deadline}\n"
        f"{get_translation(user_id, 'status')}: {status.capitalize()}\n"
    )

    # Define keyboard buttons based on the job status
    keyboard = []
    if include_actions:
        if status == "approved":
            keyboard.extend([
                [InlineKeyboardButton(get_translation(user_id, "view_applications_button"), callback_data=f"view_apps_{job_id}")],
                [InlineKeyboardButton(get_translation(user_id, "close_vacancy_button"), callback_data=f"close_{job_id}")]
            ])
        elif status == "pending":
            keyboard.append([InlineKeyboardButton(get_translation(user_id, "resubmit_vacancy_button"), callback_data=f"resubmit_{job_id}")])

    return job_text, InlineKeyboardMarkup(keyboard) if keyboard else None
def validate_job_ownership(db, job_id: int, employer_id: int):
    try:
        db.cursor.execute("""
            SELECT 
                CASE WHEN source = 'vacancy' THEN 'vacancy' ELSE 'job_post' END AS job_type,
                id AS job_id, 
                employer_id, 
                job_title, 
                employment_type, 
                gender, 
                quantity, 
                level, 
                description, 
                qualification, 
                skills, 
                salary, 
                benefits, 
                deadline, 
                status, 
                source
            FROM (
                SELECT 
                    id, 
                    employer_id, 
                    job_title, 
                    employment_type, 
                    gender, 
                    quantity, 
                    level, 
                    description, 
                    qualification, 
                    skills, 
                    salary, 
                    benefits, 
                    deadline, 
                    status, 
                    'job_post' AS source
                FROM job_posts
                WHERE id = ? AND employer_id = ?
                UNION ALL
                SELECT 
                    id, 
                    employer_id, 
                    job_title, 
                    employment_type, 
                    gender, 
                    quantity, 
                    level, 
                    description, 
                    qualification, 
                    skills, 
                    salary, 
                    benefits, 
                    application_deadline, 
                    status, 
                    'vacancy' AS source
                FROM vacancies
                WHERE id = ? AND employer_id = ?
            )
        """, (job_id, employer_id, job_id, employer_id))

        result = db.cursor.fetchone()
        if result:
            columns = [column[0] for column in db.cursor.description]
            job_data = dict(zip(columns, result))
            return job_data["source"], job_data  # Returns 'vacancy' or 'job_post'

        return None, None
    except sqlite3.Error as e:
        logging.error(f"Database error validating job ownership: {e}")
        return None, None


def get_job_status(job_data: dict) -> str:
    """Consistently determine job status considering expiration"""
    status = job_data.get('status', '').lower()

    # Only check expiration for approved jobs
    if status == 'approved' and 'deadline' in job_data and job_data['deadline']:
        try:
            deadline_date = datetime.strptime(job_data['deadline'], '%Y-%m-%d')
            if deadline_date < datetime.now():
                return 'expired'
        except ValueError:
            pass

    return status

async def handle_vacancy_actions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    # Add reply keyboard when showing action confirmation messages
    reply_keyboard = [[get_translation(user_id, 'back_to_main_menu')]]

    try:
        # More robust callback data parsing
        parts = query.data.split('_')
        if len(parts) < 2:
            raise ValueError(f"Invalid callback data format: {query.data}")

        action = parts[0]
        job_id_str = parts[-1]  # Always take last part as job ID

        try:
            job_id = int(job_id_str)
        except ValueError:
            raise ValueError(f"Invalid job ID format: {job_id_str}")

        # Validate job ownership
        job_type, job_data = validate_job_ownership(db, job_id, user_id)
        if not job_type:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "unauthorized_access")
            )
            return EMPLOYER_MAIN_MENU

        validated_job = validate_job_post(job_data)
        status = validated_job["status"].lower()

        # Check if job is expired based on deadline
        current_date = datetime.now().strftime('%Y-%m-%d')
        is_expired = job_data['deadline'] < current_date

        # Handle all possible actions
        if action == "close":
            if job_type == "vacancy" and status != "approved":
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "cannot_close_non_approved_vacancy")
                )
                return await employer_main_menu(update, context)

            context.user_data["close_job_id"] = job_id
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "confirm_close_vacancy_prompt").format(job_id=job_id),
                reply_markup=ReplyKeyboardMarkup([
                    [KeyboardButton("Yes"), KeyboardButton("No")],
                    [get_translation(user_id, 'back_to_main_menu')]
                ], resize_keyboard=True, one_time_keyboard=True)
            )
            return CONFIRM_CLOSE

        elif action == "view" and "apps" in query.data:  # Handle "view_apps_5" format
            if job_type != "vacancy":
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "cannot_view_apps_for_pending_job")
                )
                return EMPLOYER_MAIN_MENU
            return await view_applicants_list(update, context)

        elif action == "stats":
            stats = db.get_vacancy_stats(job_id)
            await context.bot.send_message(
                chat_id=user_id,
                text=format_vacancy_stats(stats, user_id),
                parse_mode="HTML"
            )
            return EMPLOYER_MAIN_MENU

        elif action == "renew":
            status = get_job_status(job_data)

            if status == 'expired':
                context.user_data["renew_job_id"] = job_id
                await show_renew_options(update, context)
                return RENEW_VACANCY
            else:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "can_only_renew_expired")
                )
                return EMPLOYER_MAIN_MENU

        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "unknown_action_detected")
            )
            return EMPLOYER_MAIN_MENU

    except ValueError as ve:
        logging.error(f"ValueError in handle_vacancy_actions: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_action_detected", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in handle_vacancy_actions: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e)),
            reply_markup=get_main_menu_keyboard(user_id)
        )

    return EMPLOYER_MAIN_MENU
def format_vacancy_stats(stats: dict, user_id: int) -> str:
    """Format stats into a readable message"""
    return (
        f"📊 <b>{get_translation(user_id, 'vacancy_statistics_title')}</b>\n\n"
        f"📨 <i>{get_translation(user_id, 'total_applications')}:</i> {stats.get('total_applications', 0)}\n"
        f"✅ <i>{get_translation(user_id, 'successful_hires')}:</i> {stats.get('hires', 0)}\n"
        f"🕒 <i>{get_translation(user_id, 'pending_reviews')}:</i> {stats.get('pending', 0)}\n"
        f"💡 <i>{get_translation(user_id, 'tip_label')}:</i> {get_translation(user_id, 'stats_improvement_tp')}"
    )

async def show_renew_options(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show renewal options for expired vacancies"""
    user_id = update.callback_query.from_user.id
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, 'select_renewal_duration'),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("30 days", callback_data="renew_30")],
            [InlineKeyboardButton("60 days", callback_data="renew_60")],
            [InlineKeyboardButton(get_translation(user_id, 'custom_option'), callback_data="renew_custom")]
        ])
    )


async def handle_renew_duration(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    choice = query.data.split("_")[1]

    job_id = context.user_data.get("renew_job_id")
    if not job_id:
        await context.bot.send_message(user_id, get_translation(user_id, 'renewal_session_expired'))
        return EMPLOYER_MAIN_MENU

    if choice in ("30", "60"):
        days = int(choice)
        new_deadline = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
        return await confirm_renewal(update, context, job_id, new_deadline)
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'enter_custom_days'),
            reply_markup=get_main_menu_keyboard(user_id)
        )
        return RENEW_VACANCY

async def handle_custom_renew_duration(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    job_id = context.user_data.get("renew_job_id")

    try:
        days = int(update.message.text.strip())
        if not 1 <= days <= 365:
            raise ValueError
        new_deadline = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
        return await confirm_renewal(update, context, job_id, new_deadline)
    except ValueError:
        await update.message.reply_text(get_translation(user_id, 'enter_number_between_1_365'))
        return RENEW_VACANCY


async def confirm_renewal(update: Update, context: ContextTypes.DEFAULT_TYPE,
                          job_id: int, new_deadline: str) -> int:
    user_id = update.message.from_user.id if update.message else update.callback_query.from_user.id

    context.user_data["renewal_data"] = {
        "job_id": job_id,
        "new_deadline": new_deadline
    }

    job_title = db.get_vacancy_title(job_id)
    await context.bot.send_message(
        chat_id=user_id,
        text=f"{get_translation(user_id, 'confirm_renewal_prompt').format(job_title=job_title, new_deadline=new_deadline)}",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(get_translation(user_id, 'confirm'), callback_data="confirm_renew")],
            [InlineKeyboardButton(get_translation(user_id, 'cancel'), callback_data="cancel_renew")]
        ])
    )
    return CONFIRM_RENEWAL


async def process_renewal_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    choice = query.data.split("_")[0]

    if choice == "cancel":
        await query.edit_message_text(get_translation(user_id, 'renewal_cancelled'))
        return EMPLOYER_MAIN_MENU

    renewal_data = context.user_data.get("renewal_data")
    if not renewal_data:
        await query.edit_message_text(get_translation(user_id, 'renewal_session_expired'))
        return EMPLOYER_MAIN_MENU

    try:
        db.renew_vacancy(
            job_id=renewal_data["job_id"],
            new_deadline=renewal_data["new_deadline"]
        )
        await query.edit_message_text(
            get_translation(user_id, 'vacancy_renewed_success').format(new_deadline=renewal_data['new_deadline']),
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Renewal failed: {e}")
        await query.edit_message_text(get_translation(user_id, 'renewal_failed'))

    return EMPLOYER_MAIN_MENU

# View Applicants

async def view_applicants_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        job_id = int(query.data.split("_")[2])
        context.user_data["selected_job_id"] = job_id

        # Get job details with stats
        job_details = db.get_vacancy_with_stats(job_id)
        if not job_details:
            await query.edit_message_text(get_translation(user_id, 'job_not_found'))
            return EMPLOYER_MAIN_MENU

        # Format deadline display
        deadline = job_details.get('application_deadline', get_translation(user_id, 'not_available'))
        if deadline != get_translation(user_id, 'not_available'):
            try:
                deadline_date = datetime.strptime(deadline, '%Y-%m-%d')
                deadline = deadline_date.strftime('%b %d, %Y')
            except ValueError:
                pass

        validated_job = {
            'job_title': escape_html(job_details.get('job_title', get_translation(user_id, 'unspecified_position'))),
            'deadline': deadline,
            'total_applications': job_details.get('total_applications', 0),
            'approved_count': job_details.get('approved_count', 0),
            'rejected_count': job_details.get('rejected_count', 0)
        }

        # Get and validate applicants
        raw_applicants = db.get_applications_for_job_with_title(job_id)
        applicants = []

        # Check if raw_applicants is None or empty
        if raw_applicants is None:
            raw_applicants = []

        for app in raw_applicants:
            if app:  # Check if app is not None
                applicants.append({
                    'application_id': app.get('application_id'),
                    'full_name': escape_html(app.get('full_name', get_translation(user_id, 'anonymous'))),
                    'status': app.get('status', 'pending'),
                    'application_date': app.get('application_date', get_translation(user_id, 'not_available')),
                    'field_of_study': app.get('field_of_study', get_translation(user_id, 'not_specified')),
                    'cv_exists': bool(app.get('cv_path')),
                    'score': app.get('match_score', 0)
                })

        if not applicants:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'no_applications_received'),
                parse_mode="HTML",
                reply_markup=get_main_menu_keyboard(user_id)
            )
            return EMPLOYER_MAIN_MENU

        # Pagination setup
        context.user_data['applicant_page'] = 0
        context.user_data['all_applicants'] = applicants
        context.user_data['page_size'] = 5  # Applicants per page

        # Send job overview
        overview_msg = (
            f"📋 <b>{get_translation(user_id, 'applications_for')}:</b> {validated_job['job_title']}\n\n"
            f"📅 {get_translation(user_id, 'deadlined')}: {validated_job['deadline']} \n "
            f"👥 {get_translation(user_id, 'total_applicants')}: {validated_job['total_applications']}\n"
            f"{get_translation(user_id, 'approved')}: {validated_job['approved_count']} | "
            f"{get_translation(user_id, 'rejected')}: {validated_job['rejected_count']}\n\n"
            f"<i>{get_translation(user_id, 'showing_top_candidates')}</i>\n"
            f"<i>{get_translation(user_id, 'export_to_excel_tip')}</i>"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text=overview_msg,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    get_translation(user_id, 'export_to_excel'),
                    callback_data=f"export_excel_{job_id}"
                )
            ]])
        )

        # Display first page
        await display_applicant_page(user_id, context)

    except Exception as e:
        logging.error(f"Error viewing applicants: {str(e)}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'failed_to_retrieve_applicants'),
            parse_mode="HTML",
            reply_markup=get_main_menu_keyboard(user_id)
        )

    return VIEW_APPLICATIONS

async def display_applicant_page(user_id: int, context: ContextTypes.DEFAULT_TYPE):
    """Display one page of applicants with navigation controls"""
    applicants = context.user_data['all_applicants']
    current_page = context.user_data['applicant_page']
    page_size = context.user_data['page_size']
    start_idx = current_page * page_size
    page_applicants = applicants[start_idx:start_idx + page_size]

    for i, applicant in enumerate(page_applicants, start_idx + 1):
        status_icon = {'pending': '🟡', 'approved': '🟢', 'rejected': '🔴'}.get(applicant['status'], '⚪')

        applicant_msg = (
            f"{i}. {status_icon} <b>{applicant['full_name']}</b>\n"
            f"📅 {get_translation(user_id, 'applied')}: {applicant['application_date']}\n"
            f"📝 {get_translation(user_id, 'status')}: {applicant['status'].capitalize()}\n"
            f"💼 {get_translation(user_id, 'field_of_study')}: {applicant['field_of_study']}"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text=applicant_msg,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    get_translation(user_id, 'review'),
                    callback_data=f"review_{applicant['application_id']}"
                )
            ]])
        )
    # Pagination controls
    total_pages = (len(applicants) + page_size - 1) // page_size
    if total_pages > 1:
        pagination_buttons = []
        if current_page > 0:
            pagination_buttons.append(InlineKeyboardButton("⬅️ Previous", callback_data="prev_page"))
        if current_page < total_pages - 1:
            pagination_buttons.append(InlineKeyboardButton("Next ➡️", callback_data="next_page"))

            # Add Export to Excel button
        pagination_buttons.append(InlineKeyboardButton("Export to Excel 📊",
                                                       callback_data=f"export_excel_{context.user_data['selected_job_id']}"))
        await context.bot.send_message(
            chat_id=user_id,
            text=f"Page {current_page + 1}/{total_pages}",
            reply_markup=InlineKeyboardMarkup([pagination_buttons])
        )
        # Add main menu button as separate message
        await context.bot.send_message(
            chat_id=user_id,
            text="You can return to main menu anytime:",
            reply_markup=get_main_menu_keyboard(user_id)
        )


# Add to your callback handler
async def handle_page_navigation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "next_page":
        context.user_data['applicant_page'] += 1
    elif query.data == "prev_page":
        context.user_data['applicant_page'] -= 1

    await display_applicant_page(user_id, context)
    return VIEW_APPLICATIONS


async def handle_applicant_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        application_id = int(query.data.split('_')[1])
        application = db.get_complete_application_details(application_id)

        if not application:  # This now properly checks for empty dict
            await query.edit_message_text("⚠️ Application not found or no longer exists")
            return VIEW_APPLICATIONS

        context.user_data["selected_applicant"] = application
        try:
            details = format_applicant_details(application, user_id)
        except Exception as e:
            logging.error(f"Error formatting applicant details: {e}")
            await query.edit_message_text("⚠️ Error displaying application details")
            return VIEW_APPLICATIONS


        # Create buttons - make sure download button has correct pattern
        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, 'accept_applicant'), callback_data="accept_applicant")],
            [InlineKeyboardButton(get_translation(user_id, 'reject_applicant'), callback_data="reject_applicant")],
        ]

        # Only show download button if CV exists
        if application.get('cv_path'):
            keyboard.append(
                [InlineKeyboardButton(get_translation(user_id, 'download_cv'),
                                      callback_data=f"download_cv_{application_id}")]
            )

        await query.edit_message_text(
            text=details,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return ACCEPT_REJECT_CONFIRMATION
    except Exception as e:
        logging.error(f"Error reviewing applicant: {e}")
        await query.edit_message_text("Failed to load application details")
        return VIEW_APPLICATIONS


def format_applicant_details(application: dict, user_id: int) -> str:
    """Completely safe applicant details formatter"""
    if not application or not isinstance(application, dict):
        return get_translation(user_id, 'application_info_not_available')

    # Helper function for safe value extraction
    def safe_get(key, default=get_translation(user_id, 'not_available')):
        value = application.get(key, default)
        return str(value) if value not in [None, ''] else default

    # Profile details
    profile = {
        'dob': safe_get('dob'),
        'qualification': safe_get('qualification'),
        'field_of_study': safe_get('field_of_study'),
        'cgpa': safe_get('cgpa'),
        'languages': safe_get('languages'),
        'profile_summary': safe_get('profile_summary', get_translation(user_id, 'not_provided'))
    }

    # Portfolio link
    portfolio_link = (
        f'<a href="{escape_html(application["portfolio_link"])}">{get_translation(user_id, "view_portfolio")}</a>'
        if application.get("portfolio_link") else get_translation(user_id, 'not_provided')
    )

    separator = "────────────────────"

    return (
        f"<b>{get_translation(user_id, 'applicant_details_title')}</b>\n\n"
        f"{separator}\n"
        f"👤 <b>{get_translation(user_id, 'name')}:</b> {escape_html(safe_get('full_name'))}\n"
        f"📅 <b>{get_translation(user_id, 'applied_on')}:</b> {safe_get('application_date')}\n"
        f"📝 <b>{get_translation(user_id, 'status')}:</b> {safe_get('status', 'pending').capitalize()}\n"
        f"📄 <b>{get_translation(user_id, 'cv')}:</b> {get_translation(user_id, 'available') if application.get('cv_path') else get_translation(user_id, 'not_provided')}\n"
        f"🔗 <b>{get_translation(user_id, 'portfolio')}:</b> {portfolio_link}\n"
        f"👫 <b>{get_translation(user_id, 'gender')}:</b> {escape_html(safe_get('gender'))}\n"
        f"📱 <b>{get_translation(user_id, 'contact')}:</b> {escape_html(safe_get('contact_number'))}\n"
        f"🎂 <b>{get_translation(user_id, 'date_of_birth')}:</b> {profile['dob']}\n"
        f"{separator}\n"
        f"<b>{get_translation(user_id, 'education')}</b>\n"
        f"🎓 <b>{get_translation(user_id, 'qualification')}:</b> {escape_html(profile['qualification'])}\n"
        f"📚 <b>{get_translation(user_id, 'field_of_study')}:</b> {escape_html(profile['field_of_study'])}\n"
        f"⭐ <b>{get_translation(user_id, 'cgpa')}:</b> {profile['cgpa']}\n"
        f"{separator}\n"
        f"<b>{get_translation(user_id, 'skills_and_languages')}</b>\n"
        f"🗣️ <b>{get_translation(user_id, 'languages')}:</b> {escape_html(profile['languages'])}\n"
        f"🛠️ <b>{get_translation(user_id, 'skills')}:</b> {escape_html(safe_get('skills_experience'))}\n"
        f"{separator}\n"
        f"<b>{get_translation(user_id, 'profile_summary')}</b>\n"
        f"{escape_html(profile['profile_summary'])}\n"
        f"{separator}\n"
        f"<b>{get_translation(user_id, 'for_position')}:</b> {escape_html(safe_get('job_title'))}\n"
        f"{separator}\n"
        f"<b>{get_translation(user_id, 'cover_letter')}</b>\n"
        f"{escape_html(safe_get('cover_letter', get_translation(user_id, 'not_provided')))}\n"
    )

async def handle_cv_download( user_id: int, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    try:
        application_id = int(query.data.split('_')[-1])
        application = context.user_data.get("selected_applicant") or db.get_application_details(application_id)

        if not application:
            await query.answer(get_translation(user_id, 'application_not_found'), show_alert=True)
            return VIEW_APPLICATIONS

        cv_file_id = application.get("cv_path")
        if not cv_file_id:
            await query.answer(get_translation(user_id, 'no_cv_available'), show_alert=True)
            return ACCEPT_REJECT_CONFIRMATION

        try:
            await context.bot.send_document(
                chat_id=query.from_user.id,
                document=cv_file_id,
                filename=f"CV_{application.get('full_name', 'Applicant').replace(' ', '_')}.pdf",
                caption=get_translation(user_id, 'cv_caption_for_applicant').format(applicant_name=application.get('full_name', 'Applicant'))
            )
            return ACCEPT_REJECT_CONFIRMATION
        except Exception as e:
            logging.error(f"Failed to send CV: {str(e)}")
            await query.answer(get_translation(user_id, 'failed_to_send_cv'), show_alert=True)
            return ACCEPT_REJECT_CONFIRMATION

    except Exception as e:
        logging.error(f"Error in CV download handler: {str(e)}")
        await query.answer(get_translation(user_id, 'error_processing_request'), show_alert=True)
        return ACCEPT_REJECT_CONFIRMATION

async def select_applicant(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip()
    # First check if user wants to go back to main menu
    if choice == get_translation(user_id, 'back_to_main_menu'):
        return await employer_main_menu(update, context)

    try:
        # Retrieve the selected job ID from context
        job_id = context.user_data.get("selected_job_id")
        if not job_id:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_job_selected_error")
            )
            return VIEW_APPLICATIONS

        # Fetch all applications for the selected job
        applications = db.get_applications_for_job_with_title(job_id)
        validated_apps = [validate_application(app) for app in applications]

        # Validate the selected index
        selected_index = int(choice) - 1
        if not (0 <= selected_index < len(validated_apps)):
            raise ValueError("Invalid applicant number")

        # Get the selected applicant's details
        selected_app = validated_apps[selected_index]
        context.user_data["selected_applicant"] = selected_app
        job_seeker_id = selected_app["job_seeker_id"]
        job_seeker_profile = db.get_user_profile(job_seeker_id)

        # Define a separator for better readability
        separator = escape_markdown("────────────────────")

        # Format applicant details with professional style
        details = (
            f"👤 *{escape_markdown(get_translation(user_id, 'applicant_details'))}*\n"
            f"{separator}\n"
            f"• *{escape_markdown(get_translation(user_id, 'name'))}:* `{escape_markdown(selected_app['full_name'])}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'application_date'))}:* `{escape_markdown(selected_app['application_date'])}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'cv'))}:* "
            f"`{'Available' if selected_app['additional_docs'] else 'Not Provided'}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'portfolio_link'))}:* " +
            (f"[{escape_markdown(get_translation(user_id, 'view_portfolio'))}]({selected_app['portfolio_link']})\n"
             if selected_app['portfolio_link'] else f"`{escape_markdown('Not Provided')}`\n") +
            f"• *{escape_markdown(get_translation(user_id, 'gender'))}:* `{escape_markdown(selected_app['gender'])}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'contact_number'))}:* `{escape_markdown(selected_app['contact_number'])}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'dob'))}:* `{escape_markdown(job_seeker_profile.get('dob', 'N/A'))}`\n"
            f"{separator}\n"
            f"🎓 *{escape_markdown(get_translation(user_id, 'education_details'))}*\n"
            f"• *{escape_markdown(get_translation(user_id, 'qualification'))}:* `{escape_markdown(job_seeker_profile.get('qualification', 'N/A'))}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'field_of_study'))}:* `{escape_markdown(job_seeker_profile.get('field_of_study', 'N/A'))}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'cgpa'))}:* `{escape_markdown(str(job_seeker_profile.get('cgpa', 'N/A')))}`\n"
            f"{separator}\n"
            f"💼 *{escape_markdown(get_translation(user_id, 'skills_experience'))}*\n"
            f"• *{escape_markdown(get_translation(user_id, 'languages'))}:* `{escape_markdown(job_seeker_profile.get('languages', 'N/A'))}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'skills'))}:* `{escape_markdown(job_seeker_profile.get('skills_experience', 'N/A'))}`\n"
            f"• *{escape_markdown(get_translation(user_id, 'profile_summary'))}:*\n"
            f"```{escape_markdown(job_seeker_profile.get('profile_summary', 'N/A'))}```\n"  
            f"{separator}\n"
            f"• *{escape_markdown(get_translation(user_id, 'cover_letter'))}:*\n"
            f"```{escape_markdown(selected_app['cover_letter'])}```\n" 
            f"{separator}\n"
            f"📝 *{escape_markdown(get_translation(user_id, 'status'))}:* `{escape_markdown(selected_app['status'].capitalize())}`\n"
        )

        # Send the formatted details message
        await context.bot.send_message(
            chat_id=user_id,
            text=details,
            parse_mode="MarkdownV2",
            disable_web_page_preview=True,
            reply_markup=get_main_menu_keyboard(user_id)
        )

        # Send CV if available
        cv_file_id = selected_app.get("additional_docs")  # Get the file_id
        if cv_file_id:
            try:
                # Send the CV file directly using the file_id
                await context.bot.send_document(
                    chat_id=user_id,
                    document=cv_file_id,
                    caption=get_translation(user_id, "applicant_cv_caption")
                )
            except Exception as e:
                logging.error(f"Error sending CV document: {e}")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "failed_to_send_cv"),
                    reply_markup=get_main_menu_keyboard(user_id)
                )

        query = update.callback_query
        application_id = int(query.data.split('_')[1])
        application = db.get_application_details(application_id)
        # Provide Accept/Reject buttons
        keyboard = [
            [InlineKeyboardButton("✅ Accept", callback_data="accept_applicant")],
            [InlineKeyboardButton("❌ Reject", callback_data="reject_applicant")]
        ]
        if application.get('cv_path'):
            keyboard.append(
                [InlineKeyboardButton("📥 Download CV", callback_data=f"download_cv_{application_id}")]
            )

    except ValueError:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_selection_error")
        )
        return VIEW_APPLICATIONS



    except Exception as e:
        logging.error(f"Error fetching applicant details: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "failed_to_retrieve_applicant_details"),
            reply_markup=get_main_menu_keyboard(user_id)
        )

    return ACCEPT_REJECT_CONFIRMATION

async def handle_accept_reject(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    action = query.data

    try:
        # Retrieve the selected applicant
        selected_app = context.user_data.get("selected_applicant")
        if not selected_app:
            await query.edit_message_text(
                text=get_translation(user_id, "no_applicant_selected_error")
            )
            return VIEW_APPLICATIONS



        job_seeker_id = selected_app["job_seeker_id"]
        job_title = selected_app["job_title"]

        if action == "accept_applicant":
            # Prompt the employer to send an optional message
            await query.edit_message_text(
                text=get_translation(user_id, "send_optional_message_prompt")
            )
            return EMPLOYER_MESSAGE_INPUT
        elif action == "reject_applicant":
            # Prompt the employer to provide a rejection reason
            await query.edit_message_text(
                text=get_translation(user_id, "provide_rejection_reason_prompt")
            )
            return REJECTION_REASON_INPUT

    except Exception as e:
        logging.error(f"Unexpected error in handle_accept_reject: {e}")
        await query.edit_message_text(
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return VIEW_APPLICATIONS

    return VIEW_APPLICATIONS

async def handle_rejection_reason_application(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    rejection_reason = update.message.text.strip()

    try:
        # Retrieve the selected applicant
        selected_app = context.user_data.get("selected_applicant")
        if not selected_app:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_applicant_selected_error")
            )
            return VIEW_APPLICATIONS

        job_seeker_id = selected_app["job_seeker_id"]
        job_title = selected_app["job_title"]

        # Save the rejection decision in the database
        db.save_decision(
            application_id=selected_app["application_id"],
            decision="rejected",
            rejection_reason=rejection_reason
        )

        # Update status to rejected in the database - now with correct arguments
        db.update_application_status(
            application_id=selected_app["application_id"],
            status="rejected",
            rejection_reason=rejection_reason
        )

        # Notify the job seeker
        await context.bot.send_message(
            chat_id=job_seeker_id,
            text=get_translation(job_seeker_id, "application_rejected").format(
                job_title=job_title,
                reason=rejection_reason if rejection_reason else "No reason provided"
            )
        )
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "applicant_rejected_confirmation")
        )

    except Exception as e:
        logging.error(f"Unexpected error in handle_rejection_reason: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return VIEW_APPLICATIONS

    return VIEW_APPLICATIONS

async def handle_employer_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    employer_message = update.message.text.strip()

    try:
        selected_app = context.user_data.get("selected_applicant")
        if not selected_app:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_applicant_selected_error")
            )
            return VIEW_APPLICATIONS

        job_seeker_id = selected_app["job_seeker_id"]
        job_title = selected_app["job_title"]

        # Save the acceptance decision in the application_decisions table
        db.save_decision(
            application_id=selected_app["application_id"],
            decision="approved",
            employer_message=employer_message
        )

        # Update the status to approved in the applications table
        db.update_application_status(
            selected_app["application_id"],
            "approved"
        )

        applicant_message = get_translation(job_seeker_id, "application_accepted").format(
            job_title=job_title,
            message=f"\n\n{employer_message if employer_message.lower() != 'skip' else 'The employer will contact you soon.'}"
        )

        # Notify the job seeker
        await context.bot.send_message(
            chat_id=job_seeker_id,
            text=applicant_message,
            parse_mode="Markdown"
        )

        # Confirm to the employer that the applicant has been accepted
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "applicant_accepted_confirmation")
        )
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "applicant_accepted_confirmation")
        )

    except Exception as e:
        logging.error(f"Unexpected error in handle_employer_message: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return EMPLOYER_MAIN_MENU

    return EMPLOYER_MAIN_MENU


async def export_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.callback_query.from_user.id
    job_id = int(update.callback_query.data.split("_")[-1])
    try:
        # Generate Excel file
        applications = db.get_applications_for_job(job_id)
        validated_apps = [validate_application(app) for app in applications]

        # Prepare Excel data
        data = []
        for app in validated_apps:
            profile = db.get_user_profile(app["job_seeker_id"])
            # Format portfolio link as hyperlink
            portfolio_link = app.get("portfolio_link", "")
            if portfolio_link:
                portfolio_cell = f'=HYPERLINK("{portfolio_link}", "View Portfolio")'
            else:
                portfolio_cell = "Not Provided"
            # Check CV availability
            cv_status = "Available" if app.get("additional_docs") else "Not Provided"
            data.append([
                app["full_name"],
                app["application_date"],
                app["gender"],
                app["contact_number"],
                profile.get("dob", "N/A"),
                profile.get("qualification", "N/A"),
                profile.get("field_of_study", "N/A"),
                profile.get("cgpa", "N/A"),
                profile.get("languages", "N/A"),
                profile.get("skills_experience", "N/A"),
                profile.get("profile_summary", "N/A"),
                app["cover_letter"],
                portfolio_cell,
                cv_status,
                app["status"].capitalize()
            ])

        # Create DataFrame
        df = pd.DataFrame(data, columns=[
            "Full Name", "Application Date", "Gender", "Contact Number", "Date of Birth",
            "Qualification", "Field of Study", "CGPA", "Languages", "Skills",
            "Profile Summary", "Cover Letter", "Portfolio Link", "CV Available", "Status"
        ])

        # Replace NaN and INF values in the DataFrame
        df.replace([float('inf'), float('-inf')], "N/A", inplace=True)  # Replace INF with "N/A"
        df.fillna("N/A", inplace=True)  # Replace NaN with "N/A"

        # Use xlsxwriter for better formatting
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name="Applicants")

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Applicants']

            # Enable nan_inf_to_errors option in xlsxwriter
            workbook.use_future_functions = True  # Required for some advanced features
            workbook.nan_inf_to_errors = True  # Handle NaN/INF values

            # Add title section
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 18,  # Larger font size for title
                'font_color': '#FFFFFF',
                'align': 'center',
                'valign': 'vcenter',
                'fg_color': '#002060'
            })
            worksheet.merge_range('A1:O2', f"📊 Applicants Report for Job ID: {job_id}", title_format)

            # Add headers with style
            header_format = workbook.add_format({
                'bold': True,
                'font_size': 14,  # Larger font size for headers
                'font_color': '#FFFFFF',
                'align': 'center',
                'valign': 'vcenter',
                'fg_color': '#002060',
                'border': 1
            })
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(2, col_num, value, header_format)

            # Add data rows with alternating colors
            even_row_format = workbook.add_format({
                'align': 'top',
                'valign': 'vcenter',
                'border': 1,
                'fg_color': '#E6E6E6',  # Slightly darker gray for better visibility
                'text_wrap': True  # Enable text wrapping
            })
            odd_row_format = workbook.add_format({
                'align': 'top',
                'valign': 'vcenter',
                'border': 1,
                'fg_color': '#FFFFFF',  # White for alternating rows
                'text_wrap': True  # Enable text wrapping
            })

            # Set column widths
            for col_num, column in enumerate(df.columns):
                if column in ["Profile Summary", "Cover Letter"]:
                    worksheet.set_column(col_num, col_num, 50)  # Fixed width for long text columns
                else:
                    # Auto-fit width for other columns based on header length
                    header_length = len(column)
                    worksheet.set_column(col_num, col_num, header_length + 5)  # Add padding

            for row_num in range(3, len(df) + 3):
                for col_num in range(len(df.columns)):
                    cell_value = df.iat[row_num - 3, col_num]
                    cell_format = even_row_format if row_num % 2 == 0 else odd_row_format

                    # Highlight "N/A" values in red and italic
                    if str(cell_value).strip() == "N/A":
                        cell_format = workbook.add_format({
                            'italic': True,
                            'font_color': '#FF0000',
                            'align': 'top',
                            'valign': 'vcenter',
                            'border': 1,
                            'text_wrap': True
                        })

                    # Format portfolio links as hyperlinks
                    if col_num == 12 and cell_value != "Not Provided":
                        worksheet.write_url(
                            row_num, col_num,
                            cell_value.split('"')[1],
                            string='View Portfolio',
                            cell_format=workbook.add_format({
                                'font_color': 'blue',
                                'underline': 1,
                                'align': 'top',
                                'valign': 'vcenter',
                                'border': 1
                            })
                        )
                    else:
                        worksheet.write(row_num, col_num, cell_value, cell_format)

                # Adjust row height based on content length
                profile_summary = df.iat[row_num - 3, 10]  # Column K (Profile Summary)
                cover_letter = df.iat[row_num - 3, 11]  # Column L (Cover Letter)
                max_lines = max(len(str(profile_summary).split('\n')), len(str(cover_letter).split('\n')))
                worksheet.set_row(row_num, 15 * max_lines)  # Adjust row height dynamically

            # Freeze headers
            worksheet.freeze_panes(3, 0)

            # Add Summary Sheet
            summary_sheet = workbook.add_worksheet("Summary")
            summary_sheet.merge_range('A1:B1', "📊 Applicants Summary Report", title_format)

            # Add summary data with safe calculations
            average_cgpa = round(df[df['CGPA'] != "N/A"]['CGPA'].astype(float).mean(), 2)
            average_cgpa = average_cgpa if not pd.isna(average_cgpa) else "N/A"

            summary_data = [
                ["Total Applicants", len(df)],
                ["Average CGPA", average_cgpa],
                ["Top 5 Qualifications", ', '.join(df['Qualification'].value_counts().head(5).index)],
                ["Most Common Languages", ', '.join(df['Languages'].value_counts().head(5).index)],
                ["Most Popular Fields of Study", ', '.join(df['Field of Study'].value_counts().head(5).index)],
                ["CVs Available", df['CV Available'].value_counts().get("Available", 0)],
                ["CVs Not Provided", df['CV Available'].value_counts().get("Not Provided", 0)]
            ]

            for row_num, row_data in enumerate(summary_data, 3):
                for col_num, value in enumerate(row_data):
                    summary_sheet.write(row_num, col_num, value, header_format if row_num == 3 else None)

            # Auto-fit summary sheet columns
            summary_sheet.set_column('A:B', 30)

        excel_file.seek(0)

        # Send Excel file
        await context.bot.send_document(
            chat_id=user_id,
            document=excel_file,
            filename=f"applicants_job_{job_id}.xlsx"
        )

        # Send individual CVs
        for app in validated_apps:
            if app.get("additional_docs"):
                try:
                    await context.bot.send_document(
                        chat_id=user_id,
                        document=app["additional_docs"],
                        caption=f"{app['full_name']}'s CV"
                    )
                except Exception as e:
                    logging.error(f"Failed to send CV for {app['full_name']}: {e}")
                    await context.bot.send_message(
                        chat_id=user_id,
                        text=f"⚠️ Failed to send CV for {app['full_name']}"
                    )

    except Exception as e:
        logging.error(f"Export error: {e}", exc_info=True)
        await context.bot.send_message(
            chat_id=user_id,
            text="Export failed. Please try again later."
        )
async def fetch_and_validate_applications(job_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE) -> list:
    """Fetch and validate applications for a specific job."""
    try:
        applications = db.get_applications_for_job(job_id)
        if not applications:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_applications_found").format(job_id=job_id)
            )
            return []
        validated_apps = [validate_application(app) for app in applications]
        return validated_apps
    except ValueError as ve:
        logging.error(f"Invalid application data for job {job_id}: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_fetching_applications")
        )
        return []
    except Exception as e:
        logging.error(f"Error fetching applications for job {job_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred")
        )
        return []

async def confirm_close(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        # Extract job ID from callback data
        job_id = validate_callback_data(query.data, "close")

        # Validate job ID exists and belongs to the employer
        job_type, job_data = validate_job_ownership(db, job_id, user_id)
        if not job_type:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "unauthorized_access")
            )
            return EMPLOYER_MAIN_MENU

        # Store job ID in user_data
        context.user_data["close_job_id"] = job_id

        # Confirm closing the vacancy
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "confirm_close_vacancy_prompt").format(job_id=job_id)
        )

    except ValueError as ve:
        logging.error(f"ValueError in confirm_close: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_action_detected", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in confirm_close: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred")
        )

    return CONFIRM_CLOSE

async def handle_close_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip().lower()

    try:
        if choice == "yes":
            job_id = context.user_data.pop("close_job_id", None)
            if not job_id:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "error_closing_vacancy")
                )
                return EMPLOYER_MAIN_MENU

            # Close the job post or vacancy
            try:
                db.update_vacancy_status(job_id, "closed")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "vacancy_closed_successfully").format(job_id=job_id)
                )
            except Exception as e:
                logging.error(f"Error closing vacancy {job_id}: {e}")
                await context.bot.send_message(
                    chat_id=user_id,
                    text=get_translation(user_id, "error_closing_vacancy")
                )
                return EMPLOYER_MAIN_MENU

        elif choice == "no":
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "closing_canceled")
            )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "invalid_choice")
            )
            return CONFIRM_CLOSE

    except Exception as e:
        logging.error(f"Unexpected error in handle_close_confirmation: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )

    # Refresh the list of vacancies
    try:
        await manage_vacancies(update, context)
    except Exception as e:
        logging.error(f"Error refreshing vacancies: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_refreshing_vacancies")
        )

    return EMPLOYER_MAIN_MENU

def validate_callback_data(callback_data: str, prefixes: tuple) -> int:
    """
    Validate and extract job_id from callback_data.
    """
    try:
        logging.debug(f"Processing callback_data: {callback_data}")  # Log the raw callback_data
        logging.debug(f"Using prefixes: {prefixes}")  # Log the prefixes being checked

        for prefix in prefixes:
            if callback_data.startswith(prefix):
                # Correctly remove the prefix without removing valid numbers
                job_id = callback_data[len(prefix):]

                logging.debug(f"Extracted job_id: {job_id}")  # Log the extracted job_id

                # Ensure job_id is numeric
                if not job_id.isdigit():
                    raise ValueError(f"Invalid job_id: {job_id}")

                return int(job_id)

        raise ValueError(f"Invalid callback_data: {callback_data}. Expected one of prefixes: {prefixes}")

    except Exception as e:
        logging.error(f"Error validating callback_data: {e}")
        return None

def validate_application(application: dict):
    """Validate the structure of an application dictionary."""
    required_fields = [
        "application_id", "job_seeker_id", "full_name", "cover_letter",
        "application_date", "status", "job_id", "additional_docs", "portfolio_link",
        "gender", "contact_number",
        "languages", "qualification", "field_of_study",
        "cgpa", "skills", "profile_summary"
    ]
    missing_fields = [field for field in required_fields if field not in application]
    if missing_fields:
        raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

    # Validate application_date format
    try:
        datetime.strptime(application["application_date"], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise ValueError(f"Invalid date format: {application['application_date']}")

    # Validate status
    if application["status"] not in ("pending", "reviewed", "approved", "rejected"):
        raise ValueError(f"Invalid status: {application['status']}")

    # Exclude withdrawn applications from further processing
    if application["status"] == "withdrawn":
        return None

    return application

async def VIEW_APPLICATIONS(job_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        # Fetch and display applicants
        await fetch_and_display_applicants(job_id, user_id, context)
    except ValueError as ve:
        logging.error(f"ValueError in VIEW_APPLICATIONS: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_data_detected", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in VIEW_APPLICATIONS: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )

async def fetch_and_display_applicants(update: Update, job_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        applications = db.get_applications_for_job(job_id)
        if not applications:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'no_applications_found_for_job').format(job_id=job_id)
            )
            return

        validated_apps = []
        for app in applications:
            validated_app = validate_application(app)
            if validated_app is not None:  # Only include non-withdrawn applications
                validated_apps.append(validated_app)

        # Format the list of applicants
        message = ""
        for idx, app in enumerate(validated_apps, start=1):
            message += f"{idx}. {app['full_name']} ({app['application_date']})\n"

        if message:
            keyboard = [
                [InlineKeyboardButton(get_translation(user_id, 'export_to_excel'), callback_data=f"export_excel_{job_id}")],
                [InlineKeyboardButton(get_translation(user_id, 'back_to_manage_vacancies'), callback_data="back_to_manage_vacancies")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)

            await context.bot.send_message(
                chat_id=user_id,
                text=message + "\n" + get_translation(user_id, 'select_applicant_by_number'),
                reply_markup=reply_markup
            )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'no_valid_applications_found_for_job').format(job_id=job_id)
            )

    except Exception as e:
        logging.error(f"Error fetching applicants for job {job_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'error_fetching_application').format(error=str(e))
        )

async def handle_view_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        # Extract job ID from callback data
        parts = query.data.split("_")
        if len(parts) < 3:
            raise ValueError("Invalid callback data format.")
        job_id = parts[2]

        # Validate job ID exists and is open/approved
        job_post = db.get_job_post_by_id(job_id)
        if not job_post or job_post["status"] not in ("approved", "open"):
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "job_not_open").format(job_id=job_id)
            )
            return EMPLOYER_MAIN_MENU

        # Fetch and display applicants
        await fetch_and_display_applicants(job_id, user_id, context)

    except ValueError as ve:
        logging.error(f"ValueError in handle_view_applications: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_data_detected", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in handle_view_applications: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )

    return VIEW_APPLICATIONS

async def confirm_resubmit(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    job_id = context.user_data.get("resubmit_job_id")

    try:
        if not job_id:
            raise ValueError("Job ID not found in user data.")

        # Validate that the job post exists and belongs to the employer
        job_type, job_data = validate_job_ownership(db, job_id, user_id)
        if not job_type or job_type != "job_post":
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "job_not_found").format(job_id=job_id)
            )
            return EMPLOYER_MAIN_MENU

        # Reset the job post status to 'pending' and clear rejection reason
        db.resubmit_job_post(job_id)
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "vacancy_resubmitted_successfully").format(job_id=job_id)
        )

    except ValueError as ve:
        logging.error(f"ValueError in confirm_resubmit: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_data_detected", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Error resubmitting job post {job_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_resubmitting_vacancy")
        )
    finally:
        # Clear user_data and refresh the list of vacancies
        context.user_data.pop("resubmit_job_id", None)
        try:
            await manage_vacancies(update, context)
        except Exception as e:
            logging.error(f"Error refreshing vacancies: {e}")
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "error_refreshing_vacancies")
            )
    return EMPLOYER_MAIN_MENU


async def fetch_and_display_vacancies(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> list:
    """
    Fetch and display all job posts (pending and approved) for the specified employer.
    """
    try:
        # Fetch job posts for the employer
        job_posts = db.get_job_posts_by_employer(user_id)
        if not job_posts:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_job_posts_found")
            )
            return []

        # Validate and process job posts
        validated_jobs = []
        for job in job_posts:
            try:
                validated_job = validate_job_post(dict(zip([col[0] for col in db.cursor.description], job)))
                validated_jobs.append(validated_job)
            except ValueError as ve:
                logging.warning(f"Skipping invalid job post: {ve}")
                continue

        # Format job titles with unique numbers and statuses
        message = get_translation(user_id, "vacancies_list_header") + "\n"
        for idx, job in enumerate(validated_jobs, start=1):
            message += f"{idx}. {job['job_title']} ({job['status'].capitalize()})\n"

        # Store validated jobs in user_data
        context.user_data["vacancies"] = validated_jobs

        # Send the list of job posts to the employer
        await context.bot.send_message(
            chat_id=user_id,
            text=message + "\n" + get_translation(user_id, "select_vacancy_prompt"),
            reply_markup=ReplyKeyboardRemove()
        )

        return validated_jobs

    except Exception as e:
        logging.error(f"Error fetching and displaying job posts: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_fetching_jobs", error=str(e))
        )
        return []

from io import BytesIO


def export_applications_to_excel(applications: list, job_title: str, user_id: int) -> BytesIO:
    """
    Export all applications for a job post to an Excel file.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = job_title

        # Define headers with localization
        headers = [
            get_translation(user_id, "applicant_name"),
            get_translation(user_id, "cover_letter"),
            get_translation(user_id, "unique_number"),
            get_translation(user_id, "application_date"),
            get_translation(user_id, "status")
        ]
        ws.append(headers)

        # Validate and process applications
        validated_apps = [validate_application(app) for app in applications]

        # Add application data
        for app in validated_apps:
            ws.append([
                app["full_name"],
                app["cover_letter"][:50],  # Truncate cover letter for better readability
                app["application_id"],
                app["application_date"],
                app["status"].capitalize()
            ])

        # Save workbook to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

    except ValueError as ve:
        logging.error(f"Invalid application data: {ve}")
        raise  # Re-raise the exception to allow the caller to handle it
    except Exception as e:
        logging.error(f"Error exporting applications to Excel: {e}")
        raise  # Re-raise the exception to allow the caller to handle it


async def view_analytics(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Get comprehensive analytics data
    analytics_data = db.get_employer_analytics(user_id)
    profile = db.get_employer_profile(user_id)
    company_name = profile.get('company_name', get_translation(user_id, 'your_company'))

    # Create visually rich analytics message
    analytics_msg = f"""
📊 <b>{company_name} - {get_translation(user_id, 'analytics_dashboard_title')}</b>
━━━━━━━━━━━━━━━━━━━━━━━━

<b>📈 {get_translation(user_id, 'performance_overview')}</b>
🟢 <b>{get_translation(user_id, 'active_vacancies')}:</b> {analytics_data.get('active_vacancies', 0)}
📨 <b>{get_translation(user_id, 'total_applications')}:</b> {analytics_data.get('total_applications', 0)}
✅ <b>{get_translation(user_id, 'hire_rate')}:</b> {analytics_data.get('hire_rate', 0)}%
⏱️ <b>{get_translation(user_id, 'avg_response_time')}:</b> {analytics_data.get('avg_response_time', 0)} {get_translation(user_id, 'days')}
📅 <b>{get_translation(user_id, 'member_since')}:</b> {analytics_data.get('member_since', get_translation(user_id, 'not_available'))}

<b>📊 {get_translation(user_id, 'application_flow')}</b>
📥 <b>{get_translation(user_id, 'new_applications')}:</b> {analytics_data.get('pending_applications', 0)}
👁️ <b>{get_translation(user_id, 'viewed_applications')}:</b> {analytics_data.get('reviewed_applications', 0)}
✅ <b>{get_translation(user_id, 'approved_applications')}:</b> {analytics_data.get('approved_applications', 0)}
❌ <b>{get_translation(user_id, 'rejected_applications')}:</b> {analytics_data.get('rejected_applications', 0)}

<b>📅 {get_translation(user_id, 'recent_activity')}</b>
{format_recent_activity(analytics_data.get('recent_activity', []))}

<b>💡 {get_translation(user_id, 'tips_for_improvement')}</b>
{get_analytics_tips(analytics_data, user_id)}
    """

    # Create interactive keyboard
    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'performance_trends'), callback_data="analytics_trends")],
        [InlineKeyboardButton(get_translation(user_id, 'candidate_demographics'), callback_data="analytics_demographics")],
        [InlineKeyboardButton(get_translation(user_id, 'response_time_analysis'), callback_data="analytics_response")],
        [InlineKeyboardButton(get_translation(user_id, 'compare_to_peers'), callback_data="analytics_benchmark")],
        [InlineKeyboardButton(get_translation(user_id, 'export_data'), callback_data="analytics_export")],
        [InlineKeyboardButton(get_translation(user_id, 'back_to_main_menu'), callback_data="go_to_employer_main_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        # Try to generate and send a chart
        chart = generate_analytics_chart(analytics_data)
        if chart:
            await context.bot.send_photo(
                chat_id=user_id,
                photo=chart,
                caption=analytics_msg,
                parse_mode="HTML",
                reply_markup=reply_markup
            )
        else:
            # Fallback to text-only message if chart generation fails
            await context.bot.send_message(
                chat_id=user_id,
                text=analytics_msg,
                parse_mode="HTML",
                reply_markup=reply_markup
            )
    except Exception as e:
        logging.error(f"Error sending analytics: {e}")
        # Fallback to simple message if anything fails
        await context.bot.send_message(
            chat_id=user_id,
            text=analytics_msg,
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    return ANALYTICS_VIEW


def generate_analytics_chart(data: dict):
    """Generate a simple bar chart of key metrics"""
    try:
        import matplotlib.pyplot as plt
        import io

        # Prepare data
        labels = ['Vacancies', 'Applications', 'Hire Rate']
        values = [
            data.get('active_vacancies', 0),
            data.get('total_applications', 0),
            data.get('hire_rate', 0)
        ]

        # Create figure
        fig, ax = plt.subplots(figsize=(10, 5))
        bars = ax.bar(labels, values)

        # Add value labels
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{height}',
                    ha='center', va='bottom')

        ax.set_title('Key Performance Metrics')
        plt.tight_layout()

        # Save to bytes buffer
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=80)
        plt.close()
        buf.seek(0)

        return buf
    except Exception as e:
        logging.error(f"Chart generation error: {e}")
        return None

def format_recent_activity(activities: list) -> str:
    """Format recent activity into readable lines"""
    if not activities:
        return "No recent activity"

    formatted = []
    icons = {
        'application': '📨',
        'approval': '✅',
        'rejection': '❌',
        'post': '📢',
        'view': '👁️'
    }

    for activity in activities[:5]:  # Show last 5 activities
        icon = icons.get(activity['type'], '⚪')
        formatted.append(f"{icon} {activity['date']}: {activity['description']}")

    return "\n".join(formatted)


def get_analytics_tips(data: dict, user_id: int) -> str:
    """Generate personalized tips based on analytics"""
    tips = []

    if data['hire_rate'] < 20:
        tips.append(f"• {get_translation(user_id, 'tip_refine_job_descriptions')}")

    if data['avg_response_time'] > 7:
        tips.append(f"• {get_translation(user_id, 'tip_improve_response_time')}")

    if data['pending_applications'] > 10:
        tips.append(f"• {get_translation(user_id, 'tip_review_pending_applications').format(pending=data['pending_applications'])}")

    if not data['active_vacancies']:
        tips.append(f"• {get_translation(user_id, 'tip_post_new_jobs')}")

    if not tips:
        return get_translation(user_id, 'tip_keep_up_good_work')

    return "\n".join(tips)

async def handle_analytics_actions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle analytics sub-menu actions"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    action = query.data

    if action == "analytics_trends":
        await show_performance_trends(update, context)
    elif action == "analytics_demographics":
        await show_demographics(update, context)
    elif action == "analytics_response":
        await show_response_analysis(update, context)
    elif action == "analytics_benchmark":
        await show_benchmark_comparison(update, context)
    elif action == "analytics_export":
        await show_export_options(update, context)
        return ANALYTICS_EXPORT
    elif action == "analytics_back":
        return await view_analytics(update, context)
    elif action.startswith("export_"):  # Add this line to handle export actions directly
        return await handle_export_format(update, context)
    elif action == "go_to_employer_main_menu":
        return await employer_main_menu(update, context)

    return ANALYTICS_VIEW



async def show_export_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show export format selection menu with comprehensive error handling"""
    query = update.callback_query
    await query.answer()

    # Store user ID for easier access
    user_id = query.from_user.id

    # Create the export options keyboard
    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'export_csv'), callback_data="export_csv")],
        [InlineKeyboardButton(get_translation(user_id, 'export_excele'), callback_data="export_excel")],
        [InlineKeyboardButton(get_translation(user_id, 'export_pdf'), callback_data="export_pdf")],
        [InlineKeyboardButton(get_translation(user_id, 'back_to_analytics'), callback_data="back_to_analytics")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # First attempt: Try to edit the existing message
    try:
        if query.message and query.message.text:
            await query.edit_message_text(
                text=get_translation(user_id, 'select_export_format'),
                reply_markup=reply_markup
            )
            return
    except Exception as edit_error:
        logging.warning(f"Message edit failed for user {user_id}: {str(edit_error)}")

    # Second attempt: Try to delete then send new message
    try:
        # Clean up old message if it exists
        if query.message:
            try:
                await query.message.delete()
            except Exception as delete_error:
                logging.debug(f"Couldn't delete old message: {str(delete_error)}")

        # Send fresh message
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'select_export_format'),
            reply_markup=reply_markup
        )
    except Exception as send_error:
        logging.error(f"Failed to send new message to user {user_id}: {str(send_error)}")
        # Ultimate fallback - send plain text instructions
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'fallback_export_instructions')
        )

async def show_performance_trends(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show performance trends over time"""
    user_id = get_user_id(update)
    trends = db.get_performance_trends(user_id)

    msg = """
📈 <b>{performance_trends_title}</b>
━━━━━━━━━━━━━━━━━━━━━━━━

<b>{applications_over_time}:</b>
{applications_chart}

<b>{hire_rate_trend}:</b>
{hire_rate_chart}

<b>{response_time_trend}:</b>
{response_time_chart}
    """.format(
        performance_trends_title=get_translation(user_id, 'performance_trends_title'),
        applications_over_time=get_translation(user_id, 'applications_over_time'),
        hire_rate_trend=get_translation(user_id, 'hire_rate_trend'),
        response_time_trend=get_translation(user_id, 'response_time_trend'),
        applications_chart=generate_sparkline(trends['applications']),
        hire_rate_chart=generate_sparkline(trends['hire_rate']),
        response_time_chart=generate_sparkline(trends['response_time'], inverse=True)
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=msg,
        parse_mode="HTML"
    )

def generate_sparkline(data: list, inverse: bool = False) -> str:
    """Generate text-based sparkline with robust error handling"""
    if not data or not isinstance(data, list):
        return "📊"  # Return simple placeholder

    try:
        # Extract values safely
        values = []
        for item in data:
            if isinstance(item, dict):
                value = item.get('count', item.get('rate', item.get('days', 0)))
            else:
                value = float(item) if str(item).replace('.', '', 1).isdigit() else 0
            values.append(value)

        if not values:
            return "📊"

        # Normalize to 0-10 scale
        min_val, max_val = min(values), max(values)
        if max_val == min_val:
            normalized = [5] * len(values)  # All values are equal, map to middle
        else:
            normalized = [int(10 * (v - min_val) / (max_val - min_val)) for v in values]

        # Generate sparkline
        spark_chars = ['▁', '▂', '▃', '▄', '▅', '▆', '▇', '█']
        if inverse:
            spark_chars = spark_chars[::-1]  # Reverse sparkline characters

        # Map normalized values to sparkline characters
        return ''.join(
            [spark_chars[min(len(spark_chars) - 1, max(0, v // (10 // len(spark_chars))))] for v in normalized])

    except Exception as e:
        logging.error(f"Sparkline generation error: {e}")
        return "📊"  # Fallback to simple emoji

def compare_metric(user_value: float, benchmark_value: float, lower_better: bool = False) -> str:
    """Generate comparison text for benchmarks"""
    if user_value == benchmark_value:
        return "→ Equal to industry average"
    elif (user_value > benchmark_value and not lower_better) or (user_value < benchmark_value and lower_better):
        diff = abs(user_value - benchmark_value)
        return f"↑ {diff:.1f} better than average"
    else:
        diff = abs(user_value - benchmark_value)
        return f"↓ {diff:.1f} below average"


def format_recent_responses(responses: list) -> str:
    """Format recent responses for display"""
    if not responses:
        return "No recent responses"

    formatted = []
    for resp in responses[:5]:  # Show last 5 responses
        try:
            days = resp['days']
            emoji = "🚀" if days < 3 else "🐢" if days > 7 else "⏱️"
            formatted.append(f"{emoji} {resp.get('job_title', 'Unknown Job')}: {days:.1f} days")
        except (KeyError, TypeError) as e:
            logging.error(f"Error formatting response: {e}")
            formatted.append("⚠️ Invalid response data")

    return '\n'.join(formatted)


async def show_demographics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show candidate demographics with error handling"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    try:
        demographics = db.get_candidate_demographics(user_id)

        if not demographics:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'no_demographic_data_available'),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(get_translation(user_id, 'back_to_analytics'),
                                          callback_data="back_to_analytics")]
                ])
            )
            return ANALYTICS_VIEW

        # Generate the chart
        chart_buffer = generate_demographics_chart(demographics)

        # Send the chart
        await context.bot.send_photo(
            chat_id=user_id,
            photo=chart_buffer,
            caption=format_demographics_message(demographics),
            parse_mode="HTML"
        )

        # Close the buffer
        chart_buffer.close()

    except Exception as e:
        logging.error(f"Error in show_demographics: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Could not generate demographics chart",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back_to_analytics'), callback_data="back_to_analytics")]
            ])
        )

    return ANALYTICS_VIEW


def format_demographics_message(data: dict) -> str:
    """Format demographics data into a nicely structured message"""
    gender_dist = data.get('gender', {})
    education_dist = data.get('education', {})
    experience_dist = data.get('experience', {})

    message = [
        "👥 <b>Candidate Demographics</b>",
        "━━━━━━━━━━━━━━━━━━━━━━━━",
        "",
        "<b>Gender Distribution:</b>",
        f"🚹 Male: {gender_dist.get('Male', 0)}%",
        f"🚺 Female: {gender_dist.get('Female', 0)}%",
        f"⚧ Other: {gender_dist.get('Other', 0)}%",
        "",
        "<b>Education Level:</b>",
        "🎓 " + format_distribution(education_dist),
        "",
        "<b>Experience Level:</b>",
        "💼 " + format_distribution(experience_dist)
    ]

    return "\n".join(message)
def format_distribution(data: dict) -> str:
    """Format distribution data for display"""
    return " | ".join([f"{k}: {v}%" for k, v in data.items()])

# Add this helper function for navigation back to analytics
async def handle_analytics_back(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle back to analytics navigation"""
    query = update.callback_query
    await query.answer()
    await view_analytics(update, context)
    return ANALYTICS_VIEW


def generate_demographics_chart(demographics: dict):
    """Generate a demographics chart image using matplotlib without file operations"""
    try:
        # Create figure with subplots
        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        fig.suptitle('Candidate Demographics Overview')

        # Gender pie chart
        if demographics.get('gender'):
            genders = list(demographics['gender'].keys())
            sizes = list(demographics['gender'].values())
            axes[0].pie(sizes, labels=genders, autopct='%1.1f%%', startangle=90)
            axes[0].set_title('Gender Distribution')

        # Education bar chart
        if demographics.get('education'):
            educations = list(demographics['education'].keys())
            values = list(demographics['education'].values())
            axes[1].bar(educations, values)
            axes[1].set_title('Education Level')
            axes[1].tick_params(axis='x', rotation=45)

        # Experience bar chart
        if demographics.get('experience'):
            experiences = list(demographics['experience'].keys())
            exp_values = list(demographics['experience'].values())
            axes[2].bar(experiences, exp_values)
            axes[2].set_title('Experience Level')
            axes[2].tick_params(axis='x', rotation=45)

        # Save to bytes buffer
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=100)
        plt.close(fig)
        buf.seek(0)

        return buf

    except Exception as e:
        logging.error(f"Error generating demographics chart: {e}")
        plt.close('all')
        raise

from telegram import InputFile


async def handle_export_format(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle export format selection and file generation"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    export_format = query.data.split('_')[1]  # csv, pdf, or excel

    try:
        # Notify user export is starting
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⏳ Preparing your {export_format.upper()} export..."
        )

        # Get analytics data
        analytics_data = db.get_employer_analytics(user_id)
        if not analytics_data:
            raise ValueError("No analytics data found")

        company_name = db.get_employer_profile(user_id).get('company_name', 'Your Company')
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

        # Generate and send the export file
        if export_format == 'csv':
            file_data = generate_csv_export(analytics_data, company_name, timestamp)
            await context.bot.send_document(
                chat_id=user_id,
                document=InputFile(file_data, filename=f"{company_name}_Analytics_{timestamp}.csv"),
                caption=f"📊 {company_name} Analytics (CSV)"
            )
        elif export_format == 'excel':
            file_data = generate_excel_export(analytics_data, company_name, timestamp)
            await context.bot.send_document(
                chat_id=user_id,
                document=InputFile(file_data, filename=f"{company_name}_Analytics_{timestamp}.xlsx"),
                caption=f"📈 {company_name} Analytics (Excel)"
            )
        elif export_format == 'pdf':
            file_data = generate_pdf_export(analytics_data, company_name, timestamp)
            await context.bot.send_document(
                chat_id=user_id,
                document=InputFile(file_data, filename=f"{company_name}_Analytics_{timestamp}.pdf"),
                caption=f"📄 {company_name} Analytics (PDF)"
            )

        # Send completion message
        await context.bot.send_message(
            chat_id=user_id,
            text=f"✅ {export_format.upper()} export completed!",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back_to_analytics'), callback_data="back_to_analytics")]
            ])
        )

    except Exception as e:
        logging.error(f"Export error for user {user_id}: {str(e)}", exc_info=True)
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Failed to generate export. Please try again later.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back_to_analytics'), callback_data="back_to_analytics")]
            ])
        )

    return ANALYTICS_VIEW


async def export_analytics_data(update: Update, context: ContextTypes.DEFAULT_TYPE, format: str):
    """Simplified export handler"""
    user_id = update.effective_user.id
    try:
        await context.bot.send_message(
            chat_id=user_id,
            text=f"📤 Exporting data as {format.upper()}...",
        )

        # Generate simple CSV as fallback
        if format == 'csv':
            data = "Metric,Value\nActive Vacancies,0\nTotal Applications,0"
            await context.bot.send_document(
                chat_id=user_id,
                document=io.BytesIO(data.encode()),
                filename=f"analytics_export_{datetime.now().strftime('%Y%m%d')}.csv"
            )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"⚠️ {format.upper()} export is currently unavailable",
            )

    except Exception as e:
        logging.error(f"Export error: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Export failed",
        )


def generate_csv_export(data: dict, company_name: str, timestamp: str) -> io.BytesIO:
    """Generate CSV export in memory"""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Header
    writer.writerow([f"{company_name} - Analytics Export ({timestamp})"])
    writer.writerow([])

    # Performance Metrics
    writer.writerow(["PERFORMANCE METRICS"])
    writer.writerow(["Active Vacancies", data.get('active_vacancies', 0)])
    writer.writerow(["Total Applications", data.get('total_applications', 0)])
    writer.writerow(["Hire Rate (%)", data.get('hire_rate', 0)])
    writer.writerow(["Avg Response Time (days)", data.get('avg_response_time', 0)])
    writer.writerow([])

    # Application Flow
    writer.writerow(["APPLICATION FLOW"])
    writer.writerow(["Pending Applications", data.get('pending_applications', 0)])
    writer.writerow(["Reviewed Applications", data.get('reviewed_applications', 0)])
    writer.writerow(["Approved Applications", data.get('approved_applications', 0)])
    writer.writerow(["Rejected Applications", data.get('rejected_applications', 0)])
    writer.writerow([])

    # Convert to BytesIO for Telegram
    mem_file = io.BytesIO(buffer.getvalue().encode('utf-8'))
    buffer.close()
    return mem_file


from xlsxwriter.utility import xl_range
import io


def generate_excel_export(data: dict, company_name: str, timestamp: str) -> io.BytesIO:
    """Generate advanced formatted Excel export with professional styling"""
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    # Create formats
    header_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'bg_color': '#4F81BD',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    title_format = workbook.add_format({
        'bold': True,
        'font_size': 20,
        'font_color': '#1F497D',
        'align': 'center',
        'valign': 'vcenter'
    })

    subtitle_format = workbook.add_format({
        'italic': True,
        'font_color': '#7F7F7F',
        'align': 'center',
        'valign': 'vcenter'
    })

    section_format = workbook.add_format({
        'bold': True,
        'font_size': 12,
        'bg_color': '#DCE6F1',
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    metric_label_format = workbook.add_format({
        'bold': True,
        'font_color': '#002060',
        'bg_color': '#F2F2F2',
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    metric_value_format = workbook.add_format({
        'num_format': '#,##0',
        'bg_color': '#FFFFFF',
        'border': 1,
        'align': 'right',
        'valign': 'vcenter'
    })

    alt_row_format = workbook.add_format({
        'bg_color': '#F8F8F8',
        'border': 1
    })

    # Add worksheet with better name
    worksheet = workbook.add_worksheet("Analytics Report")

    # Set column widths
    worksheet.set_column('A:A', 30)  # Label column
    worksheet.set_column('B:B', 15)  # Value column

    # Title section
    worksheet.merge_range('A1:B1', f"{company_name} Analytics Report", title_format)
    worksheet.merge_range('A2:B2', f"Generated on {timestamp}", subtitle_format)
    worksheet.set_row(0, 30)  # Increase title row height
    worksheet.set_row(1, 20)  # Increase subtitle row height

    # Performance Metrics section
    worksheet.write('A4', "PERFORMANCE METRICS", section_format)
    metrics = [
        ("Active Vacancies", data.get('active_vacancies', 0)),
        ("Total Applications", data.get('total_applications', 0)),
        ("Hire Rate (%)", data.get('hire_rate', 0)),
        ("Avg Response Time (days)", data.get('avg_response_time', 0))
    ]

    for row, (label, value) in enumerate(metrics, start=5):
        format = metric_label_format if row % 2 == 0 else alt_row_format
        worksheet.write(f'A{row}', label, format)
        worksheet.write(f'B{row}', value, metric_value_format)

    # Application Flow section
    worksheet.write('A9', "APPLICATION FLOW", section_format)
    flow = [
        ("Pending Applications", data.get('pending_applications', 0)),
        ("Reviewed Applications", data.get('reviewed_applications', 0)),
        ("Approved Applications", data.get('approved_applications', 0)),
        ("Rejected Applications", data.get('rejected_applications', 0))
    ]

    for row, (label, value) in enumerate(flow, start=10):
        format = metric_label_format if row % 2 == 0 else alt_row_format
        worksheet.write(f'A{row}', label, format)
        worksheet.write(f'B{row}', value, metric_value_format)

    # Add borders around all used cells
    max_row = 13  # Update based on your data length
    worksheet.conditional_format(
        f'A1:B{max_row}',
        {'type': 'formula', 'criteria': 'TRUE', 'format': workbook.add_format({'border': 1})}
    )

    # Add chart for performance metrics
    chart = workbook.add_chart({'type': 'column'})
    # In the generate_excel_export function's chart creation section:
    chart.add_series({
        'name': 'Performance Metrics',
        'categories': f"='Analytics Report'!$A$5:$A$8",  # Added quotes around sheet name
        'values': f"='Analytics Report'!$B$5:$B$8",  # Added quotes around sheet name
        'fill': {'color': '#4F81BD'},
        'data_labels': {'value': True}
    })
    chart.set_title({'name': 'Key Performance Indicators'})
    chart.set_legend({'none': True})
    worksheet.insert_chart('D4', chart)

    # Freeze panes and set zoom
    worksheet.freeze_panes(4, 1)
    worksheet.set_zoom(85)

    # Close workbook and prepare output
    workbook.close()
    output.seek(0)
    return output


from fpdf import FPDF
from fpdf.enums import XPos, YPos


def generate_pdf_export(data: dict, company_name: str, timestamp: str) -> io.BytesIO:
    """Generate PDF export in memory"""
    buffer = io.BytesIO()
    pdf = FPDF()
    pdf.add_page()

    # Set font (use Helvetica instead of Arial)
    pdf.set_font("Helvetica", size=12)

    # Header
    pdf.set_font("Helvetica", 'B', 16)
    pdf.cell(200, 10, text=f"{company_name} - Analytics Export",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font("Helvetica", '', 10)
    pdf.cell(200, 10, text=f"Generated on {timestamp}",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(10)

    # Performance Metrics
    pdf.set_font("Helvetica", 'B', 14)
    pdf.cell(200, 10, text="PERFORMANCE METRICS",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", '', 12)
    metrics = [
        ("Active Vacancies", data.get('active_vacancies', 0)),
        ("Total Applications", data.get('total_applications', 0)),
        ("Hire Rate", f"{data.get('hire_rate', 0)}%"),
        ("Avg Response Time", f"{data.get('avg_response_time', 0)} days")
    ]
    for label, value in metrics:
        pdf.cell(200, 10, text=f"{label}: {value}",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(5)

    # Application Flow
    pdf.set_font("Helvetica", 'B', 14)
    pdf.cell(200, 10, text="APPLICATION FLOW",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", '', 12)
    flow = [
        ("Pending Applications", data.get('pending_applications', 0)),
        ("Reviewed Applications", data.get('reviewed_applications', 0)),
        ("Approved Applications", data.get('approved_applications', 0)),
        ("Rejected Applications", data.get('rejected_applications', 0))
    ]
    for label, value in flow:
        pdf.cell(200, 10, text=f"{label}: {value}",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Save the PDF to the buffer
    pdf.output(buffer)
    buffer.seek(0)
    return buffer

async def show_response_analysis(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Simplified response analysis"""
    await update.callback_query.answer()
    await context.bot.send_message(
        chat_id=update.effective_user.id,
        text="⏱️ Response time analysis is currently unavailable",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 Back to Analytics", callback_data="back_to_analytics")]
        ])
    )

async def show_benchmark_comparison(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Simplified benchmark view"""
    await update.callback_query.answer()
    await context.bot.send_message(
        chat_id=update.effective_user.id,
        text="🏆 Benchmark comparison is currently unavailable",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 Back to Analytics", callback_data="back_to_analytics")]
        ])
    )


from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import ContextTypes, ConversationHandler
import sqlite3
# Function to show the broadcast options (job seekers or employers)
async def handle_broadcast_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id

    # Enhanced keyboard with emojis and better layout
    keyboard = [
        [InlineKeyboardButton("👨‍💼 Job Seekers", callback_data="job_seekers"),
         InlineKeyboardButton("👔 Employers", callback_data="employers")],
        [InlineKeyboardButton("🌍 All Users", callback_data="all")],
        [InlineKeyboardButton("📊 Stats Preview", callback_data="stats_preview")],
        [InlineKeyboardButton("❌ Cancel", callback_data="cancel")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Formatted message with Markdown
    await context.bot.send_message(
        chat_id=user_id,
        text="""📢 *Broadcast Message Settings*

Choose the audience for your broadcast:

• *Job Seekers* - All registered Job seekers
• *Employers* - All registered clients
• *Stats* - See audience metrics before sending""",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    return BROADCAST_TYPE


async def select_broadcast_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    choice = query.data

    if choice == "cancel":
        await context.bot.send_message(
            chat_id=user_id,
            text="🚫 Broadcast operation canceled.",
            reply_markup=ReplyKeyboardRemove()
        )
        return await show_admin_menu(update, context)

    if choice == "stats_preview":
        # Get stats from database
        cursor = db.cursor
        cursor.execute("SELECT COUNT(*) FROM users")
        job_seekers_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM employers")
        employers_count = cursor.fetchone()[0]

        await context.bot.send_message(
            chat_id=user_id,
            text=f"""📊 *Audience Statistics*

👨‍💼 Job Seekers: {job_seekers_count}
👔 Employers: {employers_count}
👥 Total: {job_seekers_count + employers_count}

Please select your broadcast target again:""",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("👨‍💼 Job Seekers", callback_data="job_seekers"),
                 InlineKeyboardButton("👔 Employers", callback_data="employers")],
                [InlineKeyboardButton("❌ Cancel", callback_data="cancel")]
            ])
        )
        return BROADCAST_TYPE

    context.user_data["broadcast_group"] = choice

    # Enhanced message with formatting options
    await context.bot.send_message(
        chat_id=user_id,
        text=f"""✍️ *Compose Your Broadcast Message*

You're sending to: *{'Job Seekers' if choice == 'job_seekers' else 'Employers'}*

You can use:
- *Bold text*
- _Italic text_
- `Code formatting`
- [Links](https://example.com)
- Emojis 😊

Enter your message below:""",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )
    return BROADCAST_MESSAGE

def escape_markdown_v2(text: str) -> str:
    escape_chars = r"_*[]()~`>#+-=|{}.!\\"
    return ''.join(f'\\{c}' if c in escape_chars else c for c in text)


async def get_broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id

    # Store the actual message object, not just text
    context.user_data["broadcast_message_obj"] = update.message

    # Confirmation buttons
    keyboard = [
        [InlineKeyboardButton("✅ Confirm & Send", callback_data="confirm")],
        [InlineKeyboardButton("❌ Cancel", callback_data="cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Show preview by forwarding back to admin
    await update.message.forward(
        chat_id=user_id,
        disable_notification=True
    )

    await context.bot.send_message(
        chat_id=user_id,
        text="📝 Above is your message preview. Confirm to broadcast:",
        reply_markup=reply_markup
    )

    return CONFIRM_BROADCAST


async def confirm_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    choice = query.data

    if choice == "cancel":
        await context.bot.send_message(
            chat_id=user_id,
            text="🚫 Broadcast canceled."
        )
        await show_admin_menu(update, context)
        return ADMIN_MAIN_MENU

    broadcast_group = context.user_data.get("broadcast_group")
    if not broadcast_group:
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Error: Missing broadcast data."
        )
        await show_admin_menu(update, context)
        return ADMIN_MAIN_MENU

    # Show sending animation
    sending_msg = await context.bot.send_message(
        chat_id=user_id,
        text="🔄 Sending broadcast..."
    )

    # Broadcast by forwarding
    success_count, fail_count = await broadcast_to_group(broadcast_group, context)

    # Delete the sending animation
    await context.bot.delete_message(chat_id=user_id, message_id=sending_msg.message_id)

    # Send stats
    await context.bot.send_message(
        chat_id=user_id,
        text=f"""✅ Broadcast Completed

📊 Statistics
✔️ Success: {success_count}
❌ Failed: {fail_count}"""
    )

    # Clear user data
    context.user_data.pop("broadcast_group", None)
    context.user_data.pop("broadcast_message_obj", None)

    await show_admin_menu(update, context)
    return ADMIN_MAIN_MENU


async def broadcast_to_group(group: str, context: ContextTypes.DEFAULT_TYPE) -> tuple:
    success_count = 0
    fail_count = 0
    message_obj = context.user_data.get("broadcast_message_obj")

    if not message_obj:
        return (0, 0)

    try:
        cursor = db.cursor

        if group == "job_seekers":
            cursor.execute("SELECT user_id FROM users")
        elif group == "employers":
            cursor.execute("SELECT employer_id FROM employers")
        elif group == "all":
            cursor.execute("SELECT user_id FROM users UNION SELECT employer_id FROM employers")
        else:
            return (0, 0)

        recipients = [row[0] for row in cursor.fetchall()]

        for recipient in recipients:
            try:
                # Forward the original message
                await message_obj.forward(
                    chat_id=recipient,
                    disable_notification=True
                )
                success_count += 1
            except Exception as e:
                fail_count += 1
                print(f"Failed to send to {recipient}: {e}")

        return (success_count, fail_count)

    except sqlite3.OperationalError as e:
        print(f"Database error: {e}")
        return (success_count, fail_count)

#apply Vacancy Job seeker

async def display_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    try:
        # First present the user with options to view all or filter vacancies
        reply_keyboard = [
            [get_translation(user_id, "view_all_vacancies")],
            [get_translation(user_id, "filter_vacancies")]
        ]

        await context.bot.send_message(
            chat_id=user_id,
            text=f"✨ <b>{get_translation(user_id, 'vacancy_options_title')}</b> ✨\n\n"
                 f"{get_translation(user_id, 'vacancy_options_description')}",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup(
                reply_keyboard,
                one_time_keyboard=True,
                resize_keyboard=True
            )
        )

        return VACANCY_DISPLAY_OPTION

    except Exception as e:
        logging.error(f"Error in display_vacancies initial step: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return MAIN_MENU


async def handle_vacancy_display_option(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    view_all_text = get_translation(user_id, "view_all_vacancies")
    filter_text = get_translation(user_id, "filter_vacancies")

    if choice == view_all_text:
        return await display_all_vacancies(update, context)
    elif choice == filter_text:
        return await show_filter_options(update, context)
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_choice")
        )
        return VACANCY_DISPLAY_OPTION


async def display_all_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    try:
        db = Database()
        validated_jobs = db.fetch_approved_vacancies()

        if not validated_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_vacancies_found")
            )
            return MAIN_MENU

        context.user_data["vacancies"] = validated_jobs

        # Create a more attractive and professional vacancy display
        vacancy_list = []
        for idx, vacancy in enumerate(validated_jobs, start=1):
            company_name = vacancy.get("company_name", get_translation(user_id, "not_provided"))
            job_title = vacancy.get("job_title", get_translation(user_id, "title_not_available"))
            deadline = vacancy.get("deadline", get_translation(user_id, "no_deadline"))
            employment_type = vacancy.get("employment_type", get_translation(user_id, "not_specified"))
            level = vacancy.get("level", get_translation(user_id, "not_specified"))
            salary = vacancy.get("salary", get_translation(user_id, "negotiable"))

            # Emoji mapping for employment types
            employment_emojis = {
                "full_time": "🏢",
                "part_time": "⏱",
                "remote": "🏠",
                "hybrid": "🔀",
                "freelance": "🖊"
            }

            emp_emoji = employment_emojis.get(employment_type, "💼")

            vacancy_text = (
                f"<b>🔹 {idx}. {job_title}</b>\n"
                f"   {emp_emoji} <i>{get_translation(user_id, employment_type)}</i>\n"
                f"   🏛 <b>{get_translation(user_id, 'company')}:</b> {company_name}\n"
                f"   📍 <b>{get_translation(user_id, 'level')}:</b> {get_translation(user_id, level)}\n"
                f"   💰 <b>{get_translation(user_id, 'salary')}:</b> {salary}\n"
                f"   ⏰ <b>{get_translation(user_id, 'deadline')}:</b> {deadline}\n"
                f"   {'▬' * 20}\n"  # Stylish separator
            )
            vacancy_list.append(vacancy_text)

        # Split vacancies into chunks to avoid message length limits
        chunk_size = 5
        vacancy_chunks = [vacancy_list[i:i + chunk_size] for i in range(0, len(vacancy_list), chunk_size)]

        # Send introduction message
        intro_message = (
            f"🌟 <b>{get_translation(user_id, 'available_vacancies')}</b> 🌟\n\n"
            f"{get_translation(user_id, 'vacancy_count', count=len(validated_jobs))}\n\n"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text=intro_message,
            parse_mode="HTML"
        )

        # Send vacancies in chunks
        for chunk in vacancy_chunks:
            await context.bot.send_message(
                chat_id=user_id,
                text="\n".join(chunk),
                parse_mode="HTML"
            )

        # Send prompt message with instructions
        prompt_message = (
            f"\n📌 <b>{get_translation(user_id, 'how_to_apply')}</b>\n"
            f"{get_translation(user_id, 'select_vacancy_instruction')}\n\n"
            f"🔎 {get_translation(user_id, 'filter_options_reminder')}"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text=prompt_message,
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup(
                [[get_translation(user_id, "filter_vacancies")],
                 [get_translation(user_id, "proceed_to_main_menu")]],
                resize_keyboard=True
            )
        )

        return SELECT_VACANCY

    except Exception as e:
        logging.error(f"Error displaying all vacancies: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
        return MAIN_MENU


async def show_filter_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    filter_options = [
        [get_translation(user_id, "filter_by_employment")],
        [get_translation(user_id, "filter_by_level")],
        [get_translation(user_id, "filter_by_qualification")],
        [get_translation(user_id, "filter_by_gender")],
        [get_translation(user_id, "sort_by_deadline")],
        [get_translation(user_id, "sort_by_quantity")],
        [get_translation(user_id, "cancel_filtering")]
    ]

    await context.bot.send_message(
        chat_id=user_id,
        text=f"🔍 <b>{get_translation(user_id, 'filter_options_title')}</b>\n\n"
             f"{get_translation(user_id, 'filter_options_instruction')}",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            filter_options,
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )

    return FILTER_SELECTION


async def handle_filter_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    filter_texts = {
        "employment": get_translation(user_id, "filter_by_employment"),
        "level": get_translation(user_id, "filter_by_level"),
        "qualification": get_translation(user_id, "filter_by_qualification"),
        "gender": get_translation(user_id, "filter_by_gender"),
        "deadline": get_translation(user_id, "sort_by_deadline"),
        "quantity": get_translation(user_id, "sort_by_quantity"),
        "cancel": get_translation(user_id, "cancel_filtering")
    }

    if choice == filter_texts["employment"]:
        return await show_employment_filters(update, context)
    elif choice == filter_texts["level"]:
        return await show_level_filters(update, context)
    elif choice == filter_texts["qualification"]:
        return await show_qualification_filters(update, context)
    elif choice == filter_texts["gender"]:
        return await show_gender_filters(update, context)
    elif choice == filter_texts["deadline"]:
        return await sort_by_deadline(update, context)
    elif choice == filter_texts["quantity"]:
        return await sort_by_quantity(update, context)
    elif choice == filter_texts["cancel"]:
        return await display_vacancies(update, context)
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_filter_choice")
        )
        return FILTER_SELECTION


async def show_employment_filters(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    employment_types = [
        [get_translation(user_id, "full_time")],
        [get_translation(user_id, "part_time")],
        [get_translation(user_id, "remote")],
        [get_translation(user_id, "hybrid")],
        [get_translation(user_id, "freelance")],
        [get_translation(user_id, "cancel_filtering")]
    ]

    await context.bot.send_message(
        chat_id=user_id,
        text=f"💼 <b>{get_translation(user_id, 'select_employment_type')}</b>",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            employment_types,
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )

    return EMPLOYMENT_FILTER


async def show_level_filters(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    level_options = [
        [get_translation(user_id, "entry_level")],
        [get_translation(user_id, "mid_level")],
        [get_translation(user_id, "senior_level")],
        [get_translation(user_id, "cancel_filtering")]
    ]

    await context.bot.send_message(
        chat_id=user_id,
        text=f"📊 <b>{get_translation(user_id, 'select_level')}</b>",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            level_options,
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )

    return LEVEL_FILTER


async def show_qualification_filters(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    qualification_options = [
        [get_translation(user_id, "training")],
        [get_translation(user_id, "degree")],
        [get_translation(user_id, "ma")],
        [get_translation(user_id, "phd")],
        [get_translation(user_id, "cancel_filtering")]
    ]

    await context.bot.send_message(
        chat_id=user_id,
        text=f"🎓 <b>{get_translation(user_id, 'select_qualification')}</b>",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            qualification_options,
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )

    return QUALIFICATION_FILTER


async def show_gender_filters(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    gender_options = [
        [get_translation(user_id, "male")],
        [get_translation(user_id, "female")],
        [get_translation(user_id, "both")],
        [get_translation(user_id, "cancel_filtering")]
    ]

    await context.bot.send_message(
        chat_id=user_id,
        text=f"🚻 <b>{get_translation(user_id, 'select_gender')}</b>",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            gender_options,
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )

    return GENDER_FILTER

async def handle_employment_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    cancel_text = get_translation(user_id, "cancel_filtering")

    if choice == cancel_text:
        return await show_filter_options(update, context)

    # Map the translated text back to the database value
    employment_mapping = {
        get_translation(user_id, "full_time"): "full_time",
        get_translation(user_id, "part_time"): "part_time",
        get_translation(user_id, "remote"): "remote",
        get_translation(user_id, "hybrid"): "hybrid",
        get_translation(user_id, "freelance"): "freelance"
    }

    employment_type = employment_mapping.get(choice)

    if employment_type:
        db = Database()
        filtered_jobs = db.fetch_approved_vacancies()
        filtered_jobs = [job for job in filtered_jobs if job.get("employment_type") == employment_type]

        context.user_data["vacancies"] = filtered_jobs

        if not filtered_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_vacancies_filter", filter=choice)
            )
            return await show_filter_options(update, context)

        return await display_filtered_vacancies(update, context, f"employment_type = {choice}")

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "invalid_employment_type")
    )
    return EMPLOYMENT_FILTER


async def display_filtered_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                     filter_description: str) -> int:
    user_id = get_user_id(update)
    vacancies = context.user_data.get("vacancies", [])

    if not vacancies:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "no_vacancies_found")
        )
        return MAIN_MENU

    # Similar attractive display as display_all_vacancies but with filter info
    vacancy_list = []
    for idx, vacancy in enumerate(vacancies, start=1):
        company_name = vacancy.get("company_name", get_translation(user_id, "not_provided"))
        job_title = vacancy.get("job_title", get_translation(user_id, "title_not_available"))
        deadline = vacancy.get("deadline", get_translation(user_id, "no_deadline"))
        employment_type = vacancy.get("employment_type", get_translation(user_id, "not_specified"))
        level = vacancy.get("level", get_translation(user_id, "not_specified"))

        vacancy_text = (
            f"<b>🔸 {idx}. {job_title}</b>\n"
            f"   🏢 <b>{get_translation(user_id, 'company')}:</b> {company_name}\n"
            f"   📌 <b>{get_translation(user_id, 'type')}:</b> {get_translation(user_id, employment_type)}\n"
            f"   📊 <b>{get_translation(user_id, 'level')}:</b> {get_translation(user_id, level)}\n"
            f"   ⏳ <b>{get_translation(user_id, 'deadline')}:</b> {deadline}\n"
            f"   {'━' * 15}\n"
        )
        vacancy_list.append(vacancy_text)

    # Send filter information first
    filter_message = (
        f"🔍 <b>{get_translation(user_id, 'filter_results')}</b>\n"
        f"{get_translation(user_id, 'applied_filter')}: {filter_description}\n"
        f"{get_translation(user_id, 'matching_vacancies')}: {len(vacancies)}\n\n"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=filter_message,
        parse_mode="HTML"
    )

    # Send vacancies in chunks
    chunk_size = 5
    for i in range(0, len(vacancy_list), chunk_size):
        await context.bot.send_message(
            chat_id=user_id,
            text="\n".join(vacancy_list[i:i + chunk_size]),
            parse_mode="HTML"
        )

    # Next steps
    next_steps = (
        f"\n📌 <b>{get_translation(user_id, 'next_steps')}</b>\n"
        f"{get_translation(user_id, 'select_or_filter_again')}"
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=next_steps,
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            [

                [get_translation(user_id, "apply_new_filter")],
                [get_translation(user_id, "view_all_vacancies")]
            ],
            resize_keyboard=True
        )
    )

    return SELECT_VACANCY


async def handle_level_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    cancel_text = get_translation(user_id, "cancel_filtering")

    if choice == cancel_text:
        return await show_filter_options(update, context)

    level_mapping = {
        get_translation(user_id, "entry_level"): "entry_level",
        get_translation(user_id, "mid_level"): "mid_level",
        get_translation(user_id, "senior_level"): "senior_level"
    }

    level = level_mapping.get(choice)

    if level:
        db = Database()
        filtered_jobs = db.fetch_approved_vacancies()
        filtered_jobs = [job for job in filtered_jobs if job.get("level") == level]

        context.user_data["vacancies"] = filtered_jobs

        if not filtered_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_vacancies_filter", filter=choice)
            )
            return await show_filter_options(update, context)

        return await display_filtered_vacancies(update, context, f"level = {choice}")

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "invalid_level")
    )
    return LEVEL_FILTER


async def handle_qualification_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    cancel_text = get_translation(user_id, "cancel_filtering")

    if choice == cancel_text:
        return await show_filter_options(update, context)

    qualification_mapping = {
        get_translation(user_id, "training"): "training",
        get_translation(user_id, "degree"): "degree",
        get_translation(user_id, "ma"): "ma",
        get_translation(user_id, "phd"): "phd"
    }

    qualification = qualification_mapping.get(choice)

    if qualification:
        db = Database()
        filtered_jobs = db.fetch_approved_vacancies()
        filtered_jobs = [job for job in filtered_jobs if job.get("qualification") == qualification]

        context.user_data["vacancies"] = filtered_jobs

        if not filtered_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_vacancies_filter", filter=choice)
            )
            return await show_filter_options(update, context)

        return await display_filtered_vacancies(update, context, f"qualification = {choice}")

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "invalid_qualification")
    )
    return QUALIFICATION_FILTER


async def handle_gender_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text

    cancel_text = get_translation(user_id, "cancel_filtering")

    if choice == cancel_text:
        return await show_filter_options(update, context)

    gender_mapping = {
        get_translation(user_id, "male"): "Male",
        get_translation(user_id, "female"): "Female",
        get_translation(user_id, "both"): "Any"
    }

    gender = gender_mapping.get(choice)

    if gender:
        db = Database()
        filtered_jobs = db.fetch_approved_vacancies()

        if gender == "any":
            filtered_jobs = [job for job in filtered_jobs if job.get("gender") in ["male", "female", "any"]]
        else:
            filtered_jobs = [job for job in filtered_jobs if job.get("gender") == gender]

        context.user_data["vacancies"] = filtered_jobs

        if not filtered_jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_vacancies_filter", filter=choice)
            )
            return await show_filter_options(update, context)

        return await display_filtered_vacancies(update, context, f"gender = {choice}")

    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "invalid_gender")
    )
    return GENDER_FILTER


async def sort_by_deadline(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    db = Database()
    jobs = db.fetch_approved_vacancies()

    # Sort by deadline (earliest first)
    sorted_jobs = sorted(
        jobs,
        key=lambda x: x.get("deadline", "9999-12-31")  # Default far future date for jobs without deadline
    )

    context.user_data["vacancies"] = sorted_jobs

    return await display_filtered_vacancies(
        update,
        context,
        get_translation(user_id, "sorted_by_deadline")
    )


async def sort_by_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    db = Database()
    jobs = db.fetch_approved_vacancies()

    # Sort by quantity (highest first)
    sorted_jobs = sorted(
        jobs,
        key=lambda x: int(x.get("quantity", 0)),
        reverse=True
    )

    context.user_data["vacancies"] = sorted_jobs

    return await display_filtered_vacancies(
        update,
        context,
        get_translation(user_id, "sorted_by_quantity")
    )

async def select_vacancy(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    choice = update.message.text.strip()

    # Handle navigation commands first
    navigation_commands = {
        get_translation(user_id, "view_all_vacancies"): display_all_vacancies,
        get_translation(user_id, "filter_vacancies"): show_filter_options,
        get_translation(user_id, "apply_new_filter"): show_filter_options,
        get_translation(user_id, "main_menu"): handle_main_menu,
        get_translation(user_id, "proceed_to_main_menu"): handle_main_menu,
        get_translation(user_id, "back_to_list"): display_all_vacancies
    }

    if choice in navigation_commands:
        return await navigation_commands[choice](update, context)

    try:
        # Retrieve stored job posts with enhanced validation
        job_posts = context.user_data.get("vacancies", [])

        if not job_posts:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"⚠️ {get_translation(user_id, 'no_job_selected_error')}",
                reply_markup=ReplyKeyboardMarkup(
                    [[get_translation(user_id, "main_menu")]],
                    resize_keyboard=True
                ),
                parse_mode="HTML"
            )
            return MAIN_MENU

        # Validate and convert selection
        selected_index = int(choice) - 1
        if not (0 <= selected_index < len(job_posts)):
            raise ValueError(get_translation(user_id, "invalid_vacancy_range", count=len(job_posts)))

        selected_job = job_posts[selected_index]

        # Enhanced field validation with fallbacks
        job_details = {
            "job_id": selected_job.get("job_id", ""),
            "employer_id": selected_job.get("employer_id", ""),
            "job_title": selected_job.get("job_title", get_translation(user_id, "title_not_available")),
            "company_name": selected_job.get("company_name", get_translation(user_id, "company_not_provided")),
            "employment_type": get_translation(user_id, selected_job.get("employment_type", "not_specified")),
            "deadline": selected_job.get("deadline", get_translation(user_id, "no_deadline")),
            "gender": get_translation(user_id, selected_job.get("gender", "any")),
            "quantity": selected_job.get("quantity", "N/A"),
            "level": get_translation(user_id, selected_job.get("level", "not_specified")),
            "description": selected_job.get("description", get_translation(user_id, "no_description")),
            "qualification": get_translation(user_id, selected_job.get("qualification", "not_specified")),
            "skills": selected_job.get("skills", get_translation(user_id, "no_skills")),
            "salary": selected_job.get("salary", get_translation(user_id, "negotiable")),
            "benefits": selected_job.get("benefits", get_translation(user_id, "no_benefits"))
        }

        # Store selected job details
        context.user_data["selected_job"] = job_details

        # Create visually appealing job card with emojis and formatting
        job_card = [
            f"🌟 <b>{get_translation(user_id, 'job_details_title')}</b> 🌟",
            "",
            f"📌 <b>{job_details['job_title']}</b>",
            f"🏢 {job_details['company_name']}",
            "",
            f"⏳ <b>{get_translation(user_id, 'deadline')}:</b> {job_details['deadline']}",
            f"💼 <b>{get_translation(user_id, 'type')}:</b> {job_details['employment_type']}",
            f"👥 <b>{get_translation(user_id, 'positions')}:</b> {job_details['quantity']}",
            f"📊 <b>{get_translation(user_id, 'level')}:</b> {job_details['level']}",
            f"🚻 <b>{get_translation(user_id, 'gender')}:</b> {job_details['gender']}",
            "",
            f"💰 <b>{get_translation(user_id, 'salary')}:</b> {job_details['salary']}",
            "",
            f"📝 <b>{get_translation(user_id, 'description')}:</b>",
            f"{job_details['description']}",
            "",
            f"🎓 <b>{get_translation(user_id, 'qualification')}:</b>",
            f"{job_details['qualification']}",
            "",
            f"🛠️ <b>{get_translation(user_id, 'skills')}:</b>",
            f"{job_details['skills']}",
            "",
            f"🎁 <b>{get_translation(user_id, 'benefits')}:</b>",
            f"{job_details['benefits']}",
            "",
            "▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬"
        ]

        # Create interactive buttons
        reply_markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(
                    f"✅ {get_translation(user_id, 'apply_now')}",
                    callback_data="confirm"
                ),
                InlineKeyboardButton(
                    f"❌ {get_translation(user_id, 'Cancel')}",
                    callback_data="cancel"
                )
            ]
        ])

        # Send the job card with interactive buttons
        await context.bot.send_message(
            chat_id=user_id,
            text="\n".join(job_card),
            parse_mode="HTML",
            reply_markup=reply_markup,
            disable_web_page_preview=True
        )

        return CONFIRM_SELECTION

    except ValueError as ve:
        logging.error(f"ValueError in select_vacancy: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ {get_translation(user_id, 'invalid_selection')}\n\n"
                 f"{get_translation(user_id, 'please_select_between', count=len(job_posts))}",
            reply_markup=ReplyKeyboardMarkup(
                [
                    [get_translation(user_id, "back_to_list")],
                    [get_translation(user_id, "main_menu")]
                ],
                resize_keyboard=True
            ),
            parse_mode="HTML"
        )
        return SELECT_VACANCY

    except Exception as e:
        logging.error(f"Unexpected error in select_vacancy: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ {get_translation(user_id, 'unexpected_error_occurred')}",
            reply_markup=ReplyKeyboardMarkup(
                [[get_translation(user_id, "main_menu")]],
                resize_keyboard=True
            ),
            parse_mode="HTML"
        )
        return MAIN_MENU

async def confirm_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    query = update.callback_query
    await query.answer()  # Acknowledge the callback query

    choice = query.data  # Get the callback data ("confirm" or "cancel")
    if choice == "back_to_job_detail":
        # Re-display the job details using the stored job information
        job_details = context.user_data["selected_job"]

        # Recreate the job card
        job_card = [
            f"🌟 <b>{get_translation(user_id, 'job_details_title')}</b> 🌟",
            "",
            f"📌 <b>{job_details['job_title']}</b>",
            f"🏢 {job_details['company_name']}",
            "",
            f"⏳ <b>{get_translation(user_id, 'deadline')}:</b> {job_details['deadline']}",
            f"💼 <b>{get_translation(user_id, 'type')}:</b> {job_details['employment_type']}",
            f"👥 <b>{get_translation(user_id, 'positions')}:</b> {job_details['quantity']}",
            f"📊 <b>{get_translation(user_id, 'level')}:</b> {job_details['level']}",
            f"🚻 <b>{get_translation(user_id, 'gender')}:</b> {job_details['gender']}",
            "",
            f"💰 <b>{get_translation(user_id, 'salary')}:</b> {job_details['salary']}",
            "",
            f"📝 <b>{get_translation(user_id, 'description')}:</b>",
            f"{job_details['description']}",
            "",
            f"🎓 <b>{get_translation(user_id, 'qualification')}:</b>",
            f"{job_details['qualification']}",
            "",
            f"🛠️ <b>{get_translation(user_id, 'skills')}:</b>",
            f"{job_details['skills']}",
            "",
            f"🎁 <b>{get_translation(user_id, 'benefits')}:</b>",
            f"{job_details['benefits']}",
            "",
            "▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬"
        ]

        # Recreate the interactive buttons
        reply_markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(
                    f"✅ {get_translation(user_id, 'apply_now')}",
                    callback_data="confirm"
                ),
                InlineKeyboardButton(
                    f"❌ {get_translation(user_id, 'Cancel')}",
                    callback_data="cancel"
                )
            ]
        ])

        # Edit the existing message to show the job details again
        await query.edit_message_text(
            text="\n".join(job_card),
            parse_mode="HTML",
            reply_markup=reply_markup,
            disable_web_page_preview=True
        )

        return CONFIRM_SELECTION
    try:
        # Validate that a job is selected
        selected_job = context.user_data.get("selected_job")
        if not selected_job or "job_id" not in selected_job:
            await query.edit_message_text(
                text=get_translation(user_id, "no_job_selected_error")
            )
            return MAIN_MENU

        # Handle user's choice
        if choice == "confirm":
            prompt = (
                f"✍️ <b>{get_translation(user_id, 'cover_letter_guide')}</b>\n\n"
                f"<i>{get_translation(user_id, 'cover_letter_tips')}</i>\n\n"
                f"<b>{get_translation(user_id, 'applying_for')}:</b> {selected_job['job_title']}\n"
                f"<b>{get_translation(user_id, 'at_company')}:</b> {selected_job.get('company_name', 'N/A')}\n\n"
                f"📝 <b>{get_translation(user_id, 'your_cover_letter')}:</b>"
            )

            await query.edit_message_text(
                text=prompt,
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(
                        f"⬅️ {get_translation(user_id, 'back_to_job')}",
                        callback_data="back_to_job_detail"
                    )]
                ])
            )
            return WRITE_COVER_LETTER
        elif choice == "cancel":
            # Return to main menu and clear selected job
            await query.edit_message_text(
                text=f"✅ {get_translation(user_id, 'application_canceled')}",
                parse_mode="HTML"
            )
            context.user_data.pop("selected_job", None)
            return await display_vacancies(update, context)

    except ValueError as ve:
        logging.error(f"ValueError in confirm_selection: {ve}")
        await query.edit_message_text(
            text=get_translation(user_id, "error_invalid_selection", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in confirm_selection: {e}")
        await query.edit_message_text(
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
    return MAIN_MENU

async def write_cover_letter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    try:
        # Validate job selection with enhanced error handling
        selected_job = context.user_data.get("selected_job")
        if not selected_job or "job_id" not in selected_job:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"⚠️ {get_translation(user_id, 'no_job_selected_error')}",
                reply_markup=ReplyKeyboardMarkup(
                    [[get_translation(user_id, "main_menu")]],
                    resize_keyboard=True
                ),
                parse_mode="HTML"
            )
            return MAIN_MENU

        # Check if user has already applied for this job
        if db.has_user_applied(user_id, selected_job["job_id"]):
            await context.bot.send_message(
                chat_id=user_id,
                text=f"ℹ️ {get_translation(user_id, 'already_applied_error')}",
                parse_mode="HTML"
            )
            context.user_data.pop("selected_job", None)
            return await main_menu(update, context)

        # Fetch and validate cover letter
        cover_letter = update.message.text.strip()
        if not cover_letter:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"📝 {get_translation(user_id, 'cover_letter_empty_error')}",
                parse_mode="HTML"
            )
            return WRITE_COVER_LETTER

        if len(cover_letter) < 50:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"📝 {get_translation(user_id, 'cover_letter_too_short')}",
                parse_mode="HTML"
            )
            return WRITE_COVER_LETTER

        if len(cover_letter) > 2000:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"📝 {get_translation(user_id, 'cover_letter_too_long')}",
                parse_mode="HTML"
            )
            return WRITE_COVER_LETTER

        # Save application in the database
        application_id =  db.save_application(user_id, selected_job["job_id"], cover_letter)

        # Create beautiful confirmation message
        confirmation_msg = (
            f"🎉 <b>{get_translation(user_id, 'application_submitted_successfully')}</b> 🎉\n\n"
            f"<b>🔢 {get_translation(user_id, 'application_id')}:</b> {application_id}\n"
            f"<b>📌 {get_translation(user_id, 'position')}:</b> {selected_job['job_title']}\n"
            f"<b>🏢 {get_translation(user_id, 'company')}:</b> {selected_job.get('company_name', 'N/A')}\n\n"
            f"<i>{get_translation(user_id, 'application_follow_up')}</i>"
        )

        # Send confirmation to applicant
        await context.bot.send_message(
            chat_id=user_id,
            text=confirmation_msg,
            parse_mode="HTML",


        )

        # Show formatted cover letter back to user
        await context.bot.send_message(
            chat_id=user_id,
            text=f"<b>📝 {get_translation(user_id, 'your_cover_letter')}:</b>\n\n{cover_letter}",
            parse_mode="HTML"
        )

        # Forward application to employer with enhanced details
        if "employer_id" in selected_job:
            await forward_application_to_employer(
                employer_id=selected_job["employer_id"],
                job_seeker_id=user_id,
                job_id=selected_job["job_id"],
                application_id=application_id,
                context=context
            )
        else:
            logging.warning(f"No employer_id found for job: {selected_job['job_id']}")

        # Clear selection
        context.user_data.pop("selected_job", None)

        return await main_menu(update, context)

    except ValueError as ve:
        logging.error(f"ValueError in write_cover_letter: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_invalid_selection", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in write_cover_letter: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )

async def forward_application_to_employer(
    employer_id: int,
    job_seeker_id: int,
    job_id: str,
    application_id: str,
    context: ContextTypes.DEFAULT_TYPE
):
    """Enhanced application forwarding with beautiful formatting and all necessary details"""
    try:
        # Fetch all required data
        job = db.get_job_details(job_id)
        applicant = db.get_user_profile(job_seeker_id)
        cover_letter = db.get_cover_letter_for_job(job_seeker_id, job_id)

        # Create beautifully formatted message
        application_msg = (
            f"📨 <b>{get_translation(employer_id, 'new_job_application_received')}</b> 📨\n\n"
            f"<b>🆔 {get_translation(employer_id, 'application_id')}:</b> {application_id}\n"
            f"<b>📌 {get_translation(employer_id, 'position')}:</b> {job.get('job_title', get_translation(employer_id, 'not_available'))}\n"
            f"<b>👤 {get_translation(employer_id, 'applicant_details')}:</b>\n"
            f"• <b>{get_translation(employer_id, 'name')}:</b> {applicant.get('full_name', get_translation(employer_id, 'not_available'))}\n"
            f"• <b>{get_translation(employer_id, 'gender')}:</b> {applicant.get('gender', get_translation(employer_id, 'not_available'))}\n"
            f"• <b>{get_translation(employer_id, 'phone')}:</b> {applicant.get('contact_number', get_translation(employer_id, 'not_available'))}\n\n"
            f"<b>📝 {get_translation(employer_id, 'cover_letter')}:</b>\n"
            f"{cover_letter or get_translation(employer_id, 'no_cover_letter_provided')}\n\n"
            f"<i>{get_translation(employer_id, 'review_application_in_dashboard')}</i>"
        )

        # Send to employer
        await context.bot.send_message(
            chat_id=employer_id,
            text=application_msg,
            parse_mode="HTML",

        )

    except Exception as e:
        logging.error(f"Error forwarding application to employer {employer_id}: {str(e)}")
        # Notify admin about the failure
        # return await main_menu(update, context)

async def confirm_application(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    cover_letter = update.message.text.strip()

    try:
        # Validate that a job is selected
        selected_job = context.user_data.get("selected_job")
        if not selected_job:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_job_selected_error")
            )
            return MAIN_MENU

        # Validate cover letter
        if not cover_letter:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "cover_letter_empty_error")
            )
            return WRITE_COVER_LETTER

        # Store cover letter in user_data
        context.user_data["cover_letter"] = cover_letter

        # Confirm submission with inline buttons
        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, "confirm_button"), callback_data="confirm")],
            [InlineKeyboardButton(get_translation(user_id, "cancel_button"), callback_data="cancel")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "confirm_submission_prompt"),
            reply_markup=reply_markup
        )
    except ValueError as ve:
        logging.error(f"ValueError in confirm_application: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_invalid_selection", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in confirm_application: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
    return CONFIRM_SUBMISSION


async def handle_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    action = query.data

    try:
        # Validate that a job is selected
        selected_job = context.user_data.get("selected_job")
        if not selected_job:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "no_job_selected_error")
            )
            return MAIN_MENU

        # Check if user has already applied for this job
        if db.has_user_applied(user_id, selected_job["job_id"]):
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "already_applied_error")
            )
            context.user_data.pop("selected_job", None)
            context.user_data.pop("cover_letter", None)
            return MAIN_MENU

        # Validate that a cover letter is provided
        cover_letter = context.user_data.get("cover_letter")
        if not cover_letter:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "cover_letter_empty_error")
            )
            return WRITE_COVER_LETTER

        # Handle user's action
        if action == "confirm":
            # Save application in the database
            db.save_application(user_id, selected_job["job_id"], cover_letter)

            # Notify the user that the application was submitted
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "application_submitted_successfully")
            )

            # Forward the application to the employer
            employer_id = selected_job["employer_id"]
            await forward_application_to_employer(employer_id, user_id, context)

            # Clear user data after submission
            context.user_data.pop("selected_job", None)
            context.user_data.pop("cover_letter", None)
            return MAIN_MENU

        elif action == "cancel":
            # Cancel submission and return to main menu
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "submission_canceled")
            )
            context.user_data.pop("selected_job", None)  # Clear selected job
            context.user_data.pop("cover_letter", None)  # Clear cover letter
            return MAIN_MENU

        else:
            # Invalid action
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "invalid_action_error")
            )
            return CONFIRM_SUBMISSION

    except ValueError as ve:
        logging.error(f"ValueError in handle_confirmation: {ve}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "error_invalid_selection", error=str(ve))
        )
    except Exception as e:
        logging.error(f"Unexpected error in handle_confirmation: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "unexpected_error_occurred", error=str(e))
        )
    return MAIN_MENU

#search vaccancies

async def start_job_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Initiate the job search process with filter options."""
    user_id = get_user_id(update)

    # Create an attractive intro message
    intro_message = (
        f"🔍 <b>{get_translation(user_id, 'advanced_job_search')}</b> 🔍\n\n"
        f"{get_translation(user_id, 'search_intro_message')}\n\n"
        f"✨ {get_translation(user_id, 'search_tip_message')}"
    )

    # Create inline keyboard with search options
    keyboard = [
        [
            InlineKeyboardButton(get_translation(user_id, "search_by_keyword"), callback_data="search_by_keyword"),
            InlineKeyboardButton(get_translation(user_id, "advanced_filters"), callback_data="advanced_filters")
        ],
        [
            InlineKeyboardButton(get_translation(user_id, "quick_search_recent"), callback_data="quick_search_recent"),
            InlineKeyboardButton(get_translation(user_id, "quick_search_high_paying"),
                                 callback_data="quick_search_high_paying")
        ],
        [
            InlineKeyboardButton(get_translation(user_id, "cancel_search"), callback_data="cancel_search")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=intro_message,
        parse_mode="HTML",
        reply_markup=reply_markup
    )

    return SEARCH_OPTIONS


async def handle_search_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle the user's choice of search method."""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    choice = query.data

    if choice == "search_by_keyword":
        await query.edit_message_text(
            text=get_translation(user_id, "enter_keyword_prompt"),
            parse_mode="HTML"
        )
        context.user_data["search_type"] = "keyword"
        return KEYWORD_SEARCH

    elif choice == "advanced_filters":
        return await display_advanced_filters(update, context)

    elif choice == "quick_search_recent":
        context.user_data["search_filters"] = {"sort_by": "newest"}
        return await perform_search(update, context)

    elif choice == "quick_search_high_paying":
        context.user_data["search_filters"] = {"sort_by": "salary_desc"}
        return await perform_search(update, context)

    elif choice == "cancel_search":
        await query.edit_message_text(
            text=get_translation(user_id, "search_cancelled"),
            parse_mode="HTML"
        )
        return MAIN_MENU

    return SEARCH_OPTIONS


async def display_advanced_filters(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display a menu of advanced filter options."""
    user_id = get_user_id(update)

    # Create a message explaining the filters
    filter_explanation = (
        f"⚙️ <b>{get_translation(user_id, 'advanced_filters_title')}</b> ⚙️\n\n"
        f"{get_translation(user_id, 'filter_explanation')}\n\n"
        f"📌 {get_translation(user_id, 'current_filters')}:\n"
    )

    # Get current filters if any
    current_filters = context.user_data.get("search_filters", {})
    if not current_filters:
        filter_explanation += get_translation(user_id, "no_filters_applied")
    else:
        for key, value in current_filters.items():
            filter_explanation += f"• {key}: {value}\n"

    # Create filter buttons
    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, "filter_job_type"), callback_data="filter_job_type")],
        [InlineKeyboardButton(get_translation(user_id, "filter_salary"), callback_data="filter_salary")],
        [InlineKeyboardButton(get_translation(user_id, "filter_experience"), callback_data="filter_experience")],
        [
            InlineKeyboardButton(get_translation(user_id, "apply_filters"), callback_data="apply_filters"),
            InlineKeyboardButton(get_translation(user_id, "clear_filters"), callback_data="clear_filters")
        ],
        [InlineKeyboardButton(get_translation(user_id, "back_to_search"), callback_data="back_to_search")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Handle both callback queries and regular messages
    if hasattr(update, 'callback_query') and update.callback_query:
        # Edit existing message if coming from callback
        await update.callback_query.edit_message_text(
            text=filter_explanation,
            parse_mode="HTML",
            reply_markup=reply_markup
        )
    else:
        # Send new message if coming from regular message
        await context.bot.send_message(
            chat_id=user_id,
            text=filter_explanation,
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    return ADVANCED_FILTERS

async def filter_job_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle job type filter selection."""
    user_id = get_user_id(update)

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, "full_time"), callback_data="job_type_full_time")],
        [InlineKeyboardButton(get_translation(user_id, "part_time"), callback_data="job_type_part_time")],
        [InlineKeyboardButton(get_translation(user_id, "freelance"), callback_data="job_type_freelance")],
        [InlineKeyboardButton(get_translation(user_id, "remote"), callback_data="job_type_remote")],
        [InlineKeyboardButton(get_translation(user_id, "hybrid"), callback_data="job_type_hybrid")],
        [InlineKeyboardButton(get_translation(user_id, "back_to_filters"), callback_data="back_to_filters")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.callback_query.edit_message_text(
        text=get_translation(user_id, "select_job_type_prompt"),
        parse_mode="HTML",
        reply_markup=reply_markup
    )

    return FILTER_JOB_TYPE


async def perform_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Perform the actual search based on filters."""
    user_id = get_user_id(update)
    search_filters = context.user_data.get("search_filters", {})

    try:
        db = Database()

        # Base query
        query = """
            SELECT v.id AS job_id, v.job_title, v.employment_type, v.gender, v.quantity, v.level,
                   v.description, v.qualification, v.skills, v.salary, v.benefits,
                   v.application_deadline AS deadline,
                   e.company_name, v.employer_id, v.status, 'vacancy' AS source
            FROM vacancies v
            JOIN employers e ON v.employer_id = e.employer_id
            WHERE v.status = 'approved'
        """

        params = []

        # Apply filters - updated to match posting format
        if "keyword" in search_filters:
            keyword = search_filters["keyword"]
            query += " AND (v.job_title LIKE ? OR v.description LIKE ? OR v.skills LIKE ? OR e.company_name LIKE ?)"
            # Add % for wildcard search
            param = f"%{keyword}%"
            params.extend([param, param, param, param])

        if "job_type" in search_filters:
            # Map search filter to posted job types
            type_mapping = {
                "full_time": "full_time",
                "part_time": "part_time",
                "freelance": "freelance",
                "remote": "remote",
                "hybrid": "hybrid"
            }
            query += " AND v.employment_type = ?"
            params.append(type_mapping.get(search_filters["job_type"], search_filters["job_type"]))

        if "min_salary" in search_filters:
            # Handle salary format matching (remove currency symbols, commas)
            query += """ AND CAST(REPLACE(REPLACE(REPLACE(v.salary, '$', ''), ',', ''), ' ', '') AS INTEGER) >= ?"""
            params.append(int(search_filters["min_salary"]))

        if "max_salary" in search_filters:
            query += """ AND CAST(REPLACE(REPLACE(REPLACE(v.salary, '$', ''), ',', ''), ' ', '') AS INTEGER) <= ?"""
            params.append(int(search_filters["max_salary"]))

        if "experience_level" in search_filters:
            # Map search filter to posted levels
            level_mapping = {
                "entry": "entry_level",
                "mid": "mid_level",
                "senior": "senior_level"
            }
            query += " AND v.level = ?"
            params.append(level_mapping.get(search_filters["experience_level"], search_filters["experience_level"]))

        # Sorting
        if search_filters.get("sort_by") == "newest":
            query += " ORDER BY v.id DESC"
        elif search_filters.get("sort_by") == "salary_desc":
            query += " ORDER BY CAST(REPLACE(REPLACE(REPLACE(v.salary, '$', ''), ',', ''), ' ', '') AS INTEGER) DESC"
        else:
            query += " ORDER BY v.application_deadline ASC"

        # Execute query
        db.cursor.execute(query, params)
        results = db.cursor.fetchall()
        jobs = [dict(row) for row in results]

        if not jobs:
            message = get_translation(user_id, "no_jobs_found_with_filters")
            if update.callback_query:
                await update.callback_query.edit_message_text(
                    text=message,
                    parse_mode="HTML"
                )
            else:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=message,
                    parse_mode="HTML"
                )
            return await display_advanced_filters(update, context)

        # Store results and display
        context.user_data["search_results"] = jobs
        return await display_search_results(update, context, jobs)

    except Exception as e:
        logging.error(f"Error performing search: {e}")
        message = get_translation(user_id, "search_error_message")
        if update.callback_query:
            await update.callback_query.edit_message_text(
                text=message,
                parse_mode="HTML"
            )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=message,
                parse_mode="HTML"
            )
        return MAIN_MENU



async def display_search_results(update: Update, context: ContextTypes.DEFAULT_TYPE, jobs: list) -> int:
    """Display the search results in an attractive format."""
    user_id = get_user_id(update)

    # Store the jobs for selection
    context.user_data["search_results"] = jobs

    # Prepare the results message
    results_message = f"🎯 <b>{get_translation(user_id, 'search_results_title')}</b> 🎯\n\n"
    results_message += f"📊 {get_translation(user_id, 'jobs_found')}: {len(jobs)}\n\n"

    # Show all jobs if coming from view_all_results, otherwise first 10
    display_jobs = jobs if len(jobs) <= 10 or context.user_data.get("showing_all_results") else jobs[:10]

    # Add each job to the message
    for idx, job in enumerate(display_jobs, start=1):
        results_message += (
            f"<b>{idx}. {job['job_title']}</b>\n"
            f"🏢 {job['company_name']}\n"
            f"💰 {job.get('salary', get_translation(user_id, 'salary_not_specified'))}\n"
            f"⏱ {job['employment_type']} | 📅 {job['deadline']}\n"
            f"📌 {job['level']}\n\n"
        )

    # Create action buttons
    keyboard = []
    if len(jobs) > 10 and not context.user_data.get("showing_all_results"):
        keyboard.append([InlineKeyboardButton(
            get_translation(user_id, "view_all_results"),
            callback_data="view_all_results")
        ])
        context.user_data["showing_all_results"] = False  # Reset flag

    keyboard.extend([
        [InlineKeyboardButton(
            get_translation(user_id, "refine_search"),
            callback_data="refine_search")
        ],
        [InlineKeyboardButton(
            get_translation(user_id, "new_search"),
            callback_data="new_search")
        ],
        [InlineKeyboardButton(
            get_translation(user_id, "back_to_main"),
            callback_data="back_to_main")
        ]
    ])

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send or edit the message
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=results_message,
            parse_mode="HTML",
            reply_markup=reply_markup
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=results_message,
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    # Send the selection prompt as a separate message
    await context.bot.send_message(
        chat_id=user_id,
        text=get_translation(user_id, "select_vacancy_instruction"),
        parse_mode="HTML"
        # reply_markup=ReplyKeyboardRemove()  # Remove any previous keyboards
    )


    return SELECT_SEARCH_RESULT

async def select_search_result(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    try:
        choice = update.message.text.strip()
        selected_index = int(choice) - 1

        jobs = context.user_data.get("search_results", [])
        if not jobs:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"⚠️ {get_translation(user_id, 'no_jobs_available')}",
                parse_mode="HTML"
            )
            return MAIN_MENU

        if not (0 <= selected_index < len(jobs)):
            raise ValueError(get_translation(user_id, "invalid_selection"))

        selected_job = jobs[selected_index]

        # Create job_details in the same format as in select_vacancy
        job_details = {
            "job_id": selected_job.get("job_id", ""),
            "employer_id": selected_job.get("employer_id", ""),
            "job_title": selected_job.get("job_title", get_translation(user_id, "title_not_available")),
            "company_name": selected_job.get("company_name", get_translation(user_id, "company_not_provided")),
            "employment_type": get_translation(user_id, selected_job.get("employment_type", "not_specified")),
            "deadline": selected_job.get("deadline", get_translation(user_id, "no_deadline")),
            "gender": get_translation(user_id, selected_job.get("gender", "any")),
            "quantity": selected_job.get("quantity", "N/A"),
            "level": get_translation(user_id, selected_job.get("level", "not_specified")),
            "description": selected_job.get("description", get_translation(user_id, "no_description")),
            "qualification": get_translation(user_id, selected_job.get("qualification", "not_specified")),
            "skills": selected_job.get("skills", get_translation(user_id, "no_skills")),
            "salary": selected_job.get("salary", get_translation(user_id, "negotiable")),
            "benefits": selected_job.get("benefits", get_translation(user_id, "no_benefits"))
        }

        # Save it so other functions expect consistent data
        context.user_data["selected_job"] = job_details

        # Then proceed to confirm selection just like in select_vacancy
        reply_markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(f"✅ {get_translation(user_id, 'apply_now')}", callback_data="confirm"),
                InlineKeyboardButton(f"❌ {get_translation(user_id, 'Cancel')}", callback_data="cancel")
            ]
        ])

        # Reuse the same job card formatting logic
        job_card = create_job_card(job_details, user_id)

        await context.bot.send_message(
            chat_id=user_id,
            text="\n".join(job_card),
            parse_mode="HTML",
            reply_markup=reply_markup,
            disable_web_page_preview=True
        )

        return CONFIRM_SELECTION

    except ValueError as ve:
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ {get_translation(user_id, 'invalid_selection')}\n\n"
                 f"{get_translation(user_id, 'please_select_between', count=len(jobs))}",
            parse_mode="HTML"
        )
        return SELECT_SEARCH_RESULT

    except Exception as e:
        logging.error(f"Unexpected error in select_search_result: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ {get_translation(user_id, 'unexpected_error_occurred')}",
            parse_mode="HTML"
        )
        return MAIN_MENU
def create_job_card(job_details, user_id):
    return [
        f"🌟 <b>{get_translation(user_id, 'job_details_title')}</b> 🌟",
        "",
        f"📌 <b>{job_details['job_title']}</b>",
        f"🏢 {job_details['company_name']}",
        "",
        f"⏳ <b>{get_translation(user_id, 'deadline')}:</b> {job_details['deadline']}",
        f"💼 <b>{get_translation(user_id, 'type')}:</b> {job_details['employment_type']}",
        f"👥 <b>{get_translation(user_id, 'positions')}:</b> {job_details['quantity']}",
        f"📊 <b>{get_translation(user_id, 'level')}:</b> {job_details['level']}",
        f"🚻 <b>{get_translation(user_id, 'gender')}:</b> {job_details['gender']}",
        "",
        f"💰 <b>{get_translation(user_id, 'salary')}:</b> {job_details['salary']}",
        "",
        f"📝 <b>{get_translation(user_id, 'description')}:</b>",
        f"{job_details['description']}",
        "",
        f"🎓 <b>{get_translation(user_id, 'qualification')}:</b>",
        f"{job_details['qualification']}",
        "",
        f"🛠️ <b>{get_translation(user_id, 'skills')}:</b>",
        f"{job_details['skills']}",
        "",
        f"🎁 <b>{get_translation(user_id, 'benefits')}:</b>",
        f"{job_details['benefits']}",
        "",
        "▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬"
    ]

async def handle_keyword_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle keyword search input from user."""
    user_id = get_user_id(update)
    keyword = update.message.text.strip()

    if not keyword:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "empty_keyword_error")
        )
        return KEYWORD_SEARCH

    # Store keyword and perform search
    context.user_data["search_filters"] = {"keyword": keyword}
    return await perform_search(update, context)


async def handle_advanced_filters(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle selections from the advanced filters menu."""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    choice = query.data

    if choice == "filter_job_type":
        return await filter_job_type(update, context)
    elif choice == "filter_salary":
        await query.edit_message_text(
            text=get_translation(user_id, "enter_salary_range_prompt"),
            parse_mode="HTML"
        )
        return FILTER_SALARY
    elif choice == "filter_experience":
        return await display_experience_levels(update, context)

    elif choice == "apply_filters":
        return await perform_search(update, context)
    elif choice == "clear_filters":
        context.user_data.pop("search_filters", None)
        return await display_advanced_filters(update, context)
    elif choice == "back_to_search":
        return await start_job_search(update, context)

    return ADVANCED_FILTERS


async def handle_job_type_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle job type filter selection."""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    choice = query.data

    if choice.startswith("job_type_"):
        job_type = choice.replace("job_type_", "")
        # Map to posting format
        type_mapping = {
            "full_time": "full_time",
            "part_time": "part_time",
            "freelance": "freelance",
            "remote": "remote",
            "hybrid": "hybrid"
        }
        filtered_type = type_mapping.get(job_type, job_type)

        if "search_filters" not in context.user_data:
            context.user_data["search_filters"] = {}
        context.user_data["search_filters"]["job_type"] = filtered_type

        await query.edit_message_text(
            text=get_translation(user_id, "job_type_set_success", job_type=job_type.replace("_", " ")),
            parse_mode="HTML"
        )
        return await display_advanced_filters(update, context)
    elif choice == "back_to_filters":
        return await display_advanced_filters(update, context)

    return FILTER_JOB_TYPE


async def handle_experience_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle experience level selection."""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    choice = query.data

    if choice.startswith("exp_"):
        exp_level = choice.replace("exp_", "")
        # Map to posting format
        level_mapping = {
            "entry": "entry_level",
            "mid": "mid_level",
            "senior": "senior_level"
        }
        filtered_level = level_mapping.get(exp_level, exp_level)

        if "search_filters" not in context.user_data:
            context.user_data["search_filters"] = {}
        context.user_data["search_filters"]["experience_level"] = filtered_level

        await query.edit_message_text(
            text=get_translation(user_id, "experience_level_set_success", level=exp_level.replace("_", " ")),
            parse_mode="HTML"
        )
        return await display_advanced_filters(update, context)
    elif choice == "back_to_filters":
        return await display_advanced_filters(update, context)

    return FILTER_EXPERIENCE


async def handle_salary_filter(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle salary range input from user."""
    user_id = get_user_id(update)
    salary_input = update.message.text.strip()

    try:
        # Parse salary range (format: "min-max" or "min" or "max+")
        if "-" in salary_input:
            min_salary, max_salary = map(str.strip, salary_input.split("-"))
            min_salary = int(''.join(filter(str.isdigit, min_salary)))
            max_salary = int(''.join(filter(str.isdigit, max_salary)))
        elif "+" in salary_input:
            min_salary = int(''.join(filter(str.isdigit, salary_input.replace("+", ""))))
            max_salary = 999999  # Arbitrary large number
        else:
            min_salary = int(''.join(filter(str.isdigit, salary_input)))
            max_salary = min_salary  # Exact amount

        if min_salary > max_salary:
            raise ValueError("Min salary cannot be greater than max salary")

        # Store in filters
        if "search_filters" not in context.user_data:
            context.user_data["search_filters"] = {}
        context.user_data["search_filters"]["min_salary"] = min_salary
        context.user_data["search_filters"]["max_salary"] = max_salary

        # Send confirmation message
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "salary_range_set_success",
                                 min_salary=min_salary, max_salary=max_salary),
            parse_mode="HTML"
        )

        # Return to filters menu with proper update context
        return await display_advanced_filters(update, context)

    except ValueError as e:
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_salary_format_error"),
            parse_mode="HTML"
        )
        return FILTER_SALARY


async def display_experience_levels(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display experience level options to user."""
    user_id = get_user_id(update)

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, "entry_level"), callback_data="exp_entry")],
        [InlineKeyboardButton(get_translation(user_id, "mid_level"), callback_data="exp_mid")],
        [InlineKeyboardButton(get_translation(user_id, "senior_level"), callback_data="exp_senior")],
        [InlineKeyboardButton(get_translation(user_id, "back_to_filters"), callback_data="back_to_filters")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.callback_query.edit_message_text(
        text=get_translation(user_id, "select_experience_level_prompt"),
        parse_mode="HTML",
        reply_markup=reply_markup
    )
    return FILTER_EXPERIENCE

async def handle_search_results(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle actions from search results view."""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    choice = query.data

    if choice == "view_all_results":
        # Pass all results without limiting to 10
        return await display_search_results(update, context, context.user_data["search_results"])
    elif choice == "refine_search":
        return await display_advanced_filters(update, context)
    elif choice == "new_search":
        context.user_data.pop("search_filters", None)
        context.user_data.pop("search_results", None)
        return await start_job_search(update, context)
    elif choice == "back_to_main":
        if update.callback_query:
            await update.callback_query.edit_message_text(
                text=get_translation(user_id, "returning_to_main_menu"),
                parse_mode="HTML"
            )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, "returning_to_main_menu"),
                parse_mode="HTML"
            )
        return MAIN_MENU
    elif choice.startswith("view_job_"):
        job_index = int(choice.replace("view_job_", ""))
        return await display_job_details(update, context, job_index)

    return SEARCH_RESULTS


async def display_job_details(update: Update, context: ContextTypes.DEFAULT_TYPE, job_index: int) -> int:
    """Display detailed view of a specific job."""
    user_id = get_user_id(update)
    jobs = context.user_data.get("search_results", [])

    if not jobs or job_index < 0 or job_index >= len(jobs):
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, "invalid_job_selection_error"),
            parse_mode="HTML"
        )
        return SEARCH_RESULTS

    job = jobs[job_index]

    # Format job details (similar to your existing job display format)
    details = (
        f"<b>📌 {get_translation(user_id, 'job_title')}:</b> {job['job_title']}\n"
        f"🏢 <b>{get_translation(user_id, 'employer')}:</b> {job.get('company_name', 'N/A')}\n"
        f"📅 <b>{get_translation(user_id, 'deadline')}:</b> {job['deadline']}\n"
        f"💼 <b>{get_translation(user_id, 'employment_type')}:</b> {job['employment_type']}\n"
        f"💰 <b>{get_translation(user_id, 'salary')}:</b> {job.get('salary', get_translation(user_id, 'not_specified'))}\n\n"
        f"<b>📝 {get_translation(user_id, 'description')}:</b>\n{job['description']}\n\n"
        f"<b>🎓 {get_translation(user_id, 'qualification')}:</b>\n{job['qualification']}\n\n"
        f"<b>🔑 {get_translation(user_id, 'skills')}:</b>\n{job['skills']}"
    )

    # Create action buttons
    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, "back_to_results"), callback_data="back_to_results")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:  # Check if this is a callback query update
        await update.callback_query.edit_message_text(
            text=details,
            parse_mode="HTML",
            reply_markup=reply_markup
        )
    else:  # Fallback for non-callback updates
        await context.bot.send_message(
            chat_id=user_id,
            text=details,
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    return VIEW_JOB_DETAILS


async def handle_job_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle actions from job details view."""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    choice = query.data

    if choice == "back_to_results":
        return await display_search_results(update, context, context.user_data["search_results"])

    return VIEW_JOB_DETAILS

#remove buttons
async def remove_job_seekers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Prompt for search term with clearer instructions
    await context.bot.send_message(
        chat_id=user_id,
        text="Enter search term (name, ID, or leave empty for all). Example: 'John' or '12345':"
    )
    return SEARCH_JOB_SEEKERS

async def handle_job_seeker_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1  # Start from the first page

    try:
        # Fetch paginated results
        job_seekers = db.search_job_seekers(search_term, page=page)
        total_pages = db.get_total_pages_job_seekers(search_term)

        if not job_seekers:
            await context.bot.send_message(
                chat_id=user_id,
                text="No matching job seekers found."
            )
            return await back_to_database_menu(update, context)

        # Create paginated keyboard
        keyboard = create_paginated_keyboard(job_seekers, page, total_pages, "job_seeker")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_users")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"🔍 *Search Results*: '{search_term}'\n"
                 f"📄 Page 1/{total_pages}\n\n",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Error handling job seeker search: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching job seekers. Please try again later."
        )
        return await back_to_database_menu(update, context)

    return REMOVE_JOB_SEEKERS_PAGINATED

def create_paginated_keyboard(items, current_page, total_pages, entity_type):
    keyboard = []
    # Add item buttons
    for item in items:
        if entity_type == "job_seeker":
            text = f"{item['full_name']} (ID: {item['user_id']})"
            callback_data = f"remove_seeker_{item['user_id']}"
        elif entity_type == "employer":
            text = f"{item['company_name']} (ID: {item['employer_id']})"
            callback_data = f"remove_employer_{item['employer_id']}"

        elif entity_type in ["application", "application_list"]:
            # Use get() with default values for safety
            job_title = item.get('job_title', 'Unknown Job')
            applicant_name = item.get('applicant_name', 'Unknown Applicant')  # Changed from full_name
            app_id = item.get('application_id', '?')

            text = f"{job_title} - {applicant_name} (ID: {app_id})"

            if entity_type == "application":
                callback_data = f"remove_application_{app_id}"
            else:
                callback_data = f"application_detail_{app_id}"

        elif entity_type == "job" or entity_type == "job_remove":
            # Updated to match what search_jobs() returns
            text = f"{item['job_title']} (ID: {item['id']})"
            if entity_type == "job":
                callback_data = f"job_detail_{item['id']}"
            else:
                callback_data = f"remove_job_{item['id']}"
        elif entity_type == "vacancy" or entity_type == "vacancy_remove":
            text = f"{item['job_title']} (ID: {item['id']})"
            if entity_type == "vacancy":
                callback_data = f"vacancy_detail_{item['id']}"
            else:
                callback_data = f"remove_vacancy_{item['id']}"

        # Add other entity types here...
        else:
            continue  # Skip unsupported entity types
        keyboard.append([InlineKeyboardButton(text, callback_data=callback_data)])

    # Add pagination buttons
    nav_buttons = []
    if current_page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅ Prev", callback_data=f"prev_{entity_type}_{current_page}"))
    if current_page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Next ➡", callback_data=f"next_{entity_type}_{current_page}"))

    if nav_buttons:
        keyboard.append(nav_buttons)

    return keyboard


async def handle_pagination(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle pagination for all entity types"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    data = query.data.split("_")

    if len(data) < 3:
        await query.edit_message_text("Invalid pagination action.")
        return await back_to_manage_jobs(update, context)

    action, entity_type, current_page = data[:3]
    current_page = int(current_page)
    search_term = context.user_data.get("search_term", "")

    # Calculate new page
    new_page = current_page + 1 if action == "next" else current_page - 1

    try:
        # Determine which function to call based on entity type
        if entity_type == "job":
            items = db.get_all_jobs(page=new_page)
            total_pages = db.get_total_pages_jobs()
            keyboard_type = "job"
            next_state = LIST_JOBS_PAGINATED
        elif entity_type == "job_remove":
            items = db.search_jobs(search_term, page=new_page)
            total_pages = db.get_total_pages_jobs(search_term)
            keyboard_type = "job_remove"
            next_state = REMOVE_JOBS_PAGINATED
        elif entity_type == "vacancy":
            items = db.search_vacancies(search_term, page=new_page)
            total_pages = db.get_total_pages_vacancies(search_term)
            keyboard_type = "vacancy"
            next_state = LIST_VACANCIES_PAGINATED
        elif entity_type == "vacancy_remove":
            items = db.search_vacancies(search_term, page=new_page)
            total_pages = db.get_total_pages_vacancies(search_term)
            keyboard_type = "vacancy_remove"
            next_state = REMOVE_VACANCIES_PAGINATED
        else:
            raise ValueError(f"Unknown entity type: {entity_type}")

        if not items:
            await query.edit_message_text("No more items found.")
            return next_state

        # Create paginated keyboard
        keyboard = create_paginated_keyboard(items, new_page, total_pages, keyboard_type)

        # Add appropriate back button based on context
        if "remove" in entity_type:
            keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_jobs")])
        else:
            keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_vacancies")])

        # Update the message
        prefix = "Jobs" if "job" in entity_type else "Vacancies"
        if search_term:
            text = f"{prefix} matching '{search_term}' (Page {new_page}/{total_pages}):"
        else:
            text = f"All {prefix} (Page {new_page}/{total_pages}):"

        await query.edit_message_text(
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Pagination error: {e}", exc_info=True)
        await query.edit_message_text("Error loading page. Please try again.")
        return await back_to_database_menu(update, context)

    return next_state

async def handle_pagination_ban(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    data = query.data
    action, entity_type, page = data.split('_')
    page = int(page)

    search_term = context.user_data.get("search_term", "")
    page_size = 10

    try:
        if entity_type == "employer":
            items = db.search_employers_for_ban(search_term, page=page, page_size=page_size)
            total_pages = db.get_total_pages_employers_for_ban(search_term, page_size=page_size)
            state = BAN_EMPLOYERS_PAGINATED
        else:  # job_seeker
            items = db.search_job_seekers_for_ban(search_term, page=page, page_size=page_size)
            total_pages = db.get_total_pages_job_seekers_for_ban(search_term, page_size=page_size)
            state = BAN_JOB_SEEKERS_PAGINATED

        keyboard = create_ban_paginated_keyboard(items, page, total_pages, entity_type)
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_users")])

        await query.edit_message_text(
            text=f"🔍 *Search Results*: '{search_term}'\n"
                 f"📄 Page {page}/{total_pages}\n\n"
                 f"Select an {entity_type.replace('_', ' ')} to ban:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )
        return state
    except Exception as e:
        logging.error(f"Error handling pagination: {e}")
        await query.edit_message_text(text="An error occurred. Please try again.")
        return await back_to_manage_users(update, context)
import asyncio
from threading import Lock
import logging

# Define a lock
notification_lock = Lock()

import asyncio
from threading import Lock
import logging

# Define a lock
notification_lock = Lock()

async def handle_notifications(context):
    if not notification_lock.acquire(blocking=False):  # Try to acquire the lock
        logging.debug("Notification handler is already running. Skipping this execution.")
        return

    try:
        # Fetch and process up to 5 notifications at a time
        notifications = db.fetch_notifications(limit=3)
        if notifications:
            for notification in notifications:
                user_id = notification["user_id"]
                action = notification["action"]

                if action == "removed":
                    await context.bot.send_message(
                        chat_id=user_id,
                        text="You have been removed by Admin. Please restart by clicking /start."
                    )

            # Clear processed notifications after batch processing
            db.clear_notifications(limit=3)

    except Exception as e:
        logging.error(f"Error handling notifications: {e}")

    finally:
        notification_lock.release()  # Release the lock when done




async def remove_employers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Prompt for search term with clearer instructions
    await context.bot.send_message(
        chat_id=user_id,
        text="Enter search term (company name, ID, or leave empty for all). Example: 'TechCorp' or '12345':"
    )
    return SEARCH_EMPLOYERS


async def handle_employer_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1  # Start from the first page

    try:
        # Fetch paginated results
        employers = db.search_employers(search_term, page=page)
        total_pages = db.get_total_pages_employers(search_term)

        if not employers:
            await context.bot.send_message(
                chat_id=user_id,
                text="No matching employers found."
            )
            return await back_to_database_menu(update, context)

        # Create paginated keyboard
        keyboard = create_paginated_keyboard(employers, page, total_pages, "employer")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_users")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"🔍 *Search Results*: '{search_term}'\n"
                 f"📄 {page}/{total_pages}\n\n",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Error handling employer search: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching employers. Please try again later."
        )
        return await back_to_database_menu(update, context)

    return REMOVE_EMPLOYERS_PAGINATED


async def remove_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Prompt for search term with clearer instructions
    await context.bot.send_message(
        chat_id=user_id,
        text="Enter search term (job title, seeker name, or leave empty for all). Example: 'Software Engineer' or 'John Doe':"
    )
    return SEARCH_APPLICATIONS


async def handle_application_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1

    try:
        applications = db.search_applications(search_term, page=page)
        total_pages = db.get_total_pages_applications(search_term)

        if not applications:
            await context.bot.send_message(
                chat_id=user_id,
                text="No matching applications found."
            )
            return await back_to_manage_applications(update, context)

        keyboard = create_paginated_keyboard(applications, page, total_pages, "application")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_applications")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"Applications matching '{search_term}' (Page {page}/{total_pages}):",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Error searching applications: {e}", exc_info=True)
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while searching applications. Please try again."
        )
        return await back_to_manage_applications(update, context)

    return REMOVE_APPLICATIONS_PAGINATED

async def list_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    page = 1
    search_term = ""

    try:
        applications = db.search_applications(search_term, page)
        total_pages = db.get_total_pages_applications(search_term)

        if not applications:
            await context.bot.send_message(
                chat_id=user_id,
                text="No applications found."
            )
            return await back_to_database_menu(update, context)

        keyboard = create_paginated_keyboard(applications, page, total_pages, "application_list")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_applications")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"Applications (Page {page}/{total_pages}):",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Error listing applications: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching applications. Please try again later."
        )
        return await back_to_database_menu(update, context)

    return LIST_APPLICATIONS_PAGINATED

async def remove_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle job removal with search/pagination"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    # Prompt for search term with clearer instructions
    await context.bot.send_message(
        chat_id=user_id,
        text="Enter search term (job title, employer name, or leave empty for all). Example: 'Software Engineer' or 'TechCorp':"
    )
    return SEARCH_JOBS


async def export_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Export jobs to a professionally formatted Excel file."""
    user_id = update.effective_user.id
    jobs = db.get_all_job_posts()  # Fetch all job posts

    if not jobs:
        await context.bot.send_message(chat_id=user_id, text="No jobs found.")
        return

    # Convert to DataFrame
    df = pd.DataFrame(jobs)

    # Handle missing values
    df.fillna("Not provided", inplace=True)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # Title section
    ws.merge_cells("A1:O2")
    title = ws.cell(row=1, column=1, value="📋 Job Posts Report")
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="005B96", end_color="005B96", fill_type="solid")  # Blue theme
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Headers styling
    header_fill = PatternFill(start_color="005B96", end_color="005B96", fill_type="solid")
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Adding data rows with alternating colors and formatting
    light_blue = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    for r_idx, row in enumerate(df.itertuples(index=False), 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Alternating row colors
            if (r_idx - 4) % 2 == 0:
                cell.fill = light_blue

            # Style "Not provided" entries
            if str(value).strip() == "Not provided":
                cell.font = Font(italic=True, color="FF0000")

    # Auto-fit column widths for jobs sheet
    for column_cells in ws.columns:
        if isinstance(column_cells[0], MergedCell):
            continue
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = ws.column_dimensions[column_cells[0].column_letter]
        column_letter.width = max_length + 4

    # Freeze top row
    ws.freeze_panes = 'A5'

    # Save and send file
    filename = "Job_Posts_Report.xlsx"
    wb.save(filename)

    try:
        with open(filename, "rb") as doc:
            await context.bot.send_document(
                chat_id=user_id,
                document=doc,
                filename=filename,
                caption="📋 Job Posts Report"
            )
    finally:
        if os.path.exists(filename):
            os.remove(filename)


async def handle_job_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle job search with better empty term handling"""
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1

    try:
        jobs = db.search_jobs(search_term, page=page)
        total_pages = db.get_total_pages_jobs(search_term)

        if not jobs:
            # Provide more helpful feedback
            if search_term:
                msg = f"No jobs found matching '{search_term}'"
            else:
                msg = "No jobs found in the database"

            await context.bot.send_message(
                chat_id=user_id,
                text=msg
            )
            return await back_to_manage_jobs(update, context)

        keyboard = create_paginated_keyboard(jobs, page, total_pages, "job_remove")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_jobs")])

        response_text = (f"Found {len(jobs)} jobs matching "
                         f"'{search_term if search_term else 'all criteria'}' "
                         f"(Page {page}/{total_pages})")

        await context.bot.send_message(
            chat_id=user_id,
            text=response_text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Job search error: {e}", exc_info=True)
        await context.bot.send_message(
            chat_id=user_id,
            text="Failed to search jobs. Please try different terms."
        )
        return await back_to_manage_jobs(update, context)

    return REMOVE_JOBS_PAGINATED

# 1. List Vacancies Handler
async def list_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    page = 1
    search_term = ""  # Default search term for all vacancies

    try:
        vacancies = db.search_vacancies(search_term, page)
        total_pages = db.get_total_pages_vacancies(search_term)

        if not vacancies:
            await context.bot.send_message(
                chat_id=user_id,
                text="No vacancies found."
            )
            return await back_to_database_menu(update, context)

        keyboard = create_paginated_keyboard(vacancies, page, total_pages, "vacancy")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_vacancies")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"Vacancies (Page {page}/{total_pages}):",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Error listing vacancies: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching vacancies. Please try again later."
        )
        return await back_to_database_menu(update, context)

    return LIST_VACANCIES_PAGINATED

# 2. Remove Vacancies Handler
async def remove_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    await context.bot.send_message(
        chat_id=user_id,
        text="Enter search term (job title, employer name, or leave empty for all). Example: 'Software Engineer' or 'TechCorp':"
    )
    return SEARCH_VACANCIES


async def handle_vacancy_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1  # Start from the first page

    try:
        # Fetch paginated results
        vacancies = db.search_vacancies(search_term, page)
        total_pages = db.get_total_pages_vacancies(search_term)

        if not vacancies:
            await context.bot.send_message(
                chat_id=user_id,
                text="No matching vacancies found."
            )
            return await back_to_database_menu(update, context)

        # Create paginated keyboard
        keyboard = create_paginated_keyboard(vacancies, page, total_pages, "vacancy_remove")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_vacancies")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"Vacancies matching '{search_term}' (Page {page}/{total_pages}):",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    except Exception as e:
        logging.error(f"Error handling vacancy search: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching vacancies. Please try again later."
        )
        return await back_to_database_menu(update, context)

    return REMOVE_VACANCIES_PAGINATED

# 3. Export Vacancies Handler
async def export_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Export all vacancies to a professionally formatted Excel file."""
    user_id = update.effective_user.id
    vacancies = db.get_all_vacancies_posts()  # Fetch all vacancies

    if not vacancies:
        await context.bot.send_message(chat_id=user_id, text="No vacancies found.")
        return

    # Convert to DataFrame
    df = pd.DataFrame(vacancies)

    # Handle missing values
    df.fillna("Not provided", inplace=True)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacancies"

    # Title section
    ws.merge_cells("A1:N2")  # Merge cells for the title
    title = ws.cell(row=1, column=1, value="📋 Vacancy Report")
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")  # Indigo theme
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Headers styling
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Adding data rows with alternating colors and formatting
    light_indigo = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    for r_idx, row in enumerate(df.itertuples(index=False), 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Alternating row colors
            if (r_idx - 4) % 2 == 0:
                cell.fill = light_indigo

            # Style "Not provided" entries
            if str(value).strip() == "Not provided":
                cell.font = Font(italic=True, color="FF0000")

    # Auto-fit column widths for vacancies sheet
    for column_cells in ws.columns:
        if isinstance(column_cells[0], MergedCell):
            continue
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = ws.column_dimensions[column_cells[0].column_letter]
        column_letter.width = max_length + 4

    # Freeze top row
    ws.freeze_panes = 'A5'

    # Save and send file
    filename = "Vacancy_Report.xlsx"
    wb.save(filename)

    try:
        with open(filename, "rb") as doc:
            await context.bot.send_document(
                chat_id=user_id,
                document=doc,
                filename=filename,
                caption="📋 Vacancy Report"
            )
    finally:
        if os.path.exists(filename):
            os.remove(filename)


# 4. Export All Data Handler
async def export_all_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    # Export all entity types
    await export_job_seekers(update, context)
    await export_employers(update, context)
    await export_applications(update, context)
    await export_vacancies(update, context)
    await export_jobs(update, context)

    await context.bot.send_message(
        chat_id=user_id,
        text="All data exports completed."
    )
    return await back_to_database_menu(update, context)


async def confirm_removal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    data = query.data

    # Extract the type and ID from the callback data
    if data.startswith("remove_seeker_"):
        target_id = data.split("_")[-1]
        await query.edit_message_text(text=f"Are you sure you want to remove job seeker with ID {target_id}?")
        context.user_data["action"] = ("remove_seeker", target_id)
    elif data.startswith("remove_employer_"):
        target_id = data.split("_")[-1]
        await query.edit_message_text(text=f"Are you sure you want to remove employer with ID {target_id}?")
        context.user_data["action"] = ("remove_employer", target_id)
    elif data.startswith("remove_application_"):
        target_id = data.split("_")[-1]
        await query.edit_message_text(text=f"Are you sure you want to remove application with ID {target_id}?")
        context.user_data["action"] = ("remove_application", target_id)
    else:
        await query.edit_message_text(text="Invalid action.")
        return await back_to_admin_menu(update, context)

    # Confirm removal with Yes/No buttons
    keyboard = [
        [InlineKeyboardButton("Yes", callback_data="confirm_yes")],
        [InlineKeyboardButton("No", callback_data="confirm_no")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await context.bot.send_message(chat_id=user_id, text="Confirm action:", reply_markup=reply_markup)
    return CONFIRM_REMOVAL

async def perform_removal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    action, target_id = context.user_data.get("action", (None, None))
    if query.data == "confirm_yes" and action and target_id:
        try:
            if action == "remove_seeker":
                db.remove_job_seeker(target_id)
                await query.edit_message_text(text=get_translation(user_id, "job_seeker_removed_successfully", id=target_id))
            elif action == "remove_employer":
                db.remove_employer(target_id)
                await query.edit_message_text(text=get_translation(user_id, "employer_removed_successfully", id=target_id))
            elif action == "remove_application":
                db.remove_application(target_id)
                await query.edit_message_text(text=get_translation(user_id, "application_removed_successfully", id=target_id))
        except Exception as e:
            await context.bot.send_message(chat_id=user_id, text=get_translation(user_id, "error_performing_removal", error=str(e)))
    elif query.data == "confirm_no":
        await query.edit_message_text(text=get_translation(user_id, "removal_canceled"))
    else:
        await query.edit_message_text(text=get_translation(user_id, "invalid_action"))
    return await back_to_manage_users(update, context)


async def back_to_manage_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Return to the vacancy management menu"""
    query = update.callback_query
    if query:
        await query.answer()

    # Clear any temporary data
    if 'search_term' in context.user_data:
        del context.user_data['search_term']
    if 'vacancy_to_remove' in context.user_data:
        del context.user_data['vacancy_to_remove']

    # Return to the vacancy management menu
    return await ad_manage_vacancies(update, context)


async def confirm_removal_vacancy(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Confirm before removing a vacancy"""
    query = update.callback_query
    await query.answer()
    vacancy_id = int(query.data.split('_')[-1])

    context.user_data['vacancy_to_remove'] = vacancy_id

    keyboard = [
        [InlineKeyboardButton("✅ Confirm Remove", callback_data=f"confirm_remove_vacancy_{vacancy_id}")],
        [InlineKeyboardButton("❌ Cancel", callback_data="back_to_manage_vacancies")]
    ]

    await query.edit_message_text(
        text=f"Are you sure you want to remove vacancy ID {vacancy_id}?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

    return CONFIRM_REMOVE_VACANCY


async def execute_remove_vacancy(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Actually remove the vacancy from database"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    vacancy_id = int(query.data.split('_')[-1])

    try:
        # Remove from database
        db.remove_vacancy(vacancy_id)

        await context.bot.send_message(
            chat_id=user_id,
            text=f"✅ Successfully removed vacancy ID {vacancy_id}"
        )
    except Exception as e:
        logging.error(f"Error removing vacancy: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Failed to remove vacancy. Please try again."
        )

    return await back_to_manage_vacancies(update, context)

from openpyxl.utils import get_column_letter


async def back_to_manage_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Return to the job management menu"""
    query = update.callback_query
    if query:
        await query.answer()

    # Clear temporary data
    context.user_data.pop('search_term', None)
    context.user_data.pop('job_to_remove', None)

    return await manage_jobs(update, context)


async def view_job_detail(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show detailed view of a job"""
    query = update.callback_query
    await query.answer()
    job_id = int(query.data.split('_')[-1])

    try:
        job = db.get_job_details(job_id)
        if not job:
            await query.edit_message_text("Job not found.")
            return await back_to_manage_jobs(update, context)

        # Build the detailed message
        message = (
            f"🏢 <b>{job['job_title']}</b>\n"
            f"📌 <b>Company:</b> {job['company_details']['name']}\n"
            f"📍 <b>Location:</b> {job['company_details'].get('location', 'Not specified')}\n"
            f"📅 <b>Deadline:</b> {job['application_deadline']}\n"
            f"🔄 <b>Status:</b> {job['status'].capitalize()}\n"
            f"💼 <b>Type:</b> {job['employment_details']['type']}\n"
            f"🧑‍💻 <b>Level:</b> {job['employment_details']['level']}\n"
            f"👥 <b>Gender:</b> {job['employment_details']['gender']}\n"
            f"💰 <b>Salary:</b> {job.get('salary', 'Not specified')}\n"
            f"🔢 <b>Positions:</b> {job['employment_details']['quantity']}\n"
            f"\n📝 <b>Description:</b>\n{job['description']}\n"
            f"\n🎓 <b>Qualifications:</b>\n{job['qualification']}\n"
            f"\n🛠️ <b>Skills:</b>\n{job['skills']}"
        )

        # Add more action buttons if needed
        keyboard = [

            [InlineKeyboardButton("🔙 Back to List", callback_data="back_to_manage_jobs")]
        ]

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        logging.error(f"Error viewing job details: {e}", exc_info=True)
        await query.edit_message_text(
            text="⚠️ Error loading job details. Please try again later.",
            parse_mode=ParseMode.HTML
        )

    return JOB_DETAIL_VIEW

async def confirm_job_removal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show confirmation dialog before removing a job"""
    query = update.callback_query
    await query.answer()
    job_id = int(query.data.split('_')[-1])

    # Store job ID in context for the confirmation step
    context.user_data['job_to_remove'] = job_id

    # Get job details for the confirmation message
    job = db.get_job_details(job_id)
    if not job:
        await query.edit_message_text("Job not found.")
        return await back_to_manage_jobs(update, context)

    keyboard = [
        [InlineKeyboardButton("✅ Confirm Removal", callback_data=f"execute_remove_job_{job_id}")],
        [InlineKeyboardButton("❌ Cancel", callback_data="back_to_manage_jobs")]
    ]

    await query.edit_message_text(
        text=f"Are you sure you want to permanently remove this job?\n\n"
             f"🏢 {job['job_title']}\n"
             f"📌 Company: {job.get('company_name', 'Unknown')}\n"
             f"🆔 ID: {job_id}",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

    return CONFIRM_JOB_REMOVAL


async def execute_job_removal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Actually remove the job from database after confirmation"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    job_id = context.user_data.get('job_to_remove')

    if not job_id:
        await context.bot.send_message(
            chat_id=user_id,
            text="Error: Job reference missing. Please try again."
        )
        return await back_to_manage_jobs(update, context)

    try:
        # Get job details before deletion for confirmation
        job = db.get_job_details(job_id)
        if not job:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"❌ Job ID {job_id} not found in database."
            )
            return await back_to_manage_jobs(update, context)

        # Attempt removal
        success, message = db.remove_job(job_id)

        if success:
            response = (
                f"✅ Successfully removed job:\n"
                f"🏢 {job['job_title']}\n"
                f"🆔 ID: {job_id}\n"
                f"🏭 Company: {job['company_details']['name']}"
            )
        else:
            response = (
                f"❌ Failed to remove job:\n"
                f"🏢 {job['job_title']}\n"
                f"🆔 ID: {job_id}\n"
                f"⚠️ Reason: {message}"
            )

        await context.bot.send_message(
            chat_id=user_id,
            text=response
        )

    except Exception as e:
        logging.error(f"Error removing job {job_id}: {e}", exc_info=True)
        await context.bot.send_message(
            chat_id=user_id,
            text=f"⚠️ Critical error removing job {job_id}. Please check logs."
        )

    # Clear the stored job ID
    context.user_data.pop('job_to_remove', None)
    return await back_to_manage_jobs(update, context)


async def back_to_manage_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Return to application management menu"""
    query = update.callback_query
    if query:
        await query.answer()

    # Clear search term if exists
    context.user_data.pop('search_term', None)

    return await manage_applications(update, context)


async def view_application_detail(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show detailed view of an application with proper field access"""
    query = update.callback_query
    await query.answer()
    application_id = int(query.data.split('_')[-1])

    try:
        application = db.get_application_details(application_id)
        if not application:
            await query.edit_message_text("Application not found.")
            return await back_to_manage_applications(update, context)

        # Use get() with default values for optional fields
        message = (
            f"📄 <b>Application Details</b>\n\n"
            f"🆔 <b>ID:</b> {application['application_id']}\n"
            f"👤 <b>Applicant:</b> {application.get('applicant_name', 'Unknown')}\n"
            f"🏢 <b>Job:</b> {application.get('job_title', 'Unknown')}\n"
            f"🏭 <b>Company:</b> {application.get('company_name', 'Unknown')}\n"
            f"📅 <b>Applied:</b> {application.get('application_date', 'Unknown')}\n"
            f"🔄 <b>Status:</b> {application.get('status', 'Unknown').capitalize()}\n\n"
            f"📝 <b>Cover Letter:</b>\n{application.get('cover_letter', 'Not provided')}"
        )

        keyboard = [
            [InlineKeyboardButton("🗑️ Remove", callback_data=f"remove_application_{application_id}")],
            [InlineKeyboardButton("🔙 Back", callback_data="back_to_manage_applications")]
        ]

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        logging.error(f"Error viewing application {application_id}: {e}", exc_info=True)
        await query.edit_message_text(
            text="⚠️ Error loading application details. Please try again.",
            parse_mode=ParseMode.HTML
        )

    return APPLICATION_DETAIL_VIEW


async def confirm_application_removal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Confirm application removal"""
    query = update.callback_query
    await query.answer()
    application_id = int(query.data.split('_')[-1])

    context.user_data['application_to_remove'] = application_id

    application = db.get_application_details(application_id)
    if not application:
        await query.edit_message_text("Application not found.")
        return await back_to_manage_applications(update, context)

    keyboard = [
        [InlineKeyboardButton("✅ Confirm Remove", callback_data=f"confirm_remove_application_{application_id}")],
        [InlineKeyboardButton("❌ Cancel", callback_data="back_to_manage_applications")]
    ]

    await query.edit_message_text(
        text=f"Are you sure you want to remove this application?\n\n"
             f"👤 {application['applicant_name']}\n"
             f"🏢 {application['job_title']}\n"
             f"🆔 ID: {application_id}",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

    return CONFIRM_APPLICATION_REMOVAL


async def execute_application_removal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Permanently remove an application from the database after confirmation"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    application_id = context.user_data.get('application_to_remove')

    if not application_id:
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Error: Application reference missing. Please start the removal process again."
        )
        return await back_to_manage_applications(update, context)

    try:
        # Get application details before deletion for the confirmation message
        application = db.get_application_details(application_id)
        if not application:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"❌ Application ID {application_id} not found in database."
            )
            return await back_to_manage_applications(update, context)

        # Attempt to remove the application
        success = db.remove_application(application_id)

        if success:
            response = (
                f"✅ Successfully removed application:\n"
                f"🆔 ID: {application_id}\n"
                f"👤 Applicant: {application.get('applicant_name', 'Unknown')}\n"
                f"🏢 Job: {application.get('job_title', 'Unknown')}\n"
                f"🏭 Company: {application.get('company_name', 'Unknown')}"
            )
        else:
            response = (
                f"❌ Failed to remove application:\n"
                f"🆔 ID: {application_id}\n"
                f"ℹ️ The application may have already been deleted or is referenced by other records."
            )

        await context.bot.send_message(
            chat_id=user_id,
            text=response
        )

    except Exception as e:
        logging.error(f"Error removing application {application_id}: {str(e)}", exc_info=True)
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"⚠️ Critical error removing application {application_id}!\n"
                f"Please check the logs for details."
            )
        )

    # Clear the stored application ID regardless of outcome
    context.user_data.pop('application_to_remove', None)
    return await back_to_manage_applications(update, context)

async def export_job_seekers(update, context):
    """Export job seekers to a professionally formatted Excel file."""
    user_id = get_user_id(update)
    job_seekers = db.get_all_job_seekers_details()

    if not job_seekers:
        await context.bot.send_message(chat_id=user_id, text="No job seekers found.")
        return

    # Create DataFrame
    df = pd.DataFrame(job_seekers, columns=[
        "User ID", "Full Name", "Contact Number", "DOB", "CV Path",
        "Languages", "Qualification", "Field of Study", "CGPA",
        "Skills & Experience", "Profile Summary"
    ])

    # Handle NaN values: Separate logic for numeric and non-numeric columns
    numeric_columns = ["CGPA"]  # Add other numeric columns if needed
    df[numeric_columns] = df[numeric_columns].fillna(0)  # Fill numeric columns with 0

    non_numeric_columns = df.columns.difference(numeric_columns)
    df[non_numeric_columns] = df[non_numeric_columns].fillna("Not provided")  # Fill non-numeric columns with "Not provided"

    df.replace("", "Not provided", inplace=True)  # Replace empty strings with "Not provided"

    # Replace CV Path with "Provided" if a CV is available
    df["CV Path"] = df["CV Path"].apply(lambda x: "Provided" if x != "Not provided" else "Not provided")

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Seekers"

    # Create title section
    ws.merge_cells("A1:K2")
    title = ws.cell(row=1, column=1, value="📊 Job Seekers Report")
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Adding headers with style
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Adding data rows with alternating colors
    for r_idx, row in enumerate(df.itertuples(index=False), 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')  # Enable text wrapping
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (r_idx - 4) % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Format CGPA
            if c_idx == 9 and value != "Not provided":
                try:
                    cell.value = float(value)
                    cell.number_format = '0.00'
                except ValueError:
                    pass

            # Conditional Formatting for CGPA
            if c_idx == 9 and isinstance(value, (int, float)):
                if value > 3.5:
                    cell.font = Font(color="008000", bold=True)  # Green for high CGPA
                elif value < 2.5:
                    cell.font = Font(color="FF0000", bold=True)  # Red for low CGPA

            # Style "Not provided" entries
            if str(value).strip() == "Not provided":
                cell.font = Font(italic=True, color="FF0000")

    # Auto-fit column widths for most columns, but limit width for long-text columns
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)

        # Limit width for specific columns with long text
        if column_cells[0].value in ["Skills & Experience", "Profile Summary"]:
            ws.column_dimensions[column_letter].width = 50  # Fixed width for long-text columns (wider)
        else:
            ws.column_dimensions[column_letter].width = min(max_length + 6, 40)  # Auto-fit with a wider maximum width

    # Adjust row height for long-text columns
    for row in range(5, len(df) + 5):  # Start from row 5 (data rows)
        ws.row_dimensions[row].height = 40  # Increased row height for better readability

    # Freeze top row
    ws.freeze_panes = 'A5'

    # Add Advanced Summary Sheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(["📊 Job Seekers Summary Report"])
    summary_ws.merge_cells("A1:B1")
    title_cell = summary_ws.cell(row=1, column=1)
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    summary_data = [
        ["Total Job Seekers", len(df)],
        ["Average CGPA", round(df[df['CGPA'] != "Not provided"]['CGPA'].astype(float).mean(), 2)],
        ["Top 5 Qualifications", ', '.join(df['Qualification'].value_counts().head(5).index)],
        ["Most Common Languages", ', '.join(df['Languages'].value_counts().head(5).index)],
        ["Most Popular Fields of Study", ', '.join(df['Field of Study'].value_counts().head(5).index)]
    ]

    for row in summary_data:
        summary_ws.append(row)

    for row in summary_ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
            cell.border = Border(bottom=Side(style='thin'))

    # Auto-fit column widths for summary sheet
    for column_cells in summary_ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = get_column_letter(column_cells[0].column)
        summary_ws.column_dimensions[column_letter].width = max_length + 8  # Slightly wider columns for summary

    # Save and send file
    filename = "Job_Seekers_Report.xlsx"
    wb.save(filename)

    try:
        with open(filename, "rb") as doc:
            await context.bot.send_document(chat_id=user_id, document=doc, filename=filename,
                                            caption="📊 Job Seekers Report")
    finally:
        if os.path.exists(filename):
            os.remove(filename)

async def export_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Export applications to a professionally formatted Excel file."""
    user_id = get_user_id(update)
    applications = db.get_all_applications_details()  # Ensure this returns full details

    if not applications:
        await context.bot.send_message(chat_id=user_id, text="No applications found.")
        return await back_to_admin_menu(update, context)

    # Create DataFrame with proper columns
    df = pd.DataFrame(applications, columns=[
        "Application ID", "Job Title", "Job Seeker Name",
        "Employer Name", "Application Date", "Status",
        "Cover Letter", "Vacancy ID"
    ])
    df.fillna("Not provided", inplace=True)
    df.replace("", "Not provided", inplace=True)

    # Convert dates to datetime
    df['Application Date'] = pd.to_datetime(df['Application Date']).dt.strftime('%Y-%m-%d')

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Title section
    ws.merge_cells("A1:H2")
    title = ws.cell(row=1, column=1, value="📄 Application Tracking Report")
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="007030", end_color="007030", fill_type="solid")  # Green theme
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Headers styling
    header_fill = PatternFill(start_color="007030", end_color="007030", fill_type="solid")
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Adding data rows with alternating colors and formatting
    light_green = PatternFill(start_color="E9F5DB", end_color="E9F5DB", fill_type="solid")
    for r_idx, row in enumerate(df.itertuples(index=False), 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Alternating row colors
            if (r_idx - 4) % 2 == 0:
                cell.fill = light_green

            # Style "Not provided" entries
            if str(value).strip() == "Not provided":
                cell.font = Font(italic=True, color="FF0000")

            # Conditional formatting for status
            if c_idx == 6:  # Status column
                if value.lower() == "approved":
                    cell.font = Font(color="008000", bold=True)  # Green for approved
                elif value.lower() == "rejected":
                    cell.font = Font(color="FF0000", bold=True)  # Red for rejected
                else:
                    cell.font = Font(color="FFA500", bold=True)  # Orange for pending

    # Auto-fit column widths for applications sheet
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = max_length + 4

    # Freeze top row
    ws.freeze_panes = 'A5'

    # Add Summary Sheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(["📊 Application Summary Report"])
    summary_ws.merge_cells("A1:B1")
    title_cell = summary_ws.cell(row=1, column=1)
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="007030", end_color="007030", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Generate summary data
    total_applications = len(df)
    applications_by_status = df['Status'].value_counts().to_dict()
    most_applied_jobs = df['Job Title'].value_counts().head(5).to_dict()
    active_job_seekers = df['Job Seeker Name'].nunique()
    active_employers = df['Employer Name'].nunique()

    summary_data = [
        ["Total Applications", total_applications],
        ["Approved Applications", applications_by_status.get("approved", 0)],
        ["Rejected Applications", applications_by_status.get("rejected", 0)],
        ["Pending Applications", applications_by_status.get("pending", 0)],
        ["Most Applied Jobs", ', '.join(f"{job} ({count})" for job, count in most_applied_jobs.items())],
        ["Active Job Seekers", active_job_seekers],
        ["Active Employers", active_employers]
    ]

    for row in summary_data:
        summary_ws.append(row)

    # Style summary sheet
    for row in summary_ws.iter_rows(min_row=2, max_row=len(summary_data) + 1, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E9F5DB", end_color="E9F5DB", fill_type="solid")
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-fit column widths for summary sheet
    for column_cells in summary_ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = get_column_letter(column_cells[0].column)
        summary_ws.column_dimensions[column_letter].width = max_length + 4

    # Save and send file
    filename = "Application_Tracking_Report.xlsx"
    wb.save(filename)

    try:
        with open(filename, "rb") as doc:
            await context.bot.send_document(
                chat_id=user_id,
                document=doc,
                filename=filename,
                caption="📄 Application Tracking Report"
            )
    finally:
        if os.path.exists(filename):
            os.remove(filename)


import pandas as pd
from telegram import Update
from telegram.ext import ContextTypes


import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

async def export_employers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Export employers to a professionally formatted Excel file."""
    user_id = update.effective_user.id
    employers = db.get_all_employers_details()

    if not employers:
        await context.bot.send_message(chat_id=user_id, text="No employers found.")
        return

    # Convert to DataFrame
    df = pd.DataFrame(employers, columns=[
        "Employer ID", "Company Name", "Contact Number", "City", "Employer Type",
        "About Company", "Verification Docs"
    ])
    df.fillna("Not provided", inplace=True)
    df.replace("", "Not provided", inplace=True)

    # Modify Verification Docs column: Replace paths with "Provided" or "Not provided"
    df["Verification Docs"] = df["Verification Docs"].apply(
        lambda x: "Provided" if str(x).strip() != "Not provided" else "Not provided"
    )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Employers"

    # Title section
    ws.merge_cells("A1:H2")
    title = ws.cell(row=1, column=1, value="🏢 Employers Report")
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")  # Orange theme
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Headers styling
    header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Adding data rows with alternating colors and formatting
    light_orange = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    for r_idx, row in enumerate(df.itertuples(index=False), 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Alternating row colors
            if (r_idx - 4) % 2 == 0:
                cell.fill = light_orange

            # Style "Not provided" entries
            if str(value).strip() == "Not provided":
                cell.font = Font(italic=True, color="FF0000")

            # Hyperlink for Verification Docs
            if c_idx == 7 and value == "Provided":
                cell.value = "Provided"  # Ensure the cell displays "Provided"
                cell.font = Font(color="0000FF", underline="single")  # Blue font for hyperlink-like appearance

    # Auto-fit column widths for employers sheet
    for column_cells in ws.columns:
        # Skip merged cells
        if isinstance(column_cells[0], MergedCell):
            continue

        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = ws.column_dimensions[column_cells[0].column_letter]
        column_letter.width = max_length + 4

    # Freeze top row
    ws.freeze_panes = 'A5'

    # Add Summary Sheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(["📊 Employer Summary Report"])
    summary_ws.merge_cells("A1:B1")
    title_cell = summary_ws.cell(row=1, column=1)
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Generate summary data
    total_employers = len(df)
    verified_employers = df[df['Verification Docs'] == "Provided"].shape[0]
    most_common_cities = ', '.join(df['City'].value_counts().head(5).index)
    most_common_types = ', '.join(df['Employer Type'].value_counts().head(5).index)

    summary_data = [
        ["Total Employers", total_employers],
        ["Verified Employers", verified_employers],
        ["Unverified Employers", total_employers - verified_employers],
        ["Most Common Cities", most_common_cities],
        ["Most Common Employer Types", most_common_types]
    ]

    for row in summary_data:
        summary_ws.append(row)

    # Style summary sheet
    for row in summary_ws.iter_rows(min_row=2, max_row=len(summary_data) + 1, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-fit column widths for summary sheet
    for column_cells in summary_ws.columns:
        # Skip merged cells
        if isinstance(column_cells[0], MergedCell):
            continue

        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column_letter = summary_ws.column_dimensions[column_cells[0].column_letter]
        column_letter.width = max_length + 4

    # Save and send file
    filename = "Employers_Report.xlsx"
    wb.save(filename)

    try:
        with open(filename, "rb") as doc:
            await context.bot.send_document(
                chat_id=user_id,
                document=doc,
                filename=filename,
                caption="🏢 Employers Report"
            )
    finally:
        if os.path.exists(filename):
            os.remove(filename)



# Step 1: Implement back_to_database_menu handler
async def back_to_database_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Return to the main database management menu"""
    query = update.callback_query
    if query:
        await query.answer()

    return await show_database_menu(update, context)


# Step 2: Implement export_data handler
async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show export options menu"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    keyboard = [
        [InlineKeyboardButton("Export Job Seekers", callback_data="export_job_seekers")],
        [InlineKeyboardButton("Export Employers", callback_data="export_employers")],
        [InlineKeyboardButton("Export Applications", callback_data="export_applications")],
        [InlineKeyboardButton("Export All Data", callback_data="export_all_data")],
        [InlineKeyboardButton("Back to Database Menu", callback_data="back_to_database_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text="Select data to export:",
        reply_markup=reply_markup
    )
    return EXPORT_DATA  # Make sure this state is defined in your conversation handler

async def clear_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show confirmation dialog for data clearing"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    keyboard = [
        [InlineKeyboardButton("Yes, clear all data", callback_data="confirm_clear")],
        [InlineKeyboardButton("Cancel", callback_data="back_to_database_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text="⚠️ Are you sure you want to CLEAR ALL DATA? This action cannot be undone!",
        reply_markup=reply_markup
    )
    return CLEAR_CONFIRMATION


async def perform_clear(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Execute data clearing after confirmation"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    # Clear all data
    db.clear_all_data()

    await context.bot.send_message(
        chat_id=user_id,
        text="✅ All data has been successfully cleared.",
        reply_markup=ReplyKeyboardRemove()
    )
    return await back_to_database_menu(update, context)


async def handle_cancel_clear(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle cancel action"""
    query = update.callback_query
    await query.answer()
    return await back_to_database_menu(update, context)


async def table_cleanup_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show the table cleanup submenu with all tables listed."""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    # List of all tables (you can dynamically fetch this from the database if needed)
    tables = [
        "vacancies", "applications", "users", "employers", "job_posts", "appeals", "rating_privacy",
        "review_metadata", "review_responses", "reports", "admin_notifications", "notifications", "bot_logs",
        "message_logs", "bans", "contact_messages", "account_metadata", "reviews", "bot_errors",
        "contact_categories", "review_limits", "application_decisions"
    ]

    # Create buttons for each table, arranging them 2 per row
    keyboard = []
    for i in range(0, len(tables), 2):
        row = []
        # Add the first button in the pair
        row.append(InlineKeyboardButton(f" {tables[i].capitalize()}", callback_data=f"delete_table_{tables[i]}"))

        # Add the second button in the pair if it exists
        if i + 1 < len(tables):
            row.append(
                InlineKeyboardButton(f" {tables[i + 1].capitalize()}", callback_data=f"delete_table_{tables[i + 1]}"))

        # Append the row to the keyboard
        keyboard.append(row)

    # Add a back button as the last row
    keyboard.append([InlineKeyboardButton("🔙 Back to Database Menu", callback_data="back_to_database_menu")])

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text="🗑️ <b>Select a table to delete its data:</b>\n\n⚠️ Warning: This action cannot be undone!",
        reply_markup=reply_markup,
        parse_mode="HTML"
    )
    return TABLE_CLEANUP
async def confirm_table_deletion(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Confirm the deletion of a specific table."""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    # Extract the table name from the callback data
    table_name = query.data.replace("delete_table_", "")

    # Store the table name in the context for later use
    context.user_data["table_to_delete"] = table_name

    keyboard = [
        [InlineKeyboardButton("Yes, delete data", callback_data="confirm_delete")],
        [InlineKeyboardButton("Cancel", callback_data="back_to_table_cleanup")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=f"⚠️ Are you sure you want to delete all data from the table <b>{table_name.capitalize()}</b>? This action cannot be undone!",
        reply_markup=reply_markup,
        parse_mode="HTML"
    )
    return CONFIRM_TABLE_DELETION

async def perform_table_deletion(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Execute the deletion of data from the selected table."""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id

    # Retrieve the table name from the context
    table_name = context.user_data.get("table_to_delete")

    if not table_name:
        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Error: Table name not found. Please try again.",
            reply_markup=ReplyKeyboardRemove()
        )
        return await back_to_database_menu(update, context)

    # Perform the deletion
    try:
        db.delete_table_data(table_name)  # Assuming you have a method like this in your DB class
        await context.bot.send_message(
            chat_id=user_id,
            text=f"✅ All data from the table <b>{table_name.capitalize()}</b> has been successfully deleted.",
            parse_mode="HTML"
        )
    except Exception as e:
        await context.bot.send_message(
            chat_id=user_id,
            text=f"❌ An error occurred while deleting data: {e}",
            parse_mode="HTML"
        )

    return await back_to_database_menu(update, context)

async def back_to_manage_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    await context.bot.send_message(
        chat_id=user_id,
        text="🔙 Returning to User Management Panel...",
        parse_mode="Markdown"
    )
    return await show_manage_users_menu(update, context)


async def back_to_admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    await context.bot.send_message(
        chat_id=user_id,
        text="🔙 Returning to Admin Dashboard...",
        parse_mode="Markdown"
    )
    return await show_admin_menu(update, context)


async def show_manage_users_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Organized menu sections with emojis for better visual hierarchy
    keyboard = [
        [InlineKeyboardButton("🔍 View Banned Users", callback_data="view_banned_users")],
        [
            InlineKeyboardButton("🚫 Ban Job Seekers", callback_data="ban_job_seekers"),
            InlineKeyboardButton("🚫 Ban Employers", callback_data="ban_employers")
        ],
        [InlineKeyboardButton("✅ Unban Users", callback_data="unban_users_menu")],
        [
            InlineKeyboardButton("🗑 Remove Job Seekers", callback_data="remove_job_seekers"),
            InlineKeyboardButton("🗑 Remove Employers", callback_data="remove_employers")
        ],

        [
            InlineKeyboardButton("📤 Export Job Seekers", callback_data="export_job_seekers"),
            InlineKeyboardButton("📤 Export Employers", callback_data="export_employers")
        ],
        [InlineKeyboardButton("🔙 Back to Admin Menu", callback_data="back_to_admin_menu")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text="*👤 User Management Panel*\n\n"
             "*Please select an action:*\n"
             "• *Ban/Unban* - Manage user restrictions\n"
             "• *Remove* - Delete user accounts\n"
             "• *Export* - Download user data\n"
             "• *Clear* - Wipe all system data\n\n"
             "⚠️ *Caution:* Some actions are irreversible",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    return MANAGE_USERS


async def manage_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Redirect to the user management panel"""
    return await show_manage_users_menu(update, context)


async def manage_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show application management submenu"""
    query = update.callback_query
    if query:  # Only answer if it's a callback query
        await query.answer()

    user_id = update.effective_user.id

    keyboard = [
        [InlineKeyboardButton("List Applications", callback_data="list_applications")],
        [InlineKeyboardButton("Remove Applications", callback_data="remove_applications")],
        [InlineKeyboardButton("Export Applications", callback_data="export_applications")],
        [InlineKeyboardButton("Back to Database Menu", callback_data="back_to_database_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Use edit_message_text if it's a callback query, otherwise send new message
    if query and query.message:
        await query.edit_message_text(
            text="Application Management Options:",
            reply_markup=reply_markup
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text="Application Management Options:",
            reply_markup=reply_markup
        )
    return MANAGE_APPLICATIONS
#ban method
async def ban_job_seekers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Initiate job seeker ban process"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    await context.bot.send_message(
        chat_id=user_id,
        text="🔍 *Search Job Seekers for Ban*\n\n"
             "Enter search criteria:\n"
             "- Name\n"
             "- User ID\n",

        parse_mode="Markdown"
    )
    return SEARCH_JOB_SEEKERS_FOR_BAN

async def handle_job_seeker_ban_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1
    page_size = 10

    try:
        # Fetch paginated results
        job_seekers = db.search_job_seekers_for_ban(search_term, page=page, page_size=page_size)
        total_pages = db.get_total_pages_job_seekers_for_ban(search_term, page_size=page_size)

        if not job_seekers:
            await context.bot.send_message(
                chat_id=user_id,
                text="⚠️ No matching job seekers found."
            )
            return await back_to_manage_users(update, context)

        # Create paginated keyboard using the unified function
        keyboard = create_ban_paginated_keyboard(job_seekers, page, total_pages, "job_seeker", "ban")
        keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="back_to_manage_users")])

        await context.bot.send_message(
            chat_id=user_id,
            text=f"🔍 *Search Results*: '{search_term}'\n"
                 f"📄 Page 1/{total_pages}\n\n"
                 "Select a job seeker to ban:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Error handling job seeker ban search: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="❌ An error occurred while fetching job seekers. Please try again later."
        )
        return await back_to_manage_users(update, context)

    return BAN_JOB_SEEKERS_PAGINATED

async def confirm_ban_job_seeker(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    # Log the received callback_data for debugging
    logging.info(f"Received callback_data: {query.data}")

    # Validate the callback_data format
    if not query.data.startswith("ban_job_seeker_"):
        await query.edit_message_text("❌ Invalid action.")
        return await back_to_database_menu(update, context)

    try:
        # Extract the job seeker ID (everything after the prefix)
        job_seeker_id = int(query.data.split("ban_job_seeker_")[1])  # Ensure ID is an integer
        context.user_data["target_job_seeker_id"] = job_seeker_id  # Store properly
    except ValueError:
        await query.edit_message_text("❌ Invalid job seeker ID.")
        return await back_to_manage_users(update, context)

    await query.edit_message_text(
        text="📝 *Ban Reason Required*\n\n"
             "Please provide the reason for banning this job seeker:\n\n"
             "Examples:\n"
             "- Spam messages\n"
             "- Inappropriate profile\n"
             "- Violation of terms\n\n",
        parse_mode="Markdown"
    )
    return REASON_FOR_BAN_JOB_SEEKER

async def apply_ban_job_seeker(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    reason = update.message.text.strip()

    if not reason:
        await update.message.reply_text("❌ Ban reason cannot be empty. Please enter a valid reason.")
        return REASON_FOR_BAN_JOB_SEEKER

    target_job_seeker_id = context.user_data.get("target_job_seeker_id")

    if not target_job_seeker_id:
        await update.message.reply_text("❌ Error: Job seeker ID not found. Restart the process.")
        return await back_to_database_menu(update, context)

    try:
        # Apply the ban in the database
        db.ban_user(user_id=target_job_seeker_id, reason=reason, entity_type="job_seeker")

        # Notify the banned job seeker
        await context.bot.send_message(
            chat_id=target_job_seeker_id,
            text=(
                f"🚫 *{get_translation(target_job_seeker_id, 'you_have_been_banned')}*\n\n"
                f"{get_translation(target_job_seeker_id, 'reason')}: {reason}\n\n"
                f"{get_translation(target_job_seeker_id, 'ban_appeal_instructions')}"
            ),
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton("/start")]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
        )

        # Notify the admin
        await update.message.reply_text(f"✅ Successfully banned job seeker with ID {target_job_seeker_id}.")
        logging.info(f"Job seeker {target_job_seeker_id} successfully banned for reason: {reason}")

    except ValueError as e:
        logging.error(f"Validation error in ban_user: {e}")
        await update.message.reply_text(f"❌ Error: {e}")
    except Exception as e:
        logging.error(f"Unexpected error during job seeker ban: {e}", exc_info=True)
        await update.message.reply_text("❌ A server error occurred. Please try again later.")

    return await back_to_database_menu(update, context)

async def ban_employers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    # Prompt for search term
    await context.bot.send_message(
        chat_id=user_id,
        text="🔍 *Search Employers for Ban*\n\n"
             "Enter search criteria:\n"
             "- Name\n"
             "- User ID\n",
        parse_mode="Markdown"
    )
    return SEARCH_EMPLOYERS_FOR_BAN

async def handle_employer_ban_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    search_term = update.message.text.strip()
    context.user_data["search_term"] = search_term
    page = 1
    page_size = 10  # Add page size parameter
    try:
        employers = db.search_employers_for_ban(search_term, page=page, page_size=page_size)
        total_pages = db.get_total_pages_employers_for_ban(search_term, page_size=page_size)
        if not employers:
            await context.bot.send_message(
                chat_id=user_id,
                text="⚠️ No matching employers found."
            )
            return await back_to_database_menu(update, context)
        # Create paginated keyboard using the new function
        keyboard = create_ban_paginated_keyboard(employers, page, total_pages, "employer", "ban")
        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_manage_users")])
        await context.bot.send_message(
            chat_id=user_id,
            text=f"🔍 *Search Results*: '{search_term}'\n"
                 f"📄 {page}/{total_pages}\n\n"
                 "Select an employer to ban:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Error handling employer ban search: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while fetching employers. Please try again later."
        )
        return await back_to_database_menu(update, context)
    return BAN_EMPLOYERS_PAGINATED


async def confirm_ban_employer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data.split("_")

    if len(data) != 3 or data[0] != "ban" or data[1] != "employer":
        await query.edit_message_text("Invalid action.")
        return await back_to_database_menu(update, context)

    try:
        employer_id = int(data[2])  # Ensure ID is an integer
        context.user_data["target_employer_id"] = employer_id  # Store properly
        context.user_data["ban_reason_pending"] = True  # Flag to track state
    except ValueError:
        await query.edit_message_text("Invalid employer ID.")
        return await back_to_database_menu(update, context)

    await query.edit_message_text(
        text="📝 *Ban Reason Required*\n\n"
             "Please provide the reason for banning this employer:\n\n"
             "Examples:\n"
             "- Spam messages\n"
             "- Inappropriate profile\n"
             "- Violation of terms\n\n",
        parse_mode="Markdown"
    )
    return REASON_FOR_BAN_EMPLOYER


async def apply_ban_employer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    reason = update.message.text.strip() if update.message and update.message.text else None

    # Validate reason to ensure it's a non-empty string
    if not reason:
        logging.error("Invalid or empty ban reason received.")
        await update.message.reply_text("❌ Ban reason cannot be empty. Please enter a valid reason.")
        return REASON_FOR_BAN_EMPLOYER  # Ask for reason again

    target_employer_id = context.user_data.get("target_employer_id")

    if not target_employer_id:
        logging.error("Employer ID is missing from context.user_data.")
        await update.message.reply_text("❌ Error: Employer ID not found. Restart the process.")
        return await back_to_database_menu(update, context)

    try:
        # Explicitly ensure entity_type is "employer"
        db.ban_user(
            user_id=None,
            employer_id=target_employer_id,
            reason=reason,
            entity_type="employer"
        )

        # Notify the banned employer (localized)
        await context.bot.send_message(
            chat_id=target_employer_id,
            text=(
                f"🚫 *{get_translation(target_employer_id, 'you_have_been_banned')}*\n\n"
                f"{get_translation(target_employer_id, 'reason')}: {reason}\n\n"
                f"{get_translation(target_employer_id, 'ban_appeal_instructions')}"
            ),
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton("/start")]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
        )

        # Notify the admin
        await update.message.reply_text(f"✅ Successfully banned employer with ID {target_employer_id}.")
        logging.info(f"Employer {target_employer_id} successfully banned for reason: {reason}")

    except ValueError as e:
        logging.error(f"Validation error in ban_user: {e}")
        await update.message.reply_text(f"❌ Error: {e}")

    except Exception as e:
        logging.error(f"Unexpected error during employer ban: {e}", exc_info=True)
        await update.message.reply_text("❌ A server error occurred. Please try again later.")

    return await back_to_database_menu(update, context)


async def unban_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    banned_users = db.get_banned_users()
    if not banned_users:
        await context.bot.send_message(
            chat_id=user_id,
            text="No banned users found."
        )
        return await back_to_database_menu(update, context)
    keyboard = []
    for user in banned_users:
        text = f"{user['full_name']} (ID: {user['user_id']}) - Reason: {user['reason']}"
        callback_data = f"unban_{user['user_id']}"
        keyboard.append([InlineKeyboardButton(text, callback_data=callback_data)])
    keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_database_menu")])
    await context.bot.send_message(
        chat_id=user_id,
        text="Banned Users:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return UNBAN_USERS



async def view_banned_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    banned_users = db.get_banned_users()

    if not banned_users:
        await context.bot.send_message(
            chat_id=user_id,
            text="No banned users found."
        )
        return await back_to_database_menu(update, context)

    keyboard = []
    for user in banned_users:
        if user['user_id']:  # Job seeker
            text = f"👤 {user['name']} (ID: {user['user_id']}) - Reason: {user['reason']}"
            callback_data = f"unban_user_{user['user_id']}"
        else:  # Employer
            text = f"🏢 {user['name']} (ID: {user['employer_id']}) - Reason: {user['reason']}"
            callback_data = f"unban_employer_{user['employer_id']}"

        keyboard.append([InlineKeyboardButton(text, callback_data=callback_data)])

    keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="back_to_database_menu")])

    await context.bot.send_message(
        chat_id=user_id,
        text="🚫 Banned Users:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return UNBAN_USERS


async def handle_unban(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    try:
        # Parse callback data
        parts = query.data.split('_')
        if len(parts) != 3:
            raise ValueError("Invalid callback data format")

        action, entity_type, entity_id_str = parts
        entity_id = int(entity_id_str)

        # Get ban reason before unbanning
        if entity_type == "user":
            reason = db.get_ban_reason(user_id=entity_id)
            db.unban_user(user_id=entity_id)
            entity_name = "job seeker"
        elif entity_type == "employer":
            reason = db.get_ban_reason(employer_id=entity_id)
            db.unban_employer(employer_id=entity_id)
            entity_name = "employer"
        else:
            raise ValueError("Invalid entity type")

        # Notify the unbanned user/employer
        try:
            await context.bot.send_message(
                chat_id=entity_id,
                text=f"🎉 Your ban has been lifted!\n\n"
                     f"Previous reason: {reason}\n"
                     f"You can now use the bot normally."
            )
        except Exception as e:
            logging.error(f"Could not notify unbanned {entity_type} {entity_id}: {e}")

        # Confirm to admin
        await query.edit_message_text(
            text=f"✅ Successfully unbanned {entity_name} with ID {entity_id}.\n"
                 f"They have been notified.",
            reply_markup=None
        )

    except ValueError as e:
        logging.error(f"Validation error in handle_unban: {e}")
        await query.edit_message_text("❌ Invalid request. Please try again.")
    except Exception as e:
        logging.error(f"Error in handle_unban: {e}", exc_info=True)
        await query.edit_message_text("❌ Failed to unban. Please try again.")

    return await back_to_database_menu(update, context)


async def unban_users_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Get banned users from database
    banned_users = db.get_banned_users()

    if not banned_users:
        await query.edit_message_text("No banned users found.")
        return await back_to_database_menu(update, context)

    # Create menu with unban options
    keyboard = [
        [InlineKeyboardButton("📝 Unban by selection", callback_data="unban_by_selection")],
        [InlineKeyboardButton("🚀 Unban all users", callback_data="unban_all_confirmation")],
        [InlineKeyboardButton("🔙 Back", callback_data="back_to_manage_users")]
    ]

    await query.edit_message_text(
        text="🔓 Unban Users Menu:\n\n"
             f"Total banned users: {len(banned_users)}\n"
             "Choose an unban option:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return UNBAN_USERS_MENU


async def unban_by_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    banned_users = db.get_banned_users()
    context.user_data["banned_users_list"] = banned_users

    # Create numbered list with user details
    user_list = []
    for idx, user in enumerate(banned_users, start=1):
        if user['user_id']:
            user_type = "👤 Job Seeker"
            user_id = user['user_id']
            name = user.get('full_name', 'Unknown')
        else:
            user_type = "🏢 Employer"
            user_id = user['employer_id']
            name = user.get('company_name', 'Unknown')

        user_list.append(
            f"{idx}. {user_type} - {name} (ID: {user_id})\n"
            f"   Reason: {user['reason']}\n"
        )

    await query.edit_message_text(
        text="📋 Banned Users List:\n\n" + "".join(user_list) +
             "\nEnter numbers to unban (e.g., '1,3,5') or 'all' to unban everyone:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="unban_users_menu")]])
    )
    return UNBAN_SELECTION


async def handle_unban_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    selection = update.message.text.strip().lower()
    banned_users = context.user_data.get("banned_users_list", [])

    if not banned_users:
        await update.message.reply_text("No banned users found. Please try again.")
        return await back_to_database_menu(update, context)

    try:
        if selection == "all":
            # Unban all users
            success_count = 0
            for user in banned_users:
                try:
                    if user['user_id']:
                        db.unban_user(user['user_id'])
                        entity_id = user['user_id']
                        entity_type = "job seeker"
                    else:
                        db.unban_employer(user['employer_id'])
                        entity_id = user['employer_id']
                        entity_type = "employer"

                    # Notify unbanned user
                    try:
                        await context.bot.send_message(
                            chat_id=entity_id,
                            text=f"🎉 Your ban has been lifted!\n\n"
                                 f"Reason: {user['reason']}\n"
                                 f"You can now use the bot normally."
                        )
                    except Exception as e:
                        logging.error(f"Could not notify unbanned user {entity_id}: {e}")

                    success_count += 1
                except Exception as e:
                    logging.error(f"Error unbanning user {entity_id}: {e}")

            await update.message.reply_text(
                f"✅ Successfully unbanned {success_count}/{len(banned_users)} users."
            )
        else:
            # Unban selected users
            selected_indices = [int(num.strip()) - 1 for num in selection.split(",")]
            success_count = 0

            for idx in selected_indices:
                if 0 <= idx < len(banned_users):
                    user = banned_users[idx]
                    try:
                        if user['user_id']:
                            db.unban_user(user['user_id'])
                            entity_id = user['user_id']
                            entity_type = "job seeker"
                        else:
                            db.unban_employer(user['employer_id'])
                            entity_id = user['employer_id']
                            entity_type = "employer"

                        # Notify unbanned user
                        try:
                            await context.bot.send_message(
                                chat_id=entity_id,
                                text=f"🎉 Your ban has been lifted!\n\n"
                                     f"Reason: {user['reason']}\n"
                                     f"You can now use the bot normally."
                            )
                        except Exception as e:
                            logging.error(f"Could not notify unbanned user {entity_id}: {e}")

                        success_count += 1
                    except Exception as e:
                        logging.error(f"Error unbanning user {entity_id}: {e}")

            await update.message.reply_text(
                f"✅ Successfully unbanned {success_count}/{len(selected_indices)} selected users."
            )

    except Exception as e:
        logging.error(f"Error in handle_unban_selection: {e}")
        await update.message.reply_text(
            "❌ Invalid input. Please enter numbers separated by commas (e.g., '1,3,5') or 'all'."
        )
        return UNBAN_SELECTION

    return await back_to_database_menu(update, context)


async def confirm_unban_all(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    keyboard = [
        [InlineKeyboardButton("✅ Confirm Unban All", callback_data="execute_unban_all")],
        [InlineKeyboardButton("🔙 Cancel", callback_data="unban_users_menu")]
    ]

    await query.edit_message_text(
        text="⚠️ Are you sure you want to unban ALL users?\n"
             "This action cannot be undone!",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return UNBAN_ALL_CONFIRMATION


async def execute_unban_all(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    banned_users = db.get_banned_users()
    success_count = 0

    for user in banned_users:
        try:
            if user['user_id']:
                db.unban_user(user['user_id'])
                entity_id = user['user_id']
                entity_type = "job seeker"
            else:
                db.unban_employer(user['employer_id'])
                entity_id = user['employer_id']
                entity_type = "employer"

            # Notify unbanned user
            try:
                await context.bot.send_message(
                    chat_id=entity_id,
                    text=f"🎉 Your ban has been lifted!\n\n"
                         f"Reason: {user['reason']}\n"
                         f"You can now use the bot normally."
                )
            except Exception as e:
                logging.error(f"Could not notify unbanned user {entity_id}: {e}")

            success_count += 1
        except Exception as e:
            logging.error(f"Error unbanning user {entity_id}: {e}")

    await query.edit_message_text(
        text=f"✅ Successfully unbanned {success_count}/{len(banned_users)} users.",
        reply_markup=None
    )
    return await back_to_database_menu(update, context)


def create_ban_paginated_keyboard(items, current_page, total_pages, entity_type, action="ban"):
    keyboard = []
    # Add item buttons with ban action
    for item in items:
        if entity_type == "job_seeker":
            text = f"{item['full_name']} (ID: {item['user_id']})"
            callback_data = f"ban_job_seeker_{item['user_id']}"
        elif entity_type == "employer":
            text = f"{item['company_name']} (ID: {item['employer_id']})"
            callback_data = f"{action}_employer_{item['employer_id']}"
        else:
            continue
        keyboard.append([InlineKeyboardButton(text, callback_data=callback_data)])

    # Add pagination buttons - use target page numbers instead of current page
    nav_buttons = []
    if current_page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Prev", callback_data=f"prev_{entity_type}_{current_page-1}"))
    if current_page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Next ➡️", callback_data=f"next_{entity_type}_{current_page+1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    return keyboard

async def reject_appeal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data.split("_")
    if len(data) != 3:
        await query.edit_message_text("Invalid action.")
        return await back_to_database_menu(update, context)
    _, _, user_id = data
    try:
        db.reject_appeal(user_id)
        await context.bot.send_message(
            chat_id=user_id,
            text="Your appeal has been rejected. For more details, please contact the admin."
        )
        await query.edit_message_text(
            text="Appeal has been rejected successfully."
        )
    except Exception as e:
        logging.error(f"Error rejecting appeal: {e}")
        await query.edit_message_text(
            text="An error occurred while rejecting the appeal. Please try again later."
        )
    return await back_to_database_menu(update, context)

async def confirm_ban(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data.split("_")
    if len(data) != 3:
        await query.edit_message_text("Invalid action.")
        return await back_to_database_menu(update, context)
    action, entity_type, user_id = data
    context.user_data["target_user_id"] = user_id
    await query.edit_message_text(
        text="Enter the reason for banning this user:"
    )
    return REASON_FOR_BAN

async def apply_ban(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    target_user_id = context.user_data.get("target_user_id")
    reason = update.message.text.strip()
    try:
        db.ban_user(target_user_id, reason)
        await context.bot.send_message(
            chat_id=target_user_id,
            text=f"You are temporarily banned from using this bot due to: {reason}. Please contact the admin for more details."
        )
        await context.bot.send_message(
            chat_id=user_id,
            text="User has been banned successfully."
        )
    except Exception as e:
        logging.error(f"Error applying ban: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred while banning the user. Please try again later."
        )
    return await back_to_database_menu(update, context)


async def start_ban_appeal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not db.is_user_banned(user_id=user_id) and not db.is_user_banned(employer_id=user_id):
        await query.edit_message_text("You are not currently banned.")
        return ConversationHandler.END

    # Store the original message ID for cleanup
    context.user_data["ban_message_id"] = query.message.message_id

    try:
        await query.edit_message_text(
            f"{get_translation(user_id, 'appeal_instructions')}\n\n"
            f"{get_translation(user_id, 'appeal_details_prompt')}"
        )
    except Exception as e:
        logging.error(f"Error editing message: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="✍️ Please write your appeal message (max 500 characters):\n\n"
                 "Explain why you believe the ban should be lifted. "
                 "Include any relevant details or evidence."
        )

    return APPEAL_INPUT


def get_complete_ban_reason(user_id):
    """Get ban reason checking both job seeker and employer status"""
    # Check job seeker ban first
    if db.is_user_banned(user_id=user_id, employer_id=None):
        reason = db.get_ban_reason(user_id=user_id, employer_id=None)
        if reason:
            return reason

    # Check employer ban if no job seeker ban found
    employer_profile = db.get_employer_profile(user_id)
    if employer_profile and db.is_user_banned(user_id=None, employer_id=employer_profile.get("employer_id")):
        reason = db.get_ban_reason(user_id=None, employer_id=employer_profile.get("employer_id"))
        if reason:
            return reason

    return "No reason provided"

async def handle_appeal_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.message.from_user.id
    appeal_text = update.message.text.strip()

    if len(appeal_text) > 500:
        await update.message.reply_text("❌ Appeal is too long (max 500 characters). Please shorten it.")
        return APPEAL_INPUT

    try:
        # Clean up original ban message if possible
        if "ban_message_id" in context.user_data:
            try:
                await context.bot.delete_message(
                    chat_id=user_id,
                    message_id=context.user_data["ban_message_id"]
                )
            except Exception as e:
                logging.error(f"Couldn't delete ban message: {e}")

        # Save appeal to database
        db.create_appeal(
            user_id=user_id,
            content=appeal_text
        )

        # Get user details for admin notification
        user_type = "Job Seeker"
        name = db.get_user_profile(user_id).get('full_name', 'Unknown')
        if db.is_user_banned(employer_id=user_id):  # Check if employer
            employer = db.get_employer_profile(user_id)
            user_type = "Employer"
            name = employer.get('company_name', 'Unknown')

        ban_reason = get_complete_ban_reason(user_id)

        # Notify all active admins
        for admin_id in active_admins:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=f"📩 New Ban Appeal\n\n"
                         f"User Type: {user_type}\n"
                         f"Name: {name}\n"
                         f"User ID: {user_id}\n\n"
                         f"Ban Reason: {ban_reason}\n\n"
                         f"Appeal Content:\n{appeal_text}\n\n",

                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("🔓 Lift Ban", callback_data=f"lift_ban_{user_id}"),
                         InlineKeyboardButton("🔒 Uphold Ban", callback_data=f"uphold_ban_{user_id}")],
                        [InlineKeyboardButton("ℹ️ Request Info", callback_data=f"request_info_{user_id}")]
                    ])
                )
            except Exception as e:
                logging.error(f"Could not notify admin {admin_id}: {e}")

        # Confirm to user
        await update.message.reply_text(
            "✅ Your appeal has been submitted!\n\n"
            "Our team will review it when available. "
            "You'll receive a notification when a decision is made."
        )

    except Exception as e:
        logging.error(f"Error handling appeal: {e}")
        await update.message.reply_text("❌ Failed to submit appeal. Please try again later.")

    return ConversationHandler.END


async def review_appeal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        if not context.args:
            await update.message.reply_text("Usage: /review_appeal <user_id>")
            return

        user_id = int(context.args[0])
        appeal = db.get_appeal(user_id)

        if not appeal:
            await update.message.reply_text("No pending appeal found for this user.")
            return

        # Get user details
        user_type = "Job Seeker"
        name = db.get_user_profile(user_id).get('full_name', 'Unknown')
        if db.is_user_banned(employer_id=user_id):
            employer = db.get_employer_profile(user_id)
            user_type = "Employer"
            name = employer.get('company_name', 'Unknown')

        ban_reason = get_complete_ban_reason(user_id)

        # Updated to match the new callback patterns
        keyboard = [
            [InlineKeyboardButton("🔓 Lift Ban", callback_data=f"lift_ban_{user_id}"),
             InlineKeyboardButton("🔒 Uphold Ban", callback_data=f"uphold_ban_{user_id}")],
            [InlineKeyboardButton("ℹ️ Request Info", callback_data=f"request_info_{user_id}")]
        ]

        await update.message.reply_text(
            f"📄 Appeal Review\n\n"
            f"User Type: {user_type}\n"
            f"Name: {name}\n"
            f"User ID: {user_id}\n\n"
            f"Ban Reason: {ban_reason}\n\n"
            f"Appeal Content:\n{appeal['content']}\n\n"
            f"Submitted: {appeal['review_date']}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return APPEAL_REVIEW

    except ValueError:
        await update.message.reply_text("❌ Invalid user ID. Please provide a numeric ID.")
    except Exception as e:
        logging.error(f"Error in review_appeal: {str(e)}")
        await update.message.reply_text("❌ Failed to review appeal. Please try again.")


async def handle_appeal_decision(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    try:
        # Extract action and user_id from callback_data
        action, user_id_str = query.data.split('_')[:2], query.data.split('_')[-1]
        user_id = int(user_id_str)
        action = '_'.join(action)  # Reconstructs "lift_ban", "uphold_ban", or "request_info"

        # Get ban details
        is_job_seeker = db.is_user_banned(user_id=user_id)
        is_employer = not is_job_seeker and db.is_user_banned(employer_id=user_id)

        if not (is_job_seeker or is_employer):
            await query.edit_message_text("⚠️ User is not currently banned")
            return ConversationHandler.END

        # Map actions to status values
        action_status_map = {
            "lift_ban": "ban_lifted",
            "uphold_ban": "ban_upheld",
            "request_info": "info_requested"
        }

        status = action_status_map.get(action)
        if not status:
            await query.edit_message_text("❌ Invalid action")
            return ConversationHandler.END

        # Process the action
        if action == "lift_ban":
            if is_job_seeker:
                db.unban_user(user_id)
            else:
                db.unban_employer(user_id)

            # Notify user
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'ban_lifted_success')
            )

            await query.edit_message_text(f"✅ Successfully lifted ban for user {user_id}")

        elif action == "uphold_ban":
            # Notify user
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'ban_rejected_notice')
            )

            await query.edit_message_text(f"ℹ️ Ban upheld for user {user_id}")

        elif action == "request_info":
            # Notify user
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'need_more_info_appeal')
            )

            await query.edit_message_text(f"📝 Requested more info from user {user_id}")

        # Update appeal status in database
        db.update_appeal_status(
            user_id=user_id,
            status=status
        )

    except Exception as e:
        logging.error(f"Error in handle_appeal_decision: {e}")
        await query.edit_message_text("❌ Failed to process appeal decision")

    return ConversationHandler.END
start_time = datetime.now()


async def generate_stats_message(stats: dict) -> str:
    """Generate formatted stats message from statistics dictionary."""

    def format_growth(value):
        value = float(value or 0)
        if value > 0:
            return f"📈 +{escape_markdown(value)}%"
        elif value < 0:
            return f"📉 {escape_markdown(value)}%"
        else:
            return "➖ 0%"

    # Pre-format the date to avoid backslashes in f-string
    last_updated = escape_markdown(datetime.now().strftime('%Y-%m-%d %H:%M'))

    return (
        "📊 *Database Statistics Dashboard* 📊\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "🚀 *Quick Summary*\n"
        f"├ 👥 Users: `{escape_markdown(stats['total_users'])}` ({format_growth(stats['user_growth_week'])})\n"
        f"├ 💼 Vacancies: `{escape_markdown(stats['active_jobs'])}` active / `{escape_markdown(stats['completed_jobs'])}` done\n"
        f"├ 📨 Apps: `{escape_markdown(stats['total_applications'])}` (Pending: `{escape_markdown(stats['pending_applications'])}`)\n"
        f"└ ⚠️ Errors: `{escape_markdown(stats['error_rate'])}%`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "👥 *USERS*\n"
        f"├ Total: `{escape_markdown(stats['total_users'])}`\n"
        f"├ Active: `{escape_markdown(stats['active_users'])}`\n"
        f"├ Inactive: `{escape_markdown(stats['inactive_users'])}`\n"
        f"├ New (7d): `{escape_markdown(stats['new_users_last_7_days'])}` ({format_growth(stats['user_growth_week'])})\n"
        f"└ New (30d): `{escape_markdown(stats['new_users_last_30_days'])}` ({format_growth(stats['user_growth_month'])})\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "💼 *Vacancies*\n"
        f"├ Total: `{escape_markdown(stats['total_jobs'])}`\n"
        f"├ Active: `{escape_markdown(stats['active_jobs'])}`\n"
        f"├ Rejected: `{escape_markdown(stats['rejected_jobs'])}`\n"
        f"└ Completed: `{escape_markdown(stats['completed_jobs'])}`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📨 *APPLICATIONS*\n"
        f"├ Total: `{escape_markdown(stats['total_applications'])}`\n"
        f"├ Avg/Job: `{escape_markdown(stats['average_applications_per_job'])}`\n"
        f"├ Pending: `{escape_markdown(stats['pending_applications'])}`\n"
        f"├ Approved: `{escape_markdown(stats['approved_applications'])}`\n"
        f"└ Rejected: `{escape_markdown(stats['rejected_applications'])}`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📈 *TRENDS*\n"
        f"└ Signup Rate: `{escape_markdown(stats['user_signup_rate'])}%`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "⚙️ *SYSTEM*\n"
        f"├ DB Size: `{escape_markdown(stats['db_size'])}`\n"
        f"├ Error Rate: `{escape_markdown(stats['error_rate'])}%`\n"
        f"├ CPU Usage: `{escape_markdown(str(stats['cpu_usage']))}%`\n"
        f"├ Memory Usage: `{escape_markdown(str(stats['memory_usage']))}%`\n"
        f"└ Bot Uptime: `{escape_markdown(stats['bot_uptime'])}`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🔄 *Last Updated*: `{last_updated}`\n"
    )

async def db_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer("⏳ Preparing database status... Please wait while fetching datas")
    # await query.answer()

    try:
        stats = await fetch_latest_stats()
    except Exception as e:
        logging.error(f"Error fetching database stats: {e}")
        await query.edit_message_text(" Error fetching database statistics. Please try again later.")
        return DB_STATS_VIEW

    reply_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Refresh", callback_data="refresh_stats")],
        [InlineKeyboardButton("⬅️ Back", callback_data="back_to_database_menu")]
    ])

    await query.edit_message_text(
        text=await generate_stats_message(stats),
        reply_markup=reply_markup,
        parse_mode="MarkdownV2"
    )
    return DB_STATS_VIEW

async def refresh_db_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer("🔄 Refreshing...")

    try:
        new_stats = await fetch_latest_stats()
        old_stats = context.user_data.get('last_stats', {})

        if stats_equal(old_stats, new_stats) and not context.user_data.get('force_refresh'):
            await query.answer("✅ Data is already up-to-date", show_alert=True)
            return DB_STATS_VIEW

        context.user_data['last_stats'] = new_stats
        context.user_data['force_refresh'] = False

        await query.edit_message_text(
            text=await generate_stats_message(new_stats),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="back_to_database_menu")]
            ]),
            parse_mode="MarkdownV2"
        )
        return DB_STATS_VIEW

    except Exception as e:
        logging.error(f"Refresh error: {e}")
        return DB_STATS_VIEW

async def fetch_latest_stats():
    """Fetch all current statistics from database"""
    return {
        'total_users': db.get_total_users(),
        'active_users': db.get_active_users(),
        'inactive_users': db.get_inactive_users(),
        'new_users_last_7_days': db.get_new_users(period='last_7_days'),
        'new_users_last_30_days': db.get_new_users(period='last_30_days'),
        'user_growth_week': db.get_user_growth_rate('week'),
        'user_growth_month': db.get_user_growth_rate('month'),
        'total_jobs': db.get_total_jobs(),
        'active_jobs': db.get_active_jobs(),
        'rejected_jobs': db.get_rejected_jobs(),
        'completed_jobs': db.get_completed_jobs(),
        'total_applications': db.get_total_applications(),
        'average_applications_per_job': db.get_average_applications_per_job(),
        'pending_applications': db.get_application_count_by_status('pending'),
        'approved_applications': db.get_application_count_by_status('approved'),
        'rejected_applications': db.get_application_count_by_status('rejected'),
        'db_size': db.get_database_size(),
        'error_rate': db.get_error_rate(),
        'user_signup_rate': db.get_user_signup_rate(),
        'cpu_usage': db.get_system_resources()['cpu_usage'],
        'memory_usage': db.get_system_resources()['memory_usage'],
        'bot_uptime': db.get_bot_uptime(start_time),
        'fetch_time': datetime.now()
    }

def stats_equal(old, new):
    """Compare relevant stat fields"""
    compare_fields = ['total_users', 'active_jobs', 'pending_applications']
    return all(old.get(field) == new.get(field) for field in compare_fields)
#help option
async def show_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Get a random smart tip with decorative emoji
    smart_tips = [
        f"💎 {get_translation(user_id, 'tip_portfolio')}",
        f"📢 {get_translation(user_id, 'tip_communication')}",
        f"✅ {get_translation(user_id, 'tip_profile_completion')}",
        f"🤝 {get_translation(user_id, 'tip_negotiation')}"
    ]
    random_tip = random.choice(smart_tips)

    # Create help keyboard with better visual hierarchy
    keyboard = [
        [InlineKeyboardButton(
            f"📚 {get_translation(user_id, 'faq_section')}",
            callback_data="help_faq"
        )],
        [InlineKeyboardButton(
            f"📬 {get_translation(user_id, 'contact_admin')}",
            callback_data="help_contact"
        )],
        [InlineKeyboardButton(
            f" {get_translation(user_id, 'back_to_main')}",
            callback_data="help_back"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Build beautiful help message with ASCII art and sections
    help_message = (
        f"<b>🌟 {get_translation(user_id, 'help_center_title')} 🌟</b>\n\n"
        f"┏━━━━━━━━━━━━━━━━━━━┓\n"
        f"  {get_translation(user_id, 'help_intro')}  \n"
        f"┗━━━━━━━━━━━━━━━━━━━┛\n\n"
        f"<b>💡 {get_translation(user_id, 'smart_tip')}:</b>\n"
        f"  <i> {random_tip}</i>\n\n"
        f"<b>📋 {get_translation(user_id, 'help_choose_option')}</b>\n"
        f"▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰"
    )

    # Send or update message with HTML formatting
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=help_message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=help_message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )

    return HELP_MENU

async def help_button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    data = query.data

    if data == "help_faq":
        return await show_faq_category_section(update, context)

    elif data == "help_contact":
        return await show_contact_options(update, context)
    elif data == "help_back":
        return await back_to_main_menu_report(update, context)

    return HELP_MENU




async def show_faq_category_section(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Main FAQ categories
    faq_main_categories = {
        "employer": "👨‍💼 " + get_translation(user_id, "employer_faq"),
        "job_seeker": "👤 " + get_translation(user_id, "job_seeker_faq"),
        "admin": "🔒 " + get_translation(user_id, "admin_faq")
    }

    # Check if we're coming from a sub-category selection
    if query.data.startswith("faq_main_"):  # Changed from "faq_main_" to "faq_main_"
        faq_type = "_".join(query.data.split("_")[2:])

        if faq_type == "employer":
            return await show_faq_section(update, context)
        elif faq_type == "job_seeker":
            return await show_job_seeker_faq(update, context)
        elif faq_type == "admin":
            return await show_admin_faq(update, context)
        else:
            await query.edit_message_text(
                text=get_translation(user_id, "invalid_selection_message")
            )
            return FAQ_SECTION

    # Create inline keyboard buttons for main FAQ categories
    keyboard = [
        [InlineKeyboardButton(category, callback_data=f"faq_main_{key}")]
        for key, category in faq_main_categories.items()
    ]
    # Change the back button's callback_data to "faq_back_help"
    keyboard.append([InlineKeyboardButton(
        "🔙 " + get_translation(user_id, "back_to_help"),
        callback_data="help_back")  # Updated callback data
    ])
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Build FAQ category selection message
    faq_category_message = (
        f"📚 {get_translation(user_id, 'faq_section')}\n\n"
        f"{get_translation(user_id, 'select_faq_category')}"
    )

    await query.edit_message_text(
        text=faq_category_message,
        reply_markup=reply_markup
    )

    return FAQ_SECTION

import random


async def show_faq_section(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Random helpful tips
    smart_tips = [
        get_translation(user_id, "tip_profile_completion"),
        get_translation(user_id, "tip_post_job"),
        get_translation(user_id, "tip_manage_vacancies"),
        get_translation(user_id, "tip_view_analytics"),
        get_translation(user_id, "tip_change_language")
    ]
    random_tip = random.choice(smart_tips)

    # Organize FAQs into categories for better navigation
    faq_categories = {
        "profile_account": [
            ("faq_profile_completion", "faq_profile_strength", "faq_edit_profile", "faq_delete_account"),
        ],
        "vacancies_applications": [
            ("faq_active_vacancies", "faq_post_job", "faq_manage_vacancies", "faq_view_applicants"),
        ],
        "analytics_performance": [
            ("faq_view_analytics", "faq_export_data"),
        ],
        "language_settings": [
            ("faq_change_language"),
        ],
        "tips_suggestions": [
            ("faq_employer_tips"),
        ],
        "ban_appeal": [
            ("js_faq_ban_notification"),("js_faq_appeal_process"),("js_faq_appeal_review"),("js_faq_ban_removal"),("js_faq_appeal_rejection"),("js_faq_admin_ban_criteria"), ("js_faq_ban_reason_visibility"), ("js_faq_admin_ban_tools"),
        ]
    }

    # Create inline keyboard buttons for FAQ categories
    keyboard = []
    for index, category_key in enumerate(faq_categories.keys()):
        translated_category = get_translation(user_id, category_key)
        button = InlineKeyboardButton(f"📂 {translated_category}", callback_data=f"faq_category_{index}")
        keyboard.append([button])
    keyboard.append([InlineKeyboardButton("🔙 " + get_translation(user_id, "back_to_help"), callback_data="help_back")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Build FAQ section message with a random tip
    faq_message = (
        f"📚 {get_translation(user_id, 'faq_section')}\n\n"
        f"{get_translation(user_id, 'faq_intro')}\n\n"
        f"💡 {get_translation(user_id, 'smart_tip')}: {random_tip}\n\n"
        f"{get_translation(user_id, 'choose_faq_category')}"
    )

    # Send or update message
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=faq_message,
            reply_markup=reply_markup
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=faq_message,
            reply_markup=reply_markup
        )

    return FAQ_SECTION


async def handle_faq_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    data = query.data

    try:
        # Remove the manual back button handling
        if data.startswith("faq_category_"):
            category_index = int(data.split("_")[2])
            context.user_data['current_faq_category'] = category_index
        else:
            raise ValueError("Invalid callback_data format")

        # Map category index to FAQ category
        faq_categories = {
            0: "👤 Profile & Account",
            1: "💼 Vacancies & Applications",
            2: "📊 Analytics & Performance",
            3: "🌐 Language & Settings",
            4: "💡 Tips & Suggestions",
            5: "🚫 Ban & Appeal"
        }

        if category_index not in faq_categories:
            raise KeyError(f"Invalid category index: {category_index}")

        selected_category = faq_categories[category_index]

        # Get FAQ questions for selected category
        faq_questions = {
            "profile_account": [
                "faq_profile_completion", "faq_profile_strength", "faq_edit_profile", "faq_delete_account"
            ],
            "vacancies_applications": [
                "faq_active_vacancies", "faq_post_job", "faq_manage_vacancies", "faq_view_applicants"
            ],
            "analytics_performance": [
                "faq_view_analytics", "faq_export_data"
            ],
            "language_settings": [
                "faq_change_language"
            ],
            "tips_suggestions": [
                "faq_employer_tips"
            ],
            "ban_appeal": [
                "js_faq_ban_notification",
                "js_faq_appeal_process",
                "js_faq_appeal_review",
                "js_faq_ban_removal",
                "js_faq_appeal_rejection",
                "js_faq_admin_ban_criteria",
                "js_faq_ban_reason_visibility",
                "js_faq_admin_ban_tools"
            ]
        }

        # Create keyboard with FAQ questions
        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, q), callback_data=f"faq_question_{q}")]
            for q in faq_questions[selected_category]
        ]
        keyboard.append([
            InlineKeyboardButton(
                "🔙 " + get_translation(user_id, "back_to_faq"),
                callback_data="faq_category_back"
            )
        ])

        # Build message text
        message = (
            f"📂 {selected_category}\n\n"
            f"{get_translation(user_id, 'faq_category_intro')}\n\n"
            f"{get_translation(user_id, 'choose_faq_question')}"
        )

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return FAQ_CATEGORY

    except (ValueError, IndexError) as e:
        error_msg = f"⚠️ Error: {str(e)}\n\n{get_translation(user_id, 'try_again')}"
        await query.edit_message_text(text=error_msg)
        return FAQ_SECTION

    except KeyError as e:
        error_msg = f"⚠️ Error: {str(e)}\n\n{get_translation(user_id, 'invalid_selection')}"
        await query.edit_message_text(text=error_msg)
        return FAQ_SECTION

async def handle_faq_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    try:
        # Extract question key from callback data
        prefix, action, question_key = query.data.split("_", 2)

        # Validate the prefix and action
        if prefix != "faq" or action != "question":
            raise ValueError("Invalid callback_data format")

        # Get the answer for the selected question
        faq_answers = {
            "faq_profile_completion": "faq_profile_completion_answer",
            "faq_profile_strength": "faq_profile_strength_answer",
            "faq_edit_profile": "faq_edit_profile_answer",
            "faq_delete_account": "faq_delete_account_answer",
            "faq_active_vacancies": "faq_active_vacancies_answer",
            "faq_post_job": "faq_post_job_answer",
            "faq_manage_vacancies": "faq_manage_vacancies_answer",
            "faq_view_applicants": "faq_view_applicants_answer",
            "faq_view_analytics": "faq_view_analytics_answer",
            "faq_export_data": "faq_export_data_answer",
            "faq_change_language": "faq_change_language_answer",
            "faq_employer_tips": "faq_employer_tips_answer",
            "js_faq_ban_notification": "js_faq_ban_notification_answer",
            "js_faq_appeal_process": "js_faq_appeal_process_answer",
            "js_faq_appeal_review": "js_faq_appeal_review_answer",
            "js_faq_ban_removal": "js_faq_ban_removal_answer",
            "js_faq_appeal_rejection": "js_faq_appeal_rejection_answer",
            "js_faq_admin_ban_criteria": "js_faq_admin_ban_criteria_answer",
            "js_faq_ban_reason_visibility": "js_faq_ban_reason_visibility_answer",
            "js_faq_admin_ban_tools": "js_faq_admin_ban_tools_answer"
        }

        # Ensure the question key exists in the FAQ answers
        if question_key not in faq_answers:
            raise KeyError(f"Question key '{question_key}' not found in FAQ answers")

        # Get the translated question and answer
        question = get_translation(user_id, question_key)
        answer = get_translation(user_id, faq_answers[question_key])

        category_index = context.user_data.get('current_faq_category', 0)
        keyboard = [
            [InlineKeyboardButton("🔙 " + get_translation(user_id, "back_to_faq_category"),
                                  callback_data=f"faq_question_back_{category_index}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        # Build FAQ question message
        faq_question_message = (
            f"❓ {question}\n\n"
            f"✅ {answer}"
        )

        await query.edit_message_text(
            text=faq_question_message,
            reply_markup=reply_markup
        )

        return FAQ_QUESTION

    except ValueError as e:
        # Handle invalid callback_data format
        error_message = f"Error: Invalid callback_data format. Details: {str(e)}"
        await query.edit_message_text(text=error_message)
        return FAQ_CATEGORY

    except KeyError as e:
        # Handle missing question key
        error_message = f"Error: {str(e)}"
        await query.edit_message_text(text=error_message)
        return FAQ_CATEGORY

# job seeker faq
async def show_job_seeker_faq(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Job seeker specific smart tips
    smart_tips = [
        get_translation(user_id, "js_tip_complete_profile"),
        get_translation(user_id, "js_tip_application_strategy"),
        get_translation(user_id, "js_tip_search_filters"),
        get_translation(user_id, "js_tip_document_upload"),
        get_translation(user_id, "js_tip_profile_visibility")
    ]
    random_tip = random.choice(smart_tips)

    # Organized categories for job seekers
    faq_categories = {
        "registration_profile": [
            "js_faq_registration_steps",
            "js_faq_profile_update",
            "js_faq_document_upload",
            "js_faq_optional_fields"
        ],
        "job_search_applications": [
            "js_faq_job_application_process",
            "js_faq_application_status",
            "js_faq_application_outcome",
            "js_faq_application_export"
        ],
        "search_filters": [
            "js_faq_job_search",
            "js_faq_advanced_filters",
            "js_faq_search_sorting",
            "js_faq_search_saving"
        ],
        "account_management": [
            "js_faq_account_deletion",
            "js_faq_data_visibility"
        ],
        "ban_appeal": [
            "js_faq_ban_notification",
            "js_faq_appeal_process",
            "js_faq_appeal_review",
            "js_faq_ban_removal",
            "js_faq_appeal_rejection",
            "js_faq_admin_ban_criteria",
            "js_faq_ban_reason_visibility",
            "js_faq_admin_ban_tools"
        ],
    }

    keyboard = [
        [InlineKeyboardButton(f"📂 {category}", callback_data=f"js_faq_category_{index}")]
        for index, category in enumerate(faq_categories.keys())
    ]
    keyboard.append(
        [InlineKeyboardButton("🔙 " + get_translation(user_id, "back_to_help"), callback_data="js_faq_back")])

    message = (
        f"📚 {get_translation(user_id, 'js_faq_welcome')}\n\n"
        f"{get_translation(user_id, 'js_faq_intro')}\n\n"
        f"💡 {get_translation(user_id, 'smart_tip')}: {random_tip}\n\n"
        f"{get_translation(user_id, 'choose_faq_category')}"
    )

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    return JS_FAQ_SECTION


async def handle_job_seeker_faq_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    faq_categories = {
        0: "📝 Registration & Profile",
        1: "🔍 Job Search & Applications",
        2: "📂 Search & Filters",
        3: "🔒 Account Management",
        4: "🚫 Ban & Appeal"
    }

    try:
        category_index = int(query.data.split("_")[3])
        # Store the current category index in user_data
        context.user_data['current_js_category'] = category_index
        selected_category = faq_categories[category_index]

        category_questions = {
            "registration_profile": [
                "js_faq_registration_steps",
                "js_faq_profile_update",
                "js_faq_document_upload",
                "js_faq_optional_fields"
            ],
            "job_search_applications": [
                "js_faq_job_application_process",
                "js_faq_application_status",
                "js_faq_application_outcome",
                "js_faq_application_export"
            ],
            "search_filters": [
                "js_faq_job_search",
                "js_faq_advanced_filters",
                "js_faq_search_sorting",
                "js_faq_search_saving"
            ],
            "account_management": [
                "js_faq_account_deletion",
                "js_faq_data_visibility"
            ],
            "ban_appeal": [
                "js_faq_ban_notification",
                "js_faq_appeal_process",
                "js_faq_appeal_review",
                "js_faq_ban_removal",
                "js_faq_appeal_rejection",
                "js_faq_admin_ban_criteria",
                "js_faq_ban_reason_visibility",
                "js_faq_admin_ban_tools"
            ],
        }

        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, q), callback_data=f"js_faq_q_{q}")]
            for q in category_questions[selected_category]
        ]
        keyboard.append([InlineKeyboardButton("🔙 Back", callback_data=f"js_faq_return_{category_index}")])

        message = (
            f"📂 {selected_category}\n\n"
            f"{get_translation(user_id, 'faq_category_intro')}\n\n"
            f"{get_translation(user_id, 'choose_faq_question')}"
        )

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return JS_FAQ_CATEGORY

    except Exception as e:
        await query.edit_message_text(text=f"⚠️ Error: {str(e)}")
        return JS_FAQ_SECTION


async def handle_job_seeker_faq_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Job seeker FAQ answers mapping
    faq_answers = {
        "js_faq_registration_steps": "js_faq_registration_steps_answer",
        "js_faq_profile_update": "js_faq_profile_update_answer",
        "js_faq_document_upload": "js_faq_document_upload_answer",
        "js_faq_optional_fields": "js_faq_optional_fields_answer",
        "js_faq_job_application_process": "js_faq_job_application_process_answer",
        "js_faq_application_status": "js_faq_application_status_answer",
        "js_faq_application_outcome": "js_faq_application_outcome_answer",
        "js_faq_application_export": "js_faq_application_export_answer",
        "js_faq_job_search": "js_faq_job_search_answer",
        "js_faq_advanced_filters": "js_faq_advanced_filters_answer",
        "js_faq_search_sorting": "js_faq_search_sorting_answer",
        "js_faq_search_saving": "js_faq_search_saving_answer",
        "js_faq_account_deletion": "js_faq_account_deletion_answer",
        "js_faq_data_visibility": "js_faq_data_visibility_answer",
        "js_faq_ban_notification": "js_faq_ban_notification_answer",
        "js_faq_appeal_process": "js_faq_appeal_process_answer",
        "js_faq_appeal_review": "js_faq_appeal_review_answer",
        "js_faq_ban_removal": "js_faq_ban_removal_answer",
        "js_faq_appeal_rejection": "js_faq_appeal_rejection_answer",
        "js_faq_admin_ban_criteria": "js_faq_admin_ban_criteria_answer",
        "js_faq_ban_reason_visibility": "js_faq_ban_reason_visibility_answer",
        "js_faq_admin_ban_tools": "js_faq_admin_ban_tools_answer"
    }

    try:
        question_key = query.data.split("_", 3)[3]
        answer_key = faq_answers[question_key]

        question = get_translation(user_id, question_key)
        answer = get_translation(user_id, answer_key)

        keyboard = [
            [InlineKeyboardButton("🔙 Back",
                                  callback_data=f"js_faq_return_{context.user_data.get('current_js_category')}")]
        ]

        message = (
            f"❓ {question}\n\n"
            f"✅ {answer}"
        )

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return JS_FAQ_QUESTION

    except Exception as e:
        await query.edit_message_text(text=f"⚠️ Error: {str(e)}")
        return JS_FAQ_CATEGORY


async def show_admin_faq(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Admin-specific smart tips
    smart_tips = [
        get_translation(user_id, "admin_tip_database_backup"),
        get_translation(user_id, "admin_tip_post_moderation"),
        get_translation(user_id, "admin_tip_broadcast_usage"),
        get_translation(user_id, "admin_tip_ban_management")
    ]
    random_tip = random.choice(smart_tips)

    # Organized categories for admin features
    faq_categories = {
        "🛡️ Admin Panel": [
            "admin_faq_panel_overview",
            "admin_faq_menu_options"
        ],
        "📝 Job Moderation": [
            "admin_faq_job_management",
            "admin_faq_post_sharing",
            "admin_faq_invalid_posts"
        ],
        "📂 Database Management": [
            "admin_faq_database_options",
            "admin_faq_data_deletion"
        ],
        "📢 Communications": [
            "admin_faq_broadcast_feature"
        ],
        "🚫 Ban Management": [
            "admin_faq_ban_process",
            "admin_faq_appeal_review",
            "admin_faq_ban_tools"
        ]
    }

    keyboard = [
        [InlineKeyboardButton(f"📂 {category}", callback_data=f"admin_faq_cat_{index}")]
        for index, category in enumerate(faq_categories.keys())
    ]
    keyboard.append([
        InlineKeyboardButton(
            "🔙 " + get_translation(user_id, "back_to_help"),
            callback_data="admin_faq_back_to_main"  # Correct pattern
        )
    ])
    message = (
        f"🔒 {get_translation(user_id, 'admin_faq_welcome')}\n\n"
        f"{get_translation(user_id, 'admin_faq_intro')}\n\n"
        f"💡 {get_translation(user_id, 'smart_tip')}: {random_tip}\n\n"
        f"{get_translation(user_id, 'choose_faq_category')}"
    )

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    return ADMIN_FAQ_SECTION


async def handle_admin_faq_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Category index mapping for admin
    faq_categories = {
        0: "🛡️ Admin Panel",
        1: "📝 Job Moderation",
        2: "📂 Database Management",
        3: "📢 Communications",
        4: "🚫 Ban Management"
    }

    try:
        category_index = int(query.data.split("_")[3])
        context.user_data['current_admin_category'] = category_index
        selected_category = faq_categories[category_index]

        # Question mapping for each category
        category_questions = {
            "🛡️ Admin Panel": [
                "admin_faq_panel_overview",
                "admin_faq_menu_options"
            ],
            "📝 Job Moderation": [
                "admin_faq_job_management",
                "admin_faq_post_sharing",
                "admin_faq_invalid_posts"
            ],
            "📂 Database Management": [
                "admin_faq_database_options",
                "admin_faq_data_deletion"
            ],
            "📢 Communications": [
                "admin_faq_broadcast_feature"
            ],
            "🚫 Ban Management": [
                "admin_faq_ban_process",
                "admin_faq_appeal_review",
                "admin_faq_ban_tools"
            ]
        }

        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, q), callback_data=f"admin_faq_q_{q}")]
            for q in category_questions[selected_category]
        ]
        keyboard.append([InlineKeyboardButton("🔙 Back", callback_data=f"admin_faq_return_{category_index}")])
        message = (
            f"📂 {selected_category}\n\n"
            f"{get_translation(user_id, 'faq_category_intro')}\n\n"
            f"{get_translation(user_id, 'choose_faq_question')}"
        )

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return ADMIN_FAQ_CATEGORY

    except Exception as e:
        await query.edit_message_text(text=f"⚠️ Error: {str(e)}")
        return ADMIN_FAQ_SECTION


async def handle_admin_faq_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Admin FAQ answers mapping
    faq_answers = {
        # Admin Panel
        "admin_faq_panel_overview": "admin_faq_panel_overview_ans",
        "admin_faq_menu_options": "admin_faq_menu_options_ans",

        # Job Moderation
        "admin_faq_job_management": "admin_faq_job_management_ans",
        "admin_faq_post_sharing": "admin_faq_post_sharing_ans",
        "admin_faq_invalid_posts": "admin_faq_invalid_posts_ans",

        # Database Management
        "admin_faq_database_options": "admin_faq_database_options_ans",
        "admin_faq_data_deletion": "admin_faq_data_deletion_ans",

        # Communications
        "admin_faq_broadcast_feature": "admin_faq_broadcast_feature_ans",

        # Ban Management
        "admin_faq_ban_process": "admin_faq_ban_process_ans",
        "admin_faq_appeal_review": "admin_faq_appeal_review_ans",
        "admin_faq_ban_tools": "admin_faq_ban_tools_ans"
    }

    try:
        question_key = query.data.split("_", 3)[3]
        answer_key = faq_answers[question_key]

        question = get_translation(user_id, question_key)
        answer = get_translation(user_id, answer_key)

        # Get category index from user data
        category_index = context.user_data.get('current_admin_category', 0)  # Retrieve category
        keyboard = [
            [InlineKeyboardButton("🔙 Back", callback_data=f"admin_faq_cat_{category_index}")]
        ]

        message = (
            f"❓ {question}\n\n"
            f"✅ {answer}"
        )

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        return ADMIN_FAQ_QUESTION

    except Exception as e:
        await query.edit_message_text(text=f"⚠️ Error: {str(e)}")
        return ADMIN_FAQ_CATEGORY

#contact admin
async def show_contact_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)

    # Get categories from database
    categories = db.get_contact_categories()

    # Create beautiful keyboard with categories
    keyboard = []
    for category in categories:
        keyboard.append([
            InlineKeyboardButton(
                f"{category['emoji']} {get_translation(user_id, category['name_key'])}",
                callback_data=f"contact_category_{category['id']}"
            )
        ])

    # Add back button with consistent styling
    keyboard.append([
        InlineKeyboardButton(
            f"↩️ {get_translation(user_id, 'back_to_help')}",
            callback_data="contact_back"
        )
    ])

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Build message with beautiful formatting and stats
    user_stats = db.get_user_contact_stats(user_id)

    message = (
        f"<b>📩 {get_translation(user_id, 'contact_admin_title')}</b>\n\n"
        f"┌───────────────────────┐\n"
        f"│  {get_translation(user_id, 'contact_admin_intro')}  │\n"
        f"└───────────────────────┘\n\n"
    )

    if user_stats:
        message += (
            f"<b>📊 {get_translation(user_id, 'your_contact_stats')}:</b>\n"
            f"┌───────────────────────────────┐\n"
            f"│   • {get_translation(user_id, 'total_messages')}: <b>{user_stats['total']}</b>   \n"
            f"│   • {get_translation(user_id, 'pending_messages')}: <b>{user_stats['pending']}</b>    \n"
            f"│   • {get_translation(user_id, 'answered_messages')}: <b>{user_stats['answered']}</b>  \n"
            f"└───────────────────────────────┘\n\n"
        )

    message += (
        f"<b>🔍 {get_translation(user_id, 'select_contact_category')}</b>\n"
        f"▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰"
    )

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
    else:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )

    return CONTACT_CATEGORY

async def handle_contact_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    if query.data == "contact_back":
        return await show_help(update, context)

    # Extract category ID from callback data
    category_id = int(query.data.split("_")[-1])
    context.user_data['contact_category'] = category_id

    # Create priority selection with color-coded buttons
    keyboard = [
        [
            InlineKeyboardButton(
                f"🟢 {get_translation(user_id, 'priority_normal')}",
                callback_data="contact_priority_1"
            ),
            InlineKeyboardButton(
                f"🟡 {get_translation(user_id, 'priority_high')}",
                callback_data="contact_priority_2"
            ),
            InlineKeyboardButton(
                f"🔴 {get_translation(user_id, 'priority_urgent')}",
                callback_data="contact_priority_3"
            )
        ],
        [
            InlineKeyboardButton(
                f"↩️ {get_translation(user_id, 'back_to_categories')}",
                callback_data="contact_back_to_categories"
            )
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text=(
            f"<b>❗ {get_translation(user_id, 'select_priority')}</b>\n\n"
            f"┌───────────────────────────────────┐\n"
            f"│  ℹ️ <i>{get_translation(user_id, 'priority_explanation')}</i>  │\n"
            f"└───────────────────────────────────┘\n\n"
            f"<b>🟢 Normal</b> = within 24h\n"
            f"<b>🟡 High</b> = within 12h\n"
            f"<b>🔴 Urgent</b> = within 6h (for critical issues only)\n\n"
            f"▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰"
        ),
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    return CONTACT_PRIORITY

async def handle_contact_priority(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    if query.data == "contact_back_to_categories":
        return await show_contact_options(update, context)

    # Store priority in user data
    priority = int(query.data.split("_")[-1])
    context.user_data['contact_priority'] = priority

    # Create formatted message
    await query.edit_message_text(
        text=f"✨ <b>{get_translation(user_id, 'write_your_message')}</b> ✨\n\n"
             f"📌 <b>{get_translation(user_id, 'message_guidelines')}:</b>\n"
             f"   🔹 {get_translation(user_id, 'be_specific')}\n"
             f"   🔹 {get_translation(user_id, 'include_details')}\n"
             f"   🔹 {get_translation(user_id, 'avoid_spam')}\n\n"
             f"⏱ <b>{get_translation(user_id, 'response_time')}:</b>\n"
             f"   🕒 {get_translation(user_id, 'within_24_hours')}\n\n"
             f"━━━━━━━━━━━━━━━━━━━━\n"
             f"🚫 /cancel - {get_translation(user_id, 'cancel_contact')}",
        parse_mode="HTML"
    )

    return CONTACT_MESSAGE


async def handle_contact_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = get_user_id(update)
    message_text = update.message.text

    # Check for cancel command
    if message_text.lower() == "/cancel":
        await update.message.reply_text(
            f"❌ {get_translation(user_id, 'contact_cancelled')}",
            parse_mode="HTML"
        )
        return await show_help(update, context)

    # Save to database
    category_id = context.user_data.get('contact_category')
    priority = context.user_data.get('contact_priority', 1)

    message_id = db.save_contact_message(
        user_id=user_id,
        category_id=category_id,
        message_text=message_text,
        priority=priority
    )

    # Notify admins with beautiful formatting
    active_admins = get_all_admins()
    category_name = escape_html(db.get_category_name(category_id))
    safe_message_text = escape_html(message_text)
    user_profile_link = get_user_profile_link_html(user_id)

    admin_notification = (
        f"<b>📬 NEW SUPPORT REQUEST {priority_emoji(priority)}</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>    \n"
        f"┃  👤 <b>From</b>: {user_profile_link}  \n"
        f"┃  📂 <b>Category</b>: {category_name}  \n"
        f"┃  ⚠️ <b>Priority</b>: {priority_name(priority)}  \n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
        f"<b>📝 Message:</b>\n"
        f"<code>{safe_message_text[:500]}</code>\n\n"
        f"▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰"
    )

    for admin_id in active_admins:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_notification,
                parse_mode="HTML"
            )
        except Exception as e:
            logging.error(f"Failed to notify admin {admin_id}: {e}")
            # Fallback to plain text
            await context.bot.send_message(
                chat_id=admin_id,
                text=f"New Support Ticket #{message_id}\nFrom: User #{user_id}\n\n{safe_message_text[:500]}"
            )

    # Beautiful confirmation to user
    confirmation = (
        f"<b>✅ {get_translation(user_id, 'message_received')}</b>\n\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>    \n"
        f"┃  📂 <b>Category</b>: {category_name}  \n"
        f"┃  ⚠️ <b>Priority</b>: {priority_name(priority)}  \n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
        f"<i>{get_translation(user_id, 'response_time_notice')}</i>\n\n"
        f"📬 <b>{get_translation(user_id, 'contact_follow_up_info')}</b>\n\n"
        f"▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰"
    )

    await update.message.reply_text(
        confirmation,
        parse_mode="HTML"
    )

    return await show_help(update, context)


def priority_emoji(priority):
    """Get emoji for priority level with text"""
    priority_map = {
        1: "🟢 Normal",
        2: "🟡 High",
        3: "🔴 Urgent"
    }
    return priority_map.get(priority, "⚪ Unknown")


def priority_name(priority):
    """Get formatted priority name"""
    return f"<b>{priority_emoji(priority).split()[0]} {priority_emoji(priority).split()[1]}</b>"


def get_user_profile_link_html(user_id: int) -> str:
    """Creates an HTML-formatted link to the user's profile"""
    return f'<a href="tg://user?id={user_id}">👤 User #{user_id}</a>'

async def admin_reply_to_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    admin_id = get_user_id(update)
    message_id = int(query.data.split("_")[-1])

    # Get message details
    message = db.get_contact_message(message_id)
    if not message:
        await query.edit_message_text("❌ Message not found")
        return

    # Store in context
    context.user_data['admin_reply'] = {
        'message_id': message_id,
        'user_id': message['user_id'],
        'category': message['category_name'],
        'priority': message['priority'],
        'original_text': message['message_text']
    }

    # Create beautiful message header
    header = (
        f"📬 <b>ADMIN REPLY TO TICKET #{message_id}</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>      ┃\n"
        f"┃  👤 <b>User</b>: <a href='tg://user?id={message['user_id']}'>User #{message['user_id']}</a>  ┃\n"
        f"┃  📂 <b>Category</b>: {message['category_name']}  ┃\n"
        f"┃  ⚠️ <b>Priority</b>: {priority_emoji(message['priority'])}  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Format original message with quote styling
    original_msg = (
        f"📝 <b>Original Message:</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"<i>{escape_html(message['message_text'])}</i>\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Instructions for admin
    instructions = (
        f"✍️ <b>Please compose your reply below:</b>\n"
        f"• You can use Markdown formatting\n"
        f"• Include any relevant details\n"
        f"• Be polite and professional\n\n"
        f"🚫 /cancel - Cancel this reply"
    )

    await query.edit_message_text(
        text=header + original_msg + instructions,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

    return ADMIN_REPLY_STATE

async def handle_admin_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('admin_reply'):
        await update.message.reply_text(
            "❌ No active reply session found. Please start over.",
            parse_mode="HTML"
        )
        return ConversationHandler.END

    reply_text = update.message.text
    message_data = context.user_data['admin_reply']
    message_id = message_data['message_id']
    user_id = message_data['user_id']
    admin_id = get_user_id(update)

    try:
        # Format the reply to user beautifully
        user_reply = (
            f"💌 <b>{get_translation(user_id, 'admin_response_to_ticket')} #{message_id}</b>\n\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"{reply_text}\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
            
            f"🔗 Ticket #{message_id} • {message_data['category']}"
        )

        # Send to user
        await context.bot.send_message(
            chat_id=user_id,
            text=user_reply,
            parse_mode="HTML"
        )

        # Update database
        db.update_contact_message(
            message_id=message_id,
            admin_id=admin_id,
            response=reply_text,
            status='answered'
        )

        # Beautiful confirmation to admin
        confirmation = (
            f"✅ <b>Reply Successfully Sent!</b>\n\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>      ┃\n"
            f"┃  👤 <b>To User</b>: <a href='tg://user?id={user_id}'>User #{user_id}</a>  ┃\n"
            f"┃  📝 <b>Your Reply</b>: Sent  ┃\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
            f"📊 <i>This ticket has been marked as answered in the system.</i>"
        )

        await update.message.reply_text(
            confirmation,
            parse_mode="HTML"
        )

    except Exception as e:
        logging.error(f"Failed to send admin reply: {e}")
        error_msg = (
            f"❌ <b>Failed to send reply</b>\n\n"
            f"<i>Error details:</i> <code>{escape_html(str(e))}</code>\n\n"
            f"Please try again or contact technical support."
        )
        await update.message.reply_text(
            error_msg,
            parse_mode="HTML"
        )

    # Clean up
    context.user_data.pop('admin_reply', None)
    return await show_contact_management_dashboard(update, context)


async def handle_admin_reply_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Extract message ID from callback data (format: "admin_reply_123")
    try:
        message_id = int(query.data.split('_')[-1])
    except (IndexError, ValueError):
        await query.edit_message_text("❌ Invalid message reference")
        return

    # Get the original contact message from database
    contact_message = db.get_contact_message(message_id)
    if not contact_message:
        await query.edit_message_text("❌ Message not found in database")
        return

    # Store in context for the reply flow
    context.user_data['admin_reply'] = {
        'message_id': message_id,
        'user_id': contact_message['user_id'],
        'original_text': contact_message['message_text']
    }

    user_id = get_user_id(update)

    # def get_user_profile_link_html(user_id: int) -> str:
    #     """Creates an HTML-formatted link to the user's profile"""
    #     return f'<a href="tg://user?id={user_id}">User #{user_id}</a>'
    user_profile_link = get_user_profile_link_html(user_id)

    # Ask admin for their reply text
    await query.edit_message_text(
        text=f"✍️ Replying to message #{message_id}\n\n"
             f"👤 <b>User</b>: {user_profile_link}\n"
             f"📝 Original message:\n{contact_message['message_text']}\n\n"
             f"Please write your reply below:",
        parse_mode="Markdown"
    )

    return ADMIN_REPLY_STATE

async def cancel_contact_request(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel the contact request process and return to help menu"""
    user_id = get_user_id(update)

    # Clear any stored contact data
    context.user_data.pop('contact_category', None)
    context.user_data.pop('contact_priority', None)

    await update.message.reply_text(
        get_translation(user_id, 'contact_cancelled'),
        reply_markup=ReplyKeyboardRemove()
    )
    return await show_help(update, context)


async def cancel_admin_reply(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel the admin reply process with beautiful formatting"""
    admin_id = get_user_id(update)
    reply_data = context.user_data.get('admin_reply', {})

    # Create cancellation message
    if reply_data.get('message_id'):
        cancellation_msg = (
            f"❌ <b>Reply Cancelled</b>\n\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"┃  🆔 <b>Ticket #</b>: <code>{reply_data['message_id']}</code>  ┃\n"
            f"┃  👤 <b>User</b>: <a href='tg://user?id={reply_data['user_id']}'>User #{reply_data['user_id']}</a>  ┃\n"
            f"┃  📝 <b>Status</b>: Not replied  ┃\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
            f"<i>No reply was sent to the user.</i>"
        )
    else:
        cancellation_msg = (
            f"❌ <b>Reply Process Cancelled</b>\n\n"
            f"<i>No active reply session was found.</i>"
        )

    # Clear admin reply data
    context.user_data.pop('admin_reply', None)

    await update.message.reply_text(
        cancellation_msg,
        parse_mode="HTML",
        reply_markup=ReplyKeyboardRemove()
    )

    return ConversationHandler.END


async def show_contact_management_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    # Get statistics for dashboard with safe defaults
    stats = db.get_contact_stats() or {
        'total': 0,
        'pending': 0,
        'answered': 0,
        'avg_response_time': 0
    }

    # Format statistics with fallback values
    total = stats.get('total', 0)
    pending = stats.get('pending', 0)
    answered = stats.get('answered', 0)
    avg_response = stats.get('avg_response_time', 0)

    # Create interactive buttons
    buttons = [
        [
            InlineKeyboardButton("📥 Inbox (All Messages)", callback_data="contact_inbox"),
            InlineKeyboardButton("🔄 Pending Replies", callback_data="contact_pending")
        ],
        [
            InlineKeyboardButton("✅ Answered Tickets", callback_data="contact_answered"),
            InlineKeyboardButton("📤 My Sent Replies", callback_data="contact_outbox")
        ],
        [
            InlineKeyboardButton("📈 Performance Stats", callback_data="contact_stats"),
            InlineKeyboardButton("🏠 Main Menu", callback_data="contact_back_to_menu")
        ],
    ]

    reply_markup = InlineKeyboardMarkup(buttons)

    # Create dashboard message with safe formatting
    dashboard_message = (
        f"📨 <b>CONTACT MANAGEMENT DASHBOARD</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  📊 <b>Message Statistics</b>          \n"
        f"┣━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┫\n"
        f"┃  📥 <i>Total Messages</i>: {total or 0:>5}      \n"
        f"┃  🔄 <i>Pending</i>: {pending or 0:>12}      \n"
        f"┃  ✅ <i>Answered</i>: {answered or 0:>10}      \n"
        f"┃  ⏳ <i>Avg Response</i>: {avg_response or 0:>3}h      \n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
        f"📌 <i>Select an option below to manage messages:</i>"
    )

    try:
        if update.callback_query:
            await update.callback_query.answer()
            await update.callback_query.edit_message_text(
                text=dashboard_message,
                reply_markup=reply_markup,
                parse_mode="HTML"
            )
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=dashboard_message,
                reply_markup=reply_markup,
                parse_mode="HTML"
            )
    except Exception as e:
        logging.error(f"Error showing contact dashboard: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Could not load contact dashboard. Please try again."
        )
        return USER_INTERACTIONS_MENU

    return CONTACT_MANAGEMENT

async def show_contact_inbox(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    # Get paginated messages (10 per page)
    page = context.user_data.get('contact_page', 1)
    messages = db.get_paginated_messages(status='all', page=page)
    total_messages = db.get_message_count(status='all')
    total_pages = (total_messages + 9) // 10

    # Build beautiful inbox header
    message_text = (
        f"📥 <b>MESSAGE INBOX</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  Showing {min(10, len(messages))} of {total_messages} total messages  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    buttons = []

    for msg in messages:
        # Enhanced message display information
        priority_icon = {1: "🟢", 2: "🟡", 3: "🔴"}.get(msg.get('priority', 1), "⚪")
        status_icon = "🔄" if msg.get('status') == 'pending' else "✅"

        # Ensure 'created_at' is a datetime object
        try:
            created_at = datetime.strptime(msg['created_at'], '%Y-%m-%d %H:%M:%S') if 'created_at' in msg else None
        except ValueError:
            created_at = None  # Fallback if the date format is invalid

        # Calculate days_old only if created_at is valid
        days_old = (datetime.now() - created_at).days if created_at else 0
        freshness = "🆕" if days_old == 0 else f"{days_old}d"

        # Create button with condensed but informative text
        buttons.append([
            InlineKeyboardButton(
                f"{priority_icon} #{msg['id']} • {freshness} • {msg.get('category', 'Unknown')[:10]} • {status_icon}",
                callback_data=f"contact_view_{msg['id']}"
            )
        ])
        # Add pagination controls with view-specific prefix
    buttons = add_pagination_buttons(buttons, page, total_pages, view_type="inbox")



    # Add action buttons row
    buttons.extend([

        [
            InlineKeyboardButton("🏠 Dashboard", callback_data="contact_back_to_dashboard")
        ]
    ])

    reply_markup = InlineKeyboardMarkup(buttons)

    await query.edit_message_text(
        text=message_text + f"📌 <i>Select a message to view details:</i>",
        reply_markup=reply_markup,
        parse_mode="HTML"
    )

    # Store current page and view in context
    context.user_data['contact_page'] = page
    context.user_data['current_contact_view'] = 'inbox'

    return CONTACT_INBOX


async def view_contact_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    message_id = int(query.data.split('_')[-1])
    message = db.get_contact_message_details(message_id)
    user_id = get_user_id(update)

    # Create beautiful header with ASCII art
    header = (
        f"📄 <b>TICKET DETAILS #{message_id}</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🏷️ <b>Status</b>: {status_with_icon(message['status'])}  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # User information section
    user_section = (
        f"👤 <b>User Information</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🔗 <a href='tg://user?id={message['user_id']}'>User #{message['user_id']}</a>  ┃\n"
        f"┃  📅 <b>Received</b>: {format_datetime(message['created_at'])}  ┃\n"
        f"┃  🏷️ <b>Priority</b>: {priority_with_icon(message.get('priority', 1))}  ┃\n"
        f"┃  📂 <b>Category</b>: {message['category']}  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Original message section
    original_message = (
        f"📝 <b>Original Message</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"{escape_html(message['text'])}\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Response section (if answered)
    response_section = ""
    if message['status'] == 'answered':
        response_section = (
            f"📩 <b>Admin Response</b>\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"┃  👨‍💼 <b>By</b>: {message['admin']}  ┃\n"
            f"┃  📅 <b>At</b>: {format_datetime(message['answered_at'])}  ┃\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"{escape_html(message['response'])}\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
        )

    # Prepare action buttons with better visual hierarchy
    buttons = []
    if message['status'] == 'pending':
        buttons.append([
            InlineKeyboardButton("✍️ Reply", callback_data=f"contact_reply_{message_id}"),
            InlineKeyboardButton("✅ Mark Resolved", callback_data=f"contact_close_{message_id}")
        ])
    else:
        buttons.append([
            InlineKeyboardButton("📝 Follow Up", callback_data=f"contact_followup_{message_id}")

        ])



    buttons.append([
        InlineKeyboardButton("🏠 Dashboard", callback_data="contact_back_to_dashboard")
    ])

    full_message = header + user_section + original_message + response_section
    await query.edit_message_text(
        text=full_message,
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML",
        disable_web_page_preview=True
    )

    return CONTACT_VIEW_MESSAGE


from datetime import datetime

from datetime import datetime

async def show_contact_outbox(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    page = context.user_data.get('contact_page', 1)
    messages = db.get_paginated_messages(status='answered', page=page)
    total_messages = db.get_message_count(status='answered')
    total_pages = (total_messages + 9) // 10

    # Create beautiful outbox header
    header = (
        f"📤 <b>OUTBOX - ANSWERED MESSAGES</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  Showing {min(10, len(messages))} of {total_messages} answered  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    buttons = []
    for msg in messages:
        # Convert created_at to datetime if it's a string
        created_at = msg.get('created_at')
        if isinstance(created_at, str):
            try:
                created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')  # Adjust format if needed
            except ValueError:
                created_at = None  # Fallback if parsing fails

        # Calculate days_old only if created_at is valid
        days_old = (datetime.now() - created_at).days if created_at else 0
        freshness = "🆕" if days_old == 0 else f"{days_old}d"

        # Ensure created_at is a datetime object before passing to format_date
        formatted_date = format_date(created_at)

        buttons.append([
            InlineKeyboardButton(
                f"{priority_with_icon(msg.get('priority', 1))} #{msg['id']} • {freshness} • {msg.get('category', 'Unknown')[:12]} • {formatted_date}",
                callback_data=f"contact_view_{msg['id']}"
            )
        ])

    buttons = add_pagination_buttons(buttons, page, total_pages, view_type="outbox")

    # Add action buttons
    buttons.extend([
        [
            InlineKeyboardButton("🏠 Dashboard", callback_data="contact_back_to_dashboard")
        ]
    ])

    await query.edit_message_text(
        text=header + "📌 Select a message to view details:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

    # Store current page in context
    context.user_data['contact_page'] = page
    context.user_data['current_contact_view'] = 'outbox'

    return CONTACT_OUTBOX
# Helper functions
def status_with_icon(status):
    icons = {
        'pending': '🔄 Pending',
        'answered': '✅ Answered',
        'archived': '🗄 Archived',
        'deleted': '🗑️ Deleted'
    }
    return icons.get(status, f"⚪ {status.capitalize()}")


def priority_with_icon(priority):
    icons = {
        1: '🟢 Normal',
        2: '🟡 High',
        3: '🔴 Urgent'
    }
    return icons.get(priority, '⚪ Unknown')


from datetime import datetime


def format_datetime(dt):
    if not dt:
        return "Unknown date"

    # If dt is a string, parse it into a datetime object
    if isinstance(dt, str):
        try:
            dt = datetime.strptime(dt, '%Y-%m-%d %H:%M:%S')  # Adjust the format if necessary
        except ValueError:
            return "Invalid date format"

    # Format the datetime object
    return dt.strftime("%b %d, %Y %H:%M")


def format_date(dt):
    if not dt:
        return "Unknown date"
    return dt.strftime("%b %d")


def add_pagination_buttons(buttons, current_page, total_pages=None, view_type=None):
    """Add beautiful pagination controls to button list with view-specific prefix"""
    pagination_row = []
    prefix = f"contact_{view_type}_" if view_type else "contact_page_"

    # First page button
    if current_page > 2:
        pagination_row.append(
            InlineKeyboardButton("⏮️ First", callback_data=f"{prefix}1")
        )

    # Previous page button
    if current_page > 1:
        pagination_row.append(
            InlineKeyboardButton("◀️ Prev", callback_data=f"{prefix}{current_page - 1}")
        )

    # Current page indicator
    page_display = f"📄 {current_page}"
    if total_pages:
        page_display += f"/{total_pages}"
    pagination_row.append(
        InlineKeyboardButton(page_display, callback_data="current_page")
    )

    # Next page button
    if not total_pages or current_page < total_pages:
        pagination_row.append(
            InlineKeyboardButton("Next ▶️", callback_data=f"{prefix}{current_page + 1}")
        )

    # Last page button (if total pages known)
    if total_pages and current_page < total_pages - 1:
        pagination_row.append(
            InlineKeyboardButton("Last ⏭️", callback_data=f"{prefix}{total_pages}")
        )

    if pagination_row:
        buttons.append(pagination_row)
    return buttons

# def format_date(timestamp):
#     """Format timestamp for display"""
#     if not timestamp:
#         return "N/A"
#     return datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S").strftime("%b %d, %H:%M")


from datetime import datetime


async def show_contact_pending(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    page = context.user_data.get('contact_page', 1)
    messages = db.get_paginated_messages(status='pending', page=page)
    total_messages = db.get_message_count(status='pending')
    total_pages = (total_messages + 9) // 10  # Calculate total pages

    # Create beautiful header with ASCII art
    header = (
        f"🔄 <b>PENDING MESSAGES</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  Showing {min(10, len(messages))} of {total_messages} pending  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    buttons = []
    for msg in messages:
        # Convert created_at to datetime if it's a string
        created_at = msg.get('created_at')
        if isinstance(created_at, str):
            try:
                created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')  # Adjust the format if necessary
            except ValueError:
                created_at = datetime.now()  # Fallback to current time if parsing fails

        # Calculate how old the message is
        days_old = (datetime.now() - created_at).days
        freshness = "🆕 New" if days_old == 0 else f"⏳ {days_old}d"

        buttons.append([
            InlineKeyboardButton(
                f"{priority_with_icon(msg.get('priority', 1))} #{msg.get('id', '?')} • {msg.get('category', 'Unknown')[:12]} • {freshness}",
                callback_data=f"contact_view_{msg.get('id', '')}"
            )
        ])

    # Add enhanced pagination controls
    buttons = add_pagination_buttons(buttons, page, total_pages, view_type="pending")

    # Add action buttons
    buttons.extend([

        [
            InlineKeyboardButton("🏠 Dashboard", callback_data="contact_back_to_dashboard")

        ]
    ])

    await query.edit_message_text(
        text=header + "📌 Select a message to respond:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

    # Store current page in context
    context.user_data['contact_page'] = page
    context.user_data['current_contact_view'] = 'pending'
    return CONTACT_PENDING

async def show_contact_answered(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    page = context.user_data.get('contact_page', 1)
    messages = db.get_paginated_messages(status='answered', page=page)
    total_messages = db.get_message_count(status='answered')
    total_pages = (total_messages + 9) // 10  # Calculate total pages

    # Create beautiful header with ASCII art
    header = (
        f"✅ <b>ANSWERED MESSAGES</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  Showing {min(10, len(messages))} of {total_messages} answered  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    buttons = []
    for msg in messages:
        # Calculate response time in hours
        response_time = ""
        if msg.get('created_at') and msg.get('answered_at'):
            delta = msg['answered_at'] - msg['created_at']
            response_time = f"⏱ {delta.total_seconds() // 3600}h"

        buttons.append([
            InlineKeyboardButton(
                f"{priority_with_icon(msg.get('priority', 1))} #{msg.get('id', '?')} • {msg.get('category', 'Unknown')[:10]} • {response_time}",
                callback_data=f"contact_view_{msg.get('id', '')}"
            )
        ])

    # Add enhanced pagination controls
    buttons = add_pagination_buttons(buttons, page, total_pages, view_type="answered")

    # Add action buttons
    buttons.extend([

        [
            InlineKeyboardButton("🏠 Dashboard", callback_data="contact_back_to_dashboard")

        ]
    ])

    await query.edit_message_text(
        text=header + "📌 Select a message to review:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

    # Store current page in context
    context.user_data['contact_page'] = page
    context.user_data['current_contact_view'] = 'answered'
    return CONTACT_ANSWERED


async def show_contact_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    # Get statistics with safe fallback defaults
    stats = db.get_contact_stats() or {
        'total': 0,
        'pending': 0,
        'answered': 0,
        'avg_response_time': 0
    }
    category_stats = db.get_category_stats() or []
    performance_stats = db.get_performance_stats() or {}

    # Extract fields safely
    total = stats.get('total', 0)
    pending = stats.get('pending', 0)
    answered = stats.get('answered', 0)
    avg_response = stats.get('avg_response_time', 0)

    # Create beautiful stats dashboard
    stats_message = (
        f"📈 <b>CONTACT PERFORMANCE DASHBOARD</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  📥 <b>Total Messages</b>: {total or 0:>6}      \n"
        f"┃  🔄 <b>Pending</b>: {pending or 0:>12}      \n"
        f"┃  ✅ <b>Answered</b>: {answered or 0:>10}      \n"
        f"┃  ⏱ <b>Avg Response</b>: {avg_response or 0:>5}h      \n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
        f"📋 <b>By Category</b>\n"
    )

    for cat in category_stats:
        emoji = cat.get('emoji', '📁')
        name = cat.get('name', 'Unknown')
        count = cat.get('count', 0)
        percentage = cat.get('percentage', 0)

        stats_message += (
            f"  {emoji} <b>{name}</b>: {count} ({percentage}%)\n"
        )

    # Add only the dashboard button
    buttons = [
        [InlineKeyboardButton("🏠 Dashboard", callback_data="contact_back_to_dashboard")]
    ]

    try:
        await query.edit_message_text(
            text=stats_message,
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Error showing contact stats: {e}")
        await context.bot.send_message(
            chat_id=update.effective_user.id,
            text="⚠️ Could not load contact statistics. Please try again."
        )
        return CONTACT_MANAGEMENT

    return CONTACT_STATS




def format_timedelta(td):
    """Format timedelta into human-readable string"""
    if not td:
        return "N/A"
    seconds = td.total_seconds()
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    return f"{hours}h {minutes}m"


async def start_admin_reply(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    message_id = int(query.data.split('_')[-1])
    message = db.get_contact_message_details(message_id)

    # Store reply context with additional details
    context.user_data['admin_reply'] = {
        'message_id': message_id,
        'user_id': message['user_id'],
        'category': message['category'],
        'priority': message.get('priority', 1),
        'original_text': message['text']
    }

    # Create beautiful message header
    header = (
        f"✍️ <b>ADMIN REPLY TO TICKET #{message_id}</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🏷️ <b>Status</b>: 🔄 Pending Reply  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # User information section
    user_section = (
        f"👤 <b>User</b>: <a href='tg://user?id={message['user_id']}'>User #{message['user_id']}</a>\n"
        f"📂 <b>Category</b>: {message['category']}\n"
        f"⚠️ <b>Priority</b>: {priority_with_icon(message.get('priority', 1))}\n\n"
    )

    # Original message section
    original_message = (
        f"📝 <b>Original Message</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"{escape_html(message['text'])}\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Instructions section
    instructions = (
        f"📌 <b>Please compose your reply below:</b>\n"
        f"• You can use Markdown formatting\n"
        f"• Include any relevant details\n"
        f"• Be professional and courteous"

    )

    await query.edit_message_text(
        text=header + user_section + original_message + instructions,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

    return ADMIN_REPLY_STATE


async def close_ticket(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    message_id = int(query.data.split('_')[-1])
    admin_id = get_user_id(update)

    # Update database
    success = db.update_contact_message(
        message_id=message_id,
        admin_id=admin_id,
        status='answered',
        response="Closed without reply"
    )

    if success:
        # Create beautiful confirmation message
        confirmation = (
            f"✅ <b>TICKET CLOSED</b>\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>      ┃\n"
            f"┃  👤 <b>Closed by</b>: <a href='tg://user?id={admin_id}'>Admin #{admin_id}</a>  ┃\n"
            f"┃  🏷️ <b>Status</b>: ✅ Answered      ┃\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
            f"<i>This ticket was closed without sending a reply.</i>"
        )
    else:
        confirmation = (
            f"❌ <b>FAILED TO CLOSE TICKET</b>\n\n"
            f"<i>Could not update ticket #{message_id} in the database.</i>"
        )

    await query.edit_message_text(
        text=confirmation,
        parse_mode="HTML"
    )

    return await show_contact_inbox(update, context)

async def delete_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    message_id = int(query.data.split('_')[-1])
    message = db.get_contact_message_details(message_id)

    # Store delete context with additional details
    context.user_data['delete_confirmation'] = {
        'message_id': message_id,
        'user_id': message['user_id'],
        'category': message['category'],
        'message_text': message['text'],
        'status': message['status']
    }

    # Create beautiful confirmation dialog
    confirmation = (
        f"❗ <b>CONFIRM MESSAGE DELETION</b> ❗\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>      ┃\n"
        f"┃  👤 <b>User</b>: <a href='tg://user?id={message['user_id']}'>User #{message['user_id']}</a>  ┃\n"
        f"┃  📂 <b>Category</b>: {message['category']}  ┃\n"
        f"┃  🏷️ <b>Status</b>: {status_with_icon(message['status'])}  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
        f"📝 <b>Message Preview:</b>\n"
        f"<i>{escape_html(message['text'][:200])}...</i>\n\n"
        f"⚠️ <b>This action cannot be undone!</b>"
    )

    buttons = [
        [
            InlineKeyboardButton("🗑️ CONFIRM DELETE", callback_data=f"contact_confirm_delete_{message_id}"),
            InlineKeyboardButton("🔙 CANCEL", callback_data=f"contact_view_{message_id}")
        ]
    ]

    await query.edit_message_text(
        text=confirmation,
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

    return CONTACT_CONFIRM_DELETE



async def follow_up_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    message_id = int(query.data.split('_')[-1])
    message = db.get_contact_message_details(message_id)

    # Store follow-up context with additional details
    context.user_data['admin_reply'] = {
        'message_id': message_id,
        'user_id': message['user_id'],
        'category': message['category'],
        'priority': message.get('priority', 1),
        'original_text': message['text'],
        'previous_response': message['response'],
        'is_followup': True
    }

    # Create beautiful header
    header = (
        f"✍️ <b>FOLLOW-UP ON TICKET #{message_id}</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┃  🏷️ <b>Status</b>: 🔄 Follow-up  ┃\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # User information section
    user_section = (
        f"👤 <b>User</b>: {get_user_profile_link_html(message['user_id'])}\n"
        f"📂 <b>Category</b>: {message['category']}\n"
        f"⚠️ <b>Priority</b>: {priority_with_icon(message.get('priority', 1))}\n\n"
    )

    # Original message section
    original_message = (
        f"📝 <b>Original Message</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"{escape_html(message['text'])}\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Previous response section
    previous_response = (
        f"💬 <b>Previous Response</b>\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"{escape_html(message['response'])}\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
    )

    # Instructions section
    instructions = (
        f"📌 <b>Please compose your follow-up message:</b>\n"
        f"• Continue the conversation naturally\n"
        f"• Reference previous messages as needed\n"
        f"• Keep it professional and helpful\n\n"

    )

    await query.edit_message_text(
        text=header + user_section + original_message + previous_response + instructions,
        parse_mode="HTML",
        disable_web_page_preview=True
    )

    return ADMIN_REPLY_STATE

# Helper functions (would be defined elsewhere in your code)
# def get_user_profile_link_html(user_id: int) -> str:
#     """Creates an HTML-formatted link to the user's profile"""
#     return f'<a href="tg://user?id={user_id}">User #{user_id}</a>'

async def confirm_delete_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    message_id = int(query.data.split('_')[-1])
    delete_data = context.user_data.get('delete_confirmation', {})

    if delete_data.get('message_id') != message_id:
        await query.edit_message_text("❌ Invalid delete confirmation")
        return await show_contact_inbox(update, context)

    # Actually delete from database
    success = db.delete_contact_message(message_id)

    if success:
        # Create beautiful success message
        result = (
            f"🗑️ <b>MESSAGE DELETED</b>\n"
            f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓\n"
            f"┃  🆔 <b>Ticket #</b>: <code>{message_id}</code>      ┃\n"
            f"┃  📂 <b>Category</b>: {delete_data['category']}  ┃\n"
            f"┃  🏷️ <b>Status</b>: {status_with_icon(delete_data['status'])}  ┃\n"
            f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n\n"
            f"<i>This message has been permanently deleted from the system.</i>"
        )
    else:
        result = (
            f"❌ <b>DELETE FAILED</b>\n\n"
            f"<i>Could not delete message #{message_id} from the database.</i>"
        )

    await query.edit_message_text(
        text=result,
        parse_mode="HTML"
    )

    # Clear delete context
    context.user_data.pop('delete_confirmation', None)
    return await show_contact_inbox(update, context)


async def handle_pagination_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle pagination for all contact views"""
    query = update.callback_query
    await query.answer()

    # Extract data and update page number
    data = query.data
    page = int(data.split('_')[-1])
    context.user_data['contact_page'] = page

    # Route to correct view based on callback prefix
    if data.startswith('contact_inbox_'):
        return await show_contact_inbox(update, context)
    elif data.startswith('contact_pending_'):
        return await show_contact_pending(update, context)
    elif data.startswith('contact_answered_'):
        return await show_contact_answered(update, context)
    elif data.startswith('contact_outbox_'):
        return await show_contact_outbox(update, context)

    # Fallback to inbox if unknown
    return await show_contact_inbox(update, context)# Rating

# Rate Feature
from datetime import date
today = date.today().isoformat()


async def show_rate_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show enhanced rating menu with visual improvements"""
    user_id = get_user_id(update)

    # Check rating capabilities with progress indicators
    rating_stats = db.get_user_rating_stats(user_id)
    can_rate_bot = True
    can_rate_users = db.has_any_application(user_id)

    # Build dynamic menu with visual indicators
    menu_text = f"🌟 <b>{get_translation(user_id, 'rating_center')}</b> 🌟\n\n"
    menu_text += f"{get_translation(user_id, 'your_rating_stats')}:\n"
    menu_text += f"• {get_translation(user_id, 'reviews_given_by_you')}: {rating_stats['total_reviews']}\n"
    menu_text += f"• {get_translation(user_id, 'average_rating_given')}: {rating_stats['average_rating']:.1f}⭐\n\n"
    menu_text += get_translation(user_id, 'choose_an_option')

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'rate_our_bot'), callback_data="rate_bot")] if can_rate_bot else [],
        [InlineKeyboardButton(get_translation(user_id, 'rate_users'), callback_data="rate_user")],
        [InlineKeyboardButton(get_translation(user_id, 'my_review_history'), callback_data="my_reviews")],
        [InlineKeyboardButton(get_translation(user_id, 'explore_reviews'), callback_data="search_reviews")],
        [InlineKeyboardButton(get_translation(user_id, 'review_privacy_settings'), callback_data="review_settings")],
        [InlineKeyboardButton(get_translation(user_id, 'back_to_main_men'), callback_data="back_to_main")]
    ]

    # Remove any empty rows
    keyboard = [row for row in keyboard if row]

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=menu_text,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            text=menu_text,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )

    return RATE_OPTIONS


async def start_rate_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle rating initiation with alert-style duplicate checks"""
    query = update.callback_query
    await query.answer()

    try:
        user_id = get_user_id(update)

        # Get rateable users first
        rateable_users = db.get_rateable_users(user_id)
        context.user_data["rateable_users"] = rateable_users

        if not rateable_users:
            # Check if message content changes
            new_text = get_translation(user_id, 'need_interact_before_rating')
            new_markup = InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_rate_menu")]
            ])
            if query.message.text != new_text or query.message.reply_markup != new_markup:
                await query.edit_message_text(new_text, reply_markup=new_markup)
            return RATE_OPTIONS

        # Filter out already-rated users today
        rateable_users = [
            user for user in rateable_users
            if not db.has_user_reviewed(user_id, user['id'], user['type'])
        ]

        if not rateable_users:
            await query.edit_message_text(
                get_translation(user_id, 'already_rated_all_users'),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(get_translation(user_id, 'back_to_rating_menu'), callback_data="back_to_rate_menu")]
                ])
            )
            return RATE_OPTIONS

        # Store rateable users in context
        context.user_data["rateable_users"] = rateable_users

        # Show user selection interface
        keyboard = []
        for user in rateable_users:
            button_text = f"{user['name']} ({user['type'].replace('_', ' ')})"
            callback_data = f"select_user_{user['id']}"
            keyboard.append([InlineKeyboardButton(button_text, callback_data=callback_data)])

        keyboard.append([InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_rate_menu")])

        await query.edit_message_text(
            get_translation(user_id, 'select_user_to_rate'),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return SELECT_USER_FOR_RATING

    except Exception as e:
        logging.error(f"Rating initiation error: {str(e)}")
        await query.answer(
            get_translation(user_id, 'failed_to_start_rating'),
            show_alert=True
        )
        return RATE_OPTIONS


async def handle_user_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle when a user is selected for rating"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    try:
        # Parse callback data (format: select_user_123)
        parts = query.data.split('_')
        if len(parts) != 3 or parts[0] != 'select' or parts[1] != 'user':
            raise ValueError("Invalid selection format")

        target_id = int(parts[2])

        # Determine user type (employer or job_seeker)
        target_type = db.get_user_type(target_id)
        if not target_type:
            raise ValueError("Could not determine user type")

        # Check if already rated today
        if db.has_user_reviewed(user_id, target_id, target_type):
            await query.edit_message_text(
                get_translation(user_id, 'already_reviewed_this_user'),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(
                        get_translation(user_id, 'back_to_selection'),
                        callback_data="back_to_rate_menu"
                    )]
                ])
            )
            return RATE_OPTIONS

        # Set up rating dimensions based on type
        rating_dimensions = {
            'employer': {
                'professionalism': get_translation(user_id, 'professionalism'),
                'communication': get_translation(user_id, 'communication'),
                'hiring_process': get_translation(user_id, 'hiring_process')
            },
            'job_seeker': {
                'reliability': get_translation(user_id, 'reliability'),
                'skills': get_translation(user_id, 'skills'),
                'communication': get_translation(user_id, 'communication')
            }
        }

        # Store in context
        context.user_data.update({
            'review_target': target_id,
            'target_type': target_type,
            'rating_dimensions': rating_dimensions[target_type],
            'target_name': db.get_user_name(target_id)
        })

        # Start with first dimension
        first_dim = next(iter(rating_dimensions[target_type]))
        await show_dimension_rating(update, context, first_dim)
        return RATE_DIMENSION

    except ValueError as e:
        logging.error(f"User selection format error: {str(e)}")
        await query.edit_message_text(
            get_translation(user_id, 'invalid_selection_error'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    get_translation(user_id, 'back_to_selection'),
                    callback_data="back_to_rate_menu"
                )]
            ])
        )
    except Exception as e:
        logging.error(f"User selection error: {str(e)}")
        await query.edit_message_text(
            get_translation(user_id, 'selection_error_generic'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    get_translation(user_id, 'back_to_menu'),
                    callback_data="back_to_rate_menu"
                )]
            ])
        )

    return RATE_OPTIONS

async def show_user_search_interface(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show advanced user search interface"""
    query = update.callback_query
    user_id = get_user_id(update)

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'search_by_name'), callback_data="search_by_name")],
        [InlineKeyboardButton(get_translation(user_id, 'employers_only'), callback_data="filter_reviews_employer")],
        [InlineKeyboardButton(get_translation(user_id, 'job_seekers_only'), callback_data="filter_reviews_job_seeker")],
        [InlineKeyboardButton(get_translation(user_id, 'top_rated'), callback_data="sort_top_rated")],  # Updated
        [InlineKeyboardButton(get_translation(user_id, 'most_recent'), callback_data="sort_recent")],  # Updated
        [InlineKeyboardButton(get_translation(user_id, 'back_to_selection'), callback_data="back_to_rate_menu")]
    ]

    await query.edit_message_text(
        text=f"{get_translation(user_id, 'search_user_to_rate_prompt')}\n\n"
             f"• {get_translation(user_id, 'rate_employers_you_applied_to')}\n"
             f"• {get_translation(user_id, 'rate_job_seekers_you_interacted_with')}",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def show_rating_interface(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show the actual rating interface with multi-dimensional options"""
    query = update.callback_query
    target_id = int(query.data.split("_")[-1])
    target_type = context.user_data["target_type"]

    # Store in context
    context.user_data["review_target"] = target_id
    context.user_data["target_type"] = target_type

    # Build rating dimensions based on target type
    if target_type == "bot":
        dimensions = {
            "ease_of_use": "Ease of Use",
            "features": "Feature Completeness",
            "support": "Support Quality"
        }
    elif target_type == "employer":
        dimensions = {
            "professionalism": "Professionalism",
            "communication": "Communication",
            "hiring_process": "Hiring Process"
        }
    else:  # job_seeker
        dimensions = {
            "reliability": "Reliability",
            "skills": "Skill Match",
            "communication": "Communication"
        }

    context.user_data["rating_dimensions"] = dimensions

    # Show first rating dimension
    first_dim = next(iter(dimensions))
    await show_dimension_rating(update, context, first_dim)
    return RATE_DIMENSION


async def show_dimension_rating(update: Update, context: ContextTypes.DEFAULT_TYPE, dimension: str):
    """Show rating interface that works with both string and dict dimension info"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    try:
        dimensions = context.user_data["rating_dimensions"]
        dim_info = dimensions[dimension]

        # Handle both string and dictionary formats for dim_info
        if isinstance(dim_info, str):
            # Default icons based on dimension type
            icon_map = {
                'ease_of_use': '🤖',
                'features': '💡',
                'support': '🏆',
                'professionalism': '👔',
                'communication': '💬',
                'hiring_process': '📝',
                'reliability': '⏱️',
                'skills': '🛠️'
            }
            display_text = dim_info
            icon = icon_map.get(dimension, '⭐')  # Default to star if not found
        else:
            # Already in dictionary format
            display_text = dim_info.get('display', dim_info.get('text', dimension))
            icon = dim_info.get('icon', '⭐')

        # Create responsive rating buttons
        rating_buttons = [
            [
                InlineKeyboardButton(
                    text=f"{icon} {'⭐' * i}",
                    callback_data=f"rate_{dimension}_{i}"
                ) for i in range(1, 4)  # First row: 1-3 stars
            ],
            [
                InlineKeyboardButton(
                    text=f"{icon} {'⭐' * i}",
                    callback_data=f"rate_{dimension}_{i}"
                ) for i in range(4, 6)  # Second row: 4-5 stars
            ],
            [
                InlineKeyboardButton(
                    text=f"⏩ {get_translation(user_id, 'skip')}",
                    callback_data=f"skip_{dimension}"
                )
            ]
        ]

        await query.edit_message_text(
            text=f"<b>{display_text}</b>\n\n"
                 f"Rate from 1 (Lowest) to 5 (Highest):",
            reply_markup=InlineKeyboardMarkup(rating_buttons),
            parse_mode="HTML"
        )

    except Exception as e:
        logging.error(f"Error showing rating for {dimension}: {str(e)}")
        await query.edit_message_text(
            "⚠️ Couldn't load rating options. Please try again.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Back", callback_data="back_to_rate_menu")]
            ])
        )
        return RATE_OPTIONS

async def handle_rating_submission(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Finalize review submission with all checks"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    try:
        # Verify required data exists
        if not all(key in context.user_data for key in ["review_target", "target_type"]):
            raise KeyError("Missing review data in context")

        # Anti-abuse checks
        if not db.can_user_review(
                user_id,
                context.user_data["review_target"],
                context.user_data["target_type"]
        ):
            await query.answer(get_translation(user_id, 'daily_review_limit_reached'), show_alert=True)
            return await show_rate_options(update, context)

        # Compile all ratings
        ratings = context.user_data.get("dimension_ratings", {})
        if not ratings:
            await query.answer(get_translation(user_id, 'complete_at_least_one_dimension'), show_alert=True)
            return RATE_DIMENSION

        overall = sum(ratings.values()) / len(ratings)
        context.user_data["overall_rating"] = overall

        # Build confirmation message
        confirmation_text = f"{get_translation(user_id, 'review_summary_title')}\n\n"
        confirmation_text += "\n".join(
            f"{get_translation(user_id, dim + '_dimension').capitalize()}: {'⭐' * rating}"
            for dim, rating in ratings.items()
        )
        confirmation_text += f"\n\n{get_translation(user_id, 'overall_rating')}: {overall:.1f}⭐"

        await query.edit_message_text(
            text=confirmation_text,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'confirm'), callback_data="confirm_review")],
                [InlineKeyboardButton(get_translation(user_id, 'edit_review'), callback_data="edit_review")],
                [InlineKeyboardButton(get_translation(user_id, 'bak'), callback_data="back_to_rate_menu")]
            ])
        )
        return CONFIRM_REVIEW

    except Exception as e:
        logging.error(f"Error in rating submission: {str(e)}")
        await query.edit_message_text(
            get_translation(user_id, 'error_processing_review'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'try_again'), callback_data="rate_menu")]
            ])
        )
        return RATE_OPTIONS


async def finalize_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Finalize review with complete context validation"""
    try:
        user_id = get_user_id(update)
        # Validate all required context data
        required_keys = ["review_target", "target_type", "overall_rating", "dimension_ratings"]
        if not all(key in context.user_data for key in required_keys):
            missing = [k for k in required_keys if k not in context.user_data]
            raise ValueError(f"Missing review data in context: {missing}")

        # Prepare review data
        review_data = {
            "reviewer_id": user_id,
            "target_id": context.user_data["review_target"],
            "target_type": context.user_data["target_type"],
            "rating": round(context.user_data["overall_rating"]),
            "comment": context.user_data.get("review_comment", ""),
            "dimension_ratings": context.user_data["dimension_ratings"]
        }

        # Check if this is an edit
        if "editing_review" in context.user_data:
            review_id = context.user_data["editing_review"]
            if not db.update_review(review_id, review_data):
                raise ValueError("Failed to update review in database")
            success_text = get_translation(user_id, 'review_updated_success')
        else:
            if not db.add_review(**review_data):
                raise ValueError("Failed to save review to database")
            success_text = get_translation(user_id, 'review_submitted_success')

        # Success message
        target_name = "the bot" if review_data["target_type"] == "bot" else db.get_user_name(review_data["target_id"])
        text = (
            f"{success_text}\n\n"
            f"{get_translation(user_id, 'review_of')} {target_name}\n"
            f"{get_translation(user_id, 'rating')}: {'⭐' * review_data['rating']}\n"
        )

        # Add dimension ratings if available
        if review_data["dimension_ratings"]:
            text += f"\n{get_translation(user_id, 'detailed_ratings')}:\n"
            for dim, rating in review_data["dimension_ratings"].items():
                if rating > 0:  # Only show rated dimensions
                    dim_name = get_translation(user_id, dim + '_dimension', dim=dim.replace('_', ' ').title())
                    text += f"{dim_name}: {'⭐' * rating}\n"

        if review_data["comment"]:
            text += f"\n{get_translation(user_id, 'your_comment')}: {review_data['comment']}"

        # Navigation buttons
        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, 'view_my_reviews'), callback_data="post_review_my_reviews")],
            [InlineKeyboardButton(get_translation(user_id, 'mai_menu'), callback_data="post_review_main_menu")]
        ]

        if update.callback_query:
            await update.callback_query.edit_message_text(
                text=text,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        else:
            await update.message.reply_text(
                text=text,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )

        # Clear context
        for key in required_keys + ["editing_review"]:
            context.user_data.pop(key, None)

        return POST_REVIEW

    except Exception as e:
        logging.error(f"Error finalizing review: {e}")
        error_msg = get_translation(user_id, 'failed_to_process_review')
        if update.callback_query:
            await update.callback_query.edit_message_text(error_msg)
        else:
            await update.message.reply_text(error_msg)
        return RATE_OPTIONS


async def start_rate_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:

    query = update.callback_query
    await query.answer()

    try:
        user_id = get_user_id(update)

        # Check if already rated with improved messaging
        if db.has_user_reviewed(user_id, "bot", "bot"):
            await query.edit_message_text(
                text=f"⭐ {get_translation(user_id, 'already_reviewed_bot_today')} ⭐",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(
                        text=f"← {get_translation(user_id, 'back_to_rating_menu')}",
                        callback_data="back_to_rate_menu"
                    )]
                ]),
                parse_mode="HTML"
            )
            return RATE_OPTIONS

        # Initialize rating context with clear structure
        context.user_data.update({
            "target_type": "bot",
            "review_target": "bot",
            "rating_dimensions": {
                "ease_of_use": {
                    "display": get_translation(user_id, 'ease_of_use_dimension'),
                    "icon": "🤖"
                },
                "features": {
                    "display": get_translation(user_id, 'feature_completeness_dimension'),
                    "icon": "💡"
                },
                "support": {
                    "display": get_translation(user_id, 'support_quality_dimension'),
                    "icon": "🏆"
                }
            }
        })

        # Start with first dimension using improved display
        await show_dimension_rating(update, context, "ease_of_use")
        return RATE_DIMENSION

    except Exception as e:
        logging.error(f"Bot rating initiation error: {str(e)}", exc_info=True)
        await query.edit_message_text(
            text=f"⚠️ {get_translation(user_id, 'failed_to_start_bot_rating')}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    text=f"🔄 {get_translation(user_id, 'try_again')}",
                    callback_data="rate_bot"
                )]
            ])
        )
        return RATE_OPTIONS


async def show_review_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show review search interface"""
    try:
        user_id = get_user_id(update)
        logging.info(f"Entering show_review_search for user {user_id}")

        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, 'search_by_name'), callback_data="search_by_name")],
            [InlineKeyboardButton(get_translation(user_id, 'employers_only'), callback_data="filter_reviews_employer")],
            [InlineKeyboardButton(get_translation(user_id, 'job_seekers_only'),
                                  callback_data="filter_reviews_job_seeker")],
            [InlineKeyboardButton(get_translation(user_id, 'top_rated'), callback_data="sort_reviews_top")],
            [InlineKeyboardButton(get_translation(user_id, 'most_recent'), callback_data="sort_reviews_recent")],
            [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_rate_menu")]
        ]

        text = f"{get_translation(user_id, 'search_reviews_prompt')}\n\n" \
               f"{get_translation(user_id, 'filter_or_sort_instructions')}"

        if update.callback_query:
            try:
                await update.callback_query.edit_message_text(
                    text=text,
                    reply_markup=InlineKeyboardMarkup(keyboard))
            except BadRequest as e:
                if "not modified" not in str(e):
                    raise
        else:
            await context.bot.send_message(
                chat_id=user_id,
                text=text,
                reply_markup=InlineKeyboardMarkup(keyboard))

        return SEARCH_REVIEWS

    except Exception as e:
        logging.error(f"Error in show_review_search: {str(e)}")
        return SEARCH_REVIEWS

async def filter_employers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Filter to show only employers"""
    query = update.callback_query
    await query.answer()

    context.user_data["current_filter"] = "employer"
    rateable_users = [u for u in context.user_data.get("rateable_users", [])
                      if u.get("type") == "employer"]

    if not rateable_users:
        await query.edit_message_text(
            text="No employers found.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back", callback_data="back_to_search")]
            ])
        )
        return SELECT_USER_FOR_RATING

    keyboard = [
        [InlineKeyboardButton(f"{u['name']} (ID: {u['id']})", callback_data=f"select_user_{u['id']}")]
        for u in rateable_users
    ]
    keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="back_to_search")])

    await query.edit_message_text(
        text="Select an employer to rate:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_USER_FOR_RATING


async def filter_jobseekers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Filter to show only job seekers"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    context.user_data["current_filter"] = "job_seeker"
    rateable_users = [u for u in context.user_data.get("rateable_users", [])
                      if u.get("type") == "job_seeker"]

    if not rateable_users:
        await query.edit_message_text(
            text=get_translation(user_id, 'no_job_seekers_found'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_search")]
            ])
        )
        return SELECT_USER_FOR_RATING

    keyboard = [
        [InlineKeyboardButton(f"{u['name']} (ID: {u['id']})", callback_data=f"select_user_{u['id']}")]
        for u in rateable_users
    ]
    keyboard.append([InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_search")])

    await query.edit_message_text(
        text=get_translation(user_id, 'select_job_seeker_to_rate'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_USER_FOR_RATING

async def sort_top_rated(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Sort users by their average rating"""
    user_id = get_user_id(update)
    rateable_users = context.user_data.get("rateable_users", [])

    # Add average rating to each user
    for user in rateable_users:
        user["avg_rating"] = db.get_user_avg_rating(user["id"])

    # Sort descending by rating
    sorted_users = sorted(rateable_users, key=lambda x: x.get("avg_rating", 0), reverse=True)

    keyboard = [
        [InlineKeyboardButton(
            f"{u['name']} (⭐{u.get('avg_rating', 0):.1f})",
            callback_data=f"select_user_{u['id']}"
        )]
        for u in sorted_users
    ]
    keyboard.append([InlineKeyboardButton(get_translation(user_id, 'bak'), callback_data="back_to_search")])

    await update.callback_query.edit_message_text(
        text=get_translation(user_id, 'top_rated_users'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_USER_FOR_RATING


async def sort_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle all sorting operations"""
    try:
        query = update.callback_query
        await query.answer()
        user_id = get_user_id(update)

        logging.info(f"Sorting callback received: {query.data} from user {user_id}")
        logging.info(f"Current user_data: {context.user_data}")

        # Get current state from user_data (Solution 1)
        current_state = context.user_data.get('_conversation_state', SEARCH_REVIEWS)
        logging.info(f"Current state: {current_state}")

        # Handle different sorting types
        if "top" in query.data.lower():
            sort_method = "top"
        elif "recent" in query.data.lower():
            sort_method = "recent"
        else:
            await query.edit_message_text("Invalid sort option")
            return current_state

        if current_state == SEARCH_REVIEWS:  # Use the actual variable, not string
            # Handle review sorting
            context.user_data["review_sort"] = sort_method

            if "review_results" not in context.user_data:
                context.user_data["review_results"] = db.search_reviews(
                    sort_by=sort_method
                )

            # Sort existing results
            review_results = [dict(row) for row in context.user_data["review_results"]]
            review_results.sort(
                key=lambda x: x.get('created_at' if sort_method == "recent" else 'rating', 0),
                reverse=True
            )
            context.user_data["review_results"] = review_results

            return await display_review_results(update, context)

        elif current_state == SEARCH_USER_FOR_RATING:  # Use the actual variable
            # Handle user sorting
            if sort_method == "top":
                rateable_users = context.user_data.get("rateable_users", [])
                if not rateable_users:
                    rateable_users = db.get_rateable_users(user_id)
                    context.user_data["rateable_users"] = rateable_users

                for user in rateable_users:
                    user["avg_rating"] = db.get_user_avg_rating(user["id"])

                sorted_users = sorted(
                    rateable_users,
                    key=lambda x: x.get("avg_rating", 0),
                    reverse=True
                )
                text = get_translation(user_id, 'top_rated_users')
            else:  # recent
                sorted_users = db.get_recently_interacted_users(user_id)
                context.user_data["rateable_users"] = sorted_users
                text = get_translation(user_id, 'recently_interacted_users')

            keyboard = [
                [InlineKeyboardButton(
                    f"{u.get('name', 'Unknown')} (⭐{u.get('avg_rating', 0):.1f})" if sort_method == "top" else u.get('name', 'Unknown'),
                    callback_data=f"select_user_{u['id']}"
                )]
                for u in sorted_users
                if u.get('id')
            ]
            keyboard.append([
                InlineKeyboardButton(
                    get_translation(user_id, 'back'),
                    callback_data="back_to_search"
                )
            ])

            await query.edit_message_text(
                text=text,
                reply_markup=InlineKeyboardMarkup(keyboard))
            return SELECT_USER_FOR_RATING

        else:
            logging.error(f"Unexpected state for sorting: {current_state}")
            return current_state

    except Exception as e:
        logging.error(f"Error in sort_reviews: {str(e)}", exc_info=True)
        await query.answer("Error sorting results", show_alert=True)
        return SEARCH_REVIEWS

async def show_review_filter_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show filter options for reviews"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'all_reviews'), callback_data="filter_reviews_all")],
        [InlineKeyboardButton(get_translation(user_id, 'employers_only'), callback_data="filter_reviews_employer")],
        [InlineKeyboardButton(get_translation(user_id, 'job_seekers_only'), callback_data="filter_reviews_job_seeker")],
        [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_review_results")]
    ]

    await query.edit_message_text(
        text=get_translation(user_id, 'filter_reviews_by_type'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SEARCH_REVIEWS

async def show_review_sort_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show sort options for reviews"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'most_recent'), callback_data="sort_reviews_recent")],
        [InlineKeyboardButton(get_translation(user_id, 'top_rated'), callback_data="sort_reviews_top")],
        [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_review_results")]
    ]

    await query.edit_message_text(
        text=get_translation(user_id, 'sort_reviews_by'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SEARCH_REVIEWS

# async def show_review_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
#     """Show detailed view of a specific review"""
#     try:
#         query = update.callback_query
#         await query.answer()
#         review_id = int(query.data.split('_')[2])  # view_review_123 -> 123
#         user_id = get_user_id(update)
#
#         # Get review details from database
#         review = db.get_review_details(review_id)
#         if not review:
#             await query.answer(get_translation(user_id, 'review_not_found'), show_alert=True)
#             return await display_review_results(update, context)
#
#         # Format review details
#         target_name = review.get('target_name', 'Unknown')
#         reviewer_name = review.get('reviewer_name', 'Anonymous')
#         rating = review.get('rating', 0)
#         comment = review.get('comment', get_translation(user_id, 'no_comment_provided'))
#         created_at = review.get('created_at', 'Unknown date')
#
#         text = (
#             f"📝 Review of {target_name}\n"
#             f"👤 By: {reviewer_name}\n"
#             f"⭐ Rating: {'⭐' * rating}\n"
#             f"📅 Date: {created_at}\n\n"
#             f"💬 Comment:\n{comment}"
#         )
#
#         keyboard = [
#             [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_review_results")]
#         ]
#
#         await query.edit_message_text(
#             text=text,
#             reply_markup=InlineKeyboardMarkup(keyboard)
#         )
#         return REVIEW_DETAILS
#
#     except Exception as e:
#         logging.error(f"Error showing review details: {str(e)}")
#         await query.answer(get_translation(user_id, 'error_loading_review'), show_alert=True)
#         return await display_review_results(update, context)

async def previous_review_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Navigate to previous page of review results"""
    context.user_data["current_review_page"] = max(0, context.user_data.get("current_review_page", 0) - 1)
    return await display_review_results(update, context)

async def next_review_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Navigate to next page of review results"""
    context.user_data["current_review_page"] = context.user_data.get("current_review_page", 0) + 1
    return await display_review_results(update, context)

async def handle_dimension_rating(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Store rating for a specific dimension and proceed to next"""
    query = update.callback_query
    await query.answer()

    try:
        parts = query.data.split('_')
        if parts[0] not in ['rate', 'edit_rate']:  # Handle both cases
            raise ValueError("Invalid callback prefix")

        dimension = '_'.join(parts[1:-1])
        rating_value = int(parts[-1])

        # Store in context
        if "dimension_ratings" not in context.user_data:
            context.user_data["dimension_ratings"] = {}
        context.user_data["dimension_ratings"][dimension] = rating_value

        # Rest of your existing flow...

        # Calculate overall rating (average of non-zero ratings)
        ratings = [v for v in context.user_data["dimension_ratings"].values() if v > 0]
        context.user_data["overall_rating"] = sum(ratings) / len(ratings) if ratings else 0

        # Get next unrated dimension
        dimensions = context.user_data["rating_dimensions"]
        remaining_dims = [d for d in dimensions if d not in context.user_data["dimension_ratings"]]

        if remaining_dims:
            await show_dimension_rating(update, context, remaining_dims[0])
            return RATE_DIMENSION
        else:
            await query.edit_message_text(
                text="Would you like to add an optional comment?",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ Yes", callback_data="add_comment")],
                    [InlineKeyboardButton("❌ No", callback_data="skip_comment")]
                ])
            )
            return ADD_COMMENT_OPTIONAL

    except Exception as e:
        logging.error(f"Error processing rating: {str(e)}")
        await query.edit_message_text(
            "⚠️ Sorry, there was an error processing your rating.\n\n"
            "Please try rating again or use /cancel to exit.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Try Again", callback_data="back_to_rate_menu")]
            ])
        )
        return RATE_OPTIONS


async def skip_dimension_rating(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Skip rating for current dimension with proper parsing"""
    query = update.callback_query
    await query.answer()

    try:
        # Extract dimension being skipped (format: skip_dimension_name)
        parts = query.data.split("_")
        dimension = "_".join(parts[1:])  # Handle multi-word dimensions

        # Mark as skipped (0 rating)
        if "dimension_ratings" not in context.user_data:
            context.user_data["dimension_ratings"] = {}
        context.user_data["dimension_ratings"][dimension] = 0

        # Get next dimension to rate
        dimensions = context.user_data["rating_dimensions"]
        rated_dims = set(context.user_data["dimension_ratings"].keys())
        remaining_dims = [d for d in dimensions if d not in rated_dims]

        if remaining_dims:
            next_dim = remaining_dims[0]
            await show_dimension_rating(update, context, next_dim)
            return RATE_DIMENSION
        else:
            # All dimensions processed (some may be skipped)
            ratings = [v for v in context.user_data["dimension_ratings"].values() if v > 0]
            if ratings:  # Only calculate if we have at least one rating
                context.user_data["overall_rating"] = sum(ratings) / len(ratings)
            else:
                await query.answer("Please rate at least one dimension", show_alert=True)
                return RATE_DIMENSION

            await query.edit_message_text(
                text="Would you like to add an optional comment?",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Yes", callback_data="add_comment")],
                    [InlineKeyboardButton("No", callback_data="skip_comment")]
                ])
            )
            return ADD_COMMENT_OPTIONAL

    except Exception as e:
        logging.error(f"Error skipping dimension: {e}")
        await query.answer("Error processing your request", show_alert=True)
        return RATE_DIMENSION


async def show_review_settings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display review privacy settings with message change detection"""
    user_id = get_user_id(update)
    privacy_settings = db.get_review_privacy_settings(user_id)

    # Generate content
    text = f"{get_translation(user_id, 'review_privacy_settings_title')}\n\n"
    text += f"• {get_translation(user_id, 'show_your_name')}: {'✅ ON' if privacy_settings['show_name'] else '❌ OFF'}\n"
    text += f"• {get_translation(user_id, 'show_contact_info')}: {'✅ ON' if privacy_settings['show_contact'] else '❌ OFF'}\n\n"

    keyboard = [
        [InlineKeyboardButton(get_translation(user_id, 'toggle_name_visibility'), callback_data="toggle_anonymous")],
        [InlineKeyboardButton(get_translation(user_id, 'toggle_contact_visibility'), callback_data="toggle_contact_visible")],
        [InlineKeyboardButton(get_translation(user_id, 'bak'), callback_data="back_to_rate_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Check if message needs updating
    current_hash = hash((text, str(reply_markup)))
    if context.user_data.get("settings_msg_hash") == current_hash:
        return REVIEW_SETTINGS  # No changes needed

    try:
        await update.callback_query.edit_message_text(
            text=text,
            reply_markup=reply_markup
        )
        context.user_data["settings_msg_hash"] = current_hash
    except BadRequest as e:
        if "not modified" not in str(e):
            raise  # Only suppress "not modified" errors

    return REVIEW_SETTINGS


async def toggle_privacy_setting(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Toggle privacy setting with error handling"""
    try:
        query = update.callback_query
        await query.answer()

        setting = query.data.split("_")[-1]  # "anonymous" or "contact_visible"
        user_id = get_user_id(update)

        if setting == "anonymous":
            db.toggle_setting(user_id, "show_name")
        elif setting == "contact_visible":
            db.toggle_setting(user_id, "show_contact")

        # Clear message hash to force update
        context.user_data.pop("settings_msg_hash", None)

        return await show_review_settings(update, context)
    except Exception as e:
        logging.error(f"Error toggling setting: {e}")
        await query.edit_message_text(get_translation(user_id, 'error_updating_settings'))
        return REVIEW_SETTINGS

async def handle_user_search_for_rating(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle user search query for rating"""
    search_term = update.message.text.strip()
    user_id = get_user_id(update)

    try:
        # Search both employers and job seekers
        results = db.search_users(
            search_term,
            user_type=context.user_data.get("current_filter")  # Respect current filter
        )

        if not results:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'no_matching_users_found'),
                reply_markup=ReplyKeyboardMarkup([[get_translation(user_id, 'back')]], resize_keyboard=True)
            )
            return await show_user_search_interface(update, context)

        # Store for pagination
        context.user_data["search_results"] = results
        context.user_data["current_page"] = 1

        # Show first page
        return await display_search_results_page(update, context)

    except Exception as e:
        logging.error(f"User search error: {str(e)}")
        await context.bot.send_message(
            chat_id=user_id,
            text=get_translation(user_id, 'error_occurred_during_search'),
            reply_markup=ReplyKeyboardMarkup([[get_translation(user_id, 'bak')]], resize_keyboard=True)
        )
        return await show_rate_options(update, context)

# async def display_search_results_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
#     """Display a page of search results"""
#     results = context.user_data["search_results"]
#     # current_page = context.user_data["current_page"]
#     current_page = context.user_data.get("current_page", 1)
#     per_page = 5
#     start_idx = (current_page - 1) * per_page
#     page_results = results[start_idx:start_idx + per_page]
#
#     keyboard = []
#     for user in page_results:
#         btn_text = f"{user['name']} ({user['type'].capitalize()})"
#         if user.get("avg_rating"):
#             btn_text += f" ⭐{user['avg_rating']:.1f}"
#         keyboard.append([InlineKeyboardButton(btn_text, callback_data=f"select_user_{user['id']}")])
#
#     # Add pagination controls if needed
#     if len(results) > per_page:
#         pagination_row = []
#         if current_page > 1:
#             pagination_row.append(InlineKeyboardButton("◀️ Prev", callback_data="prev_page"))
#         if len(results) > current_page * per_page:
#             pagination_row.append(InlineKeyboardButton("Next ▶️", callback_data="next_page"))
#         keyboard.append(pagination_row)
#
#     keyboard.append([InlineKeyboardButton("Back to Search", callback_data="back_to_search")])
#
#     text = f"Search Results (Page {current_page}):"
#     if update.callback_query:
#         await update.callback_query.edit_message_text(
#             text=text,
#             reply_markup=InlineKeyboardMarkup(keyboard)
#         )
#     else:
#         await context.bot.send_message(
#             chat_id=get_user_id(update),
#             text=text,
#             reply_markup=InlineKeyboardMarkup(keyboard)
#         )
#
#     return SELECT_USER_FOR_RATING

async def display_search_results_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display a page of search results (works with dict and sqlite3.Row)"""
    results = context.user_data.get("search_results", [])
    current_page = context.user_data.get("current_page", 1)
    per_page = 5

    start_idx = (current_page - 1) * per_page
    page_results = results[start_idx:start_idx + per_page]

    keyboard = []
    for user in page_results:
        # Handle both dictionary and sqlite3.Row
        if isinstance(user, dict):
            name = user.get('name', 'Unknown')
            user_type = user.get('type', 'unknown')
            avg_rating = user.get('avg_rating')
            user_id = user.get('id', 'unknown')
        else:  # Assume it's sqlite3.Row
            name = user['name'] if 'name' in user.keys() else 'Unknown'
            user_type = user['type'] if 'type' in user.keys() else 'unknown'
            avg_rating = user['avg_rating'] if 'avg_rating' in user.keys() else None
            user_id = user['id'] if 'id' in user.keys() else 'unknown'

        btn_text = f"{name} ({str(user_type).capitalize()})"
        if avg_rating is not None:
            btn_text += f" ⭐{float(avg_rating):.1f}"

        callback_data = f"select_user_{user_id}"
        keyboard.append([InlineKeyboardButton(btn_text, callback_data=callback_data)])

    # Pagination controls
    if len(results) > per_page:
        pagination_row = []
        if current_page > 1:
            pagination_row.append(InlineKeyboardButton("◀️ Prev", callback_data="prev_page"))
        if start_idx + per_page < len(results):
            pagination_row.append(InlineKeyboardButton("Next ▶️", callback_data="next_page"))
        if pagination_row:
            keyboard.append(pagination_row)

    keyboard.append([InlineKeyboardButton("Back to Search", callback_data="back_to_search")])

    text = f"Search Results (Page {current_page}):"
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await context.bot.send_message(
            chat_id=get_user_id(update),
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    return SELECT_USER_FOR_RATING

async def prompt_for_comment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user to enter an optional comment"""
    try:
        user_id = get_user_id(update)
        await update.callback_query.edit_message_text(
            text=f"{get_translation(user_id, 'optional_review_comment_title')}\n\n"
                 f"{get_translation(user_id, 'optional_review_comment_body')}\n"
                 f"• {get_translation(user_id, 'what_stood_out')}\n"
                 f"• {get_translation(user_id, 'what_could_improve')}\n"
                 f"• {get_translation(user_id, 'any_specific_feedback')}\n\n"
                 f"{get_translation(user_id, 'or_click_below_to_skip')}:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'skip_comment'), callback_data="skip_comment")]
            ])
        )
        return PROMPT_FOR_COMMENT
    except Exception as e:
        logging.error(f"Error in prompt_for_comment: {e}")
        await update.callback_query.answer(get_translation(user_id, 'error_loading_comment_prompt'), show_alert=True)
        return RATE_DIMENSION

async def submit_comment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle submitted comment"""
    try:
        user_id = get_user_id(update)
        comment = update.message.text.strip()
        if len(comment) > 500:
            await update.message.reply_text(
                get_translation(user_id, 'comment_too_long'),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(get_translation(user_id, 'skip_comment'), callback_data="skip_comment")]
                ])
            )
            return PROMPT_FOR_COMMENT

        context.user_data["review_comment"] = comment
        return await finalize_review(update, context)

    except Exception as e:
        logging.error(f"Error submitting comment: {e}")
        await update.message.reply_text(
            get_translation(user_id, 'error_processing_comment'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'skip_comment'), callback_data="skip_comment")]
            ])
        )
        return PROMPT_FOR_COMMENT

async def skip_comment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle skipping comment"""
    user_id = get_user_id(update)
    await update.callback_query.answer(get_translation(user_id, 'skipping_comment'))
    context.user_data["review_comment"] = None
    return await finalize_review(update, context)

async def cancel_comment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel comment entry"""
    user_id = get_user_id(update)
    await update.message.reply_text(get_translation(user_id, 'comment_entry_cancelled'))
    return await show_rate_options(update, context)

async def edit_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Allow user to edit their review before submission"""
    user_id = get_user_id(update)
    # Reset to first dimension
    dimensions = context.user_data["rating_dimensions"]
    first_dim = next(iter(dimensions))
    context.user_data["dimension_ratings"] = {}

    return await show_dimension_rating(update, context, first_dim)

async def show_my_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display all reviews written by the current user"""
    user_id = get_user_id(update)
    reviews = db.get_user_reviews(user_id)

    if not reviews:
        await update.callback_query.edit_message_text(
            text=get_translation(user_id, 'no_reviews_written_yet'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_rate_menu")]
            ])
        )
        return MY_REVIEWS

    keyboard = []
    for review in reviews:
        target_name = db.get_user_name(review['target_id']) if review['target_type'] != 'bot' else "JobBot"
        btn_text = f"{'⭐' * review['rating']} {target_name}"
        if review['comment']:
            btn_text += " 💬"
        keyboard.append([
            InlineKeyboardButton(btn_text, callback_data=f"review_my_{review['id']}")  # ← Updated callback format
        ])

    keyboard.append([InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_rate_menu")])

    await update.callback_query.edit_message_text(
        text=get_translation(user_id, 'your_reviews_prompt'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return MY_REVIEWS


async def show_review_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        # Parse callback data
        data_parts = update.callback_query.data.split('_')
        if data_parts[0] == 'review' and data_parts[1] == 'my':
            review_id = int(data_parts[2])
            from_my_reviews = True
        else:
            # Handle other flows like search, etc.
            review_id = int(data_parts[-1])
            from_my_reviews = False

        review = db.get_review_details(review_id)
        if not review:
            await update.callback_query.answer("Review not found", show_alert=True)
            return await show_my_reviews(update, context)

        target_name = db.get_user_name(review['target_id']) if review['target_type'] != 'bot' else "JobBot"
        created_at = review['created_at']
        if isinstance(created_at, str):
            from datetime import datetime
            created_at = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")

        text = (
            f"Review of {target_name}\n"
            f"Rating: {'⭐' * review['rating']}\n"
            f"Date: {created_at.strftime('%Y-%m-%d')}\n"
            f"{review['comment'] or 'No comment provided'}"
        )

        # Conditionally build keyboard
        keyboard = []
        if from_my_reviews:
            keyboard.append([InlineKeyboardButton("✏️ Edit", callback_data=f"edit_review_{review_id}")])
            keyboard.append([InlineKeyboardButton("🗑️ Delete", callback_data=f"delete_review_{review_id}")])

        keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="back_to_my_reviews" if from_my_reviews else "back_to_search")])

        await update.callback_query.edit_message_text(
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return REVIEW_DETAILS
    except Exception as e:
        logging.error(f"Error showing review details: {e}")
        await update.callback_query.answer("Error loading review", show_alert=True)
        return await show_my_reviews(update, context)


async def delete_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle review deletion with confirmation"""
    query = update.callback_query
    await query.answer()

    try:
        # Extract review ID from callback data (format: delete_review_123)
        review_id = int(query.data.split('_')[2])
        if not review_id:
            raise ValueError("Invalid review ID")

        # Store in context for next step
        context.user_data["review_to_delete"] = review_id

        # Show confirmation
        await query.edit_message_text(
            text="Are you sure you want to delete this review?",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Yes, delete", callback_data="confirm_delete")],
                [InlineKeyboardButton("❌ Cancel", callback_data=f"review_{review_id}")]
            ])
        )
        return REVIEW_DETAILS

    except Exception as e:
        logging.error(f"Error preparing to delete review: {e}")
        await query.answer("Failed to prepare delete operation", show_alert=True)
        return MY_REVIEWS


async def confirm_delete_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle confirmed review deletion"""
    query = update.callback_query
    await query.answer()

    try:
        # Get review ID from context
        review_id = context.user_data.get("review_to_delete")
        if not review_id:
            raise ValueError("No review ID found in context")

        # Delete from database
        success = db.delete_review(review_id)

        if success:
            # Clear cached data
            context.user_data.pop("review_to_delete", None)
            context.user_data.pop("search_results", None)
            context.user_data.pop("review_results", None)

            await query.edit_message_text(
                text="✅ Review deleted successfully!",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Back to Rating", callback_data="back_to_rate_menu")]
                ])
            )
        else:
            await query.edit_message_text(
                text="⚠️ Failed to delete review. It may not exist anymore.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Try Again", callback_data=f"delete_review_{review_id}")]
                ])
            )
            return MY_REVIEWS

        return MY_REVIEWS

    except Exception as e:
        logging.error(f"Error deleting review: {e}", exc_info=True)
        await query.edit_message_text(
            text="⚠️ An error occurred while trying to delete the review.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Back", callback_data="back_to_my_reviews")]
            ])
        )
        return MY_REVIEWS

async def delete_review_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt confirmation before deleting a review"""
    query = update.callback_query
    await query.answer()

    try:
        # Get review ID from callback data (format: admin_delete_review_123)
        parts = query.data.split('_')
        if len(parts) != 4 or parts[0] != 'admin' or parts[1] != 'delete' or parts[2] != 'review':
            raise ValueError("Invalid delete request")
        review_id = int(parts[3])

        # Store for next step
        context.user_data["review_to_delete"] = review_id

        # Show confirmation dialog
        await query.edit_message_text(
            text=f"⚠️ Are you sure you want to delete review #{review_id}?",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Confirm Delete", callback_data="confirm_deleted")],
                [InlineKeyboardButton("❌ Cancel", callback_data="admin_back_to_ratings")]
            ])
        )
        return ADMIN_DELETE_REVIEW

    except Exception as e:
        logging.error(f"Error preparing to delete review: {e}")
        await query.edit_message_text("⚠️ Failed to prepare delete operation.")
        return ADMIN_REVIEW_LIST

async def confirm_deleted_review_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle confirmed review deletion"""
    query = update.callback_query
    await query.answer()

    try:
        review_id = context.user_data.get("review_to_delete")
        if not review_id:
            raise ValueError("No review ID found in context")

        success = db.delete_review(review_id)
        if success:
            context.user_data.pop("review_to_delete", None)
            context.user_data.pop("search_results", None)
            context.user_data.pop("review_results", None)
            await query.edit_message_text(
                text="✅ Review deleted successfully.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Back to Dashboard", callback_data="admin_back_to_ratings")]
                ])
            )
        else:
            await query.edit_message_text(
                text="❌ Failed to delete review. It may not exist anymore.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Try Again", callback_data="admin_delete_review")]
                ])
            )

        # Clear stored ID
        context.user_data.pop("review_to_delete", None)
        return ADMIN_RATINGS_MENU

    except Exception as e:
        logging.error(f"Error deleting review: {e}")
        await query.edit_message_text(
            text="⚠️ An error occurred while trying to delete the review.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Back", callback_data="admin_back_to_ratings")]
            ])
        )
        return ADMIN_RATINGS_MENU


async def edit_existing_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start editing an existing review with all dimensions"""
    review_id = int(update.callback_query.data.split('_')[2])
    review = db.get_review_details(review_id)

    # Store original data in context
    context.user_data.update({
        "editing_review": review_id,
        "target_type": review['target_type'],
        "review_target": review['target_id'],
        "dimension_ratings": {}  # Start fresh
    })

    # Set up rating dimensions based on type
    rating_dimensions = {
        'employer': {
            'professionalism': 'Professionalism',
            'communication': 'Communication',
            'hiring_process': 'Hiring Process'
        },
        'job_seeker': {
            'reliability': 'Reliability',
            'skills': 'Skill Match',
            'communication': 'Communication'
        },
        'bot': {
            'ease_of_use': 'Ease of Use',
            'features': 'Feature Completeness',
            'support': 'Support Quality'
        }
    }

    context.user_data["rating_dimensions"] = rating_dimensions[review['target_type']]

    # Start with first dimension
    first_dim = next(iter(context.user_data["rating_dimensions"]))
    await show_dimension_rating(update, context, first_dim)
    return RATE_DIMENSION


async def handle_review_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle search query for reviews"""
    try:
        if update.message:  # Handle text message
            search_term = update.message.text.strip()
            user_id = update.message.from_user.id
            message_func = context.bot.send_message
        else:  # Handle callback query
            query = update.callback_query
            await query.answer()
            search_term = query.data.split('_', 1)[1] if '_' in query.data else ""
            user_id = query.from_user.id
            message_func = query.edit_message_text

        current_filter = context.user_data.get("review_filter", "all")
        sort_method = context.user_data.get("review_sort", "recent")

        results = db.search_reviews(
            search_term=search_term,
            target_type=current_filter if current_filter != "all" else None,
            sort_by=sort_method
        )

        if not results:
            await message_func(
                chat_id=user_id,
                text="No reviews found matching your search."
            )
            return await show_review_search(update, context)

        context.user_data["review_results"] = results
        context.user_data["current_review_page"] = 0

        return await display_review_results(update, context)

    except Exception as e:
        logging.error(f"Review search error: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text="An error occurred during search. Please try again."
        )
        return await show_rate_options(update, context)


async def filter_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Filter reviews by type (employer/job_seeker)"""
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Determine filter type
    if "employer" in query.data:
        target_type = "employer"
        filter_name = get_translation(user_id, 'employers_only')
    else:
        target_type = "job_seeker"
        filter_name = get_translation(user_id, 'job_seekers_only')

    # Ensure rateable_users exists in context
    if "rateable_users" not in context.user_data:
        context.user_data["rateable_users"] = db.get_rateable_users(user_id)

    # Get filtered users
    rateable_users = [u for u in context.user_data.get("rateable_users", [])
                     if u.get("type") == target_type]

    if not rateable_users:
        message = (
            f"🔍 {get_translation(user_id, 'no_users_found_filter')}\n\n"
            f"To rate {filter_name.lower()}, you need to have:\n"
            f"- Applied to jobs (for employers)\n"
            f"- Received applications (for job seekers)\n\n"
            f"Try interacting with users first!"
        )

        await query.edit_message_text(
            text=message,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_search")]
            ])
        )
        return SEARCH_REVIEWS

    # Build keyboard with filtered users using simplified format
    keyboard = []
    for user in rateable_users:
        btn_text = f"{user.get('name', 'Unknown')} ({user.get('type', 'user').replace('_', ' ')})"
        callback_data = f"select_user_{user.get('id')}"  # Simplified format
        keyboard.append([InlineKeyboardButton(btn_text, callback_data=callback_data)])

    keyboard.append([InlineKeyboardButton(get_translation(user_id, 'back'), callback_data="back_to_search")])

    await query.edit_message_text(
        text=f"{filter_name}:\n{get_translation(user_id, 'select_user_to_rate')}",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_USER_FOR_RATING


async def flag_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle flagging of inappropriate reviews"""
    review_id = int(update.callback_query.data.split('_')[-1])
    context.user_data["review_to_flag"] = review_id

    await update.callback_query.edit_message_text(
        text="Please select a reason for flagging this review:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Inappropriate Content", callback_data="flag_reason_inappropriate")],
            [InlineKeyboardButton("False Information", callback_data="flag_reason_false")],
            [InlineKeyboardButton("Conflict of Interest", callback_data="flag_reason_conflict")],
            [InlineKeyboardButton("Cancel", callback_data=f"review_{review_id}")]
        ])
    )
    return REVIEW_DETAILS



# async def display_review_results(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
#     """Display paginated review search results"""
#     try:
#         # Get raw results from context
#         raw_results = context.user_data.get("review_results", [])
#
#         # Convert if needed (i.e., if results are sqlite3.Row objects)
#         if raw_results and isinstance(raw_results[0], sqlite3.Row):
#             results = [dict(row) for row in raw_results]
#         else:
#             results = raw_results  # Already a list of dicts
#
#         page = context.user_data.get("current_review_page", 0)
#         per_page = 5
#         start_idx = page * per_page
#         page_results = results[start_idx:start_idx + per_page]
#
#         keyboard = []
#         for review in page_results:
#             # Safely access dictionary values using .get()
#             target_name = review.get('target_name', 'Unknown')
#             reviewer_name = review.get('reviewer_name', 'Anonymous')
#             rating = review.get('rating', 0)
#             review_id = review.get('id')
#
#             btn_text = f"{'⭐' * rating} {target_name} by {reviewer_name}"
#             keyboard.append([
#                 InlineKeyboardButton(btn_text, callback_data=f"view_review_{review['id']}")  # ← Different format
#             ])
#
#         # Pagination controls
#         pagination_row = []
#         if page > 0:
#             pagination_row.append(InlineKeyboardButton("◀️ Previous", callback_data="prev_review_page"))
#         if len(results) > (page + 1) * per_page:
#             pagination_row.append(InlineKeyboardButton("Next ▶️", callback_data="next_review_page"))
#         if pagination_row:
#             keyboard.append(pagination_row)
#
#         # # Filter & Sort options
#         # keyboard.append([
#         #     InlineKeyboardButton("Filter", callback_data="review_filter_menu"),
#         #     InlineKeyboardButton("Sort", callback_data="review_sort_menu")
#         # ])
#         keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_rate_menu")])
#
#         text = f"Review Results (Page {page + 1}):"
#
#         if update.callback_query:
#             try:
#                 await update.callback_query.edit_message_text(
#                     text=text,
#                     reply_markup=InlineKeyboardMarkup(keyboard)
#                 )
#             except BadRequest as e:
#                 if "not modified" not in str(e):
#                     raise
#         else:
#             await context.bot.send_message(
#                 chat_id=get_user_id(update),
#                 text=text,
#                 reply_markup=InlineKeyboardMarkup(keyboard)
#             )
#
#         return SEARCH_REVIEWS
#
#     except Exception as e:
#         logging.error(f"Error in display_review_results: {str(e)}")
#         await update.callback_query.answer("Error displaying results. Please try again.")
#         return SEARCH_REVIEWS

async def display_review_results(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display paginated review search results"""
    try:
        results = context.user_data.get("review_results", [])
        current_page = context.user_data.get("current_review_page", 1)
        per_page = 5

        start_idx = (current_page - 1) * per_page
        page_results = results[start_idx:start_idx + per_page]

        keyboard = []
        for review in page_results:
            # Safely extract data from sqlite3.Row or dict
            if isinstance(review, dict):
                review_id = review.get('id', 'N/A')
                target_name = review.get('target_name', 'Unknown')
                reviewer_name = review.get('reviewer_name', 'Anonymous')
                rating = review.get('rating', 0)
            else:
                # Handle sqlite3.Row by using dictionary-like access
                try:
                    review_id = review['id']
                except Exception:
                    review_id = 'N/A'

                target_name = review.get('target_name', 'Unknown') if hasattr(review, 'get') else 'Unknown'
                reviewer_name = review.get('reviewer_name', 'Anonymous') if hasattr(review, 'get') else 'Anonymous'
                rating = review.get('rating', 0) if hasattr(review, 'get') else 0

            btn_text = f"{'⭐' * rating} {target_name} by {reviewer_name}"
            callback_data = f"view_review_{review_id}"
            keyboard.append([InlineKeyboardButton(btn_text, callback_data=callback_data)])

        # Pagination controls
        if len(results) > per_page:
            pagination_row = []
            if current_page > 1:
                pagination_row.append(InlineKeyboardButton("◀️ Prev", callback_data="prev_review_page"))
            if start_idx + per_page < len(results):
                pagination_row.append(InlineKeyboardButton("Next ▶️", callback_data="next_review_page"))
            if pagination_row:
                keyboard.append(pagination_row)

        keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_rate_menu")])

        text = f"Review Results (Page {current_page}):"

        if update.callback_query:
            try:
                await update.callback_query.edit_message_text(
                    text=text,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            except BadRequest as e:
                if "not modified" not in str(e):
                    raise
        else:
            await context.bot.send_message(
                chat_id=get_user_id(update),
                text=text,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )

        return SEARCH_REVIEWS

    except Exception as e:
        logging.error(f"Error in display_review_results: {str(e)}")
        await update.callback_query.answer("Error displaying results. Please try again.")
        return SEARCH_REVIEWS

async def next_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle next page in search results"""
    context.user_data["current_page"] += 1
    return await display_search_results_page(update, context)

async def previous_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle previous page in search results"""
    context.user_data["current_page"] = max(1, context.user_data["current_page"] - 1)
    return await display_search_results_page(update, context)

async def handle_search_by_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt user to enter name for search"""
    await update.callback_query.edit_message_text(
        text="Please enter the name you want to search for:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔙 Back", callback_data="back_to_search")]
        ])
    )
    return SEARCH_REVIEWS


from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ContextTypes




from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

async def ratings_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Admin dashboard for managing reviews and ratings.
    Provides search, filtering, editing, deletion, and analytics.
    """
    user_id = update.effective_user.id

    # Initialize session data if not present
    if "review_filter" not in context.user_data:
        context.user_data["review_filter"] = "all"
    if "review_sort" not in context.user_data:
        context.user_data["review_sort"] = "recent"

    # Build summary text with stats
    total_reviews = db.get_total_review_count()
    avg_rating = db.get_overall_average_rating()

    summary_text = f"""<b>⭐ Ratings & Reviews Dashboard</b>
• Total Reviews: <b>{total_reviews}</b>
• Average Rating: <b>{avg_rating:.1f}⭐</b>

Choose an action below:"""

    # Dashboard menu options as inline buttons
    keyboard = [
        [
            InlineKeyboardButton("📊 View All Reviews", callback_data="admin_view_all_reviews"),
            InlineKeyboardButton("🔍 Search Reviews", callback_data="admin_search_reviews")
        ],
        [
            InlineKeyboardButton("📈 Review Statistics", callback_data="admin_review_stats")
        ],
        [
            InlineKeyboardButton("🗑️ Delete Review", callback_data="admin_delete_review"),
            InlineKeyboardButton("🔙 Back to User Interactions", callback_data="admin_back_to_user_interactions")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=user_id,
        text=summary_text,
        parse_mode="HTML",
        reply_markup=reply_markup
    )

    return ADMIN_RATINGS_MENU


import html

def escapedd_html(text: str) -> str:
    return html.escape(str(text), quote=False)
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes


async def show_all_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show paginated list of all reviews with action buttons for each"""
    user_id = update.effective_user.id
    page = context.user_data.get("admin_review_page", 1)
    per_page = 10

    reviews = db.get_all_reviews(page=page, per_page=per_page)

    if not reviews:
        await context.bot.send_message(chat_id=user_id, text="No reviews found.")
        return await ratings_dashboard(update, context)

    # Display each review with its own action buttons
    for review in reviews:
        target_id = review.get('target_id', 'N/A')
        reviewer_id = review.get('reviewer_id', 'N/A')

        target_name = db.get_user_name(target_id) if review.get('target_type') != 'bot' else "JobBot"
        reviewer_name = db.get_user_name(reviewer_id)
        rating = review.get('rating', 0)
        rating_stars = '⭐' * rating
        comment = review.get('comment', "<No comment>")
        created_at = review.get('created_at', 'Unknown date')

        msg = f"""<b>Review ID: {escapedd_html(str(review.get('id', 'N/A')))}</b>
    Reviewed: {escapedd_html(target_name)} ({escapedd_html(review.get('target_type', 'unknown'))})
    By: {escapedd_html(reviewer_name)}
    Rating: {escapedd_html(rating_stars)} ({escapedd_html(str(rating))}/5)
    Date: {escapedd_html(created_at)}
    Comment: {escapedd_html(comment)}"""

        # Action buttons for this specific review
        action_buttons = [
            [InlineKeyboardButton("🔍 Search", callback_data="admin_search_reviews")],
            [InlineKeyboardButton("🗑️ Delete", callback_data=f"admin_delete_review_{review['id']}")],
            [InlineKeyboardButton("🔙 Back", callback_data="admin_back_to_ratings")]
        ]
        markup = InlineKeyboardMarkup(action_buttons)

        await context.bot.send_message(
            chat_id=user_id,
            text=msg,
            parse_mode="HTML",
            reply_markup=markup
        )

    # Add pagination controls at the end
    pagination_row = []
    if page > 1:
        pagination_row.append(InlineKeyboardButton("⬅️ Prev", callback_data="admin_prev_page"))
    pagination_row.append(InlineKeyboardButton(f"Page {page}", callback_data="noop"))
    pagination_row.append(InlineKeyboardButton("Next ➡️", callback_data="admin_next_page"))

    pagination_markup = InlineKeyboardMarkup([pagination_row])

    await context.bot.send_message(
        chat_id=user_id,
        text="Navigate between pages:",
        reply_markup=pagination_markup
    )

    return ADMIN_REVIEW_LIST

async def prompt_review_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt the admin to enter a search term for reviews"""
    user_id = update.effective_user.id
    await context.bot.send_message(
        chat_id=user_id,
        text="🔍 Enter a keyword or phrase to search reviews:",
        reply_markup=ReplyKeyboardMarkup(
            [[KeyboardButton("🔙 Cancel")]],
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )
    return ADMIN_REVIEW_SEARCH



async def prompt_delete_review(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt admin to enter review ID for deletion"""
    user_id = update.effective_user.id
    await context.bot.send_message(
        chat_id=user_id,
        text="🗑️ Enter the Review ID you want to delete:",
        reply_markup=ReplyKeyboardMarkup(
            [[KeyboardButton("🔙 Cancel")]],
            one_time_keyboard=True,
            resize_keyboard=True
        )
    )
    return ADMIN_DELETE_REVIEW

async def prev_page_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Go to previous page of reviews (for paginated lists)"""
    current_page = context.user_data.get("admin_review_page", 1)
    context.user_data["admin_review_page"] = max(1, current_page - 1)
    return await show_all_reviews(update, context)

async def next_page_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Go to next page of reviews (for paginated lists)"""
    current_page = context.user_data.get("admin_review_page", 1)
    context.user_data["admin_review_page"] = current_page + 1
    return await show_all_reviews(update, context)

async def display_admin_review_search_results(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Display paginated results from admin review search"""
    results = context.user_data.get("search_results", [])
    current_page = context.user_data.get("current_page", 1)
    per_page = 5

    start_idx = (current_page - 1) * per_page
    page_results = results[start_idx:start_idx + per_page]

    keyboard = []
    for review in page_results:
        # Safely extract data from sqlite3.Row or dict
        if isinstance(review, dict):
            review_id = review.get('id', 'N/A')
            target_name = review.get('target_name', 'Unknown')
            reviewer_name = review.get('reviewer_name', 'Anonymous')
            rating = review.get('rating', 0)
            target_type = review.get('target_type', 'unknown')
        else:
            # Handle sqlite3.Row by using dictionary-like access
            try:
                review_id = review['id']
            except Exception:
                review_id = 'N/A'

            target_name = review.get('target_name', 'Unknown') if hasattr(review, 'get') else 'Unknown'
            reviewer_name = review.get('reviewer_name', 'Anonymous') if hasattr(review, 'get') else 'Anonymous'
            rating = review.get('rating', 0) if hasattr(review, 'get') else 0
            target_type = review.get('target_type', 'unknown') if hasattr(review, 'get') else 'unknown'

        btn_text = f"Review #{review_id} | {target_name} ({target_type}) by {reviewer_name} ⭐{rating}"
        callback_data = f"view_review_{review_id}"
        keyboard.append([InlineKeyboardButton(btn_text, callback_data=callback_data)])

    # Pagination controls
    if len(results) > per_page:
        pagination_row = []
        if current_page > 1:
            pagination_row.append(InlineKeyboardButton("◀️ Prev", callback_data="prev_page"))
        if start_idx + per_page < len(results):
            pagination_row.append(InlineKeyboardButton("Next ▶️", callback_data="next_page"))
        if pagination_row:
            keyboard.append(pagination_row)

    keyboard.append([InlineKeyboardButton("🔙 Back to Search", callback_data="back_to_rate_menu")])

    text = f"🔍 Review Search Results (Page {current_page}):"

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await context.bot.send_message(
            chat_id=update.effective_user.id,
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    return ADMIN_REVIEW_SEARCH

async def handle_admin_review_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle when an admin selects a review to view/delete"""
    query = update.callback_query
    await query.answer()

    try:
        # Callback data format: view_review_123
        parts = query.data.split("_")
        if len(parts) != 3 or parts[0] != "view" or parts[1] != "review":
            raise ValueError("Invalid admin review selection")

        review_id = int(parts[2])

        # Get review details
        review = db.get_review_details(review_id)
        if not review:
            raise ValueError(f"Review ID {review_id} not found")

        # Show full review info
        target_name = review.get('target_name', 'Unknown')
        reviewer_name = review.get('reviewer_name', 'Anonymous')
        rating = review.get('rating', 0)
        comment = html.escape(str(review.get('comment', '<No comment>')))
        created_at = review.get('created_at', 'Unknown date')

        msg = (
            f"<b>⭐ Review ID:</b> {review_id}\n"
            f"<b>Reviewed:</b> {target_name}\n"
            f"<b>Type:</b> {review.get('target_type', 'unknown').capitalize()}\n"
            f"<b>By:</b> {reviewer_name}\n"
            f"<b>Rating:</b> {'⭐' * rating} ({rating}/5)\n"
            f"<b>Date:</b> {created_at}\n"
            f"<b>Comment:</b> {comment}"
        )

        keyboard = [
            [InlineKeyboardButton("🗑️ Delete Review", callback_data=f"confirm_deleted_review_{review_id}")],
            [InlineKeyboardButton("🔙 Back", callback_data="admin_back_to_ratings")]
        ]

        await query.edit_message_text(
            text=msg,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ADMIN_REVIEW_DETAILS

    except ValueError as e:
        logging.error(f"Admin review selection error: {e}")
        await query.edit_message_text("⚠️ Invalid review selected.")
        return ADMIN_RATINGS_MENU
    except Exception as e:
        logging.error(f"Unexpected error in admin review selection: {e}")
        await query.edit_message_text("❌ Failed to load review details.")
        return ADMIN_RATINGS_MENU

async def view_review_details_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle when an admin selects a review to view/delete"""
    query = update.callback_query
    await query.answer()

    try:
        # Callback data format: view_review_123
        parts = query.data.split("_")
        if len(parts) != 3 or parts[0] != "view" or parts[1] != "review":
            raise ValueError("Invalid admin review selection")

        review_id = int(parts[2])

        # Get review details
        review = db.get_review_details(review_id)
        if not review:
            raise ValueError(f"Review ID {review_id} not found")

        # Show full review info
        target_name = review.get('target_name', 'Unknown')
        reviewer_name = review.get('reviewer_name', 'Anonymous')
        rating = review.get('rating', 0)
        comment = html.escape(str(review.get('comment', '<No comment>')))
        created_at = review.get('created_at', 'Unknown date')

        msg = (
            f"<b>⭐ Review ID:</b> {review_id}\n"
            f"<b>Reviewed:</b> {target_name}\n"
            f"<b>Type:</b> {review.get('target_type', 'unknown').capitalize()}\n"
            f"<b>By:</b> {reviewer_name}\n"
            f"<b>Rating:</b> {'⭐' * rating} ({rating}/5)\n"
            f"<b>Date:</b> {created_at}\n"
            f"<b>Comment:</b> {comment}"
        )

        keyboard = [
            [InlineKeyboardButton("🗑️ Delete Review", callback_data=f"confirm_deleted_review_{review_id}")],
            [InlineKeyboardButton("🔙 Back", callback_data="admin_review_search")]
        ]

        await query.edit_message_text(
            text=msg,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ADMIN_REVIEW_DETAILS

    except ValueError as e:
        logging.error(f"Admin review selection error: {e}")
        await query.edit_message_text("⚠️ Invalid review selected.")
        return ADMIN_RATINGS_MENU
    except Exception as e:
        logging.error(f"Unexpected error in admin review selection: {e}")
        await query.edit_message_text("❌ Failed to load review details.")
        return ADMIN_RATINGS_MENU

async def confirm_delete_review_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle confirmed review deletion"""
    query = update.callback_query
    await query.answer()

    try:
        # Get review ID from context
        review_id = context.user_data.get("review_to_delete")
        if not review_id:
            raise ValueError("No review ID found in context")

        # Delete from database
        success = db.delete_review(review_id)

        if success:
            # Clear stored data
            context.user_data.pop("review_to_delete", None)
            context.user_data.pop("search_results", None)

            await query.edit_message_text(
                text="✅ Review deleted successfully.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Back to Dashboard", callback_data="admin_back_to_ratings")]
                ])
            )
        else:
            await query.edit_message_text(
                text="❌ Failed to delete review. It may not exist anymore.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Try Again", callback_data=f"confirm_deleted_review_{review_id}")]
                ])
            )

        return ADMIN_RATINGS_MENU

    except Exception as e:
        logging.error(f"Error deleting review (admin): {e}")
        await query.edit_message_text(
            "⚠️ An error occurred while trying to delete the review.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Back", callback_data="admin_back_to_ratings")]
            ])
        )
        return ADMIN_RATINGS_MENU

async def prepare_delete_review_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle admin clicking 'Delete' on a specific review"""
    query = update.callback_query
    await query.answer()

    try:
        # Callback data format: confirm_deleted_review_123
        parts = query.data.split("_")
        if len(parts) != 4 or parts[0] != "confirm" or parts[1] != "deleted" or parts[2] != "review":
            raise ValueError("Invalid delete request")

        review_id = int(parts[3])
        context.user_data["review_to_delete"] = review_id

        # Show confirmation screen
        await query.edit_message_text(
            text=f"⚠️ Are you sure you want to delete review #{review_id}?",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Confirm Delete", callback_data="confirm_delete")],
                [InlineKeyboardButton("❌ Cancel", callback_data="admin_back_to_ratings")]
            ])
        )
        return ADMIN_DELETE_REVIEW

    except Exception as e:
        logging.error(f"Admin delete preparation error: {e}")
        await query.edit_message_text("⚠️ Failed to prepare deletion.")
        return ADMIN_RATINGS_MENU


async def start_delete_review_process(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle delete button from show_all_reviews (callback: admin_delete_review_123)"""
    query = update.callback_query
    await query.answer()
    try:
        parts = query.data.split("_")
        if len(parts) != 4 or parts[0] != "admin" or parts[1] != "delete" or parts[2] != "review":
            raise ValueError("Invalid delete request format")

        review_id = int(parts[3])
        context.user_data["review_to_delete"] = review_id

        # Show confirmation screen
        await query.edit_message_text(
            text=f"⚠️ Are you sure you want to delete review #{review_id}?",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Confirm Delete", callback_data="confirm_delete")],
                [InlineKeyboardButton("❌ Cancel", callback_data="admin_back_to_ratings")]
            ])
        )
        return ADMIN_DELETE_REVIEW
    except Exception as e:
        logging.error(f"Error starting delete review: {e}")
        await query.edit_message_text("⚠️ Failed to start deletion.")
        return ADMIN_RATINGS_MENU

async def process_review_id_for_deletion(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Process the review ID entered by the admin"""
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if text == "🔙 Cancel":
        return await ratings_dashboard(update, context)

    try:
        review_id = int(text)
        context.user_data["review_to_delete"] = review_id

        # Fetch the review details
        review = db.get_review_details(review_id)
        if not review:
            await context.bot.send_message(
                chat_id=user_id,
                text="❌ No review found with that ID.",
                reply_markup=ReplyKeyboardMarkup(
                    [[KeyboardButton("🔙 Back")]],
                    one_time_keyboard=True,
                    resize_keyboard=True
                )
            )
            return ADMIN_DELETE_REVIEW

        # Safely extract and escape all values
        target_name = db.get_user_name(review.get('target_id', 'N/A')) if review.get('target_type') != 'bot' else "JobBot"
        reviewer_name = db.get_user_name(review.get('reviewer_id', 'N/A'))
        rating_stars = '⭐' * review.get('rating', 0)
        comment = html.escape(str(review.get('comment', '<No comment>')))  # Escape HTML
        target_type = html.escape(str(review.get('target_type', 'unknown')))
        review_id_escaped = html.escape(str(review.get('id', 'N/A')))
        target_name_escaped = html.escape(str(target_name))
        reviewer_name_escaped = html.escape(str(reviewer_name))
        rating_stars_escaped = html.escape(str(rating_stars))

        confirm_text = (
            f"⚠️ Are you sure you want to delete this review?\n\n"
            f"<b>Review ID:</b> {review_id_escaped}\n"
            f"<b>Reviewed:</b> {target_name_escaped} ({target_type.capitalize()})\n"
            f"<b>By:</b> {reviewer_name_escaped}\n"
            f"<b>Rating:</b> {rating_stars_escaped}\n"
            f"<b>Comment:</b> {comment}"
        )

        keyboard = [
            [InlineKeyboardButton("✅ Confirm Delete", callback_data="confirm_deleted")],
            [InlineKeyboardButton("❌ Cancel", callback_data="admin_back_to_ratings")]
        ]

        await context.bot.send_message(
            chat_id=user_id,
            text=confirm_text,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ADMIN_DELETE_REVIEW

    except ValueError:
        await context.bot.send_message(
            chat_id=user_id,
            text="⚠️ Please enter a valid numeric Review ID.",
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton("🔙 Back")]],
                one_time_keyboard=True,
                resize_keyboard=True
            )
        )
        return ADMIN_DELETE_REVIEW

async def edit_review_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Allow admin to edit an existing review"""
    query = update.callback_query
    await query.answer()
    review_id = int(query.data.split("_")[-1])

    review = db.get_review_details(review_id)
    if not review:
        await query.edit_message_text("Review not found.")
        return await ratings_dashboard(update, context)

    context.user_data.update({
        "editing_review": True,
        "current_review_id": review_id,
        "current_review_target": review["target_id"],
        "target_type": review["target_type"]
    })

    # Build fake dimension data for editing
    rating_dimensions = {
        'professionalism': 'Professionalism',
        'communication': 'Communication',
        'hiring_process': 'Hiring Process'
    }
    context.user_data["rating_dimensions"] = rating_dimensions

    first_dim = next(iter(rating_dimensions))
    return await show_dimension_rating(update, context, first_dim)
async def back_to_ratings_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await ratings_dashboard(update, context)

async def back_to_user_interactions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await show_user_interactions_menu(update, context)

async def handle_review_search_query(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Process the search query entered by admin"""
    query = update.message.text.strip()
    if query == "🔙 Cancel":
        return await ratings_dashboard(update, context)

    results = db.search_reviews_admin(search_term=query)
    context.user_data["search_results"] = results
    context.user_data["current_search_page"] = 1

    if not results:
        await context.bot.send_message(
            chat_id=update.effective_user.id,
            text="🔍 No reviews found matching your query.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Try Again", callback_data="admin_search_reviews")],
                [InlineKeyboardButton("Back", callback_data="admin_back_to_ratings")]
            ])
        )
        return ADMIN_REVIEW_SEARCH

    return await display_admin_review_search_results(update, context)



async def apply_filter_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Apply filter to review list (employer/job_seeker)"""
    query = update.callback_query
    await query.answer()
    target_type = query.data.split("_")[-1]  # employer / job_seeker
    context.user_data["review_filter"] = target_type

    results = db.search_reviews(target_type=target_type)
    context.user_data["search_results"] = results
    context.user_data["current_search_page"] = 1

    return await display_search_results_page(update, context)

async def apply_sort_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Sort reviews by recent/top rated"""
    query = update.callback_query
    await query.answer()
    sort_method = query.data.split("_")[-1]  # recent / top_rated
    context.user_data["review_sort"] = sort_method

    results = context.user_data.get("search_results", [])
    if sort_method == "recent":
        results.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    elif sort_method == "top_rated":
        results.sort(key=lambda x: x.get("rating", 0), reverse=True)

    context.user_data["search_results"] = results
    context.user_data["current_search_page"] = 1

    return await display_search_results_page(update, context)


async def cancel_admin_rating_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel any ongoing admin action and go back to dashboard"""
    return await ratings_dashboard(update, context)

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

async def review_statistics_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show aggregated statistics about reviews in the system"""
    user_id = update.effective_user.id

    total_reviews = db.get_total_review_count()
    avg_rating = db.get_overall_average_rating()

    # Get top reviewed targets
    top_employers = db.get_top_reviewed_targets("employer")
    top_seekers = db.get_top_reviewed_targets("job_seeker")

    stats_text = f"""<b>📊 Review Statistics</b>
• Total Reviews: <b>{total_reviews}</b>
• Average Rating: <b>{avg_rating:.1f}⭐</b>

<b>Top Reviewed Employers:</b>"""

    if top_employers:
        for emp in top_employers:
            stats_text += f"\n- {emp['target_name']} ({emp['review_count']} reviews)"
    else:
        stats_text += "\nNo employer reviews yet."

    stats_text += "\n\n<b>Top Reviewed Job Seekers:</b>"

    if top_seekers:
        for js in top_seekers:
            stats_text += f"\n- {js['target_name']} ({js['review_count']} reviews)"
    else:
        stats_text += "\nNo job seeker reviews yet."

    stats_text += "\n\nUse the back button below to return to the dashboard."

    reply_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔙 Back", callback_data="admin_back_to_ratings")]
    ])

    await context.bot.send_message(
        chat_id=user_id,
        text=stats_text,
        parse_mode="HTML",
        reply_markup=reply_markup
    )

    return ADMIN_REVIEW_STATISTICS

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes



# Report Feature

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, CallbackQueryHandler, MessageHandler, filters
import logging
from typing import Dict, List, Optional

# Report Feature Constants
REPORT_REASONS = {
    "violation": "violation_reason",
    "spam": "spam_reason",
    "inappropriate": "inappropriate_content_reason",
    "fraud": "fraud_reason",
    "other": "other_reason"
}


async def handle_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Initiates the report process by presenting reporting options.
    Now with proper translation key handling and fallbacks.
    """
    user_id = get_user_id(update)

    # Clear any previous report data
    context.user_data.pop("report_data", None)

    # Get translations with fallback values
    title = get_translation(user_id, "select_entity_to_report",
                            fallback="⚠️ What would you like to report?")
    guidelines = get_translation(user_id, "report_guidelines",
                                 fallback="Please select the type of entity you wish to report.")

    # Button texts with fallbacks
    job_seeker_text = get_translation(user_id, "report_job_seeker",
                                      fallback="Report Job Seeker")
    employer_text = get_translation(user_id, "report_employer",
                                    fallback="Report Employer")
    application_text = get_translation(user_id, "report_application",
                                       fallback="Report Application")
    vacancy_text = get_translation(user_id, "report_vacancy",
                                   fallback="Report Job Posting")
    cancel_text = get_translation(user_id, "cancel_button",
                                  fallback="✖ Cancel")

    # Organized reporting options
    keyboard = [
        [
            InlineKeyboardButton(job_seeker_text, callback_data="report_job_seeker"),
            InlineKeyboardButton(employer_text, callback_data="report_employer")
        ],
        [
            InlineKeyboardButton(application_text, callback_data="report_application"),
            InlineKeyboardButton(vacancy_text, callback_data="report_vacancy")
        ],
        [
            InlineKeyboardButton(cancel_text, callback_data="back_to_main_menu")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Build message text
    message_text = f"{title}\n\n{guidelines}"

    # Send message with error handling
    try:
        await update.message.reply_text(
            text=message_text,
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )
    except BadRequest:
        # Fallback to plain text if Markdown fails
        await update.message.reply_text(
            text=message_text,
            reply_markup=reply_markup,
            parse_mode=None
        )

    return SELECT_REPORT_ENTITY


async def select_report_entity(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Handles the selection of entity type to report.
    Improved with better context management and user feedback.
    """
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    # Handle entity type assignment explicitly
    if query.data == "report_job_seeker":
        entity_type = "job_seeker"
    else:
        # Extract entity type from callback data
        try:
            entity_type = query.data.split("_")[1]
        except IndexError:
            logging.error(f"Invalid callback data format for user {user_id}: {query.data}")
            await handle_report_error(update, context, user_id)
            return await back_to_main_menu_report(update, context)

    # Validate entity type
    valid_entity_types = ["job_seeker", "employer", "application", "vacancy"]
    if entity_type not in valid_entity_types:
        logging.error(f"Unsupported entity type: {entity_type} for user {user_id}")
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)

    # Store entity type in context
    context.user_data["report_data"] = {
        "entity_type": entity_type,
        "step": "entity_selected"
    }

    # Provide clear instructions for the next step
    instruction_text = {
        "job_seeker": get_translation(user_id, "enter_job_seeker_name_instruction"),
        "employer": get_translation(user_id, "enter_employer_name_instruction"),
        "application": get_translation(user_id, "enter_application_reference_instruction"),
        "vacancy": get_translation(user_id, "enter_vacancy_reference_instruction")
    }.get(entity_type, get_translation(user_id, "enter_entity_name_generic"))

    await query.edit_message_text(
        text=f"🔍 {get_translation(user_id, 'search_entity_to_report')}\n"
             f"{instruction_text}\n"
             f"*{get_translation(user_id, 'search_tip')}*",
        reply_markup=None,
        parse_mode="Markdown"
    )
    return SEARCH_REPORT_ENTITY


async def search_report_entity(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Searches for entities based on user input with enhanced error handling and pagination.
    """
    user_id = get_user_id(update)
    search_term = update.message.text.strip()
    report_data = context.user_data.get("report_data", {})
    entity_type = report_data.get("entity_type")

    if not entity_type:
        logging.error(f"Missing entity_type in context for user {user_id}")
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)



    # Store search term for pagination
    context.user_data["report_data"]["search_term"] = search_term
    context.user_data["report_data"]["current_page"] = 1

    try:
        # Get search results
        search_results = await perform_entity_search(
            entity_type,
            search_term,
            page=1,
            user_id=user_id
        )

        if not search_results["items"]:
            # No results found - offer to try again
            keyboard = [
                [InlineKeyboardButton(get_translation(user_id, "try_again_button"),
                                      callback_data=f"report_{entity_type}")],
                [InlineKeyboardButton(get_translation(user_id, "cancel_button"), callback_data="back_to_main_menu")]
            ]

            await update.message.reply_text(
                text=f"❌ {get_translation(user_id, 'no_results_found', search_term=search_term)}\n\n"
                     f"{get_translation(user_id, 'search_try_again_advice')}",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return SELECT_REPORT_ENTITY

            # Extract count and total for dynamic calculation
        count = len(search_results["items"])
        total = search_results["total_items"]

        # Display search results with pagination
        keyboard = create_search_results_keyboard(
            search_results["items"],
            entity_type,
            current_page=1,
            total_pages=search_results["total_pages"],
            user_id=user_id
        )

        await update.message.reply_text(
            text=f"🔎 {get_translation(user_id, 'search_results_header')}\n\n"
                 f"*{count} of {total} results*\n\n"
                 f"{get_translation(user_id, 'select_from_results_below')}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )

        return SELECT_REPORT_RESULT

    except Exception as e:
        logging.error(f"Error searching {entity_type} for user {user_id}: {str(e)}")
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)

async def perform_entity_search(entity_type: str, search_term: str, page: int, user_id: int) -> Dict:
    """
    Centralized entity search function with standardized return format.
    """
    page_size = 5  # Number of items per page

    try:
        if entity_type == "job_seeker":
            items = db.search_job_seekers_report(search_term, page=page, page_size=page_size)
            total_items = db.get_total_pages_job_seekers(search_term)
        elif entity_type == "employer":
            items = db.search_employers_report(search_term, page=page, page_size=page_size)
            total_items = db.get_total_pages_employers(search_term)
        elif entity_type == "application":
            items = db.search_applications(search_term, page=page, page_size=page_size)
            total_items = db.get_total_pages_applications(search_term)
        elif entity_type == "vacancy":
            items = db.search_vacancies_report(search_term, page=page, page_size=page_size)
            total_items = db.get_total_pages_vacancies(search_term)
        else:
            raise ValueError(f"Unsupported entity type: {entity_type}")

        total_pages = (total_items + page_size - 1) // page_size  # Calculate total pages

        return {
            "items": items,
            "total_items": total_items,
            "total_pages": total_pages,
            "current_page": page
        }
    except Exception as e:
        logging.error(f"Error in perform_entity_search for {entity_type}: {str(e)}")
        raise

def create_search_results_keyboard(items: List, entity_type: str, current_page: int,
                                   total_pages: int, user_id: int) -> List[List[InlineKeyboardButton]]:
    """
    Creates a paginated keyboard for search results with improved layout.
    """
    keyboard = []

    # Add items as buttons
    for item in items:
        # Convert to dict if it's a Row object
        if hasattr(item, 'keys'):  # It's a SQLite Row
            item = dict(item)

        if entity_type == "job_seeker":
            full_name = item.get('full_name') if isinstance(item, dict) else item['full_name']
            user_id_val = item.get('user_id') if isinstance(item, dict) else item['user_id']
            text = f"👤 {full_name or 'N/A'} (ID: {user_id_val or '?'})"
            callback_data = f"report_seeker_{user_id_val}"
        elif entity_type == "employer":
            company_name = item.get('company_name') if isinstance(item, dict) else item['company_name']
            employer_id = item.get('employer_id') if isinstance(item, dict) else item['employer_id']
            text = f"🏢 {company_name or 'N/A'} (ID: {employer_id or '?'})"
            callback_data = f"report_employer_{employer_id}"
        elif entity_type == "application":
            job_title = item.get('job_title') if isinstance(item, dict) else item['job_title']
            full_name = item.get('full_name') if isinstance(item, dict) else item['full_name']
            app_id = item.get('application_id') if isinstance(item, dict) else item['application_id']
            text = f"📄 {job_title or 'N/A'} - {full_name or 'N/A'}"
            callback_data = f"report_application_{app_id}"
        elif entity_type == "vacancy":
            job_title = item.get('job_title') if isinstance(item, dict) else item['job_title']
            vacancy_id = item.get('id') if isinstance(item, dict) else item['id']
            text = f"📌 {job_title or 'N/A'} (ID: {vacancy_id or '?'})"
            callback_data = f"report_vacancy_{vacancy_id}"
        else:
            continue

        keyboard.append([InlineKeyboardButton(text, callback_data=callback_data)])
    # Add pagination controls if needed
    if total_pages > 1:
        pagination_row = []

        if current_page > 1:
            pagination_row.append(
                InlineKeyboardButton("⬅️ " + get_translation(user_id, "prev_button"),
                                     callback_data=f"prev_report_{entity_type}_{current_page}")
            )

        if current_page < total_pages:
            pagination_row.append(
                InlineKeyboardButton(get_translation(user_id, "next_button") + " ➡️",
                                     callback_data=f"next_report_{entity_type}_{current_page}")
            )

        if pagination_row:
            keyboard.append(pagination_row)

    # Add navigation buttons
    keyboard.append([
        InlineKeyboardButton(get_translation(user_id, "new_search_button"),
                             callback_data=f"report_{entity_type}"),
        InlineKeyboardButton(get_translation(user_id, "cancel_button"),
                             callback_data="back_to_report_menu")
    ])

    return keyboard

async def handle_pagination_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Handles pagination for search results with smooth transitions.
    """
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    try:
        # Parse callback data
        action, entity_type, current_page = query.data.split("_")[0:3]
        current_page = int(current_page)
        new_page = current_page + 1 if action == "next" else current_page - 1

        # Get search term from context
        search_term = context.user_data.get("report_data", {}).get("search_term", "")

        if not search_term:
            raise ValueError("Missing search term in context")

        # Perform search for the new page
        search_results = await perform_entity_search(
            entity_type,
            search_term,
            page=new_page,
            user_id=user_id
        )

        # Update context with current page
        context.user_data["report_data"]["current_page"] = new_page

        # Extract count and total for dynamic calculation
        count = len(search_results["items"])
        total = search_results["total_items"]

        # Create updated keyboard
        keyboard = create_search_results_keyboard(
            search_results["items"],
            entity_type,
            current_page=new_page,
            total_pages=search_results["total_pages"],
            user_id=user_id
        )

        # Update message with new results
        await query.edit_message_text(
            text=f"🔎 {get_translation(user_id, 'search_results_header')}\n\n"
                 f"*{count} of {total} results*\n\n"
                 f"{get_translation(user_id, 'select_from_results_below')}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )

        return SELECT_REPORT_RESULT

    except Exception as e:
        logging.error(f"Error handling pagination for user {user_id}: {str(e)}")
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)


async def select_report_result(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Handles selection of a specific entity to report with detailed confirmation.
    """
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)

    try:
        # Parse callback data to get entity type and ID
        parts = query.data.split("_")
        if len(parts) < 3:
            raise ValueError("Invalid callback data format")

        action = parts[0]  # "report"
        entity_type_part = parts[1]  # "seeker", "employer", "application", or "vacancy"
        entity_id = parts[2]

        # Map to actual entity types used in DB
        entity_type_map = {
            "seeker": "job_seeker",
            "employer": "employer",
            "application": "application",
            "vacancy": "vacancy"
        }
        entity_type = entity_type_map.get(entity_type_part)

        if not entity_type:
            raise ValueError(f"Unknown entity type part: {entity_type_part}")

        # Get entity details from DB
        entity = None
        if entity_type == "job_seeker":
            entity = db.get_job_seeker_by_id(entity_id)
        elif entity_type == "employer":
            entity = db.get_employer_by_id(entity_id)
        elif entity_type == "application":
            entity = db.get_application_by_id(entity_id)
        elif entity_type == "vacancy":
            entity = db.get_vacancy_by_id_report(entity_id)

        if not entity:
            raise ValueError(f"Entity not found: {entity_type} ID {entity_id}")

        # Convert to dict if it's a Row object
        if hasattr(entity, 'keys'):
            entity = dict(entity)

        # Prepare entity details text based on type
        entity_details = ""
        display_name = "Unknown"

        if entity_type == "job_seeker":
            display_name = entity.get('full_name', 'Unknown Job Seeker')
            entity_details = (
                f"👤 Job Seeker: {display_name}\n"
                f"🆔 ID: {entity.get('user_id', 'N/A')}"
            )
        elif entity_type == "employer":
            display_name = entity.get('company_name', 'Unknown Employer')
            entity_details = (
                f"🏢 Employer: {display_name}\n"
                f"🆔 ID: {entity.get('employer_id', 'N/A')}"
            )
        elif entity_type == "application":
            job_title = entity.get('job_title', 'Unknown Position')
            full_name = entity.get('full_name', 'Unknown Applicant')
            display_name = f"{full_name} → {job_title}"
            entity_details = (
                f"📄 Application for: {job_title}\n"
                f"👤 Applicant: {full_name}\n"
                f"🆔 Application ID: {entity.get('application_id', 'N/A')}"
            )
        elif entity_type == "vacancy":
            display_name = entity.get('job_title', 'Unknown Vacancy')
            entity_details = (
                f"📌 Vacancy: {display_name}\n"
                f"🆔 ID: {entity.get('id', 'N/A')}\n"
                f"🏢 Company ID: {entity.get('employer_id', 'N/A')}"
            )

        # Store entity info in context
        context.user_data["report_data"] = {
            "entity_id": entity_id,
            "entity_name": display_name,
            "entity_details": entity,
            "entity_type": entity_type
        }

        # Create reason selection keyboard
        keyboard = [
            [InlineKeyboardButton(get_translation(user_id, "violation_reason"), callback_data="reason_violation")],
            [InlineKeyboardButton(get_translation(user_id, "spam_reason"), callback_data="reason_spam")],
            [InlineKeyboardButton(get_translation(user_id, "inappropriate_content_reason"),
                                  callback_data="reason_inappropriate")],
            [InlineKeyboardButton(get_translation(user_id, "fraud_reason"), callback_data="reason_fraud")],
            [InlineKeyboardButton(get_translation(user_id, "other_reason"), callback_data="reason_other")],
            [InlineKeyboardButton(get_translation(user_id, "cancel_button"), callback_data="reason_cancel")]
        ]

        # Show entity summary and request reason
        await query.edit_message_text(
            text=f"⚠️ {get_translation(user_id, 'confirm_report_entity')}\n\n"
                 f"*{get_translation(user_id, 'entity_details')}:*\n"
                 f"{entity_details}\n\n"
                 f"🔹 *{get_translation(user_id, 'select_report_reason')}*",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )

        return CONFIRM_REPORT_DETAILS

    except Exception as e:
        logging.error(f"Error selecting report result for user {user_id}: {str(e)}", exc_info=True)
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)


async def get_entity_details(entity_type: str, entity_id: str, user_id: int) -> Optional[Dict]:
    """
    Retrieves detailed information about an entity for reporting purposes.
    Now handles both dictionary and SQLite Row objects.
    """
    try:
        entity = None
        details = {}

        if entity_type == "job_seeker":
            entity = db.get_job_seeker_by_id(entity_id)
            if entity:
                if hasattr(entity, 'keys'):  # It's a SQLite Row
                    entity = dict(entity)
                details = {
                    "display_name": entity.get("full_name", f"Job Seeker {entity_id}"),
                    "details": f"👤 {entity.get('full_name', 'N/A')}\n"
                               f"🆔 ID: {entity_id}"

                }

        elif entity_type == "employer":
            entity = db.get_employer_by_id(entity_id)
            if entity:
                if hasattr(entity, 'keys'):  # It's a SQLite Row
                    entity = dict(entity)
                details = {
                    "display_name": entity.get("company_name", f"Employer {entity_id}"),
                    "details": f"🏢 {entity.get('company_name', 'N/A')}\n"
                               f"🆔 ID: {entity_id}"

                }

        elif entity_type == "application":
            entity = db.get_application_by_id(entity_id)
            if entity:
                if hasattr(entity, 'keys'):  # It's a SQLite Row
                    entity = dict(entity)
                details = {
                    "display_name": f"Application {entity_id}",
                    "details": f"📄 Application ID: {entity_id}\n"
                               f"📌 Job: {entity.get('job_title', 'N/A')}\n"
                               f"👤 Candidate: {entity.get('full_name', 'N/A')}\n"
                               f"📅 Date: {entity.get('application_date', 'Unknown')}"
                }

        elif entity_type == "vacancy":
            entity = db.get_vacancy_by_id_report(entity_id)
            if entity:
                if hasattr(entity, 'keys'):  # It's a SQLite Row
                    entity = dict(entity)
                details = {
                    "display_name": entity.get("job_title", f"Vacancy {entity_id}"),
                    "details": f"📌 {entity.get('job_title', 'N/A')}\n"
                               f"🆔 ID: {entity_id}\n"
                               f"🏢 Company: {entity.get('company_name', 'N/A')}"

                }

        return details if details else None

    except Exception as e:
        logging.error(f"Error getting details for {entity_type} {entity_id}: {str(e)}")
        return None


async def submit_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Handles report reason selection and initiates report submission.
    """
    query = update.callback_query
    await query.answer()
    user_id = get_user_id(update)
    reason = query.data.split("_")[1]

    if reason == "cancel":
        await query.edit_message_text(
            text=f"✅ {get_translation(user_id, 'report_canceled')}",
            reply_markup=None
        )
        return await back_to_main_menu_report(update, context)

    # Store reason in context
    context.user_data["report_data"]["reason"] = reason

    if reason == "other":
        # Request additional information for "other" reason
        await query.edit_message_text(
            text=f"✍️ {get_translation(user_id, 'provide_additional_info_header')}\n\n"
                 f"{get_translation(user_id, 'provide_additional_info_instructions')}\n\n"
                 f"*{get_translation(user_id, 'max_500_chars_warning')}*",
            reply_markup=None,
            parse_mode="Markdown"
        )
        return SUBMIT_REPORT

    # For predefined reasons, proceed to finalize
    return await finalize_report(update, context)

async def provide_additional_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Handles additional information provided by the user for the report.
    """
    user_id = get_user_id(update)
    additional_info = update.message.text.strip()

    # Validate length
    if len(additional_info) > 500:
        await update.message.reply_text(
            text=f"❌ {get_translation(user_id, 'additional_info_too_long')}\n\n"
                 f"{get_translation(user_id, 'please_shorten_message')}",
            reply_markup=None
        )
        return SUBMIT_REPORT

    # Store additional info
    context.user_data["report_data"]["additional_info"] = additional_info

    # Proceed to finalize
    return await finalize_report(update, context)

async def finalize_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Finalizes the report submission with comprehensive validation and feedback.
    """
    user_id = get_user_id(update)
    report_data = context.user_data.get("report_data", {})

    # Validate required data
    required_fields = ["entity_type", "entity_id", "reason"]
    missing_fields = [field for field in required_fields if field not in report_data]

    if missing_fields:
        logging.error(f"Missing report data for user {user_id}: {missing_fields}")
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)

    try:
        # Prepare report data - now compatible with both approaches
        report = {
            "reporter_id": user_id,
            "reported_entity_type": report_data["entity_type"],
            "reported_entity_id": report_data["entity_id"],
            "reason": report_data["reason"],
            "additional_info": report_data.get("additional_info", ""),
            "status": "pending",
            "timestamp": datetime.now().isoformat()
        }

        # Save to database - now using the dictionary approach
        report_id = db.insert_report(report)
        if not report_id:
            raise ValueError("Failed to save report to database")

            # Add the report ID to the report dictionary
        report['id'] = report_id

        if not report_id:
            raise ValueError("Failed to save report to database")

        # Notify admins
        await notify_admins_about_report(context, report, user_id)

        # Send confirmation to user
        confirmation_text = (
            f"✅ *{get_translation(user_id, 'report_submitted_successfully')}*\n\n"
            f"🆔 Report ID: `{report_id}`\n"
            f"📌 Entity: {report_data.get('entity_name', 'Unknown')}\n"
            f"📝 Reason: {get_translation(user_id, REPORT_REASONS.get(report['reason'], report['reason']))}\n\n"
            f"{get_translation(user_id, 'report_follow_up_info')}"
        )

        await send_message_to_user(
            update,
            context,
            user_id,
            confirmation_text,
            parse_mode="Markdown"
        )

        # Clear report data from context
        context.user_data.pop("report_data", None)

        return await back_to_main_menu_report(update, context)

    except Exception as e:
        logging.error(f"Error finalizing report for user {user_id}: {str(e)}")
        await handle_report_error(update, context, user_id)
        return await back_to_main_menu_report(update, context)

async def notify_admins_about_report(context: ContextTypes.DEFAULT_TYPE, report: Dict, reporter_id: int):
    """
    Notifies all active admins about a new report with detailed information.
    """
    try:
        # Get reporter info
        reporter = db.get_user_profile(reporter_id)
        reporter_name = reporter.get("full_name", f"User {reporter_id}")

        def escaped_markdown(text):
            if not isinstance(text, str):
                return text
            return text.replace('*', '\\*').replace('_', '\\_').replace('`', '\\`')

        # Get entity details directly (bypassing get_entity_details)
        entity = None
        if report["reported_entity_type"] == "job_seeker":
            entity = db.get_job_seeker_by_id(report["reported_entity_id"])
            if entity:
                if hasattr(entity, 'keys'):  # Convert Row to dict if needed
                    entity = dict(entity)

                # Format date of birth if available
                dob = entity.get('dob', 'N/A')
                if dob and dob != 'N/A':
                    dob = dob.strftime('%Y-%m-%d') if hasattr(dob, 'strftime') else dob

                # Format creation date
                created_at = entity.get('created_at', 'Unknown')
                if created_at and created_at != 'Unknown':
                    created_at = created_at.strftime('%Y-%m-%d %H:%M') if hasattr(created_at,
                                                                                  'strftime') else created_at

                # Escape any Markdown special characters in text fields
                def escaped_markdown(text):
                    if not isinstance(text, str):
                        return text
                    return text.replace('*', '\\*').replace('_', '\\_').replace('`', '\\`')

                # Build the detailed profile with escaped Markdown
                details = (
                    "🔹 *Job Seeker Profile Report*\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    "👤 *Basic Information*\n"
                    f"• Full Name: {escaped_markdown(entity.get('full_name', 'N/A'))}\n"
                    f"• ID: {report['reported_entity_id']}\n"
                    f"• Contact: {escaped_markdown(entity.get('contact_number', 'N/A'))}\n"
                    f"• Date of Birth: {dob}\n"
                    f"• Gender: {escaped_markdown(entity.get('gender', 'N/A'))}\n\n"

                    "🎓 *Education*\n"
                    f"• Qualification: {escaped_markdown(entity.get('qualification', 'N/A'))}\n"
                    f"• Field of Study: {escaped_markdown(entity.get('field_of_study', 'N/A'))}\n"
                    f"• CGPA: {entity.get('cgpa', 'N/A')}\n\n"

                    "💼 *Professional Details*\n"
                    f"• Skills & Experience:\n{escaped_markdown(entity.get('skills_experience', 'N/A'))}\n"
                    f"• Profile Summary:\n{escaped_markdown(entity.get('profile_summary', 'N/A'))}\n\n"

                    "🌐 *Additional Information*\n"
                    f"• Languages: {escaped_markdown(entity.get('languages', 'N/A'))}\n"
                    f"• CV: {'Available' if entity.get('cv_path') else 'Not provided'}\n"
                    f"• Portfolio: {escaped_markdown(entity.get('portfolio_link', 'N/A'))}\n\n"

                    "📅 *Account Information*\n"
                    f"• Member since: {created_at}\n"
                    f"• Registration Type: {escaped_markdown(entity.get('registration_type', 'N/A'))}\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
                )
        elif report["reported_entity_type"] == "employer":
            entity = db.get_employer_by_id(report["reported_entity_id"])
            if entity:
                if hasattr(entity, 'keys'):
                    entity = dict(entity)

                created_at = entity.get('created_at', 'Unknown')
                if created_at and created_at != 'Unknown':
                    created_at = created_at.strftime('%Y-%m-%d %H:%M') if hasattr(created_at,
                                                                                  'strftime') else created_at

                details = (
                    "🔹 *Employer Profile Report*\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    "🏢 *Company Information*\n"
                    f"• Company Name: {escaped_markdown(entity.get('company_name', 'N/A'))}\n"
                    f"• ID: {report['reported_entity_id']}\n"
                    f"• Employer Type: {escaped_markdown(entity.get('employer_type', 'N/A'))}\n"
                    f"• Location: {escaped_markdown(entity.get('city', 'N/A'))}\n"
                    f"• Contact: {escaped_markdown(entity.get('contact_number', 'N/A'))}\n\n"

                    "📝 *About Company*\n"
                    f"{escaped_markdown(entity.get('about_company', 'No description provided'))}\n\n"

                    "📑 *Verification*\n"
                    f"• Documents: {'Provided' if entity.get('verification_docs') else 'Not provided'}\n\n"

                    "📅 *Account Information*\n"
                    f"• Member since: {created_at}\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
                )

        elif report["reported_entity_type"] == "application":
            entity = db.get_application_by_id(report["reported_entity_id"])
            if entity:
                if hasattr(entity, 'keys'):
                    entity = dict(entity)

                app_date = entity.get('application_date', 'Unknown')
                if app_date and app_date != 'Unknown':
                    app_date = app_date.strftime('%Y-%m-%d') if hasattr(app_date, 'strftime') else app_date

                deadline = entity.get('application_deadline', 'N/A')
                if deadline and deadline != 'N/A':
                    deadline = deadline.strftime('%Y-%m-%d') if hasattr(deadline, 'strftime') else deadline

                details = (
                    "🔹 *Application Report*\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    "📄 *Application Details*\n"
                    f"• Application ID: {report['reported_entity_id']}\n"
                    f"• Status: {escaped_markdown(entity.get('status', 'N/A'))}\n"
                    f"• Date: {app_date}\n"
                    f"• Additional Docs: {'Attached' if entity.get('additional_docs') else 'None'}\n\n"

                    "👨‍💼 *Candidate Information*\n"
                    f"• Name: {escaped_markdown(entity.get('full_name', 'N/A'))}\n"
                    f"• Contact: {escaped_markdown(entity.get('job_seeker_contact', 'N/A'))}\n\n"

                    "💼 *Job Details*\n"
                    f"• Title: {escaped_markdown(entity.get('job_title', 'N/A'))}\n"
                    f"• Company: {escaped_markdown(entity.get('employer_name', 'N/A'))}\n"
                    f"• Location: {escaped_markdown(entity.get('employer_city', 'N/A'))}\n"
                    f"• Type: {escaped_markdown(entity.get('employment_type', 'N/A'))}\n"
                    f"• Level: {escaped_markdown(entity.get('level', 'N/A'))}\n"
                    f"• Deadline: {deadline}\n\n"

                    "📝 *Cover Letter*\n"
                    f"{escaped_markdown(entity.get('cover_letter', 'No cover letter provided'))}\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
                )

        elif report["reported_entity_type"] == "vacancy":
            entity = db.get_vacancy_by_id_report(report["reported_entity_id"])
            if entity:
                if hasattr(entity, 'keys'):
                    entity = dict(entity)

                deadline = entity.get('application_deadline', 'N/A')
                if deadline and deadline != 'N/A':
                    deadline = deadline.strftime('%Y-%m-%d') if hasattr(deadline, 'strftime') else deadline

                details = (
                    "🔹 *Vacancy Report*\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    "📌 *Position Details*\n"
                    f"• Job Title: {escaped_markdown(entity.get('job_title', 'N/A'))}\n"
                    f"• Vacancy ID: {report['reported_entity_id']}\n"
                    f"• Status: {escaped_markdown(entity.get('status', 'N/A'))}\n"
                    f"• Applications: {entity.get('application_count', 0)}\n"
                    f"• Quantity Needed: {entity.get('quantity', 'N/A')}\n"
                    f"• Employment Type: {escaped_markdown(entity.get('employment_type', 'N/A'))}\n"
                    f"• Level: {escaped_markdown(entity.get('level', 'N/A'))}\n"
                    f"• Gender: {escaped_markdown(entity.get('gender', 'Any'))}\n"
                    f"• Deadline: {deadline}\n\n"

                    "🏢 *Company Information*\n"
                    f"• Company: {escaped_markdown(entity.get('company_name', 'N/A'))}\n"
                    f"• Location: {escaped_markdown(entity.get('city', 'N/A'))}\n"
                    f"• Contact: {escaped_markdown(entity.get('employer_contact', 'N/A'))}\n\n"

                    "💰 *Compensation*\n"
                    f"• Salary: {escaped_markdown(entity.get('salary', 'Not specified'))}\n"
                    f"• Benefits:\n{escaped_markdown(entity.get('benefits', 'None specified'))}\n\n"

                    "📝 *Requirements*\n"
                    f"• Qualification: {escaped_markdown(entity.get('qualification', 'N/A'))}\n"
                    f"• Skills:\n{escaped_markdown(entity.get('skills', 'N/A'))}\n\n"

                    "📄 *Job Description*\n"
                    f"{escaped_markdown(entity.get('description', 'No description provided'))}\n"
                    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
                )

        # Prepare admin notification message
        message = (
            f"🚨 *New Report Received* 🚨\n\n"
            f"📌 *Report ID:* `{report['id']}`\n"  # Now using the actual report ID
            f"👤 *Reporter:* {reporter_name} (ID: {reporter_id})\n"
            f"🔄 *Entity Type:* {report['reported_entity_type'].replace('_', ' ').title()}\n"
            f"🆔 *Entity ID:* {report['reported_entity_id']}\n"
            f"📝 *Reason:* {REPORT_REASONS.get(report['reason'], report['reason']).replace('_', ' ').title()}\n\n"
            f"*Entity Details:*\n"
            f"{details if entity else 'Not available'}\n\n"
            f"*Additional Info:*\n"
            f"{report['additional_info'] or 'None provided'}\n\n"
            f"⏱️ *Timestamp:* {report['timestamp']}"
        )

        # Send to all active admins
        for admin_id in active_admins:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=message,
                    parse_mode="Markdown"
                )
            except Exception as e:
                logging.error(f"Failed to notify admin {admin_id}: {str(e)}")

    except Exception as e:
        logging.error(f"Error in notify_admins_about_report: {str(e)}")
async def handle_report_error(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int):
    """
    Handles errors during the report process with user-friendly messages.
    """
    error_message = (
        f"⚠️ *{get_translation(user_id, 'error_occurred_title')}*\n\n"
        f"{get_translation(user_id, 'report_error_message')}\n\n"
        f"{get_translation(user_id, 'try_again_later_or_contact_support')}"
    )

    await send_message_to_user(
        update,
        context,
        user_id,
        error_message,
        parse_mode="Markdown"
    )

async def send_message_to_user(update: Update, context: ContextTypes.DEFAULT_TYPE,
                               user_id: int, text: str, **kwargs):
    """
    Enhanced message sender that handles both callback queries and direct messages.
    """
    try:
        if update.callback_query:
            await update.callback_query.message.reply_text(text, **kwargs)
        else:
            await context.bot.send_message(chat_id=user_id, text=text, **kwargs)
    except Exception as e:
        logging.error(f"Failed to send message to user {user_id}: {str(e)}")
        # Fallback to simple message if formatting fails
        try:
            if update.callback_query:
                await update.callback_query.message.reply_text(text)
            else:
                await context.bot.send_message(chat_id=user_id, text=text)
        except Exception as e2:
            logging.error(f"Fallback message also failed for user {user_id}: {str(e2)}")

async def back_to_main_menu_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Enhanced back to menu function with user-specific redirection.
    """
    user_id = get_user_id(update)

    try:
        # Clear temporary report data while keeping essential user data
        context.user_data.pop("report_data", None)

        # Get user profile to determine menu type
        profile = db.get_user_profile(user_id)
        registration_type = profile.get("registration_type")

        # Add a small delay for better UX
        if update.callback_query:
            await update.callback_query.answer()
            await asyncio.sleep(0.1)  # Small delay for smooth transition

        # Redirect based on user type
        if registration_type == "employer":
            return await employer_main_menu(update, context)
        else:
            # Default to job seeker menu
            return await main_menu(update, context)

    except Exception as e:
        logging.error(f"Error returning to main menu for user {user_id}: {str(e)}")
        # Fallback to standard main menu
        return await main_menu(update, context)


async def violation_reports_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    # Fetch summary statistics
    total_reports = db.get_total_report_count()
    pending_reports = db.get_report_count_by_status('pending')
    resolved_reports = db.get_report_count_by_status('resolved')
    dismissed_reports = db.get_report_count_by_status('dismissed')

    # Get top reported entities
    top_reported_users = db.get_top_reported_entities('job_seeker', limit=5)
    top_reported_employers = db.get_top_reported_entities('employer', limit=5)
    top_reported_vacancies = db.get_top_reported_entities('vacancy', limit=5)
    top_reported_applications = db.get_top_reported_entities('application', limit=5)

    # Build message
    message_text = (
            "📊 *Violation Reports Dashboard*\n\n"
            f"• Total Reports: `{total_reports}`\n"
            f"• Pending: `{pending_reports}`\n"
            f"• Resolved: `{resolved_reports}`\n"
            f"• Dismissed: `{dismissed_reports}`\n\n"

            "🚨 *Top Reported Entities:*\n"
            "_Users:_\n" +
            "\n".join([f"  • {u['full_name']} (`{u['user_id']}`) - `{u['report_count']}` reports" for u in
                       top_reported_users]) + "\n\n"

                                              "_Employers:_\n" +
            "\n".join([f"  • {e['company_name']} (`{e['employer_id']}`) - `{e['report_count']}` reports" for e in
                       top_reported_employers]) + "\n\n"

                                              "_Vacancies:_\n" +
            "\n".join([f"  • {v['job_title']} (`{v['id']}`) - `{v['report_count']}` reports" for v in
                       top_reported_vacancies]) + "\n\n"

                                                  "_Applications:_\n" +
            "\n".join([
                          f"  • {a['job_title']} by {a['job_seeker_name']} (`{a['application_id']}`) - `{a['report_count']}` reports"
                          for a in top_reported_applications])
    )

    # Create keyboard
    keyboard = [
        [InlineKeyboardButton("🔍 View All Reports", callback_data="view_all_reports")],
        [InlineKeyboardButton("📈 Export to Excel", callback_data="export_reports_excel")],
        [InlineKeyboardButton("👤 Reported Users", callback_data="view_reported_users")],
        [InlineKeyboardButton("👔  Reported Employers", callback_data="view_reported_employers")],
        [InlineKeyboardButton("📋 Reported Vacancies", callback_data="view_reported_vacancies")],
        [InlineKeyboardButton("📄 Reported Applications", callback_data="view_reported_applications")],
        [InlineKeyboardButton("❌ Back to Menu", callback_data="back_to_admin_menu")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    # Check if the update is from a callback query
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        try:
            # Edit the existing message with the dashboard content
            await query.edit_message_text(
                text=message_text,
                reply_markup=reply_markup,
                parse_mode="Markdown"
            )
        except Exception as e:

            logging.warning(f"Failed to edit message: {e}")
            await context.bot.send_message(
                chat_id=user_id,
                text=message_text,
                reply_markup=reply_markup,
                parse_mode="Markdown"
            )
    else:
        # If the update is from a message, send a new message
        await update.message.reply_text(
            text=message_text,
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

    return VIOLATION_REPORTS_DASHBOARD




async def export_reports_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("⏳ Preparing advanced report... Please wait")
    try:
        # Fetch all reports
        reports = db.get_all_reports()
        # Create Excel file in memory
        excel_file = io.BytesIO()
        with pd.ExcelWriter(excel_file, engine='xlsxwriter', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
            workbook = writer.book
            # ===== MAIN REPORTS SHEET =====
            df_reports = pd.DataFrame(reports)
            # Check which entity types we actually have reports for
            existing_entity_types = df_reports['reported_entity_type'].unique()
            # Add human-readable entity names
            entity_details = []
            for _, report in df_reports.iterrows():
                detail = report.get('entity_name', 'Unknown')
                if pd.isna(detail):
                    detail = "Unknown"
                entity_details.append(f"{detail} ({report['reported_entity_type']})")
            df_reports['entity_details'] = entity_details
            # Reorder columns
            columns_order = [
                'report_id', 'timestamp', 'reporter_id', 'reported_entity_type',
                'entity_details', 'reported_entity_id', 'reason', 'additional_info',
                'report_status'
            ]
            # Only include columns that exist in the DataFrame
            columns_to_include = [col for col in columns_order if col in df_reports.columns]
            df_reports = df_reports[columns_to_include]
            # Write to Excel with formatting
            df_reports.to_excel(writer, sheet_name='All Reports', index=False)
            worksheet = writer.sheets['All Reports']
            # Add Excel formatting
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#4472C4',
                'font_color': 'white',
                'border': 1
            })
            # Format columns differently based on content
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})
            wrap_format = workbook.add_format({'text_wrap': True})
            status_format = workbook.add_format({'bold': True}) if 'report_status' in df_reports.columns else None
            # Apply formatting
            for col_num, value in enumerate(df_reports.columns.values):
                worksheet.write(0, col_num, value, header_format)
            # Set column widths dynamically based on content
            column_widths = {
                'report_id': 10,
                'timestamp': 18,
                'reporter_id': 15,
                'reported_entity_type': 12,
                'entity_details': 35,
                'reported_entity_id': 15,
                'reason': 25,
                'description': 40,
                'admin_notes': 40,
                'resolution_timestamp': 18,
                'report_status': 12
            }
            for col_num, col_name in enumerate(df_reports.columns):
                width = column_widths.get(col_name, 15)  # Default width 15
                cell_format = None
                if col_name == 'timestamp' or col_name == 'resolution_timestamp':
                    cell_format = date_format
                elif col_name in ['entity_details', 'reason', 'description', 'admin_notes']:
                    cell_format = wrap_format
                elif col_name == 'report_status':
                    cell_format = status_format
                worksheet.set_column(col_num, col_num, width, cell_format)
            # Freeze header row
            worksheet.freeze_panes(1, 0)
            # Add autofilter
            worksheet.autofilter(0, 0, 0, len(df_reports.columns) - 1)
            # Add conditional formatting for status if the column exists
            if 'report_status' in df_reports.columns:
                status_col = df_reports.columns.get_loc('report_status')
                green_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
                yellow_format = workbook.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C6500'})
                red_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
                worksheet.conditional_format(1, status_col, len(df_reports), status_col, {
                    'type': 'text',
                    'criteria': 'containing',
                    'value': 'resolved',
                    'format': green_format
                })
                worksheet.conditional_format(1, status_col, len(df_reports), status_col, {
                    'type': 'text',
                    'criteria': 'containing',
                    'value': 'pending',
                    'format': yellow_format
                })
                worksheet.conditional_format(1, status_col, len(df_reports), status_col, {
                    'type': 'text',
                    'criteria': 'containing',
                    'value': 'dismissed',
                    'format': red_format
                })
            # ===== STATISTICS SHEET =====
            stats_sheet = workbook.add_worksheet('Statistics')
            # Summary stats
            total_reports = len(df_reports)
            # Calculate status counts if the column exists
            if 'report_status' in df_reports.columns:
                status_counts = df_reports['report_status'].value_counts().to_dict()
                resolved_count = status_counts.get('resolved', 0)
                pending_count = status_counts.get('pending', 0)
                dismissed_count = status_counts.get('dismissed', 0)
            else:
                resolved_count = pending_count = dismissed_count = 0
            stats_data = [
                ['Total Reports', total_reports],
                ['Pending Reports', pending_count],
                ['Resolved Reports', resolved_count],
                ['Dismissed Reports', dismissed_count]
            ]
            if total_reports > 0 and 'report_status' in df_reports.columns:
                stats_data.append(['Resolution Rate', f"{(resolved_count / total_reports):.1%}"])
            # Write stats
            stats_sheet.write_row(0, 0, ['Metric', 'Value'], header_format)
            for row_num, row_data in enumerate(stats_data, 1):
                stats_sheet.write_row(row_num, 0, row_data)
            # Add a pie chart if we have status data
            if 'report_status' in df_reports.columns and total_reports > 0:
                chart = workbook.add_chart({'type': 'pie'})
                chart.add_series({
                    'name': 'Report Status',
                    'categories': ['Statistics', 1, 0, len(stats_data) - 1, 0],
                    'values': ['Statistics', 1, 1, len(stats_data) - 1, 1],
                    'data_labels': {'percentage': True, 'leader_lines': True}
                })
                chart.set_title({'name': 'Report Status Distribution'})
                stats_sheet.insert_chart('D2', chart)

            # ===== ENTITY-SPECIFIC SHEETS =====
            # Moved OUTSIDE of any conditionals
            # Create a dictionary mapping entity types to their display names and data functions
            entity_sheets = {
                'job_seeker': {
                    'name': 'Reported Job Seekers',
                    'data_func': db.get_job_seeker_by_id,
                    'columns': [
                        ('user_id', 'ID', 10),
                        ('full_name', 'Full Name', 25),
                        ('contact_number', 'Contact', 15),
                        ('gender', 'Gender', 10),
                        ('qualification', 'Qualification', 20),
                        ('skills_experience', 'Skills', 40),
                        ('created_at', 'Registered On', 18)
                    ]
                },
                'employer': {
                    'name': 'Reported Employers',
                    'data_func': db.get_employer_by_id,
                    'columns': [
                        ('employer_id', 'ID', 10),
                        ('company_name', 'Company', 30),
                        ('city', 'City', 15),
                        ('contact_number', 'Contact', 15),
                        ('employer_type', 'Type', 15),
                        ('about_company', 'About', 50),
                        ('created_at', 'Registered On', 18)
                    ]
                },
                'vacancy': {
                    'name': 'Reported Vacancies',
                    'data_func': db.get_vacancy_by_id_report,
                    'columns': [
                        ('id', 'ID', 10),
                        ('job_title', 'Job Title', 30),
                        ('company_name', 'Company', 25),
                        ('employment_type', 'Type', 15),
                        ('quantity', 'Openings', 10),
                        ('salary', 'Salary', 15),
                        ('application_deadline', 'Deadline', 18),
                        ('status', 'Status', 12)
                    ]
                },
                'application': {
                    'name': 'Reported Applications',
                    'data_func': db.get_application_by_id,
                    'columns': [
                        ('application_id', 'ID', 10),
                        ('job_title', 'Job Title', 30),
                        ('job_seeker_name', 'Applicant', 25),
                        ('employer_name', 'Company', 25),
                        ('application_date', 'Applied On', 18),
                        ('status', 'Status', 12),
                        ('cover_letter', 'Cover Letter', 50)
                    ]
                }
            }
            logging.info(f"Creating entity sheets for: {existing_entity_types}")
            # Create sheets only for entity types that exist in reports
            for entity_type in existing_entity_types:
                if entity_type in entity_sheets:
                    config = entity_sheets[entity_type]
                    logging.info(f"Processing entity type: {entity_type} ({config['name']})")
                    entity_ids = df_reports[df_reports['reported_entity_type'] == entity_type][
                        'reported_entity_id'].unique()
                    sheet_data = []
                    for entity_id in entity_ids:
                        entity = config['data_func'](entity_id)
                        if entity:
                            # Add report count to each entity
                            report_count = len(df_reports[
                                                   (df_reports['reported_entity_type'] == entity_type) &
                                                   (df_reports['reported_entity_id'] == entity_id)
                                                   ])
                            # Prepare row data
                            row = {}
                            for col in config['columns']:
                                row[col[1]] = entity.get(col[0], '')
                            row['Report Count'] = report_count
                            sheet_data.append(row)
                    if sheet_data:
                        df_entity = pd.DataFrame(sheet_data)
                        df_entity.to_excel(writer, sheet_name=config['name'], index=False)
                        sheet = writer.sheets[config['name']]
                        # Apply formatting
                        header_format = workbook.add_format({
                            'bold': True,
                            'text_wrap': True,
                            'fg_color': '#4472C4',
                            'font_color': 'white',
                            'border': 1
                        })
                        wrap_format = workbook.add_format({'text_wrap': True})
                        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})
                        # Write headers and set column widths
                        for col_num, col_config in enumerate(config['columns']):
                            sheet.write(0, col_num, col_config[1], header_format)
                            sheet.set_column(col_num, col_num, col_config[2],
                                             wrap_format if col_config[0] in ['skills_experience', 'about_company',
                                                                              'cover_letter']
                                             else date_format if 'date' in col_config[0] or 'created_at' in
                                                                 col_config[0]
                                             else None)
                        # Add report count column
                        sheet.write(0, len(config['columns']), 'Report Count', header_format)
                        sheet.set_column(len(config['columns']), len(config['columns']), 12)
                        sheet.autofilter(0, 0, 0, len(config['columns']))

            # ===== FINAL TOUCHES =====
            # Add a cover sheet
            cover_sheet = workbook.add_worksheet('Report Summary')
            cover_sheet.center_horizontally()
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 20,
                'align': 'center',
                'valign': 'vcenter'
            })
            subtitle_format = workbook.add_format({
                'italic': True,
                'align': 'center',
                'valign': 'vcenter'
            })
            cover_sheet.merge_range('A1:D4', 'VIOLATION REPORTS ANALYSIS', title_format)
            cover_sheet.merge_range('A5:D6', f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                    subtitle_format)
            # Add document properties
            workbook.set_properties({
                'title': 'Job Platform Violation Reports',
                'subject': 'Comprehensive violation report analysis',
                'author': 'Telegram Admin Bot',
                'company': 'Your Job Platform'
            })
        excel_file.seek(0)
        # Send document with progress updates
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.UPLOAD_DOCUMENT)
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=excel_file,
            filename=f"Violation_Reports_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx",
            caption="📊 Here is your violation report "
        )
    except Exception as e:
        logging.error(f"Error generating Excel report: {str(e)}", exc_info=True)
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="❌ Failed to generate advanced report. Please try again later."
        )
async def view_all_reports(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Displays all violation reports in a paginated format.
    """
    query = update.callback_query
    await query.answer()

    # Fetch all reports from the database
    page = context.user_data.get("current_page", 1)
    reports, total_pages = db.get_paginated_reports(page=page, page_size=10)

    if not reports:
        await query.edit_message_text(
            text="⚠️ No violation reports found.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
            ])
        )
        return VIEW_ALL_REPORTS

    # Format reports into a readable message
    report_text = "📋 *All Violation Reports*\n\n"
    for report in reports:
        report_text += (
            f"• ID: `{report['report_id']}`\n"
            f"  Entity: `{report['reported_entity_type']}`\n"
            f"  Reason: `{report['reason']}`\n"
            f"  Status: `{report['report_status']}`\n"
            f"  Timestamp: `{report['timestamp']}`\n\n"
        )

    # Create pagination keyboard
    keyboard = []
    if page > 1:
        keyboard.append([InlineKeyboardButton("⬅️ Previous", callback_data=f"prev_all_reports_{page}")])
    if page < total_pages:
        keyboard.append([InlineKeyboardButton("Next ➡️", callback_data=f"next_all_reports_{page}")])

    keyboard.append([InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")])

    await query.edit_message_text(
        text=report_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return VIEW_ALL_REPORTS


async def view_reported_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Displays reported users with their details.
    """
    query = update.callback_query
    await query.answer()

    # Fetch reported users from the database
    reported_users = db.get_top_reported_entities(entity_type="job_seeker", limit=10)

    if not reported_users:
        await query.edit_message_text(
            text="⚠️ No reported users found.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
            ])
        )
        return VIEW_REPORTED_USERS

    # Format users into a readable message
    user_text = "👤 *Reported Users*\n\n"
    for user in reported_users:
        user_text += (
            f"• Name: `{user['full_name']}`\n"
            f"  ID: `{user['user_id']}`\n"
            f"  Reports: `{user['report_count']}`\n\n"
        )

    keyboard = [
        [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
    ]

    await query.edit_message_text(
        text=user_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return VIEW_REPORTED_USERS

async def view_reported_employers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Displays reported users with their details.
    """
    query = update.callback_query
    await query.answer()

    # Fetch reported users from the database
    reported_employers = db.get_top_reported_entities(entity_type="employer", limit=10)

    if not reported_employers:
        await query.edit_message_text(
            text="⚠️ No reported employers found.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
            ])
        )
        return VIEW_REPORTED_EMPLOYERS

    # Format users into a readable message
    employer_text = "👤 *Reported Employers*\n\n"
    for employer in reported_employers:
        employer_text += (
            f"• Name: `{employer['company_name']}`\n"
            f"  ID: `{employer['employer_id']}`\n"
            f"  Reports: `{employer['report_count']}`\n\n"
        )

    keyboard = [
        [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
    ]

    await query.edit_message_text(
        text=employer_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return VIEW_REPORTED_EMPLOYERS

async def view_reported_vacancies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Displays reported vacancies with their details.
    """
    query = update.callback_query
    await query.answer()

    # Fetch reported vacancies from the database
    reported_vacancies = db.get_top_reported_entities(entity_type="vacancy", limit=10)

    if not reported_vacancies:
        await query.edit_message_text(
            text="⚠️ No reported vacancies found.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
            ])
        )
        return VIEW_REPORTED_VACANCIES

    # Format vacancies into a readable message
    vacancy_text = "📋 *Reported Vacancies*\n\n"
    for vacancy in reported_vacancies:
        vacancy_text += (
            f"• Title: `{vacancy['job_title']}`\n"
            f"  ID: `{vacancy['id']}`\n"
            f"  Reports: `{vacancy['report_count']}`\n\n"
        )

    keyboard = [
        [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
    ]

    await query.edit_message_text(
        text=vacancy_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return VIEW_REPORTED_VACANCIES


async def view_reported_applications(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Displays reported applications with their details.
    """
    query = update.callback_query
    await query.answer()

    # Fetch reported applications from the database
    reported_applications = db.get_top_reported_entities(entity_type="application", limit=10)

    if not reported_applications:
        await query.edit_message_text(
            text="⚠️ No reported applications found.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
            ])
        )
        return VIEW_REPORTED_APPLICATIONS

    # Format applications into a readable message
    application_text = "📄 *Reported Applications*\n\n"
    for application in reported_applications:
        application_text += (
            f"• Job Title: `{application['job_title']}`\n"
            f"  Applicant: `{application['full_name']}`\n"
            f"  ID: `{application['application_id']}`\n"
            f"  Reports: `{application['report_count']}`\n\n"
        )

    keyboard = [
        [InlineKeyboardButton("🔙 Back to Dashboard", callback_data="back_to_violation_dashboard")]
    ]

    await query.edit_message_text(
        text=application_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return VIEW_REPORTED_APPLICATIONS


async def back_to_violation_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Returns to the violation reports dashboard.
    """
    query = update.callback_query
    await query.answer()
    return await violation_reports_dashboard(update, context)


async def handle_pagination_all_reports(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Handles pagination for the "View All Reports" feature.
    """
    query = update.callback_query
    await query.answer()

    # Parse callback data to determine action and current page
    action, _, current_page = query.data.split("_")
    current_page = int(current_page)
    new_page = current_page + 1 if action == "next" else current_page - 1

    # Update the current page in context
    context.user_data["current_page"] = new_page

    # Re-fetch and display reports for the new page
    return await view_all_reports(update, context)


#error handling
import traceback
from typing import Optional
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes


async def advanced_error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Advanced error handler with logging and admin notifications"""
    # Prepare error details
    error = context.error
    tb_list = traceback.format_exception(None, error, error.__traceback__)
    tb_string = "".join(tb_list)

    # Extract user and chat information if available
    user_id: Optional[int] = None
    chat_id: Optional[int] = None
    command: Optional[str] = None

    if update and hasattr(update, 'effective_user'):
        user_id = update.effective_user.id
    if update and hasattr(update, 'effective_chat'):
        chat_id = update.effective_chat.id
    if update and hasattr(update, 'message') and update.message and update.message.text:
        command = update.message.text.split()[0] if update.message.text else None

    # Prepare error data for database
    error_data = {
        "user_id": user_id,
        "chat_id": chat_id,
        "command": command,
        "error_type": error.__class__.__name__,
        "error_message": str(error),
        "traceback": tb_string,
        "context_data": dict(context.user_data) if context.user_data else None,
        "update_data": update.to_dict() if update and hasattr(update, 'to_dict') else None
    }

    # Log error to database
    error_id = db.log_error(error_data)

    # Notify user (if possible)
    if user_id:
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=get_translation(user_id, 'notifying_unexpected_error_occurred')
            )
        except Exception as e:
            logging.error(f"Could not notify user about error: {e}")

    # Notify all admins
    try:
        await notify_admins_about_error(context, error_id, error_data)
    except Exception as e:
        logging.error(f"Could not notify admins about error: {e}")

    # Return to main menu report
    await back_to_main_menu_report(update, context)


async def notify_admins_about_error(context: ContextTypes.DEFAULT_TYPE, error_id: str, error_data: dict) -> None:
    """Notify all admins about the error"""
    admin_ids = get_all_admins()
    if not admin_ids:
        return

    # Safely get error message
    error_message = str(error_data.get('error_message', 'Unknown error'))
    short_error = f"{error_data.get('error_type', 'Error')}: {error_message[:200]}"

    for admin_id in admin_ids:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=f"🚨 New Bot Error ({error_id[:8]}):\n\n{short_error}"
            )
        except Exception as e:
            logging.error(f"Could not notify admin {admin_id} about error: {str(e)}")

async def view_system_errors(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show list of system errors"""
    query = update.callback_query
    await query.answer()

    errors = db.get_errors(limit=10)

    if not errors:
        await query.edit_message_text(text="No errors found in the system.")
        return DATABASE_MANAGEMENT

    keyboard = []
    for error in errors:
        # Format timestamp
        try:
            timestamp = datetime.fromisoformat(error.get('timestamp', ''))
            time_display = timestamp.strftime('%m/%d %H:%M')
        except (ValueError, TypeError):
            time_display = "Unknown time"

        short_msg = f"{error.get('error_type', 'Error')}: {error.get('error_message', '')[0:30]}..."
        btn = InlineKeyboardButton(
            f"{time_display} - {short_msg}",
            callback_data=f"error_detail_{error.get('error_id', '')}"
        )
        keyboard.append([btn])

    keyboard.append([InlineKeyboardButton("Back", callback_data="back_to_database_menu")])

    await query.edit_message_text(
        text="Recent System Errors:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return VIEW_ERRORS

import html
async def handle_error_detail(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show detailed error information with expandable traceback"""
    query = update.callback_query
    await query.answer()

    error_id = query.data.split("_")[-1]
    error = db.get_error_by_id(error_id)

    if not error:
        await query.edit_message_text(text="Error details not found.")
        return VIEW_ERRORS

    # Format basic error info
    try:
        timestamp = datetime.fromisoformat(error['timestamp'])
        time_display = timestamp.strftime('%Y-%m-%d %H:%M:%S')
    except (ValueError, KeyError):
        time_display = "Unknown time"

    basic_info = (
        f"🚨 <b>Error ID</b>: <code>{html.escape(str(error.get('error_id', 'Unknown')))}</code>\n"
        f"⏰ <b>Time</b>: {html.escape(time_display)}\n"
        f"👤 <b>User</b>: <code>{html.escape(str(error.get('user_id', 'N/A')))}</code>\n"
        f"📝 <b>Command</b>: <code>{html.escape(str(error.get('command', 'N/A')))}</code>\n"
        f"🔧 <b>Status</b>: {html.escape(error.get('status', 'unresolved'))}\n\n"
        f"💥 <b>Error Type</b>: <code>{html.escape(str(error.get('error_type', 'Unknown')))}</code>\n"
        f"📄 <b>Message</b>:\n<code>{html.escape(str(error.get('error_message', 'No message')))}</code>\n\n"
    )

    # Prepare context info
    context_info = "<b>🔍 Context</b>:\n"
    if error.get('context_data'):
        # Convert context data to JSON and escape it for HTML
        context_data = json.dumps(error.get('context_data'), indent=2)[:1000]
        escaped_context_data = html.escape(context_data)
        context_info += f"<pre>{escaped_context_data}</pre>\n\n"
    else:
        context_info += "No context data\n\n"

    # Prepare traceback info
    traceback_info = ""
    if error.get('traceback'):
        traceback_lines = error['traceback'].split('\n')
        short_traceback = "\n".join(traceback_lines[:10])  # Show first 10 lines initially
        traceback_info = (
            f"<b>🔎 Traceback (first 10 lines):</b>\n"
            f"<pre>{html.escape(short_traceback)}</pre>\n\n"
            f"<i>Full traceback available below</i>\n\n"
        )

    # Create keyboard with expand/collapse options
    keyboard = [
        [InlineKeyboardButton("📜 Show Full Traceback", callback_data=f"show_traceback_{error_id}")],
        [InlineKeyboardButton("📋 Show Update Data", callback_data=f"show_update_{error_id}")],
        [
            InlineKeyboardButton("✅ Mark Fixed", callback_data=f"resolve_error_{error_id}"),
            InlineKeyboardButton("🔙 Back to List", callback_data="view_system_errors")
        ]
    ]

    # Send initial message
    await query.edit_message_text(
        text=basic_info + context_info + traceback_info,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )
    return ERROR_DETAIL


async def show_full_traceback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show full error traceback"""
    query = update.callback_query
    await query.answer()

    error_id = query.data.split("_")[-1]
    error = db.get_error_by_id(error_id)

    if not error or not error.get('traceback'):
        await query.answer("No traceback available")
        return ERROR_DETAIL

    # Escape HTML special characters
    full_traceback = html.escape(error['traceback'])

    # Create keyboard with back button - using a pattern that's handled in ERROR_DETAIL state
    keyboard = [
        [InlineKeyboardButton("🔙 Back to Error Details", callback_data=f"back_to_detail_{error_id}")]
    ]

    # Split into multiple messages if too long
    if len(full_traceback) > 4000:
        parts = [full_traceback[i:i + 4000] for i in range(0, len(full_traceback), 4000)]
        for part in parts:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"<pre>{part}</pre>",
                parse_mode="HTML"
            )
        # Send the back button in a separate message
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="<b>Full traceback sent above</b>",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )
        await query.answer("Sent full traceback in multiple messages")
    else:
        await query.edit_message_text(
            text=f"<b>Full Traceback:</b>\n<pre>{full_traceback}</pre>",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )

    return ERROR_DETAIL
async def show_update_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle the 'Show Update Data' button press"""
    query = update.callback_query
    await query.answer()

    error_id = query.data.split('_')[-1]  # Extract error_id from callback_data

    try:
        # 1. Get error details from database
        error = db.get_error_by_id(error_id)
        if not error or not error.get('update_data'):
            await query.answer("⚠️ No update data available", show_alert=True)
            return ERROR_DETAIL

        # 2. Format the JSON data
        update_data = error['update_data']
        if isinstance(update_data, str):
            try:
                update_data = json.loads(update_data)
            except json.JSONDecodeError:
                pass

        pretty_data = json.dumps(update_data, indent=2, ensure_ascii=False)

        # 3. Send the data (split if too large)
        if len(pretty_data) > 4000:
            await query.answer("Sending large update data in parts...")
            for i in range(0, len(pretty_data), 4000):
                part = pretty_data[i:i + 4000]
                await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text=f"<pre>{html.escape(part)}</pre>",
                    parse_mode="HTML"
                )
        else:
            await query.edit_message_text(
                text=f"📋 <b>Update Data for Error {error_id[:8]}:</b>\n<pre>{html.escape(pretty_data)}</pre>",
                parse_mode="HTML"
            )

    except Exception as e:
        logging.error(f"Error showing update data: {str(e)}")
        await query.answer("❌ Failed to load update data", show_alert=True)

    return ERROR_DETAIL

async def resolve_error(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Mark an error as resolved"""
    query = update.callback_query
    await query.answer()

    error_id = query.data.split("_")[-1]
    success = db.update_error_status(error_id, "fixed")

    if success:
        await query.edit_message_text(text=f"✅ Error {error_id[:8]} marked as fixed.")
    else:
        await query.edit_message_text(text="❌ Failed to update error status.")

    return await view_system_errors (update, context)

from telegram.error import BadRequest

async def edit_message_safely(query, text, markup):
    try:
        await query.edit_message_text(text=text, reply_markup=markup)
    except BadRequest as e:
        if "message is not modified" in str(e).lower():
            # Ignore redundant edits
            return
        raise  # Re-raise other errors

#system configuration
async def show_system_configurations_menu(update, context):
    user_id = update.effective_user.id

    welcome_text = (
        f"📊 *System Configurations Menu*\n\n"
        f"Monitor and optimize your database:\n"
        f"• Analyze database storage usage\n"
        f"• View size of individual tables\n"
        f"• Optimize and clean up the database\n"
        f"Select an option below:"
    )

    keyboard = [
        ["💾 Database Storage Overview"],
        ["⚙️ Optimize Database", "🔄 Vacuum Database"],
        ["🔙 Back to Admin Menu"]
    ]

    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        one_time_keyboard=True,
        resize_keyboard=True
    )

    await context.bot.send_message(
        chat_id=user_id,
        text=welcome_text,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )

    return SYSTEM_CONFIGURATIONS_MENU
async def handle_system_configurations_choice(update, context):
    choice = update.message.text

    if choice == "💾 Database Storage Overview":
        return await show_database_storage_overview(update, context)
    elif choice == "⚙️ Optimize Database":
        return await optimize_database(update, context)
    elif choice == "🔄 Vacuum Database":
        return await vacuum_database(update, context)
    # elif choice == "🔍 Query Performance Insights":
    #     return await show_query_performance_insights(update, context)
    # elif choice == "🚨 Error Logs":
    #     return await view_system_errors(update, context)
    elif choice == "🔙 Back to Admin Menu":
        return await show_admin_menu(update, context)


async def show_database_storage_overview(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    # Show loading message
    loading_msg = await context.bot.send_message(
        chat_id=user_id,
        text="🔄 Gathering database storage metrics...",
        parse_mode="Markdown"
    )

    try:
        # Fetch database metrics - ensure these return numbers, not strings
        db_size = float(db.get_database_siz())  # Convert to float if not already
        table_sizes = db.get_table_sizes()
        total_rows = sum(int(size['rows']) for size in table_sizes.values() if size['rows'])

        # Generate storage visualization
        storage_chart = generate_storage_chart(table_sizes)

        # Prepare detailed overview with proper numeric formatting
        overview_text = (
            f"💾 *Database Storage Overview*\n\n"
            f"📦 *Total Size:* `{db_size:.2f} MB`\n"
            f"📊 *Total Tables:* `{len(table_sizes)}`\n"
            f"📝 *Total Rows:* `{total_rows:,}`\n\n"
            f"{storage_chart}\n\n"
            f"📋 *Table Details (Size | Rows | Indexes)*:\n"
        )


        # Sort tables by size (descending)
        sorted_tables = sorted(table_sizes.items(),
                               key=lambda x: x[1].get('size_mb', 0),
                               reverse=True)

        for table, metrics in sorted_tables:
            size_mb = metrics.get('size_mb', 0)
            percentage = (size_mb / db_size) * 100 if db_size > 0 else 0
            rows = metrics.get('rows', 0)
            indexes = metrics.get('indexes', 0)

            overview_text += (
                f"▫️ `{table}`: "
                f"{size_mb:.2f}MB ({percentage:.1f}%) | "
                f"{rows:,} rows | "
                f"{indexes} indexes\n"
            )

        overview_text += (
            f"\n⏱ *Last Analyzed:* {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"💡 *Tip:* Consider archiving tables with low activity but high storage."
        )

        keyboard = [[InlineKeyboardButton("🔙 Back to System Configurations", callback_data="back_to_system_config")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        # Edit the loading message with final content
        await context.bot.edit_message_text(
            chat_id=user_id,
            message_id=loading_msg.message_id,
            text=overview_text,  # Fixed the variable name here
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

    except Exception as e:
        error_msg = (
            "⚠️ *Error fetching storage data*\n\n"
            f"```{str(e)}```\n\n"
            "Please try again later or check server logs."
        )
        await context.bot.edit_message_text(
            chat_id=user_id,
            message_id=loading_msg.message_id,
            text=error_msg,
            parse_mode="Markdown"
        )
        logging.error(f"Database storage overview error: {str(e)}", exc_info=True)

    return DATABASE_STORAGE_OVERVIEW


def generate_storage_chart(table_sizes):
    """Generate accurate storage visualization"""
    if not table_sizes:
        return ""

    # Filter out tiny tables for better visualization
    significant_tables = {t: d for t, d in table_sizes.items() if d['size_mb'] > 0.01}
    if not significant_tables:
        return ""

    max_size = max(d['size_mb'] for d in significant_tables.values())
    chart = "          Storage Distribution          \n"
    chart += "----------------------------------------\n"

    for table, metrics in sorted(significant_tables.items(),
                                 key=lambda x: x[1]['size_mb'],
                                 reverse=True)[:5]:  # Top 5 tables
        bar_length = int((metrics['size_mb'] / max_size) * 20)
        chart += (
            f"{table[:15]:<15} | "
            f"{'█' * bar_length}{' ' * (20 - bar_length)} | "
            f"{metrics['size_mb']:.2f}MB ({metrics['percentage']:.1f}%)\n"
        )

    return f"```\n{chart}```"


async def optimize_database(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    try:
        db.optimize_database()  # Assume this triggers index rebuild and cleanup
        success_message = "✅ Database optimization completed successfully."
    except Exception as e:
        success_message = f"⚠️ Database optimization failed: {str(e)}"

    # Add back button
    keyboard = [["🔙 Back to System Configurations"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    await context.bot.send_message(
        chat_id=user_id,
        text=success_message,
        reply_markup=reply_markup
    )

    return OPTIMIZE_DATABASE
async def vacuum_database(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    try:
        db.vacuum_database()  # Assume this triggers the VACUUM command
        success_message = "✅ Database vacuuming completed successfully."
    except Exception as e:
        success_message = f"⚠️ Database vacuuming failed: {str(e)}"

    # Add back button
    keyboard = [["🔙 Back to System Configurations"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    await context.bot.send_message(
        chat_id=user_id,
        text=success_message,
        reply_markup=reply_markup
    )

    return VACUUM_DATABASE
async def handle_back_system(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()  # Acknowledge the button press

    # Redirect to the SYSTEM_CONFIGURATIONS_MENU
    await query.edit_message_text(text="Returning to System Configurations...")
    return SYSTEM_CONFIGURATIONS_MENU
# Conversation Handler
def main():
    application = Application.builder().token("7567203189:AAEu1NDdQ0-b8dI39zOwFIdoQ8SUyF5K5p0").build()

    # Conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start),
            CommandHandler("admin", admin_start),
            CommandHandler("review_appeal", review_appeal)],
        states={
            LANGUAGE: [CallbackQueryHandler(set_language)],
            MOBILE: [MessageHandler(filters.CONTACT, save_mobile)],
            REGISTRATION_TYPE: [CallbackQueryHandler(registration_type)],
            JOB_SEEKER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_full_name),
            ],
            JOB_SEEKER_DOB: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_dob),
            ],
            JOB_SEEKER_GENDER: [
                CallbackQueryHandler(job_seeker_gender, pattern="^(male|female)$"),
            ],
            JOB_SEEKER_CONTACT_NUMBERS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_contact_numbers),
            ],
            JOB_SEEKER_LANGUAGES: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_languages),
            ],
            JOB_SEEKER_QUALIFICATION: [
                CallbackQueryHandler(job_seeker_qualification,
                                     pattern="^(certificate|diploma|degree|ma|phd|other|skip)$"),
            ],
            JOB_SEEKER_FIELD_OF_STUDY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_field_of_study),
                CallbackQueryHandler(job_seeker_field_of_study, pattern="^(skip)$"),
            ],
            JOB_SEEKER_CGPA: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_cgpa),
                CallbackQueryHandler(job_seeker_cgpa, pattern="^(skip)$"),
            ],
            JOB_SEEKER_SKILLS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_skills),
                CallbackQueryHandler(job_seeker_skills, pattern="^(skip)$"),
            ],
            JOB_SEEKER_PROFILE_SUMMARY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_profile_summary),
                CallbackQueryHandler(job_seeker_profile_summary, pattern="^(skip)$"),
            ],
            JOB_SEEKER_SUPPORTING_DOCUMENTS: [
                MessageHandler(filters.Document.ALL, job_seeker_supporting_documents),
                CallbackQueryHandler(job_seeker_supporting_documents, pattern="^(skip)$"),
            ],
            JOB_SEEKER_PORTFOLIO_LINK: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, job_seeker_portfolio_link),
                CallbackQueryHandler(job_seeker_portfolio_link, pattern="^(skip)$"),
            ],
            MAIN_MENU: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_main_menu),
            ],
            PROFILE_MENU: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_profile_actions),
            ],
            VIEW_PROFILE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, view_profile),
            ],

            CONFIRM_DELETE_ACCOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_delete_confirmation),
            ],
            CONFIRM_CHANGE_LANGUAGE: [
                CallbackQueryHandler(change_language_confirmed, pattern="^change_language_confirmed$"),
                CallbackQueryHandler(cancel_change_language, pattern="^cancel_change_language$")
            ],
            SELECT_LANGUAGE: [
                CallbackQueryHandler(handle_job_seeker_language_selection)
            ],
            EDIT_PROFILE: [
                CallbackQueryHandler(handle_edit_profile_field, pattern="^edit_.+$"),
                CallbackQueryHandler(handle_edit_profile_field, pattern="^view_completion$"),
                CallbackQueryHandler(handle_edit_profile_field, pattern="^back_to_editing$"),
                CallbackQueryHandler(handle_edit_profile_field, pattern="^back_to_main_menu$")
            ],
            EDIT_PROFILE_FIELD_VALUE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_edited_field),
                MessageHandler(filters.Document.ALL, handle_document_upload),
                CallbackQueryHandler(cancel_editing_profile, pattern="^cancel_editing$")
            ],
            PROFILE_COMPLETION: [
                CallbackQueryHandler(edit_profile, pattern="^back_to_editing$")
            ],
            SAVE_EDITED_FIELD: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_edited_field)],
            EDIT_FIELD_VALUE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_edited_field),
                MessageHandler(filters.Document.ALL, save_edited_field)
            ],
            EMPLOYER_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_employer_name)
            ],
            EMPLOYER_LOCATION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_employer_location)
            ],
            EMPLOYER_TYPE: [
                CallbackQueryHandler(save_employer_type)
            ],
            ABOUT_COMPANY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_about_company),
                CallbackQueryHandler(save_about_company, pattern="^(skip)$")
            ],
            VERIFICATION_DOCUMENTS: [
                MessageHandler(filters.Document.ALL, upload_verification_documents),
                CallbackQueryHandler(upload_verification_documents, pattern="^(skip)$")
            ],
            EMPLOYER_MAIN_MENU: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_employer_main_menu),
                CallbackQueryHandler(handle_pagination_callback, pattern=r"^manage_vacancies_(first|prev|next|last)_"),

                CallbackQueryHandler(
                    handle_vacancy_actions,
                    pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"
                ),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(
                    view_analytics,
                    pattern="^view_analytics$"
                )
            ],
            EMPLOYER_PROFILE_MENU: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_employer_profile_actions),
            ],
            CONFIRM_DELETE_MY_ACCOUNT: [MessageHandler(filters.TEXT, handle_my_delete_confirmation)],
            SELECT_EMPLOYER_LANGUAGE: [CallbackQueryHandler(handle_employer_language_selection)],
            CONFIRM_CHANGE_EMPLOYER_LANGUAGE: [CallbackQueryHandler(change_employer_language_confirmed)],

            SELECT_JOB_TO_MANAGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, select_job_to_manage),
                CallbackQueryHandler(handle_pagination_callback, pattern=r"^manage_vacancies_(first|prev|next|last)_"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$")
            ],
            EMPLOYER_MANAGE_VACANCIES: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, manage_vacancies),
                CallbackQueryHandler(handle_pagination_callback, pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$")
            ],

            HANDLE_JOB_ACTIONS: [
                CallbackQueryHandler(handle_job_actions, pattern=r"^(view_apps|close|stats|renew)_\d+$"),
                CallbackQueryHandler(handle_pagination_callback, pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$")
            ],

            VIEW_APPLICATIONS: [
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, select_applicant),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$")
            ],
            ACCEPT_REJECT_CONFIRMATION: [
                CallbackQueryHandler(handle_accept_reject, pattern="^accept_applicant$|^reject_applicant$"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),

            ],
            REJECTION_REASON_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_rejection_reason_application),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_cv_download, pattern=r"^download_cv_\d+$")
            ],
            EMPLOYER_MESSAGE_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_employer_message),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
            ],

            CONFIRM_CLOSE: [
                MessageHandler(filters.Regex("^(Yes|No)$"), handle_close_confirmation),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
            ],
            RESUBMIT_CONFIRMATION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, confirm_resubmit),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
            ],
            RENEW_VACANCY: [
                CallbackQueryHandler(handle_renew_duration, pattern=r"^renew_(30|60|custom)$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_custom_renew_duration),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
            ],
            CONFIRM_RENEWAL: [
                CallbackQueryHandler(process_renewal_confirmation, pattern="^(confirm|cancel)_renew$"),
                CallbackQueryHandler(handle_pagination_callback,
                                     pattern=r"^manage_vacancies_(first|prev|next|last)_[a-zA-Z]+$"),
                CallbackQueryHandler(export_to_excel, pattern=r"^export_excel_"),
                CallbackQueryHandler(view_applicants_list, pattern=r"^view_apps_"),
                CallbackQueryHandler(preview_job_details, pattern=r"^preview_\d+$"),
                CallbackQueryHandler(handle_back_to_job, pattern=r"^back_to_job_\d+$"),
                CallbackQueryHandler(handle_applicant_review, pattern=r"^review_\d+$"),
                CallbackQueryHandler(handle_vacancy_actions, pattern=r"^(view_apps|close|resubmit|stats|renew)_\d+$"),
            ],
            EDIT_EMPLOYER_PROFILE: [
                CallbackQueryHandler(handle_edit_employer_field),
                CallbackQueryHandler(cancel_editing, pattern="^cancel_editing$"),
                CallbackQueryHandler(show_profile_completion_details, pattern="^view_completion$")
            ],
            EDIT_EMPLOYER_FIELD_VALUE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_updated_employer_field),
                MessageHandler(filters.Document.ALL, save_updated_employer_field),
                CallbackQueryHandler(cancel_editing, pattern="^cancel_editing$")
            ],
            PROFILE_COMPLETION_VIEW: [
                CallbackQueryHandler(edit_employer_profile, pattern="^back_to_editing$")
            ],
            #  new states for analytics
            ANALYTICS_VIEW: [
                CallbackQueryHandler(
                    handle_analytics_actions,
                    pattern=r"^(analytics_trends|analytics_demographics|analytics_response|analytics_benchmark|analytics_export|analytics_back|go_to_employer_main_menu)$"
                )
            ],
            ANALYTICS_TRENDS: [
                CallbackQueryHandler(
                    handle_analytics_back,
                    pattern="^back_to_analytics$"
                )
            ],
            ANALYTICS_DEMOGRAPHICS: [
                CallbackQueryHandler(
                    handle_analytics_back,
                    pattern="^back_to_analytics$"
                )
            ],
            ANALYTICS_RESPONSE: [
                CallbackQueryHandler(
                    handle_analytics_back,
                    pattern="^back_to_analytics$"
                )
            ],
            ANALYTICS_BENCHMARK: [
                CallbackQueryHandler(
                    handle_analytics_back,
                    pattern="^back_to_analytics$"
                )
            ],
            ANALYTICS_EXPORT: [
                CallbackQueryHandler(
                    handle_export_format,
                    pattern=r"^export_(csv|pdf|excel)$"
                ),
                CallbackQueryHandler(
                    handle_analytics_back,
                    pattern="^back_to_analytics$"
                )
            ],
            HELP_MENU: [
                CallbackQueryHandler(help_button_handler, pattern="^help_"),
                CallbackQueryHandler(handle_admin_reply_callback, pattern="^admin_reply_"),
            ],


            # New FAQ states
            FAQ_SECTION: [
                CallbackQueryHandler(show_faq_category_section, pattern="^faq_main_"),
                # Handles all main FAQ categories
                CallbackQueryHandler(handle_faq_category, pattern=r"^faq_category_\d+"),
                CallbackQueryHandler(show_help, pattern=r"^help_back$")  # New handler
            ],
            FAQ_CATEGORY: [
                CallbackQueryHandler(handle_faq_question, pattern="^faq_question_"),
                CallbackQueryHandler(show_faq_section, pattern="^faq_category_back$"),  # Back to FAQ Categories
            ],
            FAQ_QUESTION: [
                CallbackQueryHandler(handle_faq_category, pattern=r"^faq_question_back_\d+$"),
            ],
            # Job seeker FAQ states
            JS_FAQ_SECTION: [
                CallbackQueryHandler(handle_job_seeker_faq_category, pattern=r"^js_faq_category_\d+"),
                CallbackQueryHandler(show_faq_category_section, pattern=r"^js_faq_back$")
            ],
            JS_FAQ_CATEGORY: [
                CallbackQueryHandler(handle_job_seeker_faq_question, pattern=r"^js_faq_q_"),
                CallbackQueryHandler(show_job_seeker_faq, pattern=r"^js_faq_return_\d+$")
            ],
            JS_FAQ_QUESTION: [
                CallbackQueryHandler(handle_job_seeker_faq_category, pattern=r"^js_faq_return_\d+$")
            ],
            ADMIN_FAQ_SECTION: [
                CallbackQueryHandler(handle_admin_faq_category, pattern=r"^admin_faq_cat_\d+"),
                CallbackQueryHandler(show_faq_category_section, pattern=r"^admin_faq_back_to_main")
            ],
            ADMIN_FAQ_CATEGORY: [
                CallbackQueryHandler(handle_admin_faq_question, pattern=r"^admin_faq_q_"),
                CallbackQueryHandler(show_admin_faq, pattern=r"^admin_faq_return_\d+")  # Matches category back button
            ],
            ADMIN_FAQ_QUESTION: [
                CallbackQueryHandler(handle_admin_faq_category, pattern=r"^admin_faq_cat_\d+")
            ],
            RATE_OPTIONS: [
                CallbackQueryHandler(show_rate_options, pattern="^back_to_rate_menu$"),
                CallbackQueryHandler(start_rate_bot, pattern="^rate_bot$"),
                CallbackQueryHandler(start_rate_user, pattern="^rate_user$"),
                CallbackQueryHandler(show_my_reviews, pattern="^my_reviews$"),
                CallbackQueryHandler(show_review_search, pattern="^search_reviews$"),
                CallbackQueryHandler(show_review_settings, pattern="^review_settings$"),
                CallbackQueryHandler(handle_user_selection, pattern=r"^select_user_\d+$"),
                CallbackQueryHandler(main_menu, pattern="^back_to_main$")
            ],

            SEARCH_USER_FOR_RATING: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_review_search),
                CallbackQueryHandler(filter_employers, pattern="^filter_employers$"),
                CallbackQueryHandler(filter_jobseekers, pattern="^filter_jobseekers$"),
                CallbackQueryHandler(sort_reviews, pattern="^sort_top_rated$"),
                CallbackQueryHandler(sort_reviews, pattern="^sort_recent$"),
                CallbackQueryHandler(show_rate_options, pattern="^back_to_rate_menu$"),
                CallbackQueryHandler(handle_search_by_name, pattern="^search_by_name$"),
                CallbackQueryHandler(previous_page, pattern="^prev_page$"),
                CallbackQueryHandler(next_page, pattern="^next_page$")
            ],

            SELECT_USER_FOR_RATING: [
                CallbackQueryHandler(handle_user_selection, pattern=r"^select_user_\d+$"),
                CallbackQueryHandler(start_rate_user, pattern="^back_to_rate_menu$")
            ],

            RATE_DIMENSION: [
                CallbackQueryHandler(handle_dimension_rating, pattern=r"^rate_[a-z_]+_\d$"),
                CallbackQueryHandler(skip_dimension_rating, pattern=r"^skip_[a-z_]+$"),
                CallbackQueryHandler(show_rate_options, pattern="^cancel_rating$")
            ],
            PROMPT_FOR_COMMENT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, submit_comment),
                CallbackQueryHandler(skip_comment, pattern="^skip_comment$"),
                CommandHandler("cancel", cancel_comment)
            ],

            ADD_COMMENT_OPTIONAL: [
                CallbackQueryHandler(prompt_for_comment, pattern="^add_comment$"),
                CallbackQueryHandler(skip_comment, pattern="^skip_comment$")
            ],

            CONFIRM_REVIEW: [
                CallbackQueryHandler(finalize_review, pattern="^confirm_review$"),
                CallbackQueryHandler(edit_review, pattern="^edit_review$"),
                CallbackQueryHandler(show_rate_options, pattern="^cancel_review$")
            ],

            REVIEW_SETTINGS: [
                CallbackQueryHandler(toggle_privacy_setting, pattern=r"^toggle_(anonymous|contact_visible)$"),
                CallbackQueryHandler(show_rate_options, pattern="^back_to_rate_menu$")
            ],

            MY_REVIEWS: [
                CallbackQueryHandler(show_review_details, pattern=r"^review_my_\d+$"),
                CallbackQueryHandler(delete_review, pattern=r"^delete_review_\d+$"),
                CallbackQueryHandler(edit_existing_review, pattern=r"^edit_review_\d+$"),
                CallbackQueryHandler(show_rate_options, pattern="^back_to_rate_menu$")
            ],

            REVIEW_DETAILS: [
                CallbackQueryHandler(show_review_details, pattern=r"^(view_review|review_my)_\d+$"),
                CallbackQueryHandler(edit_existing_review, pattern=r"^edit_review_\d+$"),
                CallbackQueryHandler(delete_review, pattern=r"^delete_review_\d+$"),
                CallbackQueryHandler(confirm_delete_review, pattern="^confirm_delete$"),
                CallbackQueryHandler(show_my_reviews, pattern="^back_to_my_reviews$"),
                CallbackQueryHandler(show_review_search, pattern="^back_to_search$")
            ],

            SEARCH_REVIEWS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_review_search),
                CallbackQueryHandler(filter_reviews, pattern=r"^filter_reviews_employer$"),
                CallbackQueryHandler(filter_reviews, pattern=r"^filter_reviews_job_seeker$"),
                CallbackQueryHandler(sort_reviews, pattern=r"^sort_reviews_top$"),
                CallbackQueryHandler(sort_reviews, pattern=r"^sort_reviews_recent$"),
                CallbackQueryHandler(show_rate_options, pattern="^back_to_rate_menu$"),
                CallbackQueryHandler(handle_search_by_name, pattern="^search_by_name$"),
                CallbackQueryHandler(show_review_search, pattern="^back_to_search$"),
                CallbackQueryHandler(show_review_filter_menu, pattern=r"^review_filter_menu$"),
                CallbackQueryHandler(show_review_sort_menu, pattern=r"^review_sort_menu$"),
                CallbackQueryHandler(show_review_details, pattern=r"^view_review_\d+$"),
                CallbackQueryHandler(previous_review_page, pattern=r"^prev_review_page$"),
                CallbackQueryHandler(next_review_page, pattern=r"^next_review_page$")
            ],
            POST_REVIEW: [
                CallbackQueryHandler(show_my_reviews, pattern="^post_review_my_reviews$"),
                CallbackQueryHandler(show_rate_options, pattern="^post_review_main_menu$"),
                CallbackQueryHandler(show_rate_options, pattern="^post_review_back$")
            ],
            ADMIN_RATINGS_MENU: [
                CallbackQueryHandler(show_all_reviews, pattern="^admin_view_all_reviews$"),
                CallbackQueryHandler(prompt_review_search, pattern="^admin_search_reviews$"),
                CallbackQueryHandler(review_statistics_dashboard, pattern="^admin_review_stats$"),
                CallbackQueryHandler(prompt_delete_review, pattern="^admin_delete_review$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^admin_back_to_ratings$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^cancel_delete_review$"),
                CallbackQueryHandler(back_to_user_interactions, pattern="^admin_back_to_user_interactions$"),
            ],
            ADMIN_REVIEW_LIST: [
                CallbackQueryHandler(prev_page_reviews, pattern="^admin_prev_page$"),
                CallbackQueryHandler(next_page_reviews, pattern="^admin_next_page$"),
                CallbackQueryHandler(start_delete_review_process, pattern=r"^admin_delete_review_\d+$"),
                CallbackQueryHandler(edit_review_details, pattern="^admin_edit_review_\\d+$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^admin_back_to_ratings$"),
            ],
            ADMIN_REVIEW_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_review_search_query),
                CallbackQueryHandler(display_admin_review_search_results, pattern="^prev_page$"),
                CallbackQueryHandler(display_admin_review_search_results, pattern="^next_page$"),
                CallbackQueryHandler(handle_admin_review_selection, pattern=r"^view_review_\d+$"),  # ← NEW!
                CallbackQueryHandler(apply_filter_reviews, pattern="^filter_reviews_(employer|job_seeker)$"),
                CallbackQueryHandler(apply_sort_reviews, pattern="^sort_reviews_(recent|top_rated)$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^admin_back_to_ratings$"),
            ],
            ADMIN_REVIEW_DETAILS: [
                CallbackQueryHandler(prepare_delete_review_admin, pattern=r"^confirm_deleted_review_\d+$"),
                CallbackQueryHandler(ratings_dashboard, pattern="^admin_back_to_ratings$"),
                CallbackQueryHandler(delete_review_confirmation, pattern=r"^delete_review_\d+$"),
            ],
            ADMIN_REVIEW_STATISTICS: [
                CallbackQueryHandler(back_to_ratings_menu, pattern="^admin_back_to_ratings$"),
            ],
            ADMIN_FLAGGED_REVIEWS: [
                # CallbackQueryHandler(view_flagged_review_details, pattern="^view_flagged_review_\\d+$"),
                # CallbackQueryHandler(resolve_flagged_review, pattern="^resolve_flagged_review_\\d+$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^admin_back_to_ratings$"),
            ],
            ADMIN_DELETE_REVIEW: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_review_id_for_deletion),
                CallbackQueryHandler(confirm_delete_review_admin, pattern="^confirm_delete$"),
                CallbackQueryHandler(confirm_deleted_review_admin, pattern="^confirm_deleted$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^cancel_delete_review$"),
                CallbackQueryHandler(back_to_ratings_menu, pattern="^admin_back_to_ratings$"),
            ],
            # ===== NEW Report STATES =====
            # SELECT_REPORT_ENTITY: [
            #     CallbackQueryHandler(select_report_entity,
            #                          pattern=r"^report_(job_seeker|employer|application|vacancy)$"),
            #     CallbackQueryHandler(back_to_main_menu_report, pattern=r"^back_to_main_menu$")
            # ],
            # SEARCH_REPORT_ENTITY: [
            #     MessageHandler(filters.TEXT, search_report_entity)
            # ],
            # SELECT_REPORT_RESULT: [
            #     CallbackQueryHandler(handle_pagination_report,
            #                          pattern=r"^(next|prev)_report_(job_seeker|employer|application|vacancy)_\d+$"),
            #     CallbackQueryHandler(select_report_result,
            #                          pattern=r"^report_(seeker|employer|application|vacancy)_\d+$"),
            #     CallbackQueryHandler(back_to_main_menu_report, pattern=r"^back_to_report_menu$")
            # ],
            # CONFIRM_REPORT_DETAILS: [
            #     CallbackQueryHandler(submit_report, pattern=r"^reason_(violation|spam|other|cancel)$"),
            #     MessageHandler(filters.TEXT, provide_additional_info)
            # ],
            # SUBMIT_REPORT: [
            #     MessageHandler(filters.TEXT, finalize_report)
            # ],
            SELECT_REPORT_ENTITY: [
                CallbackQueryHandler(
                    select_report_entity,
                    pattern=r"^report_(job_seeker|employer|application|vacancy)$"
                ),
                CallbackQueryHandler(
                    back_to_main_menu_report,
                    pattern=r"^back_to_main_menu$"
                )
            ],
            SEARCH_REPORT_ENTITY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    search_report_entity
                )
            ],
            SELECT_REPORT_RESULT: [
                CallbackQueryHandler(
                    handle_pagination_report,
                    pattern=r"^(next|prev)_report_(job_seeker|employer|application|vacancy)_\d+$"
                ),
                CallbackQueryHandler(
                    select_report_result,
                    pattern=r"^report_(seeker|employer|application|vacancy)_\d+$"
                ),
                CallbackQueryHandler(
                    back_to_main_menu_report,
                    pattern=r"^back_to_report_menu$"
                ),
                CallbackQueryHandler(  # New handler for "New Search"
                    select_report_entity,
                    pattern=r"^report_(job_seeker|employer|application|vacancy)$"
                )
            ],
            CONFIRM_REPORT_DETAILS: [
                CallbackQueryHandler(
                    submit_report,
                    pattern=r"^reason_(violation|spam|inappropriate|fraud|other|cancel)$"
                )
            ],
            SUBMIT_REPORT: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    provide_additional_info
                )
            ],
            VIOLATION_REPORTS_DASHBOARD: [
                CallbackQueryHandler(
                    view_all_reports,
                    pattern=r"^view_all_reports$"
                ),
                CallbackQueryHandler(
                    export_reports_to_excel,
                    pattern=r"^export_reports_excel$"
                ),
                CallbackQueryHandler(
                    view_reported_users,
                    pattern=r"^view_reported_users$"
                ),
                CallbackQueryHandler(
                    view_reported_employers,
                    pattern=r"^view_reported_employers$"
                ),
                CallbackQueryHandler(
                    view_reported_vacancies,
                    pattern=r"^view_reported_vacancies$"
                ),
                CallbackQueryHandler(
                    view_reported_applications,
                    pattern=r"^view_reported_applications$"
                ),
                CallbackQueryHandler(
                    back_to_admin_menu,
                    pattern=r"^back_to_admin_menu$"
                )
            ],
            VIEW_ALL_REPORTS: [
                CallbackQueryHandler(
                    back_to_violation_dashboard,
                    pattern=r"^back_to_violation_dashboard$"
                )
            ],
            VIEW_REPORTED_USERS: [
                CallbackQueryHandler(
                    back_to_violation_dashboard,
                    pattern=r"^back_to_violation_dashboard$"
                )
            ],
            VIEW_REPORTED_EMPLOYERS: [
                CallbackQueryHandler(
                    back_to_violation_dashboard,
                    pattern=r"^back_to_violation_dashboard$"
                )
            ],
            VIEW_REPORTED_VACANCIES: [
                CallbackQueryHandler(
                    back_to_violation_dashboard,
                    pattern=r"^back_to_violation_dashboard$"
                )
            ],
            VIEW_REPORTED_APPLICATIONS: [
                CallbackQueryHandler(
                    back_to_violation_dashboard,
                    pattern=r"^back_to_violation_dashboard$"
                )
            ],
            # ===== NEW CONTACT ADMIN STATES =====
            CONTACT_CATEGORY: [
                CallbackQueryHandler(handle_contact_category, pattern=r"^contact_category_\d+$"),
                CallbackQueryHandler(show_help, pattern=r"^contact_back$"),
            ],
            CONTACT_PRIORITY: [
                CallbackQueryHandler(handle_contact_priority, pattern=r"^contact_priority_[123]$"),
                CallbackQueryHandler(show_contact_options, pattern=r"^contact_back_to_categories$"),
            ],
            CONTACT_MESSAGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_contact_message),
                CommandHandler("cancel", cancel_contact_request),
            ],

            # Admin reply state
            ADMIN_REPLY_STATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_reply),
                CommandHandler("cancel", cancel_admin_reply),
            ],


            #Admin STates
            ADMIN_LOGIN: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, check_admin_credentials)
            ],

            ADMIN_MAIN_MENU: [
                MessageHandler(
                    filters.Text(ADMIN_MENU_OPTIONS),
                    handle_admin_menu_choice
                ),
                MessageHandler(
                    filters.Regex(r"^(⬅️ Previous|Next ➡️|🔄 Refresh)$"),
                    handle_pagination_job
                ),
                MessageHandler(
                    filters.Text(["🔙 Back"]),  # Handle the Back button
                    show_admin_menu
                ),
                # Other handlers remain the same
                CallbackQueryHandler(handle_admin_job_approval, pattern="^approve_|^reject_"),
                CallbackQueryHandler(handle_job_preview, pattern=r"^preview_"),
                CallbackQueryHandler(handle_docs_view, pattern=r"^docs_"),
                CallbackQueryHandler(handle_contact_employer, pattern=r"^contact_"),
                CallbackQueryHandler(go_back_to_pending_jobs, pattern=r"^back_to_pending_jobs$"),
            ],
            USER_INTERACTIONS_MENU: [
                MessageHandler(
                    filters.Text(USER_INTERACTION_OPTIONS),
                    handle_user_interactions_choice
                ),
                MessageHandler(
                    filters.Text(["🔙 Back"]),
                    show_admin_menu
                ),

                CallbackQueryHandler(handle_contact_employer, pattern=r"^contact_"),
            ],
            SHARE_JOBS_NAVIGATION: [
                MessageHandler(
                    filters.TEXT & (
                        filters.Regex('^(⬅️ Back|➡️ Forward|🏠 Main Menu)$')
                    ),
                    handle_share_jobs_navigation
                )
            ],
            #contact feature
            CONTACT_MANAGEMENT: [
                CallbackQueryHandler(show_contact_inbox, pattern=r"^contact_inbox$"),
                CallbackQueryHandler(show_contact_outbox, pattern=r"^contact_outbox$"),
                CallbackQueryHandler(show_contact_pending, pattern=r"^contact_pending$"),
                CallbackQueryHandler(show_contact_answered, pattern=r"^contact_answered$"),
                CallbackQueryHandler(show_contact_stats, pattern=r"^contact_stats$"),
                CallbackQueryHandler(show_admin_menu, pattern=r"^contact_back_to_menu$"),
            ],

            CONTACT_CONFIRM_DELETE: [
                CallbackQueryHandler(confirm_delete_message, pattern=r"^contact_confirm_delete_\d+$"),
                CallbackQueryHandler(view_contact_message, pattern=r"^contact_view_\d+$"),
            ],

            CONTACT_INBOX: [
                CallbackQueryHandler(handle_pagination_contact, pattern=r"^contact_inbox_\d+$"),
                CallbackQueryHandler(show_contact_management_dashboard, pattern=r"^contact_back_to_dashboard$"),
                CallbackQueryHandler(view_contact_message, pattern=r"^contact_view_\d+$"),

            ],

            CONTACT_VIEW_MESSAGE: [
                CallbackQueryHandler(start_admin_reply, pattern=r"^contact_reply_\d+$"),
                CallbackQueryHandler(close_ticket, pattern=r"^contact_close_\d+$"),
                CallbackQueryHandler(delete_message, pattern=r"^contact_delete_\d+$"),
                CallbackQueryHandler(follow_up_message, pattern=r"^contact_followup_\d+$"),
                CallbackQueryHandler(show_contact_inbox, pattern=r"^contact_back_to_inbox$"),
                CallbackQueryHandler(show_contact_outbox, pattern=r"^contact_back_to_outbox$"),
                CallbackQueryHandler(show_contact_pending, pattern=r"^contact_back_to_pending$"),
                CallbackQueryHandler(show_contact_answered, pattern=r"^contact_back_to_answered$"),
                CallbackQueryHandler(show_contact_management_dashboard, pattern=r"^contact_back_to_dashboard$")
            ],
            CONTACT_OUTBOX: [
                CallbackQueryHandler(handle_pagination_contact, pattern=r"^contact_page_\d+$"),
                CallbackQueryHandler(show_contact_management_dashboard, pattern=r"^contact_back_to_dashboard$"),
                CallbackQueryHandler(view_contact_message, pattern=r"^contact_view_\d+$"),
            ],
            CONTACT_PENDING: [
                CallbackQueryHandler(handle_pagination_contact, pattern=r"^contact_pending_\d+$"),
                CallbackQueryHandler(show_contact_management_dashboard, pattern=r"^contact_back_to_dashboard$"),
                CallbackQueryHandler(view_contact_message, pattern=r"^contact_view_\d+$"),
            ],

            CONTACT_ANSWERED: [
                CallbackQueryHandler(handle_pagination_contact, pattern=r"^contact_answered_\d+$"),
                CallbackQueryHandler(show_contact_management_dashboard, pattern=r"^contact_back_to_dashboard$"),
                CallbackQueryHandler(view_contact_message, pattern=r"^contact_view_\d+$"),
            ],

            CONTACT_STATS: [
                CallbackQueryHandler(show_contact_management_dashboard, pattern="^contact_back_to_dashboard$"),
            ],
            REJECT_JOB_REASON: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_rejection_reason)
            ],
            SELECT_JOB_POSTS_TO_SHARE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_share_job_posts),
            ],

            BROADCAST_TYPE: [
                CallbackQueryHandler(select_broadcast_type, pattern="^(job_seekers|employers|all|stats_preview|cancel)$")
            ],
            BROADCAST_MESSAGE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, get_broadcast_message)
            ],
            CONFIRM_BROADCAST: [
                CallbackQueryHandler(confirm_broadcast, pattern="^(confirm|cancel)$")
            ],
            POST_JOB_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_job_title)],
            POST_EMPLOYMENT_TYPE: [CallbackQueryHandler(post_employment_type)],
            POST_GENDER: [CallbackQueryHandler(post_gender)],
            POST_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_quantity)],
            POST_LEVEL: [CallbackQueryHandler(post_level)],
            POST_DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_description)],
            POST_QUALIFICATION: [CallbackQueryHandler(post_qualification)],
            POST_SKILLS: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_skills)],
            POST_SALARY: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_salary)],
            POST_BENEFITS: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_benefits)],
            POST_DEADLINE: [MessageHandler(filters.TEXT & ~filters.COMMAND, post_deadline)],
            JOB_PREVIEW: [MessageHandler(filters.TEXT & ~filters.COMMAND, job_preview)],
            CONFIRM_POST: [CallbackQueryHandler(confirm_post)],
            # DISPLAY_VACANCIES: [
            #     MessageHandler(filters.TEXT & ~filters.COMMAND, select_vacancy)
            # ],
            VACANCY_DISPLAY_OPTION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_vacancy_display_option)
            ],
            FILTER_SELECTION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_filter_selection)
            ],
            EMPLOYMENT_FILTER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_employment_filter)
            ],
            LEVEL_FILTER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_level_filter)
            ],
            QUALIFICATION_FILTER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_qualification_filter)
            ],
            GENDER_FILTER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_gender_filter)
            ],
            DISPLAY_VACANCIES: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, display_all_vacancies)
            ],
            SELECT_VACANCY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, select_vacancy)
            ],
            CONFIRM_SELECTION: [
                CallbackQueryHandler(confirm_selection, pattern="^(confirm|cancel|back_to_job_detail)$"),
            ],
            WRITE_COVER_LETTER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, write_cover_letter),
                CallbackQueryHandler(confirm_selection, pattern="^(back_to_job_detail)$"),
            ],
            CONFIRM_SUBMISSION: [
                CallbackQueryHandler(handle_confirmation, pattern="^confirm$|^cancel$")
            ],

            MANAGE_USERS: [
                CallbackQueryHandler(ban_job_seekers, pattern="^ban_job_seekers$"),
                CallbackQueryHandler(ban_employers, pattern="^ban_employers$"),
                CallbackQueryHandler(unban_users_menu, pattern="^unban_users_menu$"),                CallbackQueryHandler(view_banned_users, pattern="^view_banned_users$"),
                CallbackQueryHandler(remove_job_seekers, pattern="^remove_job_seekers$"),
                CallbackQueryHandler(remove_employers, pattern="^remove_employers$"),
                CallbackQueryHandler(remove_applications, pattern="^remove_applications$"),
                CallbackQueryHandler(export_job_seekers, pattern="^export_job_seekers$"),
                CallbackQueryHandler(export_employers, pattern="^export_employers$"),
                CallbackQueryHandler(export_applications, pattern="^export_applications$"),
                CallbackQueryHandler(back_to_admin_menu, pattern="^back_to_admin_menu$")
            ],
            UNBAN_USERS_MENU: [
                CallbackQueryHandler(unban_by_selection, pattern="^unban_by_selection$"),
                CallbackQueryHandler(confirm_unban_all, pattern="^unban_all_confirmation$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            UNBAN_SELECTION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_unban_selection),
                CallbackQueryHandler(unban_users_menu, pattern="^unban_users_menu$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            UNBAN_ALL_CONFIRMATION: [
                CallbackQueryHandler(execute_unban_all, pattern="^execute_unban_all$"),
                CallbackQueryHandler(unban_users_menu, pattern="^unban_users_menu$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            BAN_JOB_SEEKERS: [MessageHandler(filters.TEXT, handle_job_seeker_ban_search),
                            CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            BAN_EMPLOYERS: [MessageHandler(filters.TEXT, handle_employer_ban_search)],
            BAN_EMPLOYERS_PAGINATED: [
                CallbackQueryHandler(handle_pagination_ban, pattern=r"^(next|prev)_employer_(\d+)$"),
                CallbackQueryHandler(confirm_ban_employer, pattern=r"^ban_employer_\d+$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            SEARCH_JOB_SEEKERS_FOR_BAN: [MessageHandler(filters.TEXT, handle_job_seeker_ban_search),
                                         CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            BAN_JOB_SEEKERS_PAGINATED: [
                CallbackQueryHandler(handle_pagination_ban, pattern=r"^(next|prev)_job_seeker_(\d+)$"),
                CallbackQueryHandler(confirm_ban_job_seeker, pattern=r"^ban_job_seeker_\d+$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            REASON_FOR_BAN_JOB_SEEKER: [MessageHandler(filters.TEXT, apply_ban_job_seeker),
                                        CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            SEARCH_EMPLOYERS_FOR_BAN: [MessageHandler(filters.TEXT, handle_employer_ban_search),
                                       CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            REASON_FOR_BAN_EMPLOYER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, apply_ban_employer),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            UNBAN_USERS: [
                CallbackQueryHandler(handle_unban, pattern=r"^unban_(user|employer)_\d+$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            VIEW_BANNED_USERS: [CallbackQueryHandler(handle_appeal_decision, pattern="^review_appeals$"),
                                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            REASON_FOR_BAN: [MessageHandler(filters.TEXT, apply_ban)],

            # Appeal system states
            BANNED_STATE: [
                CallbackQueryHandler(start_ban_appeal, pattern="^appeal_start$"),
            ],
            APPEAL_START: [
                CallbackQueryHandler(start_ban_appeal, pattern="^appeal_start$"),
            ],
            APPEAL_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_appeal_input),
                # CommandHandler("cancel", cancel_appeal),
            ],
            APPEAL_REVIEW: [
                CallbackQueryHandler(handle_appeal_decision, pattern=r"^lift_ban_\d+$"),
                CallbackQueryHandler(handle_appeal_decision, pattern=r"^uphold_ban_\d+$"),
                CallbackQueryHandler(handle_appeal_decision, pattern=r"^request_info_\d+$"),
            ],

            REMOVE_JOB_SEEKERS: [CallbackQueryHandler(confirm_removal)],
            REMOVE_EMPLOYERS: [CallbackQueryHandler(confirm_removal)],
            REMOVE_APPLICATIONS: [CallbackQueryHandler(confirm_removal)],
            CONFIRM_REMOVAL: [CallbackQueryHandler(perform_removal)],
            CLEAR_CONFIRMATION: [
                CallbackQueryHandler(perform_clear, pattern="^confirm_clear$"),
                CallbackQueryHandler(handle_cancel_clear, pattern="^back_to_database_menu$"),
            ],
            TABLE_CLEANUP: [
                CallbackQueryHandler(confirm_table_deletion, pattern="^delete_table_"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$")
            ],
            CONFIRM_TABLE_DELETION: [
                CallbackQueryHandler(perform_table_deletion, pattern="^confirm_delete$"),
                CallbackQueryHandler(table_cleanup_menu, pattern="^back_to_table_cleanup$")
            ],

            DATABASE_MANAGEMENT: [
                CallbackQueryHandler(manage_users, pattern="^manage_users$"),
                CallbackQueryHandler(manage_jobs, pattern="^manage_jobs$"),
                CallbackQueryHandler(ad_manage_vacancies, pattern="^ad_manage_vacancies$"),  # Fixed pattern
                CallbackQueryHandler(manage_applications, pattern="^manage_applications$"),
                CallbackQueryHandler(export_data, pattern="^export_data$"),
                CallbackQueryHandler(clear_data, pattern="^clear_data$"),
                CallbackQueryHandler(table_cleanup_menu, pattern="^table_cleanup$"),
                CallbackQueryHandler(db_stats, pattern="^db_stats$"),
                CallbackQueryHandler(view_system_errors, pattern="^view_system_errors$"),
                CallbackQueryHandler(refresh_db_stats, pattern="^refresh_stats$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$"),
                CallbackQueryHandler(back_to_admin_menu, pattern="^back_to_admin_menu$")
            ],
            DB_STATS_VIEW: [
                CallbackQueryHandler(db_stats, pattern="^refresh_stats$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$"),
                CallbackQueryHandler(back_to_admin_menu, pattern="^back_to_admin_menu$")
            ],
            MANAGE_JOBS: [
                CallbackQueryHandler(list_jobs, pattern="^list_jobs$"),
                CallbackQueryHandler(remove_jobs, pattern="^remove_jobs$"),
                CallbackQueryHandler(export_jobs, pattern="^export_jobs$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$"),
                CallbackQueryHandler(back_to_manage_jobs, pattern="^back_to_manage_jobs$")
            ],
            LIST_JOBS: [
                CallbackQueryHandler(view_job_detail, pattern=r"^job_detail_\d+$"),
                CallbackQueryHandler(back_to_manage_jobs, pattern="^back_to_manage_jobs$")
            ],
            CONFIRM_JOB_REMOVAL: [
                CallbackQueryHandler(execute_job_removal, pattern=r"^execute_remove_job_\d+$"),
                CallbackQueryHandler(back_to_manage_jobs, pattern="^back_to_manage_jobs$")
            ],
            VIEW_ERRORS: [
                CallbackQueryHandler(handle_error_detail, pattern=r"^error_detail_"),
                CallbackQueryHandler(show_database_menu, pattern="^back_to_database_menu$")
            ],
            ERROR_DETAIL: [
                CallbackQueryHandler(show_full_traceback, pattern=r"^show_traceback_"),
                CallbackQueryHandler(show_update_data, pattern=r"^show_update_"),
                CallbackQueryHandler(resolve_error, pattern=r"^resolve_error_"),
                CallbackQueryHandler(view_system_errors, pattern="^view_system_errors$"),
                CallbackQueryHandler(handle_error_detail, pattern=r"^back_to_detail_")
            ],
            SEARCH_JOBS: [
                MessageHandler(filters.TEXT, handle_job_search)
            ],
            LIST_JOBS_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_job_\d+$"),
                CallbackQueryHandler(view_job_detail, pattern=r"^job_detail_\d+$"),
                CallbackQueryHandler(back_to_manage_jobs, pattern="^back_to_manage_jobs$")
            ],
            REMOVE_JOBS_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_job_remove_\d+$"),
                CallbackQueryHandler(confirm_job_removal, pattern=r"^remove_job_\d+$"),
                CallbackQueryHandler(back_to_manage_jobs, pattern="^back_to_manage_jobs$")
            ],
            JOB_DETAIL_VIEW: [
                CallbackQueryHandler(back_to_manage_jobs, pattern="^back_to_manage_jobs$")
            ],
            AD_MANAGE_VACANCIES: [
                CallbackQueryHandler(list_vacancies, pattern="^list_vacancies$"),
                CallbackQueryHandler(remove_vacancies, pattern="^remove_vacancies$"),
                CallbackQueryHandler(export_vacancies, pattern="^export_vacancies$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$"),
                CallbackQueryHandler(back_to_manage_vacancies, pattern="^back_to_manage_vacancies$")
            ],
            SEARCH_VACANCIES: [
                MessageHandler(filters.TEXT, handle_vacancy_search)
            ],
            LIST_VACANCIES_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_vacancy_\d+$"),
                CallbackQueryHandler(back_to_manage_vacancies, pattern="^back_to_manage_vacancies$")
            ],
            REMOVE_VACANCIES_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_vacancy_remove_\d+$"),
                CallbackQueryHandler(confirm_removal_vacancy, pattern=r"^remove_vacancy_\d+$"),
                CallbackQueryHandler(back_to_manage_vacancies, pattern="^back_to_manage_vacancies$")
            ],
            CONFIRM_REMOVE_VACANCY: [
                CallbackQueryHandler(execute_remove_vacancy, pattern=r"^confirm_remove_vacancy_\d+$"),
                CallbackQueryHandler(back_to_manage_vacancies, pattern="^back_to_manage_vacancies$")
            ],
            SEARCH_JOB_SEEKERS: [MessageHandler(filters.TEXT, handle_job_seeker_search),
                                 CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            REMOVE_JOB_SEEKERS_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_job_seeker_\d+$"),
                CallbackQueryHandler(confirm_removal, pattern=r"^remove_seeker_\d+$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            SEARCH_EMPLOYERS: [MessageHandler(filters.TEXT, handle_employer_search),
                               CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            REMOVE_EMPLOYERS_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_employer_\d+$"),
                CallbackQueryHandler(confirm_removal, pattern=r"^remove_employer_\d+$"),
                CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),
            ],
            SEARCH_APPLICATIONS: [MessageHandler(filters.TEXT, handle_application_search),
                                  CallbackQueryHandler(back_to_manage_users, pattern="^back_to_manage_users$"),],
            REMOVE_APPLICATIONS_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_application_\d+$"),
                CallbackQueryHandler(confirm_application_removal, pattern=r"^remove_application_\d+$"),
                CallbackQueryHandler(back_to_manage_applications, pattern="^back_to_manage_applications$")
            ],
            APPLICATION_DETAIL_VIEW: [
                CallbackQueryHandler(back_to_manage_applications, pattern="^back_to_manage_applications$"),
                CallbackQueryHandler(confirm_application_removal, pattern=r"^remove_application_\d+$")
            ],
            CONFIRM_APPLICATION_REMOVAL: [
                CallbackQueryHandler(execute_application_removal, pattern=r"^confirm_remove_application_\d+$"),
                CallbackQueryHandler(back_to_manage_applications, pattern="^back_to_manage_applications$")
            ],
            MANAGE_APPLICATIONS: [
                CallbackQueryHandler(list_applications, pattern="^list_applications$"),
                CallbackQueryHandler(remove_applications, pattern="^remove_applications$"),
                CallbackQueryHandler(export_applications, pattern="^export_applications$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$"),
                CallbackQueryHandler(back_to_manage_applications, pattern="^back_to_manage_applications$")
            ],
            LIST_APPLICATIONS_PAGINATED: [
                CallbackQueryHandler(handle_pagination, pattern=r"^(next|prev)_application_list_\d+$"),
                CallbackQueryHandler(view_application_detail, pattern=r"^application_detail_\d+$"),
                CallbackQueryHandler(back_to_manage_applications, pattern="^back_to_manage_applications$")
            ],
            EXPORT_DATA: [
                CallbackQueryHandler(export_job_seekers, pattern="^export_job_seekers$"),
                CallbackQueryHandler(export_employers, pattern="^export_employers$"),
                CallbackQueryHandler(export_applications, pattern="^export_applications$"),
                CallbackQueryHandler(export_all_data, pattern="^export_all_data$"),
                CallbackQueryHandler(back_to_database_menu, pattern="^back_to_database_menu$")
            ],
            #search vaccancy
            SEARCH_OPTIONS: [
                CallbackQueryHandler(handle_search_options)
            ],
            KEYWORD_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_keyword_search)
            ],
            ADVANCED_FILTERS: [
                CallbackQueryHandler(handle_advanced_filters)
            ],
            FILTER_JOB_TYPE: [
                CallbackQueryHandler(handle_job_type_filter)
            ],
            FILTER_SALARY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_salary_filter)
            ],
            FILTER_EXPERIENCE: [
                CallbackQueryHandler(handle_experience_filter)
            ],
            SEARCH_RESULTS: [
                CallbackQueryHandler(handle_search_results,
                                     pattern="^(view_all_results|refine_search|new_search|back_to_main)$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, select_search_result)
            ],
            SELECT_SEARCH_RESULT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, select_search_result),
                CallbackQueryHandler(handle_search_results,
                                     pattern="^(view_all_results|refine_search|new_search|back_to_main)$"),
            ],
            VIEW_JOB_DETAILS: [
                CallbackQueryHandler(handle_job_details)
            ],
            VIEWING_APPLICATIONS: [
                CallbackQueryHandler(handle_application_actions)
            ],
            APPLICATION_DETAILS: [
                CallbackQueryHandler(handle_application_actions)
            ],
            EXPORTING_APPLICATIONS: [
                CallbackQueryHandler(handle_export_request)
            ],
            #system configuration
            SYSTEM_CONFIGURATIONS_MENU: [
                MessageHandler(filters.Text(), handle_system_configurations_choice),
                CallbackQueryHandler(handle_back_system, pattern="^back_to_system_config$"),
            ],
            DATABASE_STORAGE_OVERVIEW: [
                CallbackQueryHandler(handle_back_system, pattern="^back_to_system_config$"),
            ],
            TABLE_SIZE_ANALYSIS: [MessageHandler(filters.Text(), show_system_configurations_menu)],
            OPTIMIZE_DATABASE: [MessageHandler(filters.Text(), show_system_configurations_menu)],
            VACUUM_DATABASE: [MessageHandler(filters.Text(), show_system_configurations_menu)],
            QUERY_PERFORMANCE_INSIGHTS: [MessageHandler(filters.Text(), show_system_configurations_menu)],
            ERROR_LOGS: [MessageHandler(filters.Text(), show_system_configurations_menu)],

        },

        #fallbacks=[CommandHandler('cancel', lambda update, context: update.message.reply_text("Canceled."))]
        fallbacks=[
            CommandHandler('start', start),  # Allow /start to reset the conversation
        ],
        allow_reentry=True  # Optional: Allows re-entry into the conversation

    )

    # Add the conversation handler to the application
    application.add_handler(conv_handler)
    application.add_error_handler(advanced_error_handler)
    application.add_handler(
        CallbackQueryHandler(
            handle_appeal_decision,
            pattern=r"^(lift_ban|uphold_ban|request_info)_\d+$"
        )
    )

    # Add the job with coalesce enabled and a longer interval
    application.job_queue.run_repeating(
        handle_notifications,
        interval=30,  # Increased interval
        first=0
    )

    # Start the notification handler
    application.job_queue.run_repeating(handle_notifications, interval=30, first=0)

    # Start polling and run the bot
    application.run_polling()


if __name__ == "__main__":
    main()


